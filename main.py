from fastapi import FastAPI, HTTPException, Header, Query, Depends, Path, Request, Response
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, validator
from typing import Optional, List, Dict, Union, Literal, Callable, Any
import json
import pathlib
from datetime import datetime, date, timezone, timedelta
from zoneinfo import ZoneInfo
from contextlib import contextmanager
import psycopg2
from psycopg2.extras import RealDictCursor
import calendar
from psycopg2 import pool
import os
import re
import logging
import secrets
import traceback
import math
import uuid
import io
import csv
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def to_decimal(value) -> Decimal:
    """Convert a value to Decimal safely, rounding to 4 decimal places."""
    return Decimal(str(value)).quantize(Decimal('0.0001'), rounding=ROUND_HALF_UP)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Custom JSON encoder to handle Decimal from NUMERIC columns
class DecimalSafeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, uuid.UUID):
            return str(obj)
        return super().default(obj)

class DecimalSafeJSONResponse(JSONResponse):
    def render(self, content) -> bytes:
        return json.dumps(
            content,
            ensure_ascii=False,
            allow_nan=False,
            indent=None,
            separators=(",", ":"),
            cls=DecimalSafeEncoder,
        ).encode("utf-8")

app = FastAPI(title="Factory Ledger System", version="3.1.1", default_response_class=DecimalSafeJSONResponse)
from fastapi.middleware.cors import CORSMiddleware

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ═══════════════════════════════════════════════════════════════
# UNIFORM WRITE-RESPONSE CONTRACT (additive envelope)
# Every JSON response to a mutating request carries `success: bool`;
# failures additionally carry `error_detail: {code, message}`. Existing
# keys are never removed, renamed, or overwritten — endpoints that already
# set `success` (the 9 core inventory ops) pass through untouched.
# Registered AFTER CORSMiddleware, so it is the OUTERMOST middleware: it
# wraps the output of every route, exception handler, and any middleware
# registered earlier (CORS headers are already on the response it rebuilds).
# ═══════════════════════════════════════════════════════════════

_WRITE_ENVELOPE_METHODS = {"POST", "PUT", "PATCH", "DELETE"}


def _structured_error(payload: dict, status_code: int) -> dict:
    """Derive {code, message} from whatever error shape the response used:
    dict detail (error_code/message), string detail, validation-error list,
    or the legacy {"error": str} body."""
    detail = payload.get("detail")
    if isinstance(detail, dict):
        return {
            "code": detail.get("error_code") or f"HTTP_{status_code}",
            "message": detail.get("message") or str(detail),
        }
    if isinstance(detail, list):  # FastAPI request-validation errors (422)
        msgs = "; ".join(
            str(item.get("msg", item)) if isinstance(item, dict) else str(item)
            for item in detail
        )
        return {"code": "VALIDATION_ERROR", "message": msgs or "Validation error"}
    if isinstance(detail, str):
        return {"code": f"HTTP_{status_code}", "message": detail}
    if isinstance(payload.get("error"), str):
        return {"code": f"HTTP_{status_code}", "message": payload["error"]}
    return {"code": f"HTTP_{status_code}", "message": "Request failed"}


@app.middleware("http")
async def write_response_envelope(request, call_next):
    response = await call_next(request)
    if request.method not in _WRITE_ENVELOPE_METHODS:
        return response
    if "application/json" not in (response.headers.get("content-type") or ""):
        return response

    body = b""
    async for chunk in response.body_iterator:
        body += chunk

    headers = dict(response.headers)
    headers.pop("content-length", None)
    headers.pop("content-type", None)

    try:
        payload = json.loads(body)
    except Exception:
        payload = None
    if not isinstance(payload, dict):
        # Non-object JSON (lists, scalars) — pass through unchanged.
        return Response(
            content=body,
            status_code=response.status_code,
            headers=headers,
            media_type="application/json",
        )

    if response.status_code < 400:
        payload.setdefault("success", True)
    else:
        payload.setdefault("success", False)
        payload.setdefault("error_detail", _structured_error(payload, response.status_code))

    return Response(
        content=json.dumps(
            payload, ensure_ascii=False, allow_nan=False,
            separators=(",", ":"), cls=DecimalSafeEncoder,
        ).encode("utf-8"),
        status_code=response.status_code,
        headers=headers,
        media_type="application/json",
    )


# ═══════════════════════════════════════════════════════════════
# GLOBAL EXCEPTION HANDLER — uniform JSON envelope + readonly tripwire
# Fires on exceptions that escape a route. Routes whose per-route
# `except Exception` would otherwise swallow a read-only error have a
# `if _is_readonly_error(e): raise` line that lets the error bubble here
# so we can run the PR #6 probe and leave a structured receipt.
#
# Registration placement (Starlette): a handler keyed on bare `Exception`
# is hoisted into ServerErrorMiddleware — the OUTERMOST layer, beyond
# write_response_envelope — so its responses skip the envelope. Handlers
# keyed on a specific class run in ExceptionMiddleware, the INNERMOST
# layer, whose output the envelope post-processes. Readonly failures are
# psycopg2 errors, so the psycopg2.Error registration below is the
# production path (503 + diagnostics, envelope adds error_detail); the
# bare-Exception registration is the safety net for non-psycopg2 escapes
# and adds error_detail itself because the envelope never sees its output.
# ═══════════════════════════════════════════════════════════════

async def _exception_receipt_response(request: Request, exc: Exception) -> JSONResponse:
    if _is_readonly_error(exc):
        diagnostics = _capture_readonly_diagnostics()
        logger.error(
            "READONLY_TRIPWIRE: "
            + json.dumps(
                {"path": request.url.path, "method": request.method,
                 "error": str(exc), "diagnostics": diagnostics},
                default=str,
            )
        )
        return JSONResponse(
            status_code=503,
            content={
                "success": False,
                "error_code": "READONLY_TRANSACTION",
                "error": str(exc),
                "diagnostics": diagnostics,
                "retryable": True,
                "message": "Database is temporarily read-only (likely Supabase failover). Retry the same request in a few seconds.",
            },
        )

    logger.error(
        f"Unhandled exception on {request.method} {request.url.path}: {exc}\n"
        + traceback.format_exc()
    )
    return JSONResponse(
        status_code=500,
        content={"success": False, "error_code": "INTERNAL_SERVER_ERROR", "error": str(exc)},
    )


@app.exception_handler(psycopg2.Error)
async def db_exception_handler(request: Request, exc: Exception):
    # Runs in ExceptionMiddleware (innermost): write_response_envelope
    # post-processes this response, adding error_detail {code, message}.
    return await _exception_receipt_response(request, exc)


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    # Runs in ServerErrorMiddleware (outermost): the envelope never sees
    # this response, so replicate its error_detail contract here. Uses the
    # receipt's own error_code (vs the envelope's HTTP_<status>), which also
    # lets tests tell the two paths apart.
    response = await _exception_receipt_response(request, exc)
    payload = json.loads(bytes(response.body))
    payload.setdefault(
        "error_detail",
        {
            "code": payload.get("error_code") or f"HTTP_{response.status_code}",
            "message": payload.get("message") or payload.get("error") or "Request failed",
        },
    )
    return JSONResponse(status_code=response.status_code, content=payload)


DATABASE_URL = (os.getenv("DATABASE_URL") or "").strip()
API_KEY = (os.getenv("API_KEY") or "").strip()
# Second, SCOPED key for the browser dashboard. Accepted only on the routes in
# DASHBOARD_KEY_ALLOWLIST (see verify_api_key); everything else -> 403.
# Env-only (no literal fallback) — startup() raises if it is missing.
DASHBOARD_API_KEY = (os.getenv("DASHBOARD_API_KEY") or "").strip()

# Timezone configuration
PLANT_TIMEZONE = ZoneInfo("America/New_York")
TIMEZONE_LABEL = "ET"

# ═══════════════════════════════════════════════════════════════
# SKU PROTECTION — Private-Label Guardrails
# ═══════════════════════════════════════════════════════════════

MERGE_KEYWORDS = ["merge", "deprecat", "consolidat", "migrat"]

# Odoo codes for verified private-label finished goods + exclusive batches
PRIVATE_LABEL_ODOO_CODES = [
    '893',    # CQ Coconut Sweetened Flake 10 LB (Chef Quality)
    '1614',   # CQ Granola 10 LB (Chef Quality)
    '67470',  # Coconut Sweetened Fancy UNIPRO 10 LB
    '67473',  # Coconut Sweetened Medium UNIPRO 10 LB
    '67476',  # Coconut Sweetened Flake UNIPRO 10 LB
    '70002',  # Granola SS Original 12x10 OZ Case (Sunshine)
    '70003',  # Granola SS Chocolate Chip 12x10 OZ Case (Sunshine)
    '70010',  # Granola SS Original Low Carb 12x10 OZ Case (Sunshine)
    '70011',  # Granola SS Cranberry 12x10 OZ Case (Sunshine)
    '70056',  # Granola Setton Cocoa Crunch 25 LB
    '70070',  # Granola SS Chocolate Chip Low Carb 12x10 OZ Case (Sunshine)
    '70073',  # BS Granola – Peanut Butter Banana – 6x7 OZ Case (Blue Stripes)
    '70074',  # BS Granola – Dark Chocolate – 6x7 OZ Case (Blue Stripes)
    '70077',  # Granola Setton Cinnamon Spice Almond 25 LB
    '70078',  # Granola Setton Morning Latte Crunch 25 LB
    '70079',  # BS Almond Butter Granola – 6x7 OZ Case (Blue Stripes)
    '70080',  # BS Granola – Hazelnut Butter – 6x7 OZ Case (Blue Stripes)
    '70081',  # Granola Setton Good Ol 25 LB
    '70082',  # Granola Setton French Vanilla 25 LB
]


def check_private_label_merge(product_name: str, label_type: str, reason: str, quantity: float):
    """Check if an adjustment would violate private-label SKU protection.
    Returns a warning message string if blocked, or None if allowed."""
    if label_type != 'private_label':
        return None
    reason_lower = reason.lower()
    is_merge_reason = any(kw in reason_lower for kw in MERGE_KEYWORDS)
    if is_merge_reason and quantity < 0:
        return (
            f"BLOCKED: Cannot merge/deprecate a private-label SKU. "
            f"'{product_name}' is identity-protected. "
            f"If this is a physical repack, use reason 'repack' instead."
        )
    return None

# ═══════════════════════════════════════════════════════════════
# LOT BALANCE GUARD — Prevent Negative Inventory at Write-Time
# ═══════════════════════════════════════════════════════════════

# Tolerance for floating point dust (e.g., 9.99999999997 treated as 10.0)
BALANCE_EPSILON = 0.0001

# ── Void semantics: single source of truth for all balance math ──────────────
# Only lines whose parent transaction has status='posted' count toward any
# on-hand / balance / availability figure. Voided transactions are excluded
# everywhere; POST /void flips status and does NOT post reversal lines.
# Every balance query must read transaction lines through POSTED_LINES (or the
# lot_on_hand() helper) — never from the raw transaction_lines table.
POSTED_LINES = (
    "(SELECT tl.* FROM ledger_current_transaction_lines tl"
    " JOIN transactions _pt ON _pt.id = tl.transaction_id"
    " JOIN ledger_current_transactions _ct ON _ct.id = _pt.id"
    " WHERE _ct.effective_status = 'posted')"
)


def lot_on_hand(cur, lot_id: int) -> float:
    """Posted-only on-hand balance for a single lot, in lb."""
    cur.execute(
        f"SELECT COALESCE(SUM(quantity_lb), 0) as balance FROM {POSTED_LINES} tl WHERE lot_id = %s",
        (lot_id,)
    )
    return float(cur.fetchone()['balance'])


# ── FIFO lot order: single source of truth ───────────────────────────────────
# Every allocation path (ship single/multi-lot, order ship, make, pack) picks
# lots by this exact rule: posted-only balance per lot, only lots with a
# positive balance, oldest first by COALESCE(received_at, created_at). New
# read paths that need "the lots in allocation order" must call
# fifo_lot_balances() rather than re-inline the query. (l.id is only a
# deterministic tie-breaker for identical timestamps; the allocation paths
# leave ties to the planner.)
FIFO_LOT_ORDER_SQL = "COALESCE(l.received_at, l.created_at) ASC, l.id ASC"


def fifo_lot_balances(cur, product_id: int, include_empty: bool = False) -> list:
    """Lots for one product in FIFO allocation order with their posted-only
    balances (lb / product uom).

    Returns dicts: id, lot_code, received_at, created_at, lot_date (the FIFO
    key), entry_source, supplier_lot_code, status, available.
    include_empty=False (default) mirrors the allocation paths' HAVING > 0;
    include_empty=True also returns depleted / negative lots for audit views.
    """
    having = "" if include_empty else "HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0"
    cur.execute(
        f"""
        SELECT l.id, l.lot_code, l.received_at, l.created_at,
               COALESCE(l.received_at, l.created_at) AS lot_date,
               l.entry_source, l.supplier_lot_code, l.status,
               COALESCE(SUM(tl.quantity_lb), 0) AS available
        FROM lots l
        LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
        WHERE l.product_id = %s
        GROUP BY l.id
        {having}
        ORDER BY {FIFO_LOT_ORDER_SQL}
        """,
        (product_id,),
    )
    return [dict(r) for r in cur.fetchall()]


# ── Sales-order allocation write helpers (migration 044 / PR 3) ─────────────

def _allocations_enforced() -> bool:
    """Whether foreign sales-order reservations are a hard stock gate.

    Read the environment on every call so tests and operators can flip the
    rollout switch without changing allocation lifecycle behavior.  The
    production-safe default is observe-only.
    """
    return os.getenv("ALLOCATIONS_ENFORCED", "false").strip().lower() in {
        "1", "true", "yes", "on",
    }


def _allocation_error(code: str, message: str, status_code: int = 409, **fields):
    detail = {"error_code": code, "message": message, **fields}
    raise HTTPException(status_code=status_code, detail=detail)


def _lock_allocation_product(cur, product_id: int):
    """Serialize every allocation-affecting write for one product.

    Lock order is normative: all product lots by id, then all active allocation
    rows by id. Locking the lots also serializes the first allocation, when no
    allocation row exists yet.
    """
    cur.execute(
        "SELECT id FROM lots WHERE product_id = %s ORDER BY id FOR UPDATE",
        (product_id,),
    )
    cur.fetchall()
    cur.execute(
        "SELECT id FROM sales_order_allocations "
        "WHERE product_id = %s AND status = 'active' ORDER BY id FOR UPDATE",
        (product_id,),
    )
    cur.fetchall()


def _expire_auto_fifo_allocations(cur, product_id: int, released_by: Optional[str] = None) -> list:
    """Persist elapsed auto-FIFO TTLs on a product-locking write only."""
    cur.execute(
        """UPDATE sales_order_allocations
              SET status = 'released', released_at = clock_timestamp(),
                  released_by = %s, release_reason = 'expired'
            WHERE product_id = %s
              AND status = 'active'
              AND source = 'auto_fifo'
              AND expires_at IS NOT NULL
              AND expires_at <= clock_timestamp()
        RETURNING id, sales_order_line_id, lot_id, quantity_lb""",
        (released_by, product_id),
    )
    return [dict(row) for row in cur.fetchall()]


def _product_on_hand(cur, product_id: int) -> float:
    cur.execute(
        f"""SELECT COALESCE(SUM(tl.quantity_lb), 0) AS on_hand
              FROM lots l
              LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
             WHERE l.product_id = %s""",
        (product_id,),
    )
    return float(cur.fetchone()["on_hand"] or 0)


def _line_shipped_effective(cur, line_id: int, product_id: int) -> float:
    cur.execute(
        """SELECT COALESCE(SUM(ABS(tl.quantity_lb)), 0) AS shipped_lb
              FROM sales_order_shipments sos
              JOIN ledger_current_transactions ct
                ON ct.id = sos.transaction_id
               AND ct.effective_status = 'posted'
               AND ct.type = 'ship'
              JOIN ledger_current_transaction_lines tl
                ON tl.transaction_id = sos.transaction_id
               AND tl.product_id = %s
             WHERE sos.sales_order_line_id = %s""",
        (product_id, line_id),
    )
    return float(cur.fetchone()["shipped_lb"] or 0)


def _active_allocation_rows(cur, product_id: int) -> list:
    cur.execute(
        """SELECT *
              FROM sales_order_allocations
             WHERE product_id = %s
               AND status = 'active'
               AND (expires_at IS NULL OR expires_at > clock_timestamp())
             ORDER BY created_at ASC, id ASC""",
        (product_id,),
    )
    return [dict(row) for row in cur.fetchall()]


def available_lots_for_product(
    cur,
    product_id: int,
    so_line_id: Optional[int] = None,
    *,
    lock: bool = False,
    persist_expired: bool = False,
    released_by: Optional[str] = None,
) -> list:
    """Return deterministic FIFO lots after exclusive-reservation subtraction.

    Lot pins are removed from each lot's unpinned pool. Foreign SKU-level
    reservations then shadow-consume that pool in FIFO order. A caller's own
    lot pins are reported separately and are consumed before ``takeable`` by
    the sales-order ship plan.
    """
    if lock:
        _lock_allocation_product(cur, product_id)
    if persist_expired:
        if not lock:
            raise ValueError("persist_expired requires the product write lock")
        _expire_auto_fifo_allocations(cur, product_id, released_by)

    allocations = _active_allocation_rows(cur, product_id)
    lot_reserved = {}
    lot_reserved_others = {}
    lot_reserved_this_line = {}
    foreign_sku = 0.0
    for row in allocations:
        qty = float(row["quantity_lb"] or 0)
        line_id = int(row["sales_order_line_id"])
        if row["lot_id"] is None:
            if so_line_id is None or line_id != int(so_line_id):
                foreign_sku += qty
            continue
        lot_id = int(row["lot_id"])
        lot_reserved[lot_id] = lot_reserved.get(lot_id, 0.0) + qty
        if so_line_id is not None and line_id == int(so_line_id):
            lot_reserved_this_line[lot_id] = lot_reserved_this_line.get(lot_id, 0.0) + qty
        else:
            lot_reserved_others[lot_id] = lot_reserved_others.get(lot_id, 0.0) + qty

    result = []
    for lot in fifo_lot_balances(cur, product_id):
        item = dict(lot)
        lot_id = int(item["id"])
        on_hand = float(item["available"] or 0)
        reserved_all = lot_reserved.get(lot_id, 0.0)
        unpinned = max(0.0, on_hand - reserved_all)
        shadowed = min(foreign_sku, unpinned)
        foreign_sku -= shadowed
        item.update({
            "lot_id": lot_id,
            "on_hand": on_hand,
            "reserved_lot_lb": reserved_all,
            "reserved_others_lot": lot_reserved_others.get(lot_id, 0.0),
            "reserved_this_line_lot": lot_reserved_this_line.get(lot_id, 0.0),
            "takeable_unpinned": unpinned,
            "foreign_sku_shadow_lb": shadowed,
            "takeable": max(0.0, unpinned - shadowed),
        })
        result.append(item)
    return result


def _allocation_reservation_summary(
    cur, product_id: int, so_line_id: Optional[int] = None
) -> dict:
    """Summarize live, unexpired reservations owned by other SO lines.

    A sibling line is deliberately "other" even when it belongs to the same
    order.  Standalone inventory deductions pass ``so_line_id=None``, so every
    live reservation is foreign to them.
    """
    cur.execute(
        """SELECT so.id AS sales_order_id, so.order_number,
                  SUM(soa.quantity_lb) AS quantity_lb
             FROM sales_order_allocations soa
             JOIN sales_orders so ON so.id = soa.sales_order_id
            WHERE soa.product_id = %s
              AND soa.status = 'active'
              AND (soa.expires_at IS NULL OR soa.expires_at > clock_timestamp())
              AND (%s::bigint IS NULL OR soa.sales_order_line_id <> %s)
            GROUP BY so.id, so.order_number
            ORDER BY so.id""",
        (product_id, so_line_id, so_line_id),
    )
    rows = [dict(row) for row in cur.fetchall()]
    return {
        "reserved_others_lb": sum(float(row["quantity_lb"] or 0) for row in rows),
        "reserved_by_orders": [
            {
                "order_number": row["order_number"],
                "quantity_lb": float(row["quantity_lb"] or 0),
            }
            for row in rows
        ],
    }


def _takeable_deduction_plan(
    lots: list,
    requested_lb: float,
) -> dict:
    """Plan a standalone deduction without stealing until takeable is spent.

    PR 4 remains observe-only: after all unreserved capacity is planned, a
    physical-balance fallback may use reserved pounds.  PR 5 can reject that
    fallback without changing the deterministic lot plan.

    Pin fidelity (owner ruling 2026-08-20): when the request names a lot, the
    CALLER restricts ``lots`` to the pinned lot(s) before planning. The plan
    never reorders or widens the list it is given, so pinned pounds can never
    spill to a different physical lot to avoid reserved stock — observe mode
    takes the pinned lot's reserved pounds instead (with the warning + shrink).
    """
    ordered = list(lots)

    requested = max(0.0, float(requested_lb))
    remaining = requested
    planned_by_lot = {}

    # First consume only unreserved capacity, retaining FIFO order within
    # the caller-provided pool.
    for lot in ordered:
        if remaining <= BALANCE_EPSILON:
            break
        take = min(remaining, float(lot["takeable"] or 0))
        if take <= BALANCE_EPSILON:
            continue
        planned_by_lot[int(lot["lot_id"])] = take
        remaining -= take

    takeable_planned = requested - remaining

    # Observe-mode fallback: physical stock may still move even though another
    # order reserved it.  The caller must surface the warning; standalone ship
    # also repairs the resulting over-allocation after posting.
    for lot in ordered:
        if remaining <= BALANCE_EPSILON:
            break
        lot_id = int(lot["lot_id"])
        already = planned_by_lot.get(lot_id, 0.0)
        physical_left = max(0.0, float(lot["on_hand"] or 0) - already)
        take = min(remaining, physical_left)
        if take <= BALANCE_EPSILON:
            continue
        planned_by_lot[lot_id] = already + take
        remaining -= take

    plan = [
        {"lot": lot, "quantity_lb": planned_by_lot[int(lot["lot_id"])]}
        for lot in ordered
        if planned_by_lot.get(int(lot["lot_id"]), 0.0) > BALANCE_EPSILON
    ]
    planned_lb = requested - remaining
    return {
        "lots": plan,
        "planned_lb": planned_lb,
        "short_lb": max(0.0, remaining),
        "can_take_lb": min(requested, takeable_planned),
        "reserved_taken_lb": max(0.0, planned_lb - takeable_planned),
        "total_on_hand_lb": sum(float(lot["on_hand"] or 0) for lot in lots),
        "total_takeable_lb": sum(float(lot["takeable"] or 0) for lot in lots),
    }


def _allocation_observe_warning(
    action: str,
    requested_lb: float,
    can_take_lb: float,
    reserved_taken_lb: float,
    summary: dict,
    *,
    preview: bool = False,
) -> Optional[dict]:
    if _allocations_enforced():
        if (
            reserved_taken_lb <= BALANCE_EPSILON
            and (
                not preview
                or float(summary["reserved_others_lb"]) <= BALANCE_EPSILON
            )
        ):
            return None
        if reserved_taken_lb > BALANCE_EPSILON:
            message = (
                f"{action} would use {reserved_taken_lb:.4f} lb reserved for other "
                "sales orders. Allocation enforcement is on, so commit will be blocked."
            )
        else:
            message = (
                f"{action} has {float(summary['reserved_others_lb']):.4f} lb reserved "
                "for other sales orders. This preview stays within takeable stock; "
                "a larger commit may be blocked while allocation enforcement is on."
            )
        return {
            "warning_code": "STOCK_ALLOCATED",
            "message": message,
            "requested_lb": float(requested_lb),
            "can_take_lb": float(can_take_lb),
            "reserved_taken_lb": float(reserved_taken_lb),
            "reserved_others_lb": float(summary["reserved_others_lb"]),
            "reserved_by_orders": summary["reserved_by_orders"],
        }
    if reserved_taken_lb <= BALANCE_EPSILON:
        return None
    return {
        "warning_code": "RESERVED_STOCK_OBSERVE_ONLY",
        "message": (
            f"{action} uses {reserved_taken_lb:.4f} lb reserved for other sales "
            "orders. Allocation enforcement is off, so the write proceeds."
        ),
        "requested_lb": float(requested_lb),
        "can_take_lb": float(can_take_lb),
        "reserved_taken_lb": float(reserved_taken_lb),
        "reserved_others_lb": float(summary["reserved_others_lb"]),
        "reserved_by_orders": summary["reserved_by_orders"],
    }


def _enforce_allocation_takeable(
    action: str,
    requested_lb: float,
    can_take_lb: float,
    reserved_taken_lb: float,
    summary: dict,
    **fields,
):
    """Raise the PR-5 steal envelope before a stock-reducing write."""
    if not _allocations_enforced() or reserved_taken_lb <= BALANCE_EPSILON:
        return
    _allocation_error(
        "STOCK_ALLOCATED",
        (
            f"{action} would use {reserved_taken_lb:.4f} lb reserved for other "
            "sales orders. Allocation enforcement is on, so no stock was moved."
        ),
        requested_lb=float(requested_lb),
        can_take_lb=float(can_take_lb),
        reserved_taken_lb=float(reserved_taken_lb),
        reserved_others_lb=float(summary["reserved_others_lb"]),
        reserved_by_orders=summary["reserved_by_orders"],
        **fields,
    )


def _copy_allocation_row(
    cur,
    row: dict,
    quantity_lb: float,
    status: str,
    *,
    split_from_id: Optional[int],
    ship_transaction_id: Optional[int] = None,
    last_ship_transaction_id: Optional[int] = None,
    released_by: Optional[str] = None,
    release_reason: Optional[str] = None,
) -> int:
    released_at_sql = "clock_timestamp()" if status == "released" else "NULL"
    cur.execute(
        f"""INSERT INTO sales_order_allocations
               (sales_order_id, sales_order_line_id, product_id, lot_id,
                quantity_lb, status, source, ship_transaction_id,
                last_ship_transaction_id, split_from_id, created_by,
                released_at, released_by, release_reason, expires_at, note)
             VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                     {released_at_sql}, %s, %s, %s, %s)
          RETURNING id""",
        (
            row["sales_order_id"], row["sales_order_line_id"], row["product_id"],
            row["lot_id"], quantity_lb, status, row["source"],
            ship_transaction_id, last_ship_transaction_id, split_from_id,
            row.get("created_by"), released_by, release_reason,
            row.get("expires_at"), row.get("note"),
        ),
    )
    return int(cur.fetchone()["id"])


def _consume_allocation_row(cur, allocation_id: int, quantity_lb: float, transaction_id: int) -> dict:
    """Convert a live reservation slice to shipped, splitting when partial."""
    cur.execute(
        "SELECT * FROM sales_order_allocations WHERE id = %s AND status = 'active' FOR UPDATE",
        (allocation_id,),
    )
    row = cur.fetchone()
    if not row:
        _allocation_error(
            "ALLOCATION_NOT_ACTIVE",
            f"Allocation #{allocation_id} is no longer active",
        )
    row = dict(row)
    available = float(row["quantity_lb"])
    take = float(quantity_lb)
    if take <= BALANCE_EPSILON or take > available + BALANCE_EPSILON:
        raise ValueError(f"invalid allocation consume {take} from {available}")

    if take + BALANCE_EPSILON >= available:
        cur.execute(
            """UPDATE sales_order_allocations
                  SET status = 'shipped', ship_transaction_id = %s,
                      last_ship_transaction_id = %s, release_reason = NULL
                WHERE id = %s""",
            (transaction_id, transaction_id, allocation_id),
        )
        return {"allocation_id": allocation_id, "shipped_id": allocation_id, "leftover_id": None, "quantity_lb": available}

    cur.execute(
        """UPDATE sales_order_allocations
              SET status = 'superseded', release_reason = 'split_on_ship'
            WHERE id = %s""",
        (allocation_id,),
    )
    leftover_id = _copy_allocation_row(
        cur,
        row,
        available - take,
        "active",
        split_from_id=allocation_id,
        # A new leftover has never shipped.  In particular, it must not retain
        # an earlier voided transaction marker; that would be false audit
        # attribution. Restore demand comes only from the void-time record.
        last_ship_transaction_id=None,
    )
    shipped_id = _copy_allocation_row(
        cur,
        row,
        take,
        "shipped",
        split_from_id=allocation_id,
        ship_transaction_id=transaction_id,
        last_ship_transaction_id=transaction_id,
    )
    return {"allocation_id": allocation_id, "shipped_id": shipped_id, "leftover_id": leftover_id, "quantity_lb": take}


def _sales_order_ship_plan(
    cur,
    product_id: int,
    line_id: int,
    requested_lb: float,
    *,
    released_by: Optional[str] = None,
    lock: bool = True,
    persist_expired: bool = True,
) -> dict:
    """Lock and plan one SO line: own lot pins first, then takeable FIFO."""
    lots = available_lots_for_product(
        cur,
        product_id,
        line_id,
        lock=lock,
        persist_expired=persist_expired,
        released_by=released_by,
    )
    by_lot = {int(lot["lot_id"]): lot for lot in lots}
    allocations = _active_allocation_rows(cur, product_id)
    own_lot_rows = [
        row for row in allocations
        if int(row["sales_order_line_id"]) == int(line_id) and row["lot_id"] is not None
    ]
    own_sku_rows = [
        row for row in allocations
        if int(row["sales_order_line_id"]) == int(line_id) and row["lot_id"] is None
    ]

    remaining = float(requested_lb)
    lot_takes = []
    allocation_takes = []

    def add_lot_take(lot: dict, qty: float):
        if qty <= BALANCE_EPSILON:
            return
        existing = next((item for item in lot_takes if item["lot_id"] == int(lot["lot_id"])), None)
        if existing:
            existing["quantity_lb"] += qty
        else:
            lot_takes.append({
                "lot_id": int(lot["lot_id"]),
                "lot_code": lot["lot_code"],
                "quantity_lb": qty,
            })

    for row in own_lot_rows:
        if remaining <= BALANCE_EPSILON:
            break
        lot = by_lot.get(int(row["lot_id"]))
        if not lot:
            continue
        protected_for_others = float(lot["reserved_others_lot"] or 0)
        own_physical_cover = max(0.0, float(lot["on_hand"]) - protected_for_others)
        take = min(remaining, float(row["quantity_lb"]), own_physical_cover)
        if take <= BALANCE_EPSILON:
            continue
        add_lot_take(lot, take)
        allocation_takes.append({"allocation_id": int(row["id"]), "quantity_lb": take})
        remaining -= take

    for lot in lots:
        if remaining <= BALANCE_EPSILON:
            break
        take = min(remaining, float(lot["takeable"] or 0))
        if take <= BALANCE_EPSILON:
            continue
        add_lot_take(lot, take)
        remaining -= take

    actual_ship = max(0.0, float(requested_lb) - remaining)
    lot_allocated_take = sum(item["quantity_lb"] for item in allocation_takes)
    sku_cover_needed = max(0.0, actual_ship - lot_allocated_take)
    for row in own_sku_rows:
        if sku_cover_needed <= BALANCE_EPSILON:
            break
        take = min(sku_cover_needed, float(row["quantity_lb"]))
        allocation_takes.append({"allocation_id": int(row["id"]), "quantity_lb": take})
        sku_cover_needed -= take

    return {
        "requested_lb": float(requested_lb),
        "actual_ship_lb": actual_ship,
        "lots": lot_takes,
        "allocation_takes": allocation_takes,
    }


def _consume_sales_order_allocations(cur, plan: dict, transaction_id: int) -> list:
    return [
        _consume_allocation_row(cur, item["allocation_id"], item["quantity_lb"], transaction_id)
        for item in plan["allocation_takes"]
        if item["quantity_lb"] > BALANCE_EPSILON
    ]


def _release_active_allocations(
    cur,
    *,
    reason: str,
    released_by: Optional[str] = None,
    order_id: Optional[int] = None,
    line_id: Optional[int] = None,
    allocation_id: Optional[int] = None,
) -> list:
    clauses = ["status = 'active'"]
    params = [released_by]
    if order_id is not None:
        clauses.append("sales_order_id = %s")
        params.append(order_id)
    if line_id is not None:
        clauses.append("sales_order_line_id = %s")
        params.append(line_id)
    if allocation_id is not None:
        clauses.append("id = %s")
        params.append(allocation_id)
    params.append(reason)
    cur.execute(
        f"""UPDATE sales_order_allocations
               SET status = 'released', released_at = clock_timestamp(),
                   released_by = %s, release_reason = %s
             WHERE {' AND '.join(clauses)}
         RETURNING id, product_id, sales_order_line_id, lot_id, quantity_lb""",
        [params[0], params[-1], *params[1:-1]],
    )
    return [dict(row) for row in cur.fetchall()]


def _shrink_active_allocations(cur, rows: list, excess_lb: float, reason: str, released_by: Optional[str]) -> list:
    """Release newest/least-urgent slices until ``excess_lb`` is removed."""
    remaining = float(excess_lb)
    changed = []
    for raw in rows:
        if remaining <= BALANCE_EPSILON:
            break
        row = dict(raw)
        qty = float(row["quantity_lb"])
        take = min(qty, remaining)
        if take + BALANCE_EPSILON >= qty:
            cur.execute(
                """UPDATE sales_order_allocations
                      SET status = 'released', released_at = clock_timestamp(),
                          released_by = %s, release_reason = %s
                    WHERE id = %s""",
                (released_by, reason, row["id"]),
            )
            released_id = int(row["id"])
        else:
            cur.execute(
                "UPDATE sales_order_allocations SET quantity_lb = quantity_lb - %s WHERE id = %s",
                (take, row["id"]),
            )
            released_id = _copy_allocation_row(
                cur,
                row,
                take,
                "released",
                split_from_id=int(row["id"]),
                released_by=released_by,
                release_reason=reason,
            )
        changed.append({"allocation_id": released_id, "quantity_lb": take, "reason": reason})
        remaining -= take
    return changed


def _void_ship_allocations(cur, transaction_id: int, operator_id: Optional[str]) -> list:
    cur.execute(
        """SELECT DISTINCT product_id
              FROM sales_order_allocations
             WHERE ship_transaction_id = %s AND status = 'shipped'
             ORDER BY product_id""",
        (transaction_id,),
    )
    product_ids = [int(row["product_id"]) for row in cur.fetchall()]
    restored = []
    for product_id in product_ids:
        _lock_allocation_product(cur, product_id)
        _expire_auto_fifo_allocations(cur, product_id, operator_id)
        cur.execute(
            """SELECT * FROM sales_order_allocations
                 WHERE product_id = %s AND ship_transaction_id = %s
                   AND status = 'shipped'
                 ORDER BY id FOR UPDATE""",
            (product_id, transaction_id),
        )
        for shipped_raw in cur.fetchall():
            shipped = dict(shipped_raw)
            cur.execute(
                """SELECT * FROM sales_order_allocations
                     WHERE sales_order_line_id = %s
                       AND lot_id IS NOT DISTINCT FROM %s
                       AND status = 'active'
                     ORDER BY id LIMIT 1 FOR UPDATE""",
                (shipped["sales_order_line_id"], shipped["lot_id"]),
            )
            leftover = cur.fetchone()
            if leftover:
                cur.execute(
                    """UPDATE sales_order_allocations
                          SET quantity_lb = quantity_lb + %s,
                              last_ship_transaction_id = %s
                        WHERE id = %s""",
                    (shipped["quantity_lb"], transaction_id, leftover["id"]),
                )
                cur.execute(
                    """UPDATE sales_order_allocations
                          SET status = 'superseded', ship_transaction_id = NULL,
                              last_ship_transaction_id = %s,
                              release_reason = 'void_coalesced'
                        WHERE id = %s""",
                    (transaction_id, shipped["id"]),
                )
                live_id = int(leftover["id"])
                coalesced = True
            else:
                cur.execute(
                    """UPDATE sales_order_allocations
                          SET status = 'active', ship_transaction_id = NULL,
                              last_ship_transaction_id = %s,
                              release_reason = NULL
                        WHERE id = %s""",
                    (transaction_id, shipped["id"]),
                )
                live_id = int(shipped["id"])
                coalesced = False
            restored.append({
                "allocation_id": int(shipped["id"]),
                "live_allocation_id": live_id,
                "sales_order_line_id": int(shipped["sales_order_line_id"]),
                "quantity_lb": float(shipped["quantity_lb"]),
                "coalesced": coalesced,
            })
    return restored


def _record_void_allocation_reactivations(
    cur,
    transaction_id: int,
    correction_id,
    restored: list,
):
    """Persist one explicit void-time reactivation quantity per shipped SO line."""
    totals = {}
    for row in restored:
        line_id = int(row["sales_order_line_id"])
        totals[line_id] = totals.get(line_id, Decimal("0")) + Decimal(
            str(row["quantity_lb"])
        )

    cur.execute(
        """SELECT DISTINCT sales_order_line_id
              FROM sales_order_shipments
             WHERE transaction_id = %s
             ORDER BY sales_order_line_id""",
        (transaction_id,),
    )
    line_ids = [int(row["sales_order_line_id"]) for row in cur.fetchall()]
    for line_id in line_ids:
        cur.execute(
            """INSERT INTO sales_order_allocation_reactivations
                       (transaction_id, sales_order_line_id, quantity_lb,
                        correction_id)
                 VALUES (%s, %s, %s, %s)
            ON CONFLICT (transaction_id, sales_order_line_id) DO UPDATE
                    SET quantity_lb = EXCLUDED.quantity_lb,
                        correction_id = EXCLUDED.correction_id,
                        created_at = clock_timestamp()""",
            (
                transaction_id,
                line_id,
                totals.get(line_id, Decimal("0")),
                correction_id,
            ),
        )


def _preflight_restore_ship_stock(cur, transaction_id: int) -> list:
    """Lock original ship lots and reject a restore that would go negative."""
    cur.execute(
        """SELECT tl.product_id, tl.lot_id, l.lot_code,
                  SUM(ABS(tl.quantity_lb)) AS quantity_lb
             FROM ledger_current_transaction_lines tl
             JOIN lots l ON l.id = tl.lot_id
            WHERE tl.transaction_id = %s
            GROUP BY tl.product_id, tl.lot_id, l.lot_code
            ORDER BY tl.product_id, tl.lot_id""",
        (transaction_id,),
    )
    ship_lots = [dict(row) for row in cur.fetchall()]
    product_ids = sorted({int(row["product_id"]) for row in ship_lots})
    for product_id in product_ids:
        _lock_allocation_product(cur, product_id)

    for row in ship_lots:
        required = float(row["quantity_lb"] or 0)
        available = lot_on_hand(cur, int(row["lot_id"]))
        if available + BALANCE_EPSILON < required:
            _allocation_error(
                "RESTORE_STOCK_MISSING",
                f"Cannot restore ship transaction #{transaction_id}: posted on-hand cannot cover the restored deduction",
                transaction_id=transaction_id,
                lot_id=int(row["lot_id"]),
                lot_code=row["lot_code"],
                required_lb=required,
                available_lb=available,
            )
    return product_ids


def _preflight_restore_ship_takeable(cur, transaction_id: int):
    """Reject a restore that would re-take another line's reserved pounds.

    A restore re-posts its original fixed-lot deductions.  The owning SO line's
    live lot pins are entitlement on that lot, and its SKU-level reservation is
    already represented in ``takeable`` because only foreign SKU claims are
    shadow-consumed.  Standalone ships have no owning line, so every live
    reservation is foreign.
    """
    if not _allocations_enforced():
        return

    cur.execute(
        """SELECT DISTINCT sol.product_id, sos.sales_order_line_id
             FROM sales_order_shipments sos
             JOIN sales_order_lines sol ON sol.id = sos.sales_order_line_id
            WHERE sos.transaction_id = %s
            ORDER BY sol.product_id, sos.sales_order_line_id""",
        (transaction_id,),
    )
    lines_by_product = {}
    for row in cur.fetchall():
        lines_by_product.setdefault(int(row["product_id"]), []).append(
            int(row["sales_order_line_id"])
        )

    cur.execute(
        """SELECT tl.product_id, tl.lot_id, l.lot_code,
                  SUM(ABS(tl.quantity_lb)) AS quantity_lb
             FROM ledger_current_transaction_lines tl
             JOIN lots l ON l.id = tl.lot_id
            WHERE tl.transaction_id = %s
            GROUP BY tl.product_id, tl.lot_id, l.lot_code
            ORDER BY tl.product_id, tl.lot_id""",
        (transaction_id,),
    )
    ship_lots = [dict(row) for row in cur.fetchall()]
    lot_views = {}
    summaries = {}
    for row in ship_lots:
        product_id = int(row["product_id"])
        owner_lines = lines_by_product.get(product_id, [])
        # The production ship path creates one transaction per physical line.
        # If a historical/manual transaction is ambiguous, do not guess an
        # entitlement: treat its reservations as foreign.
        owner_line_id = owner_lines[0] if len(owner_lines) == 1 else None
        cache_key = (product_id, owner_line_id)
        if cache_key not in lot_views:
            lot_views[cache_key] = {
                int(lot["lot_id"]): lot
                for lot in available_lots_for_product(
                    cur, product_id, owner_line_id, lock=False,
                    persist_expired=False,
                )
            }
            summaries[cache_key] = _allocation_reservation_summary(
                cur, product_id, owner_line_id
            )

        lot = lot_views[cache_key].get(int(row["lot_id"]))
        required = float(row["quantity_lb"] or 0)
        if not lot:
            # The always-on stock preflight runs first and owns this condition.
            continue
        own_lot_cover = 0.0
        if owner_line_id is not None:
            own_lot_cover = min(
                float(lot["reserved_this_line_lot"] or 0),
                max(
                    0.0,
                    float(lot["on_hand"] or 0)
                    - float(lot["reserved_others_lot"] or 0),
                ),
            )
        can_take = min(
            required,
            float(lot["takeable"] or 0) + own_lot_cover,
        )
        reserved_taken = max(0.0, required - can_take)
        extra_fields = {
            "transaction_id": int(transaction_id),
            "product_id": product_id,
            "lot_id": int(row["lot_id"]),
            "lot_code": row["lot_code"],
        }
        if owner_line_id is not None:
            extra_fields["sales_order_line_id"] = owner_line_id
        _enforce_allocation_takeable(
            "Restore ship transaction",
            required,
            can_take,
            reserved_taken,
            summaries[cache_key],
            **extra_fields,
        )


def _prepare_restore_ship_allocations(
    cur,
    transaction_id: int,
    operator_id: Optional[str],
) -> dict:
    """Load recorded demand, lock live coverage, and prepare atomic consumes."""
    cur.execute(
        """SELECT DISTINCT sos.sales_order_line_id,
                  sol.sales_order_id, sol.product_id,
                  COALESCE(reactivation.quantity_lb, 0) AS quantity_lb,
                  (reactivation.transaction_id IS NULL) AS reactivation_unknown
             FROM sales_order_shipments sos
             JOIN sales_order_lines sol
               ON sol.id = sos.sales_order_line_id
        LEFT JOIN sales_order_allocation_reactivations reactivation
               ON reactivation.transaction_id = sos.transaction_id
              AND reactivation.sales_order_line_id = sos.sales_order_line_id
            WHERE sos.transaction_id = %s
            ORDER BY sol.product_id, sos.sales_order_line_id""",
        (transaction_id,),
    )
    targets = [dict(row) for row in cur.fetchall()]
    product_ids = sorted({int(row["product_id"]) for row in targets})

    for product_id in product_ids:
        _lock_allocation_product(cur, product_id)
        _expire_auto_fifo_allocations(cur, product_id, operator_id)

    plans = []
    for target in targets:
        needed = float(target["quantity_lb"] or 0)
        if needed <= BALANCE_EPSILON:
            continue
        cur.execute(
            """SELECT id, quantity_lb
                  FROM sales_order_allocations
                 WHERE sales_order_line_id = %s
                   AND product_id = %s
                   AND status = 'active'
                   AND (expires_at IS NULL OR expires_at > clock_timestamp())
                 ORDER BY (lot_id IS NULL) ASC, created_at ASC, id ASC
                 FOR UPDATE""",
            (target["sales_order_line_id"], target["product_id"]),
        )
        live_rows = [dict(row) for row in cur.fetchall()]
        available = sum(float(row["quantity_lb"]) for row in live_rows)
        if available + BALANCE_EPSILON < needed:
            _allocation_error(
                "RESTORE_SPLIT_MISSING",
                f"Cannot restore ship transaction #{transaction_id}: its live allocation coverage is missing or too small",
                transaction_id=transaction_id,
                sales_order_line_id=int(target["sales_order_line_id"]),
                required_lb=needed,
                available_lb=available,
            )
        remaining = needed
        takes = []
        for row in live_rows:
            if remaining <= BALANCE_EPSILON:
                break
            take = min(remaining, float(row["quantity_lb"]))
            takes.append((int(row["id"]), take))
            remaining -= take
        plans.append((target, takes))

    return {
        "plans": plans,
        "product_ids": product_ids,
        "unknown_line_ids": sorted(
            int(row["sales_order_line_id"])
            for row in targets
            if row["reactivation_unknown"]
        ),
    }


def _restore_ship_allocations(
    cur,
    transaction_id: int,
    operator_id: Optional[str],
    prepared: Optional[dict] = None,
) -> list:
    """Consume recorded void-time quantities from this line's live rows only."""
    prepared = prepared or _prepare_restore_ship_allocations(
        cur, transaction_id, operator_id
    )
    reshipped = []
    for target, takes in prepared["plans"]:
        for allocation_id, take in takes:
            change = _consume_allocation_row(
                cur, allocation_id, take, transaction_id
            )
            change["sales_order_line_id"] = int(target["sales_order_line_id"])
            reshipped.append(change)
    return reshipped


def _shrink_overallocated_products(
    cur,
    product_ids: list,
    operator_id: Optional[str],
    release_reason: str = "inventory_voided",
) -> list:
    changed = []
    for product_id in sorted(set(int(pid) for pid in product_ids)):
        _lock_allocation_product(cur, product_id)
        _expire_auto_fifo_allocations(cur, product_id, operator_id)

        # A named-lot pin must remain covered by that lot even when other lots
        # keep the product-level total positive. Repair lot deficits first.
        cur.execute(
            """SELECT DISTINCT lot_id
                  FROM sales_order_allocations
                 WHERE product_id = %s AND lot_id IS NOT NULL
                   AND status = 'active'
                   AND (expires_at IS NULL OR expires_at > clock_timestamp())
                 ORDER BY lot_id""",
            (product_id,),
        )
        lot_ids = [int(row["lot_id"]) for row in cur.fetchall()]
        for lot_id in lot_ids:
            cur.execute(
                """SELECT soa.*
                      FROM sales_order_allocations soa
                      JOIN sales_orders so ON so.id = soa.sales_order_id
                     WHERE soa.product_id = %s AND soa.lot_id = %s
                       AND soa.status = 'active'
                       AND (soa.expires_at IS NULL OR soa.expires_at > clock_timestamp())
                     ORDER BY so.requested_ship_date DESC NULLS LAST,
                              soa.sales_order_id DESC, soa.created_at DESC, soa.id DESC
                     FOR UPDATE OF soa""",
                (product_id, lot_id),
            )
            lot_rows = cur.fetchall()
            lot_allocated = sum(float(row["quantity_lb"]) for row in lot_rows)
            lot_excess = max(0.0, lot_allocated - lot_on_hand(cur, lot_id))
            if lot_excess > BALANCE_EPSILON:
                changed.extend(_shrink_active_allocations(
                    cur, lot_rows, lot_excess, release_reason, operator_id
                ))

        on_hand = _product_on_hand(cur, product_id)
        cur.execute(
            """SELECT soa.*
                  FROM sales_order_allocations soa
                  JOIN sales_orders so ON so.id = soa.sales_order_id
                 WHERE soa.product_id = %s
                   AND soa.status = 'active'
                   AND (soa.expires_at IS NULL OR soa.expires_at > clock_timestamp())
                 ORDER BY so.requested_ship_date DESC NULLS LAST,
                          soa.sales_order_id DESC, soa.created_at DESC, soa.id DESC
                 FOR UPDATE OF soa""",
            (product_id,),
        )
        rows = cur.fetchall()
        allocated = sum(float(row["quantity_lb"]) for row in rows)
        excess = max(0.0, allocated - on_hand)
        if excess > BALANCE_EPSILON:
            changed.extend(_shrink_active_allocations(cur, rows, excess, release_reason, operator_id))
    return changed


def _coalesce_lot_allocations(cur, product_id: int, source_lot_id: int, target_lot_id: int) -> list:
    """Move live pins to a surviving lot without violating the live unique key."""
    _lock_allocation_product(cur, product_id)
    cur.execute(
        """SELECT * FROM sales_order_allocations
             WHERE product_id = %s AND lot_id = %s AND status = 'active'
             ORDER BY id FOR UPDATE""",
        (product_id, source_lot_id),
    )
    moved = []
    for source_raw in cur.fetchall():
        source = dict(source_raw)
        cur.execute(
            """SELECT id FROM sales_order_allocations
                 WHERE sales_order_line_id = %s AND lot_id = %s
                   AND status = 'active'
                 ORDER BY id LIMIT 1 FOR UPDATE""",
            (source["sales_order_line_id"], target_lot_id),
        )
        target = cur.fetchone()
        if target:
            cur.execute(
                "UPDATE sales_order_allocations SET quantity_lb = quantity_lb + %s WHERE id = %s",
                (source["quantity_lb"], target["id"]),
            )
            cur.execute(
                """UPDATE sales_order_allocations
                      SET status = 'superseded', release_reason = 'lot_merged'
                    WHERE id = %s""",
                (source["id"],),
            )
            moved.append({"allocation_id": int(source["id"]), "target_allocation_id": int(target["id"]), "coalesced": True})
        else:
            cur.execute(
                "UPDATE sales_order_allocations SET lot_id = %s WHERE id = %s",
                (target_lot_id, source["id"]),
            )
            moved.append({"allocation_id": int(source["id"]), "target_allocation_id": int(source["id"]), "coalesced": False})

    cur.execute(
        """UPDATE sales_order_allocations
              SET lot_id = %s
            WHERE product_id = %s AND lot_id = %s
              AND status IN ('shipped', 'superseded')""",
        (target_lot_id, product_id, source_lot_id),
    )
    return moved


def _load_allocatable_line(cur, order_id: int, line_id: int) -> dict:
    cur.execute(
        """SELECT so.id AS sales_order_id, so.order_number, so.status AS order_status,
                  sol.id AS line_id, sol.product_id, sol.quantity_lb,
                  sol.line_status, p.name AS product_name, p.odoo_code AS sku,
                  COALESCE(p.is_service, false) AS is_service
             FROM sales_orders so
             JOIN sales_order_lines sol ON sol.sales_order_id = so.id
             JOIN products p ON p.id = sol.product_id
            WHERE so.id = %s AND sol.id = %s
            FOR UPDATE OF so, sol""",
        (order_id, line_id),
    )
    row = cur.fetchone()
    if not row:
        _allocation_error(
            "LINE_NOT_FOUND",
            f"Line #{line_id} was not found on order #{order_id}",
            status_code=404,
            order_id=order_id,
            line_id=line_id,
        )
    line = dict(row)
    if line["is_service"]:
        _allocation_error(
            "SERVICE_LINE_NOT_ALLOCATABLE",
            f"Service line #{line_id} does not represent inventory",
            status_code=422,
            order_id=order_id,
            line_id=line_id,
        )
    if line["order_status"] in {"cancelled", "invoiced"} or line["line_status"] in {"cancelled", "fulfilled"}:
        _allocation_error(
            "ORDER_NOT_ALLOCATABLE",
            f"Order {line['order_number']} / line #{line_id} is not open for allocation",
            order_id=order_id,
            line_id=line_id,
            order_status=line["order_status"],
            line_status=line["line_status"],
        )
    return line


def _allocation_totals(cur, product_id: int, line_id: int) -> tuple[float, float]:
    cur.execute(
        """SELECT COALESCE(SUM(quantity_lb), 0) AS product_allocated,
                  COALESCE(SUM(quantity_lb) FILTER (WHERE sales_order_line_id = %s), 0) AS line_allocated
             FROM sales_order_allocations
            WHERE product_id = %s AND status = 'active'
              AND (expires_at IS NULL OR expires_at > clock_timestamp())""",
        (line_id, product_id),
    )
    row = cur.fetchone()
    return float(row["product_allocated"] or 0), float(row["line_allocated"] or 0)


def _upsert_live_allocation(
    cur,
    *,
    line: dict,
    lot_id: Optional[int],
    quantity_lb: float,
    source: str,
    expires_at: Optional[datetime],
    note: Optional[str],
    created_by: Optional[str],
) -> dict:
    cur.execute(
        """SELECT * FROM sales_order_allocations
             WHERE sales_order_line_id = %s
               AND lot_id IS NOT DISTINCT FROM %s
               AND status = 'active'
             ORDER BY id LIMIT 1 FOR UPDATE""",
        (line["line_id"], lot_id),
    )
    existing = cur.fetchone()
    if existing:
        existing = dict(existing)
        # A deliberate/manual pin is stronger than an expiring auto claim.
        # Promote auto -> manual/staged when the caller makes that commitment;
        # never make an existing manual/staged slice expire because auto-FIFO
        # adds to the same unique live key.
        precedence = {"auto_fifo": 0, "manual": 1, "staged_lot": 2}
        merged_source = max((existing["source"], source), key=lambda value: precedence[value])
        merged_expiry = expires_at if merged_source == "auto_fifo" else None
        cur.execute(
            """UPDATE sales_order_allocations
                  SET quantity_lb = quantity_lb + %s,
                      source = %s, expires_at = %s,
                      note = COALESCE(%s, note)
                WHERE id = %s
            RETURNING *""",
            (quantity_lb, merged_source, merged_expiry, note, existing["id"]),
        )
    else:
        cur.execute(
            """INSERT INTO sales_order_allocations
                   (sales_order_id, sales_order_line_id, product_id, lot_id,
                    quantity_lb, source, expires_at, note, created_by)
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
              RETURNING *""",
            (
                line["sales_order_id"], line["line_id"], line["product_id"], lot_id,
                quantity_lb, source, expires_at, note, created_by,
            ),
        )
    return dict(cur.fetchone())


def _validate_allocation_addition(
    cur,
    line: dict,
    quantity_lb: float,
    *,
    lot_id: Optional[int] = None,
) -> dict:
    if quantity_lb <= BALANCE_EPSILON:
        _allocation_error(
            "INVALID_ALLOCATION_QUANTITY",
            "quantity_lb must be greater than zero",
            status_code=422,
            requested_lb=quantity_lb,
        )
    product_id = int(line["product_id"])
    product_allocated, line_allocated = _allocation_totals(cur, product_id, int(line["line_id"]))
    shipped_effective = _line_shipped_effective(cur, int(line["line_id"]), product_id)
    remaining = max(0.0, float(line["quantity_lb"]) - shipped_effective)
    line_coverable = max(0.0, remaining - line_allocated)
    on_hand = _product_on_hand(cur, product_id)
    product_coverable = max(0.0, on_hand - product_allocated)
    coverable = min(line_coverable, product_coverable)

    lot_coverable = None
    lot_code = None
    if lot_id is not None:
        cur.execute(
            "SELECT id, lot_code, product_id FROM lots WHERE id = %s",
            (lot_id,),
        )
        lot = cur.fetchone()
        if not lot:
            _allocation_error(
                "LOT_NOT_FOUND",
                f"Lot #{lot_id} not found",
                status_code=404,
                lot_id=lot_id,
            )
        if int(lot["product_id"]) != product_id:
            _allocation_error(
                "LOT_PRODUCT_MISMATCH",
                f"Lot #{lot_id} does not belong to line #{line['line_id']}'s product",
                status_code=422,
                lot_id=lot_id,
                product_id=product_id,
            )
        lot_code = lot["lot_code"]
        cur.execute(
            """SELECT COALESCE(SUM(quantity_lb), 0) AS allocated
                 FROM sales_order_allocations
                WHERE lot_id = %s AND status = 'active'
                  AND (expires_at IS NULL OR expires_at > clock_timestamp())""",
            (lot_id,),
        )
        lot_allocated = float(cur.fetchone()["allocated"] or 0)
        lot_coverable = max(0.0, lot_on_hand(cur, lot_id) - lot_allocated)
        coverable = min(coverable, lot_coverable)

    if quantity_lb > coverable + BALANCE_EPSILON:
        sku = line.get("sku") or line["product_name"]
        _allocation_error(
            "OVER_ALLOCATION",
            f"{sku}: requested {quantity_lb:g} lb, coverable {coverable:g} lb",
            on_hand_lb=on_hand,
            allocated_others_lb=product_allocated,
            coverable_lb=coverable,
            requested_lb=quantity_lb,
            remaining_effective_lb=remaining,
            line_allocated_lb=line_allocated,
            lot_id=lot_id,
            lot_code=lot_code,
            lot_coverable_lb=lot_coverable,
        )
    return {
        "on_hand_lb": on_hand,
        "product_allocated_lb": product_allocated,
        "line_allocated_lb": line_allocated,
        "remaining_effective_lb": remaining,
        "coverable_lb": coverable,
    }


def validate_lot_deduction(cur, lot_id: int, lot_code: str, requested_lb: float):
    """Validate that deducting requested_lb from a lot won't push it negative.

    Must be called within a DB transaction AFTER the lot row is locked (FOR UPDATE).
    Uses BALANCE_EPSILON tolerance to handle floating point dust.

    Returns the current available balance (epsilon-adjusted).
    Raises HTTPException(400) if the deduction would cause a genuine negative balance.

    NOT used for adjust transactions — adjustments are the escape hatch for corrections.
    """
    balance = lot_on_hand(cur, lot_id)

    # Snap near-zero balances to zero
    if abs(balance) < BALANCE_EPSILON:
        balance = 0.0

    # Allow if requested is within epsilon of available
    if balance + BALANCE_EPSILON >= requested_lb:
        return balance

    raise HTTPException(
        status_code=400,
        detail=f"Insufficient balance on lot {lot_code}: available {balance:.4f} lb, requested {requested_lb:.4f} lb"
    )


# Connection pool (initialized on startup)
db_pool = None


@app.on_event("startup")
async def startup():
    global db_pool
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL env var required — app cannot start without a database")
    if not API_KEY:
        raise RuntimeError("API_KEY env var required — app cannot start without authentication")
    if not DASHBOARD_API_KEY:
        raise RuntimeError("DASHBOARD_API_KEY env var required — the dashboard's scoped key has no fallback")
    if DASHBOARD_API_KEY == API_KEY:
        logger.warning("DASHBOARD_API_KEY equals API_KEY — dashboard key scoping is ineffective")
    try:
        db_pool = pool.ThreadedConnectionPool(minconn=2, maxconn=20, dsn=DATABASE_URL)
        logger.info("Database connection pool created")
    except Exception as e:
        logger.error(f"Failed to create connection pool: {e}")
        raise

    # Migration: Add label_type column for SKU protection
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                # Add column if it doesn't exist
                cur.execute("""
                    ALTER TABLE products ADD COLUMN IF NOT EXISTS label_type TEXT DEFAULT 'house'
                """)

                # Set private-label flags on verified finished goods by odoo_code
                cur.execute("""
                    UPDATE products SET label_type = 'private_label'
                    WHERE odoo_code = ANY(%s) AND COALESCE(label_type, 'house') != 'private_label'
                """, (PRIVATE_LABEL_ODOO_CODES,))
                updated_by_code = cur.rowcount

                # Set private-label flags on Blue Stripes exclusive batch products
                cur.execute("""
                    UPDATE products SET label_type = 'private_label'
                    WHERE name ILIKE 'Batch BS %%'
                      AND COALESCE(label_type, 'house') != 'private_label'
                """)
                updated_by_name = cur.rowcount

                # Set private-label flags on Setton batch products
                cur.execute("""
                    UPDATE products SET label_type = 'private_label'
                    WHERE name ILIKE 'Batch Setton %%'
                      AND COALESCE(label_type, 'house') != 'private_label'
                """)
                updated_by_name += cur.rowcount

                conn.commit()
                if updated_by_code + updated_by_name > 0:
                    logger.info(f"SKU protection migration: flagged {updated_by_code + updated_by_name} products as private_label")
                else:
                    logger.info("SKU protection: label_type column up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"SKU protection migration warning (non-fatal): {e}")

    # Migration 004: Add exclude_from_inventory flag to batch_formulas
    # Allows utility ingredients (e.g. Water) to remain visible in formulas
    # without blocking production or creating phantom inventory shortages.
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    ALTER TABLE batch_formulas
                    ADD COLUMN IF NOT EXISTS exclude_from_inventory BOOLEAN DEFAULT false
                """)

                # Flag Water as excluded in all formulas
                cur.execute("""
                    UPDATE batch_formulas bf
                    SET exclude_from_inventory = true
                    FROM products p
                    WHERE p.id = bf.ingredient_product_id
                      AND LOWER(p.name) = 'water'
                      AND bf.exclude_from_inventory = false
                """)
                water_rows = cur.rowcount

                conn.commit()
                if water_rows > 0:
                    logger.info(f"Migration 004: flagged {water_rows} Water formula row(s) as exclude_from_inventory")
                else:
                    logger.info("Migration 004: exclude_from_inventory column up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 004 warning (non-fatal): {e}")

    # Migration 005: Add yield_multiplier to products
    # Allows products that gain/lose weight during processing (e.g. coconut hydration)
    # to record an expected yield factor. Default 1.0 = no change.
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    ALTER TABLE products
                    ADD COLUMN IF NOT EXISTS yield_multiplier FLOAT DEFAULT 1.0
                """)
                conn.commit()
                logger.info("Migration 005: yield_multiplier column up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 005 warning (non-fatal): {e}")

    # Migration 006: Add case_size_lb to products
    # Stores the weight per sellable unit (e.g., 25 for "25 LB" case, 10 for "10 LB" case).
    # Required for correct line_value calculation: cases * unit_price (not lb * unit_price).
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    ALTER TABLE products
                    ADD COLUMN IF NOT EXISTS case_size_lb NUMERIC(10,2)
                """)

                # Auto-populate from product names — LB patterns
                cur.execute("UPDATE products SET case_size_lb = 25 WHERE name LIKE '%25 LB%' AND case_size_lb IS NULL")
                updated_25 = cur.rowcount
                cur.execute("UPDATE products SET case_size_lb = 10 WHERE name LIKE '%10 LB%' AND case_size_lb IS NULL")
                updated_10 = cur.rowcount
                cur.execute("UPDATE products SET case_size_lb = 50 WHERE name LIKE '%50 LB%' AND case_size_lb IS NULL")
                updated_50 = cur.rowcount

                # Auto-populate from product names — OZ patterns (e.g., "12x10 OZ", "6x7 OZ")
                cur.execute("""
                    UPDATE products SET case_size_lb = ROUND(
                        (substring(name FROM '(\d+)\s*x\s*\d+'))::numeric
                        * (substring(name FROM '\d+\s*x\s*(\d+\.?\d*)\s*OZ'))::numeric
                        / 16.0, 2)
                    WHERE name ~* '\d+\s*x\s*\d+\.?\d*\s*OZ'
                      AND case_size_lb IS NULL
                """)
                updated_oz = cur.rowcount

                conn.commit()
                total = updated_25 + updated_10 + updated_50 + updated_oz
                if total > 0:
                    logger.info(f"Migration 006: case_size_lb populated for {total} products (25lb:{updated_25}, 10lb:{updated_10}, 50lb:{updated_50}, oz:{updated_oz})")
                else:
                    logger.info("Migration 006: case_size_lb column up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 006 warning (non-fatal): {e}")

    # Migration 007: Migrate legacy 'new' orders to 'confirmed'
    # Phase 3 changed default status to 'confirmed', but pre-existing orders may still be 'new'.
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("UPDATE sales_orders SET status = 'confirmed' WHERE status = 'new'")
                migrated = cur.rowcount
                conn.commit()
                if migrated > 0:
                    logger.info(f"Migration 007: Migrated {migrated} orders from 'new' to 'confirmed'")
                else:
                    logger.info("Migration 007: No legacy 'new' orders to migrate")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 007 warning (non-fatal): {e}")

    # Migration 008: Lot merge support columns
    # Adds status, merged_into_lot_id, merged_at, merge_reason to lots table
    # for controlled lot merge operations (POST /admin/lots/merge).
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("ALTER TABLE lots ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'active'")
                cur.execute("ALTER TABLE lots ADD COLUMN IF NOT EXISTS merged_into_lot_id INTEGER REFERENCES lots(id)")
                cur.execute("ALTER TABLE lots ADD COLUMN IF NOT EXISTS merged_at TIMESTAMPTZ")
                cur.execute("ALTER TABLE lots ADD COLUMN IF NOT EXISTS merge_reason TEXT")
                conn.commit()
                logger.info("Migration 008: lot merge columns up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 008 warning (non-fatal): {e}")

    # Migration 009: Reclassify internal packing transactions
    # Before /pack/commit existed, internal packing was done via /ship/commit
    # with customer_name='Internal Packaging'. Fix those to type='pack'.
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT to_regclass('public.ledger_corrections')")
                append_only_installed = cur.fetchone()[0] is not None
                if append_only_installed:
                    updated = 0
                    logger.info("Migration 009: skipped; append-only ledger is installed")
                else:
                    cur.execute("""
                        UPDATE transactions
                        SET type = 'pack'
                        WHERE type = 'ship'
                          AND LOWER(COALESCE(customer_name, '')) = 'internal packaging'
                    """)
                    updated = cur.rowcount
                conn.commit()
                if updated > 0:
                    logger.info(f"Migration 009: reclassified {updated} internal packaging transactions from 'ship' to 'pack'")
                else:
                    logger.info("Migration 009: no internal packaging transactions to reclassify")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 009 warning (non-fatal): {e}")

    # Migration 010: Customer aliases table
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS customer_aliases (
                        id SERIAL PRIMARY KEY,
                        customer_id INTEGER NOT NULL REFERENCES customers(id) ON DELETE CASCADE,
                        alias TEXT NOT NULL,
                        created_at TIMESTAMPTZ DEFAULT now()
                    )
                """)
                cur.execute("""
                    CREATE UNIQUE INDEX IF NOT EXISTS idx_customer_aliases_lower_alias
                        ON customer_aliases (LOWER(alias))
                """)
                # Seed known customer aliases
                cur.execute("""
                    INSERT INTO customer_aliases (customer_id, alias)
                    SELECT c.id, alias_name
                    FROM customers c,
                         (VALUES ('Setton Farms', 'Setton International'),
                                 ('Setton Farms', 'Setton Intl'),
                                 ('QUALI-PACK USA', 'Quali-Pack'),
                                 ('QUALI-PACK USA', 'Quali Pack'),
                                 ('QUALI-PACK USA', 'QualiPack')
                         ) AS seed(canonical, alias_name)
                    WHERE LOWER(c.name) = LOWER(seed.canonical)
                      AND NOT EXISTS (
                          SELECT 1 FROM customer_aliases ca
                          WHERE ca.customer_id = c.id AND LOWER(ca.alias) = LOWER(seed.alias_name)
                      )
                """)
                seeded = cur.rowcount
                conn.commit()
                if seeded > 0:
                    logger.info(f"Migration 010: customer_aliases table up to date, seeded {seeded} alias(es)")
                else:
                    logger.info("Migration 010: customer_aliases table up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 010 warning (non-fatal): {e}")


    # Migration 011: Supplier lot code fields + lot_supplier_codes table
    # Adds supplier_lot_code, lot_type, received_at to lots table
    # Creates lot_supplier_codes table for commingled receipt breakdowns
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("ALTER TABLE lots ADD COLUMN IF NOT EXISTS supplier_lot_code TEXT")
                cur.execute("ALTER TABLE lots ADD COLUMN IF NOT EXISTS lot_type TEXT")
                cur.execute("ALTER TABLE lots ADD COLUMN IF NOT EXISTS received_at TIMESTAMPTZ")
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS lot_supplier_codes (
                        id SERIAL PRIMARY KEY,
                        lot_id INTEGER NOT NULL REFERENCES lots(id) ON DELETE CASCADE,
                        supplier_lot_code TEXT,
                        supplier_name TEXT,
                        quantity_lb NUMERIC,
                        notes TEXT,
                        created_at TIMESTAMPTZ DEFAULT now()
                    )
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_lot_supplier_codes_lot_id
                        ON lot_supplier_codes (lot_id)
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_lot_supplier_codes_supplier_lot
                        ON lot_supplier_codes (LOWER(supplier_lot_code))
                """)
                cur.execute("""
                    CREATE INDEX IF NOT EXISTS idx_lots_supplier_lot_code
                        ON lots (LOWER(supplier_lot_code))
                """)
                conn.commit()
                logger.info("Migration 011: supplier lot columns and lot_supplier_codes table up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 011 warning (non-fatal): {e}")

    # Migration 012: Add parent_batch_product_id to products table for pack safeguard
    # Links FG products to their expected source batch product
    try:
        conn = db_pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute("ALTER TABLE products ADD COLUMN IF NOT EXISTS parent_batch_product_id INTEGER REFERENCES products(id)")
                # Populate known FG → batch mappings
                mappings = [
                    (152, 123),   # BS Almond Butter 6x7 OZ (70079) → Batch BS Almond Butter (95001)
                    (207, 123),   # BS Almond Butter 6x8 OZ (70086) → Batch BS Almond Butter (95001)
                    (150, 121),   # BS Dark Chocolate 6x7 OZ (70074) → Batch BS Dark Chocolate (95002)
                    (208, 121),   # BS Dark Chocolate 6x8 OZ (70087) → Batch BS Dark Chocolate (95002)
                    (153, 124),   # BS Hazelnut Butter 6x7 OZ (70080) → Batch BS Hazelnut Butter (95003)
                    (206, 124),   # BS Hazelnut Butter 6x8 OZ (70085) → Batch BS Hazelnut Butter (95003)
                    (151, 122),   # BS PB Banana 6x7 OZ (70073) → Batch BS PB Banana (95005)
                    (209, 122),   # BS PB Banana 6x8 OZ (70088) → Batch BS PB Banana (95005)
                ]
                for fg_id, batch_id in mappings:
                    cur.execute(
                        "UPDATE products SET parent_batch_product_id = %s WHERE id = %s AND parent_batch_product_id IS NULL",
                        (batch_id, fg_id)
                    )
                conn.commit()
                logger.info("Migration 012: parent_batch_product_id column and FG→batch mappings up to date")
        finally:
            db_pool.putconn(conn)
    except Exception as e:
        logger.warning(f"Migration 012 warning (non-fatal): {e}")


@app.on_event("shutdown")
async def shutdown():
    global db_pool
    if db_pool:
        db_pool.closeall()
        logger.info("Database connection pool closed")


@contextmanager
def get_db_connection():
    conn = db_pool.getconn()
    try:
        yield conn
        conn.commit()
    except Exception as exc:
        if _is_readonly_error(exc):
            _discard_readonly_connection(conn, "get_db_connection")
            conn = None
            raise
        conn.rollback()
        raise
    finally:
        if conn is not None:
            db_pool.putconn(conn)


@contextmanager
def get_transaction():
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            yield cur


READONLY_PROBE_SQL = """
SELECT
  current_setting('default_transaction_read_only') AS default_ro,
  current_setting('transaction_read_only')          AS txn_ro,
  inet_server_addr()::text                          AS server_ip,
  current_database()                                AS db,
  current_user                                      AS usr,
  pg_is_in_recovery()                               AS is_replica,
  version()                                         AS pg_version
"""


def _is_readonly_error(exc: Exception) -> bool:
    return "read-only transaction" in str(exc).lower()


class ReadOnlyRecoveryError(RuntimeError):
    """Raised when a fresh connection still proves read-only before retry."""


READONLY_WRITABLE_CHECK_SQL = """
SELECT
  current_setting('transaction_read_only') AS txn_ro,
  pg_is_in_recovery()                      AS is_replica
"""


def _discard_readonly_connection(conn, context: str):
    """Remove a known/suspect read-only connection from the psycopg2 pool."""
    try:
        conn.rollback()
    except Exception as rollback_exc:
        logger.warning(
            f"READONLY_RECOVERY_ROLLBACK_FAILED: context={context} error={rollback_exc}"
        )
    try:
        db_pool.putconn(conn, close=True)
        logger.warning(f"READONLY_RECOVERY_DISCARD: context={context} close=True")
    except Exception as put_exc:
        logger.error(f"READONLY_RECOVERY_DISCARD_FAILED: context={context} error={put_exc}")


def _verify_connection_writable(conn, context: str) -> dict:
    """Verify a fresh connection is on a writable primary before retrying a write."""
    with conn.cursor(cursor_factory=RealDictCursor) as cur:
        cur.execute(READONLY_WRITABLE_CHECK_SQL)
        row = dict(cur.fetchone() or {})
    conn.rollback()
    txn_ro = str(row.get("txn_ro", "")).lower()
    is_replica = bool(row.get("is_replica"))
    if txn_ro in ("on", "true", "1") or is_replica:
        raise ReadOnlyRecoveryError(
            f"fresh database connection is still in a read-only transaction; context={context}; "
            f"txn_ro={row.get('txn_ro')}; is_replica={row.get('is_replica')}"
        )
    return row


def _run_db_write(operation_name: str, work: Callable[[Any], Any], *, verify_writable: bool = False):
    conn = db_pool.getconn()
    try:
        if verify_writable:
            state = _verify_connection_writable(conn, operation_name)
            logger.info(
                "READONLY_RECOVERY_WRITABLE_CHECK: "
                + json.dumps({"operation": operation_name, "state": state}, default=str)
            )
        result = work(conn)
        conn.commit()
        return result
    except Exception as exc:
        if _is_readonly_error(exc):
            _discard_readonly_connection(conn, operation_name)
            conn = None
            raise
        conn.rollback()
        raise
    finally:
        if conn is not None:
            db_pool.putconn(conn)


def run_idempotent_write_with_readonly_retry(operation_name: str, work: Callable[[Any], Any]):
    """Retry an explicitly safe/idempotent write once after discarding read-only conn."""
    try:
        return _run_db_write(operation_name, work)
    except Exception as exc:
        if not _is_readonly_error(exc):
            raise
        logger.warning(f"READONLY_RECOVERY_RETRY_ONCE: operation={operation_name}")
        return _run_db_write(operation_name, work, verify_writable=True)


def _capture_readonly_diagnostics() -> dict:
    """Probe DB session state on a fresh pooled connection.
    Wrapped in its own try/except so a probe failure can't mask the original error."""
    conn = None
    try:
        conn = db_pool.getconn()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(READONLY_PROBE_SQL)
            row = cur.fetchone()
        conn.rollback()
        return dict(row) if row else {"probe_error": "no row"}
    except Exception as probe_exc:
        return {"probe_error": str(probe_exc)}
    finally:
        if conn is not None:
            try:
                db_pool.putconn(conn)
            except Exception:
                pass


# ─────────────────────────────────────────────────────────────────
# API-key authorization
#
# Two keys:
#   * API_KEY (master)          — full access to every authenticated route.
#   * DASHBOARD_API_KEY (scoped) — accepted ONLY on (METHOD, route-template)
#     pairs in DASHBOARD_KEY_ALLOWLIST; any other route returns 403.
#
# The allowlist is keyed on the *route template* (request.scope["route"].path,
# e.g. "/sales/orders/{order_id}"), not the raw URL, so path params can't be
# used to smuggle a request onto an unlisted route.
# ─────────────────────────────────────────────────────────────────

# Read-only GET routes the browser dashboard uses, plus the specific
# ship/receive endpoints. Anything not listed here (admin/*, /make, /pack,
# /adjust, /void, deletes, migrations, etc.) is master-key only.
DASHBOARD_KEY_ALLOWLIST = frozenset({
    # Legacy dashboard summaries
    ("GET", "/dashboard/inventory"),
    ("GET", "/dashboard/low-stock"),
    ("GET", "/dashboard/today"),
    ("GET", "/dashboard/lots"),
    ("GET", "/dashboard/production"),
    ("GET", "/sales/dashboard"),
    # Products / inventory / lots (read-only)
    ("GET", "/products/search"),
    ("GET", "/products/{product_id}"),
    ("GET", "/inventory/current"),
    ("GET", "/inventory/lookup"),
    ("GET", "/inventory/{item_name}"),
    ("GET", "/lots/by-code/{lot_code}"),
    ("GET", "/lots/by-supplier-lot/{supplier_lot_code}"),
    ("GET", "/lots/{lot_id}"),
    ("GET", "/bom/products"),
    ("GET", "/bom/batches/{batch_id}/formula"),
    ("GET", "/reason-codes"),
    # History / traceability (read-only; traceability.html, process-flow.html)
    ("GET", "/transactions/history"),
    # FR-12 Recent Entries: authenticated global audit feed for the dashboard.
    ("GET", "/ledger/recent"),
    ("GET", "/trace/batch/{lot_code}"),
    ("GET", "/trace/ingredient/{lot_code}"),
    ("GET", "/trace/supplier-lot/{supplier_lot_code}"),
    ("GET", "/records/late"),
    ("GET", "/records/late.csv"),
    ("GET", "/records/certifications/{business_date}"),
    # Customers / sales orders (read-only + the order edits the dashboard does)
    ("GET", "/customers"),
    ("GET", "/customers/search"),
    ("GET", "/sales/orders"),
    ("GET", "/sales/orders/fulfillment-check"),
    ("GET", "/sales/orders/{order_id}"),
    ("GET", "/sales/orders/{order_id}/allocations"),
    ("GET", "/export/orders-matrix.xlsx"),
    ("PATCH", "/sales/orders/{order_id}"),
    ("PATCH", "/sales/orders/{order_id}/status"),
    ("PATCH", "/sales/orders/{order_id}/lines/{line_id}/update"),
    ("POST", "/sales/orders/{order_id}/allocations"),
    ("POST", "/sales/orders/{order_id}/allocations/{allocation_id}/release"),
    ("PATCH", "/lots/{lot_id}/received-at"),
    ("POST", "/sales-orders/{so_number}/ready"),
    # Production planning (read-only)
    ("GET", "/production/requirements"),
    ("GET", "/production/day-summary"),
    ("GET", "/production/today-tile"),
    # Dashboard notes (the dashboard's own CRUD; GET is public)
    ("POST", "/dashboard/api/notes"),
    ("PUT", "/dashboard/api/notes/{note_id}"),
    ("PUT", "/dashboard/api/notes/{note_id}/toggle"),
    ("DELETE", "/dashboard/api/notes/{note_id}"),
    # Ship / receive — the specific endpoints the dashboard is allowed to hit
    ("POST", "/sales/orders/{order_id}/ship/preview"),
    ("POST", "/sales/orders/{order_id}/ship/commit"),
    ("POST", "/receive/preview"),
    ("POST", "/receive/commit"),
    # FR-2 expected receipts + supplier list (dashboard Expected Receipts tab)
    ("GET", "/suppliers"),
    ("POST", "/suppliers"),
    ("GET", "/expected-receipts"),
    ("POST", "/expected-receipts"),
    ("GET", "/expected-receipts/{expected_receipt_id}"),
    ("PATCH", "/expected-receipts/{expected_receipt_id}"),
    # Supplies (dashboard-only: packaging/consumables inventory + request queue)
    ("GET", "/supplies/inventory"),
    ("GET", "/supplies/inventory/{product_id}/lots"),
    ("GET", "/supply-requests"),
    ("POST", "/supply-requests"),
    ("PATCH", "/supply-requests/{supply_request_id}"),
})


def _route_key(request: Request):
    """(METHOD, route template) for the matched route, e.g. ("GET", "/lots/{lot_id}")."""
    route = request.scope.get("route")
    path = getattr(route, "path", None) or request.url.path
    return (request.method.upper(), path)


def _authorize_api_key(provided_key: str, request: Request, invalid_status: int = 403) -> bool:
    """Shared check for both dependencies. Master key -> always OK. Dashboard
    key -> OK only if the matched route is on DASHBOARD_KEY_ALLOWLIST."""
    if not provided_key:
        raise HTTPException(status_code=401, detail="API key required")
    if secrets.compare_digest(provided_key, API_KEY):
        return True
    if DASHBOARD_API_KEY and secrets.compare_digest(provided_key, DASHBOARD_API_KEY):
        if _route_key(request) in DASHBOARD_KEY_ALLOWLIST:
            return True
        raise HTTPException(status_code=403, detail="API key not authorized for this endpoint")
    raise HTTPException(status_code=invalid_status, detail="Invalid API key")


def verify_api_key(request: Request, x_api_key: str = Header(None, alias="X-API-Key")):
    return _authorize_api_key(x_api_key, request, invalid_status=403)


def verify_api_key_flexible(
    request: Request,
    x_api_key: str = Header(None, alias="X-API-Key"),
    key: str = Query(None)
):
    """Accept API key from either header or query parameter (packing slip browser access)."""
    return _authorize_api_key(x_api_key or key, request, invalid_status=401)


def resolve_order_id(order_id: str = Path(...)) -> int:
    """Accept either numeric DB id or order_number string (e.g. 'SO-260323-001').
    Returns the integer DB id."""
    try:
        return int(order_id)
    except ValueError:
        pass
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT id FROM sales_orders WHERE order_number = %s", (order_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(
                    status_code=404,
                    detail={
                        "error_code": "ORDER_NOT_FOUND",
                        "message": f"Order '{order_id}' not found",
                        "input": order_id,
                        "suggestions": [],
                    }
                )
            return row['id']


def format_timestamp(dt):
    if dt is None:
        return None, None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    local_dt = dt.astimezone(PLANT_TIMEZONE)
    return local_dt.strftime("%Y-%m-%d"), local_dt.strftime("%I:%M %p") + f" {TIMEZONE_LABEL}"


def get_plant_now():
    return datetime.now(PLANT_TIMEZONE)


INVENTORY_OCCURRED_AT_FUTURE_GRACE = timedelta(minutes=5)
INVENTORY_OCCURRED_AT_STANDARD_WINDOW = timedelta(days=14)
INVENTORY_BACKFILL_SOURCE = "api_backfill"


def validate_inventory_occurred_at(
    occurred_at: Optional[datetime], backfill: bool = False
) -> tuple[Optional[datetime], Optional[str]]:
    """Normalize and validate an optional inventory event timestamp.

    Offset-free ISO timestamps are plant-local. Offset-aware timestamps are
    converted to the plant timezone before comparison/storage. A NULL return
    deliberately preserves the legacy insert path, where migration 039 fills
    occurred_at from the transaction timestamp.
    """
    if occurred_at is None:
        return None, None

    if occurred_at.tzinfo is None:
        event_time = occurred_at.replace(tzinfo=PLANT_TIMEZONE)
    else:
        event_time = occurred_at.astimezone(PLANT_TIMEZONE)

    now = get_plant_now()
    if event_time > now + INVENTORY_OCCURRED_AT_FUTURE_GRACE:
        raise HTTPException(
            status_code=400,
            detail={
                "error_code": "OCCURRED_AT_IN_FUTURE",
                "message": "occurred_at cannot be more than 5 minutes in the future.",
            },
        )
    if event_time < now - INVENTORY_OCCURRED_AT_STANDARD_WINDOW and not backfill:
        raise HTTPException(
            status_code=400,
            detail={
                "error_code": "OCCURRED_AT_BACKFILL_REQUIRED",
                "message": (
                    "occurred_at is more than 14 days old. "
                    "Set backfill=true to record an intentional historical entry."
                ),
            },
        )

    return event_time, INVENTORY_BACKFILL_SOURCE if backfill else None


def generate_confirmation_code(transaction_id: int) -> str:
    """Generate a short unique confirmation code from a transaction ID."""
    import hashlib
    hash_input = f"txn-{transaction_id}-cns"
    short_hash = hashlib.sha256(hash_input.encode()).hexdigest()[:6].upper()
    return f"TXN-{short_hash}"


def get_daily_production_summary(cur, target_date=None):
    """Query all make+pack transactions for a given date (plant timezone).
    Returns dict with 'production' list and 'adjustments' list."""
    if target_date is None:
        target_date = get_plant_now().date()

    # Build timezone-aware start/end for the target date
    day_start = datetime(target_date.year, target_date.month, target_date.day,
                         tzinfo=PLANT_TIMEZONE)
    day_end = day_start + timedelta(days=1)

    # Production + packing summary
    cur.execute("""
        SELECT p.name as product_name, p.type as product_type,
               t.type as transaction_type,
               SUM(tl.quantity_lb) FILTER (WHERE tl.quantity_lb > 0) as output_lb,
               COUNT(DISTINCT t.id) as transaction_count
        FROM ledger_current_transactions t
        JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
        JOIN products p ON p.id = tl.product_id
        WHERE t.type IN ('make', 'pack')
          AND t.effective_status = 'posted'
          AND t.timestamp >= %s AND t.timestamp < %s
        GROUP BY p.id, p.name, p.type, t.type
        ORDER BY t.type, p.name
    """, (day_start, day_end))
    prod_rows = cur.fetchall()

    production = [
        {
            "product_name": r['product_name'],
            "product_type": r['product_type'],
            "transaction_type": r['transaction_type'],
            "total_lb": float(r['output_lb'] or 0),
            "transaction_count": r['transaction_count']
        }
        for r in prod_rows
    ]

    # Adjustments for the day
    cur.execute("""
        SELECT p.name as product_name, l.lot_code,
               tl.quantity_lb as adjustment_lb,
               t.adjust_reason as reason
        FROM ledger_current_transactions t
        JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
        JOIN products p ON p.id = tl.product_id
        JOIN lots l ON l.id = tl.lot_id
        WHERE t.type = 'adjust'
          AND t.effective_status = 'posted'
          AND t.timestamp >= %s AND t.timestamp < %s
        ORDER BY t.timestamp
    """, (day_start, day_end))
    adj_rows = cur.fetchall()

    adjustments = [
        {
            "product_name": r['product_name'],
            "lot_code": r['lot_code'],
            "adjustment_lb": float(r['adjustment_lb']),
            "reason": r['reason']
        }
        for r in adj_rows
    ]

    return {"production": production, "adjustments": adjustments}


# ═══════════════════════════════════════════════════════════════
# BILINGUAL SUPPORT HELPERS
# ═══════════════════════════════════════════════════════════════

def validate_bilingual(english_val, spanish_val, field_name: str):
    """Validate bilingual field pair: English required when Spanish is provided."""
    if spanish_val and not english_val:
        raise HTTPException(400,
            f"English version required. Provide '{field_name}' along with '{field_name}_es'."
        )


def bilingual_response(english_val, spanish_val, field_name: str) -> dict:
    """Return bilingual fields for a response dict. Only includes _es if it has a value."""
    result = {field_name: english_val}
    if spanish_val:
        result[f"{field_name}_es"] = spanish_val
    return result


# ═══════════════════════════════════════════════════════════════
# PYDANTIC MODELS
# ═══════════════════════════════════════════════════════════════

class CommandRequest(BaseModel):
    raw_text: str


class InventoryWriteRequest(BaseModel):
    occurred_at: Optional[datetime] = None
    backfill: bool = False


class ReceiveRequest(InventoryWriteRequest):
    mode: Literal["preview", "commit"] = "preview"
    product_name: str
    cases: int
    case_size_lb: float
    shipper_name: str
    bol_reference: str
    shipper_code_override: Optional[str] = None
    # Lot Identity Policy: If lot_code is provided, find-or-create by (product_id, lot_code).
    # Only auto-generate if lot_code is omitted.
    lot_code: Optional[str] = None
    # LAT Code Policy v1.1 fields
    supplier_lot_code: Optional[str] = None
    lot_type: Optional[str] = None  # "single_supplier" or "commingled"
    supplier_lot_entries: Optional[List[Dict]] = None  # For commingled receipts

class SupplierCreate(BaseModel):
    name: str
    active: bool = True


class ExpectedReceiptCreate(BaseModel):
    """FR-2: one expected delivery. supplier_name is resolved server-side against
    the suppliers table (case/whitespace-insensitive); it is NEVER auto-created.
    created_by is an interim (pre-FR-15) caller source tag — see caller_source_tag()."""
    product_id: Optional[int] = None
    product_name: Optional[str] = None
    supplier_name: str
    expected_qty: float  # lb
    expected_date: Optional[date] = None
    reference_number: Optional[str] = None
    notes: Optional[str] = None
    created_by: Optional[str] = None


class ExpectedReceiptUpdate(BaseModel):
    """PATCH body. Omitted fields are untouched; fields sent as null are cleared
    (expected_date / reference_number / notes). status may only move to
    'closed' or 'cancelled', and only while the record is open."""
    expected_qty: Optional[float] = None
    expected_date: Optional[date] = None
    reference_number: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[Literal["closed", "cancelled"]] = None


class SupplyRequestCreate(BaseModel):
    """Supplies: one "we need X" request from the floor/office. Exactly one of
    product_id (a catalogued product) or item_text (free text for anything not
    in products) — never both, never neither. qty is optional (product uom /
    whatever the text implies). requested_by is a plain name/source tag."""
    product_id: Optional[int] = None
    item_text: Optional[str] = None
    qty: Optional[float] = None
    note: Optional[str] = None
    requested_by: str


class SupplyRequestUpdate(BaseModel):
    """PATCH body. The only transition is open -> done (sets done_at)."""
    status: Literal["done"]


class ShipRequest(InventoryWriteRequest):
    mode: Literal["preview", "commit"] = "preview"
    product_name: str
    quantity_lb: float
    customer_name: str
    customer_address: Optional[str] = None
    order_reference: str
    lot_code: Optional[str] = None
    force_standalone: bool = False
    force_create_customer: bool = False

class MakeRequest(InventoryWriteRequest):
    mode: Literal["preview", "commit"] = "preview"
    product_name: str
    batches: int
    # Lot Identity Policy: If lot_code is provided, find-or-create by (product_id, lot_code).
    # Only auto-generate if lot_code is omitted.
    lot_code: Optional[str] = None
    ingredient_lot_overrides: Optional[Union[Dict[str, str], str]] = None
    excluded_ingredients: Optional[List[int]] = None
    confirmed_sku: Optional[bool] = None  # Must be True when sibling SKUs exist
    
    def get_lot_overrides(self) -> Optional[Dict[str, str]]:
        """Parse ingredient_lot_overrides whether it's a dict or JSON string"""
        if self.ingredient_lot_overrides is None:
            return None
        if isinstance(self.ingredient_lot_overrides, str):
            try:
                parsed = json.loads(self.ingredient_lot_overrides)
                if isinstance(parsed, dict):
                    return parsed
                return None
            except (json.JSONDecodeError, TypeError):
                return None
        return self.ingredient_lot_overrides

class PackLotAllocation(BaseModel):
    lot_code: str
    quantity_lb: float

class PackRequest(InventoryWriteRequest):
    mode: Literal["preview", "commit"] = "preview"
    source_product: str          # Batch product name or code (e.g., "Batch Classic Granola #9" or "90002")
    target_product: str          # Finished good name or code (e.g., "CQ Granola 10 LB" or "1614")
    cases: int
    case_weight_lb: Optional[float] = None  # Override; defaults to target product's case_size_lb
    lot_allocations: Optional[List[PackLotAllocation]] = None  # Explicit lot splits; FIFO if omitted
    # Lot Identity Policy: If target_lot_code is provided, find-or-create by (product_id, lot_code).
    # Only auto-generate (inherit from batch lot) if target_lot_code is omitted.
    target_lot_code: Optional[str] = None

class AdjustRequest(InventoryWriteRequest):
    mode: Literal["preview", "commit"] = "preview"
    product_name: str
    lot_code: str
    adjustment_lb: float
    reason: str
    reason_es: Optional[str] = None

class QuickCreateProductRequest(BaseModel):
    product_name: str
    product_type: str
    uom: str = "lb"
    storage_type: str = "ambient"
    name_confidence: str = "exact"
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    performed_by: str = "system"

class QuickCreateBatchProductRequest(BaseModel):
    product_name: str
    category: str
    production_context: str
    name_confidence: str = "exact"
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    performed_by: str = "system"

class LotReassignmentRequest(BaseModel):
    to_product_id: int
    reason_code: str
    reason_notes: Optional[str] = None
    reason_notes_es: Optional[str] = None
    performed_by: str = "system"

class AddFoundInventoryRequest(InventoryWriteRequest):
    product_id: int
    quantity: float
    uom: str = "lb"
    reason_code: str
    found_location: Optional[str] = None
    estimated_age: str = "unknown"
    suspected_supplier: Optional[str] = None
    suspected_bol: Optional[str] = None
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    performed_by: str = "system"
    # Lot Identity Policy: If lot_code is provided, find-or-create by (product_id, lot_code).
    # Only auto-generate if lot_code is omitted.
    lot_code: Optional[str] = None

class AddFoundInventoryWithNewProductRequest(InventoryWriteRequest):
    product_name: str
    product_type: str
    quantity: float
    reason_code: str
    uom: str = "lb"
    storage_type: str = "ambient"
    found_location: Optional[str] = None
    estimated_age: str = "unknown"
    suspected_supplier: Optional[str] = None
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    performed_by: str = "system"
    # Lot Identity Policy: If lot_code is provided, find-or-create by (product_id, lot_code).
    # Only auto-generate if lot_code is omitted.
    lot_code: Optional[str] = None

class VerifyProductRequest(BaseModel):
    action: str
    verified_name: Optional[str] = None
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    performed_by: str = "system"


# ═══════════════════════════════════════════════════════════════
# NOTES / TO-DOS / REMINDERS PYDANTIC MODELS
# ═══════════════════════════════════════════════════════════════

class NoteCreate(BaseModel):
    category: str  # 'note', 'todo', 'reminder'
    title: str
    body: Optional[str] = ""
    priority: Optional[str] = "normal"  # 'low', 'normal', 'high'
    due_date: Optional[str] = None  # YYYY-MM-DD
    entity_type: Optional[str] = None  # 'product', 'lot', 'customer', 'supplier'
    entity_id: Optional[str] = None

    @validator("category")
    def validate_category(cls, v):
        if v not in ("note", "todo", "reminder"):
            raise ValueError("category must be 'note', 'todo', or 'reminder'")
        return v

    @validator("priority")
    def validate_priority(cls, v):
        if v not in ("low", "normal", "high"):
            raise ValueError("priority must be 'low', 'normal', or 'high'")
        return v

class NoteUpdate(BaseModel):
    title: Optional[str] = None
    body: Optional[str] = None
    priority: Optional[str] = None
    status: Optional[str] = None  # 'open', 'done', 'dismissed'
    due_date: Optional[str] = None  # YYYY-MM-DD or empty string to clear
    entity_type: Optional[str] = None
    entity_id: Optional[str] = None

    @validator("priority")
    def validate_priority(cls, v):
        if v is not None and v not in ("low", "normal", "high"):
            raise ValueError("priority must be 'low', 'normal', or 'high'")
        return v

    @validator("status")
    def validate_status(cls, v):
        if v is not None and v not in ("open", "done", "dismissed"):
            raise ValueError("status must be 'open', 'done', or 'dismissed'")
        return v


class SalesOrderReadyFlagRequest(BaseModel):
    ready: bool
    by: Optional[str] = "floor"
    note: Optional[str] = None


# ═══════════════════════════════════════════════════════════════
# SALES PYDANTIC MODELS (v2.3.0)
# ═══════════════════════════════════════════════════════════════

class CustomerCreate(BaseModel):
    name: str
    contact_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    notes_es: Optional[str] = None

class CustomerUpdate(BaseModel):
    name: Optional[str] = None
    contact_name: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    active: Optional[bool] = None
    aliases: Optional[List[str]] = None

class OrderLineInput(BaseModel):
    product_name: str
    quantity: Optional[float] = None
    unit: Optional[str] = None               # None = not explicitly provided
    case_weight_lb: Optional[float] = None
    quantity_lb: Optional[float] = None
    unit_price: Optional[float] = None
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    _unit_explicitly_set: bool = False        # internal tracking flag

    class Config:
        underscore_attrs_are_private = True

    @validator('unit', pre=True, always=True)
    def track_unit_explicit(cls, v):
        return v  # actual tracking done in calculate_quantity_lb

    @validator('quantity_lb', always=True, pre=True)
    def calculate_quantity_lb(cls, v, values):
        unit = values.get('unit')
        quantity = values.get('quantity')
        case_weight = values.get('case_weight_lb')

        # If quantity_lb is directly provided, use it (backward compatible)
        if v is not None:
            return v

        # Default unit to lb if not provided
        if unit is None:
            unit = 'lb'

        # If quantity + unit provided, calculate
        if quantity is not None:
            if unit == 'lb':
                return quantity
            elif unit in ('cases', 'bags', 'boxes'):
                if case_weight is None:
                    raise ValueError(f"case_weight_lb is required when unit is '{unit}'")
                return quantity * case_weight
            else:
                raise ValueError(f"Unknown unit: {unit}. Use 'lb', 'cases', 'bags', or 'boxes'")

        raise ValueError("Either quantity_lb or (quantity + unit) must be provided")

class OrderCreate(BaseModel):
    customer_name: str
    customer_address: Optional[str] = None
    requested_ship_date: Optional[str] = None
    lines: List[OrderLineInput]
    notes: Optional[str] = None
    notes_es: Optional[str] = None

class OrderStatusUpdate(BaseModel):
    status: str

class OrderHeaderUpdate(BaseModel):
    requested_ship_date: Optional[str] = None
    notes: Optional[str] = None
    notes_es: Optional[str] = None
    customer_id: Optional[int] = None

class AddOrderLines(BaseModel):
    lines: List[OrderLineInput]

class ShipOrderLineRequest(BaseModel):
    line_id: int
    quantity_lb: float

class ShipOrderRequest(InventoryWriteRequest):
    mode: Literal["preview", "commit"] = "preview"
    ship_all: Optional[bool] = False
    lines: Optional[List[ShipOrderLineRequest]] = None


class SalesOrderAllocationCreate(BaseModel):
    """Dashboard allocation write. ``auto_fifo`` chooses/pins lots; manual
    accepts an optional lot_id (NULL means SKU-level)."""
    mode: Literal["manual", "auto_fifo"] = "manual"
    line_id: int
    quantity_lb: Optional[float] = None
    lot_id: Optional[int] = None
    source: Optional[Literal["manual", "staged_lot"]] = None
    expires_at: Optional[datetime] = None
    note: Optional[str] = None

class CommitShipOrderRequest(InventoryWriteRequest):
    """Mode-less request for the dedicated sales-order shipment commit route."""
    ship_all: Optional[bool] = False
    lines: Optional[List[ShipOrderLineRequest]] = None

    class Config:
        extra = "forbid"


# ═══════════════════════════════════════════════════════════════
# SALES HELPER FUNCTIONS (v2.3.0)
# ═══════════════════════════════════════════════════════════════

def resolve_product_id(cur, product_name: str) -> tuple:
    """Find product by name using 3-tier search. Returns (product_id, product_name).
    For sales orders: auto-resolves high confidence, raises on ambiguous/none."""
    results = _tiered_product_search(cur, product_name, limit=5)
    if not results:
        raise HTTPException(
            status_code=404,
            detail={
                "error_code": "PRODUCT_NOT_FOUND",
                "message": f"Product not found: '{product_name}'",
                "input": product_name,
                "suggestions": [],
            }
        )

    best = results[0]
    tier = best['match_tier']

    # High confidence: exact or single keyword match → auto-resolve
    if tier == 'exact' or (tier == 'keyword' and len(results) == 1):
        return best['id'], best['name']

    # Multiple keyword matches → ambiguous
    if tier == 'keyword' and len(results) > 1:
        suggestions = [
            {"product_id": r['id'], "name": r['name'], "odoo_code": r['odoo_code']}
            for r in results
        ]
        raise HTTPException(
            status_code=409,
            detail={
                "error_code": "PRODUCT_AMBIGUOUS",
                "message": f"Multiple products match '{product_name}'.",
                "input": product_name,
                "suggestions": suggestions,
            }
        )

    # Trigram: accept if similarity is high enough and only one strong match
    if tier == 'trigram':
        if best['similarity'] > 0.4 and (len(results) == 1 or results[0]['similarity'] - results[1]['similarity'] > 0.15):
            return best['id'], best['name']
        suggestions = [
            {"product_id": r['id'], "name": r['name'], "odoo_code": r['odoo_code'],
             "similarity": round(float(r['similarity']), 2)}
            for r in results
        ]
        raise HTTPException(
            status_code=409,
            detail={
                "error_code": "PRODUCT_UNCERTAIN",
                "message": f"No confident match for '{product_name}'.",
                "input": product_name,
                "suggestions": suggestions,
            }
        )

    raise HTTPException(
        status_code=404,
        detail={
            "error_code": "PRODUCT_NOT_FOUND",
            "message": f"Product not found: '{product_name}'",
            "input": product_name,
            "suggestions": [],
        }
    )


def resolve_product_full(cur, product_name: str) -> dict:
    """Find product by name/odoo_code using 3-tier search. Returns full row dict.
    Used by receive/ship/make endpoints that need extra columns."""
    results = _tiered_product_search(cur, product_name, limit=5)
    if not results:
        raise HTTPException(
            status_code=404,
            detail={
                "error_code": "PRODUCT_NOT_FOUND",
                "message": f"Product not found: '{product_name}'",
                "input": product_name,
                "suggestions": [],
            }
        )

    best = results[0]
    tier = best['match_tier']

    # High confidence: exact or single keyword match → auto-resolve
    if tier == 'exact' or (tier == 'keyword' and len(results) == 1):
        product_id = best['id']
    elif tier == 'keyword' and len(results) > 1:
        suggestions = [
            {"product_id": r['id'], "name": r['name'], "odoo_code": r['odoo_code']}
            for r in results
        ]
        raise HTTPException(
            status_code=409,
            detail={
                "error_code": "PRODUCT_AMBIGUOUS",
                "message": f"Multiple products match '{product_name}'.",
                "input": product_name,
                "suggestions": suggestions,
            }
        )
    elif tier == 'trigram':
        if best['similarity'] > 0.4 and (len(results) == 1 or results[0]['similarity'] - results[1]['similarity'] > 0.15):
            product_id = best['id']
        else:
            suggestions = [
                {"product_id": r['id'], "name": r['name'], "odoo_code": r['odoo_code'],
                 "similarity": round(float(r['similarity']), 2)}
                for r in results
            ]
            raise HTTPException(
                status_code=409,
                detail={
                    "error_code": "PRODUCT_UNCERTAIN",
                    "message": f"No confident match for '{product_name}'.",
                    "input": product_name,
                    "suggestions": suggestions,
                }
            )
    else:
        raise HTTPException(
            status_code=404,
            detail={
                "error_code": "PRODUCT_NOT_FOUND",
                "message": f"Product not found: '{product_name}'",
                "input": product_name,
                "suggestions": [],
            }
        )

    # Fetch full product row
    cur.execute(
        """SELECT id, name, odoo_code, default_batch_lb, case_size_lb,
                  COALESCE(yield_multiplier, 1.0) as yield_multiplier,
                  verification_notes, verification_notes_es
           FROM products WHERE id = %s""",
        (product_id,)
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(
            status_code=404,
            detail={
                "error_code": "PRODUCT_NOT_FOUND",
                "message": f"Product not found: '{product_name}'",
                "input": product_name,
                "suggestions": [],
            }
        )
    return dict(row)


# Noise words filtered out during keyword search
_SEARCH_NOISE_WORDS = {'the', 'a', 'an', 'in', 'for', 'lb', 'lbs', 'case', 'cases',
                       'oz', 'bag', 'bags', 'box', 'boxes', 'of', 'and', 'with', 'per'}


def _tiered_product_search(cur, query: str, limit: int = 5) -> list:
    """3-tier product search: exact → keyword → trigram.
    Returns list of dicts with keys: id, name, odoo_code, match_tier, similarity."""
    q = query.strip()
    if not q:
        return []

    # --- Tier 1: Exact match ---
    # Try odoo_code if input looks numeric
    if q.isdigit():
        cur.execute(
            """SELECT id, name, odoo_code FROM products
               WHERE odoo_code = %s AND COALESCE(active, true) = true""",
            (q,)
        )
        rows = cur.fetchall()
        if rows:
            return [dict(r, match_tier='exact', similarity=1.0) for r in rows][:limit]

    # Try exact name match
    cur.execute(
        """SELECT id, name, odoo_code FROM products
           WHERE LOWER(name) = LOWER(%s) AND COALESCE(active, true) = true""",
        (q,)
    )
    rows = cur.fetchall()
    if rows:
        return [dict(r, match_tier='exact', similarity=1.0) for r in rows][:limit]

    # --- Tier 2: Keyword match (word-order independent) ---
    words = [w for w in re.split(r'\s+', q.lower()) if w and w not in _SEARCH_NOISE_WORDS]
    if words:
        patterns = [f"%{w}%" for w in words]
        cur.execute(
            """SELECT id, name, odoo_code FROM products
               WHERE name ILIKE ALL(%s) AND COALESCE(active, true) = true
               ORDER BY length(name), name
               LIMIT %s""",
            (patterns, limit)
        )
        rows = cur.fetchall()
        if rows:
            return [dict(r, match_tier='keyword', similarity=0.8) for r in rows]

    # --- Tier 3: Trigram similarity fallback ---
    cur.execute(
        """SELECT id, name, odoo_code,
                  similarity(LOWER(name), LOWER(%s)) AS sim
           FROM products
           WHERE similarity(LOWER(name), LOWER(%s)) > 0.25
             AND COALESCE(active, true) = true
           ORDER BY sim DESC
           LIMIT %s""",
        (q, q, limit)
    )
    rows = cur.fetchall()
    return [dict(r, match_tier='trigram', similarity=float(r['sim'])) for r in rows]


def _resolve_single_product(cur, raw_name: str) -> dict:
    """Resolve a single raw product name string using 3-tier search.
    Returns a dict with: input, match, match_tier, confidence, alternatives."""
    results = _tiered_product_search(cur, raw_name, limit=5)

    if not results:
        return {
            "input": raw_name,
            "match": None,
            "match_tier": None,
            "confidence": "none",
            "suggestions": []
        }

    best = results[0]
    tier = best['match_tier']
    sim = best['similarity']

    # Determine confidence
    if tier == 'exact':
        confidence = 'high'
    elif tier == 'keyword':
        confidence = 'high' if len(results) == 1 else 'medium'
    else:  # trigram
        if sim > 0.4:
            confidence = 'medium'
        else:
            confidence = 'low'

    match_data = {"id": best['id'], "name": best['name'], "odoo_code": best['odoo_code']}
    result = {
        "input": raw_name,
        "match": match_data,
        "match_tier": tier,
        "confidence": confidence,
    }

    # Include alternatives if there are multiple matches at tier 2/3
    if len(results) > 1 and tier in ('keyword', 'trigram'):
        result["alternatives"] = [
            {"id": r['id'], "name": r['name'], "odoo_code": r['odoo_code']}
            for r in results[1:]
        ]

    return result


class BulkResolveRequest(BaseModel):
    names: List[str]


def get_sibling_skus(cur, product_id: int) -> list:
    """Find other finished-good products that share the exact same BOM ingredients.
    If product A and product B both have identical ingredient sets in batch_formulas,
    they are 'siblings' — same batch source, different labels/packaging.
    Returns list of dicts: [{id, name, odoo_code}, ...] (excluding the given product)."""
    # Get the set of ingredient product IDs for this product
    cur.execute(
        "SELECT ingredient_product_id FROM batch_formulas WHERE product_id = %s ORDER BY ingredient_product_id",
        (product_id,)
    )
    my_ingredients = [row['ingredient_product_id'] for row in cur.fetchall()]

    if not my_ingredients:
        return []  # No formula → no siblings

    # Find all products that have a formula, grouped by their ingredient set
    cur.execute("""
        SELECT bf.product_id, ARRAY_AGG(bf.ingredient_product_id ORDER BY bf.ingredient_product_id) as ingredients
        FROM batch_formulas bf
        JOIN products p ON p.id = bf.product_id AND COALESCE(p.active, true) = true
        WHERE bf.product_id != %s
        GROUP BY bf.product_id
        HAVING ARRAY_AGG(bf.ingredient_product_id ORDER BY bf.ingredient_product_id) = %s::int[]
    """, (product_id, my_ingredients))
    sibling_ids = [row['product_id'] for row in cur.fetchall()]

    if not sibling_ids:
        return []

    cur.execute(
        "SELECT id, name, odoo_code FROM products WHERE id = ANY(%s) ORDER BY name",
        (sibling_ids,)
    )
    return [dict(r) for r in cur.fetchall()]


def _pick_by_address(cur, candidate_ids: list, address: str) -> Optional[dict]:
    """Given >1 candidate customer rows and an incoming address, return the single
    best match by trigram similarity on customers.address — but only if the top
    candidate is clearly ahead. Returns {id, name} or None (no confident pick)."""
    if not candidate_ids or not address or not address.strip():
        return None
    cur.execute(
        """SELECT id, name, COALESCE(address, '') AS address,
                  similarity(LOWER(COALESCE(address, '')), LOWER(%s)) AS addr_sim
           FROM customers
           WHERE id = ANY(%s)
           ORDER BY addr_sim DESC""",
        (address, candidate_ids)
    )
    ranked = cur.fetchall()
    if not ranked:
        return None
    top = ranked[0]
    top_sim = float(top['addr_sim'])
    # Require strong absolute match AND clear gap vs runner-up (or no runner-up)
    if top_sim < 0.6:
        return None
    if len(ranked) > 1:
        second_sim = float(ranked[1]['addr_sim'])
        if top_sim - second_sim < 0.2:
            return None
    return {"id": top['id'], "name": top['name']}


def resolve_customer_id(cur, customer_name: str, auto_create: bool = True,
                        force_create: bool = False, address: Optional[str] = None) -> tuple:
    """Find or create customer by name/alias. Returns (customer_id, canonical_name).

    Resolution order:
    1. Exact match on canonical name
    2. Exact match on alias
    3. Fuzzy LIKE across names + aliases → single match returns, multiple → address
       tiebreaker (if address provided) → still ambiguous → 409
    4. No match → first-word prefix check before auto-create → address tiebreaker
       (if address provided) → still ambiguous → 409
    """
    # Step 1: Exact match on canonical name
    cur.execute(
        "SELECT id, name FROM customers WHERE LOWER(name) = LOWER(%s) AND active = true",
        (customer_name,)
    )
    row = cur.fetchone()
    if row:
        return row['id'], row['name']

    # Step 2: Exact match on alias
    cur.execute(
        """SELECT c.id, c.name FROM customers c
           JOIN customer_aliases ca ON ca.customer_id = c.id
           WHERE LOWER(ca.alias) = LOWER(%s) AND c.active = true""",
        (customer_name,)
    )
    row = cur.fetchone()
    if row:
        return row['id'], row['name']

    # Step 3: Fuzzy match across canonical names AND aliases
    cur.execute(
        """SELECT DISTINCT c.id, c.name FROM customers c
           LEFT JOIN customer_aliases ca ON ca.customer_id = c.id
           WHERE c.active = true
             AND (LOWER(c.name) LIKE LOWER(%s) OR LOWER(ca.alias) LIKE LOWER(%s))
           ORDER BY c.name LIMIT 5""",
        (f"%{customer_name}%", f"%{customer_name}%")
    )
    rows = cur.fetchall()
    if len(rows) == 1:
        return rows[0]['id'], rows[0]['name']
    elif len(rows) > 1:
        picked = _pick_by_address(cur, [r['id'] for r in rows], address) if address else None
        if picked:
            return picked['id'], picked['name']
        suggestions = [{"customer_id": r['id'], "name": r['name']} for r in rows]
        raise HTTPException(
            status_code=409,
            detail={
                "error_code": "CUSTOMER_AMBIGUOUS",
                "message": f"Multiple customers match '{customer_name}'. Please use the exact canonical name.",
                "input": customer_name,
                "suggestions": suggestions,
                "hint": "Set force_create_customer=true to create a new customer, or use an existing name from suggestions."
            }
        )

    # Step 4: No match — check for first-word/prefix collisions before auto-create
    if not force_create:
        first_word = customer_name.strip().split()[0] if customer_name.strip() else ""
        if first_word and len(first_word) >= 3:
            cur.execute(
                """SELECT DISTINCT c.id, c.name FROM customers c
                   LEFT JOIN customer_aliases ca ON ca.customer_id = c.id
                   WHERE c.active = true
                     AND (LOWER(c.name) LIKE LOWER(%s) OR LOWER(ca.alias) LIKE LOWER(%s))
                   LIMIT 5""",
                (f"{first_word}%", f"{first_word}%")
            )
            prefix_rows = cur.fetchall()
            if prefix_rows:
                picked = _pick_by_address(cur, [r['id'] for r in prefix_rows], address) if address else None
                if picked:
                    return picked['id'], picked['name']
                suggestions = [{"customer_id": r['id'], "name": r['name']} for r in prefix_rows]
                raise HTTPException(
                    status_code=409,
                    detail={
                        "error_code": "CUSTOMER_AMBIGUOUS",
                        "message": f"No exact match for '{customer_name}', but similar customers exist. Did you mean one of these?",
                        "input": customer_name,
                        "suggestions": suggestions,
                        "hint": "Set force_create_customer=true to create a new customer, or use an existing name from suggestions."
                    }
                )

    if auto_create or force_create:
        cur.execute(
            "INSERT INTO customers (name, address) VALUES (%s, %s) RETURNING id, name",
            (customer_name, address)
        )
        row = cur.fetchone()
        return row['id'], row['name']
    raise HTTPException(
        status_code=404,
        detail={
            "error_code": "CUSTOMER_NOT_FOUND",
            "message": f"Customer not found: '{customer_name}'",
            "input": customer_name,
            "suggestions": [],
        }
    )


# ═══════════════════════════════════════════════════════════════
# DASHBOARD ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/dashboard/inventory")
def dashboard_inventory(_: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT * FROM inventory_summary WHERE on_hand > 0")
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Dashboard inventory failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/dashboard/low-stock")
def dashboard_low_stock(_: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT * FROM low_stock_alerts")
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Dashboard low-stock failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/dashboard/today")
def dashboard_today(_: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT * FROM todays_transactions")
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Dashboard today failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/dashboard/lots")
def dashboard_lots(_: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT * FROM lot_balances LIMIT 100")
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Dashboard lots failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})

@app.get("/dashboard/production")
def dashboard_production(_: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT * FROM production_history LIMIT 50")
            return cur.fetchall()
    except Exception as e:
        logger.error(f"Dashboard production failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# HEALTH & ROOT ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/")
def root():
    return {
        "name": "Factory Ledger System",
        "version": app.version,
        "status": "online",
        "features": ["receive", "ship", "make", "adjust", "trace", "bom", "quick-create", "lot-reassign", "found-inventory", "ingredient-exclusion", "ingredient-lot-override", "dashboard", "sales-orders", "customers", "fulfillment-check", "bilingual", "sku-disambiguation", "production-scheduling"]
    }


@app.get("/health")
def health_check():
    try:
        with get_transaction() as cur:
            cur.execute("SELECT 1")
        pool_status = f"active, {db_pool.minconn}-{db_pool.maxconn} connections" if db_pool else "not initialized"
        return {"status": "healthy", "database": "connected", "pool": pool_status}
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(status_code=500, content={"status": "unhealthy", "error": str(e)})


# ═══════════════════════════════════════════════════════════════
# PRODUCT SEARCH ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/products/search")
def search_products(
    q: str = Query(..., min_length=1),
    limit: int = Query(default=20, ge=1, le=100),
    _: bool = Depends(verify_api_key)
):
    """Search products using 3-tier matching: exact → keyword → trigram."""
    try:
        with get_transaction() as cur:
            results = _tiered_product_search(cur, q, limit=limit)
            products = []
            for r in results:
                # Fetch full product details for each match
                cur.execute(
                    """SELECT id, name, odoo_code, type, uom, active,
                              COALESCE(verification_status, 'verified') as verification_status,
                              case_size_lb, default_batch_lb
                       FROM products WHERE id = %s""",
                    (r['id'],)
                )
                row = cur.fetchone()
                if row:
                    prod = dict(row)
                    prod['match_tier'] = r['match_tier']
                    products.append(prod)
        return {"count": len(products), "products": products}
    except Exception as e:
        logger.error(f"Product search failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/products/missing-case-size")
def products_missing_case_size(_: bool = Depends(verify_api_key)):
    """Return active, non-service, non-ingredient SKUs where case_size_lb is NULL or 0."""
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT id, name, odoo_code, type, uom, case_size_lb, default_batch_lb
                FROM products
                WHERE COALESCE(active, true) = true
                  AND COALESCE(is_service, false) = false
                  AND type != 'ingredient'
                  AND (case_size_lb IS NULL OR case_size_lb = 0)
                ORDER BY name
            """)
            products = cur.fetchall()
        return {"count": len(products), "products": products}
    except Exception as e:
        logger.error(f"Missing case size query failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/products/resolve")
def resolve_products_bulk(req: BulkResolveRequest, _: bool = Depends(verify_api_key)):
    """Bulk-resolve raw product name strings against the database.
    Uses 3-tier matching: exact → keyword (word-order independent) → trigram similarity."""
    try:
        with get_transaction() as cur:
            resolved_list = []
            resolved_count = 0
            for name in req.names:
                result = _resolve_single_product(cur, name)
                resolved_list.append(result)
                if result['match'] is not None:
                    resolved_count += 1

        return {
            "resolved": resolved_list,
            "summary": {
                "total": len(req.names),
                "resolved": resolved_count,
                "unresolved": len(req.names) - resolved_count,
            }
        }
    except Exception as e:
        logger.error(f"Bulk product resolve failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# REVIEW QUEUE ENDPOINTS (MUST BE BEFORE /products/{product_id})
# ═══════════════════════════════════════════════════════════════

@app.get("/products/unverified")
def get_unverified_products(limit: int = Query(default=50, ge=1, le=200), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT id, name, type, uom, 
                       COALESCE(verification_status, 'verified') as verification_status,
                       verification_notes, created_via
                FROM products
                WHERE COALESCE(verification_status, 'verified') = 'unverified'
                  AND COALESCE(active, true) = true
                ORDER BY id DESC
                LIMIT %s
            """, (limit,))
            products = cur.fetchall()
        return {"count": len(products), "products": products}
    except Exception as e:
        logger.error(f"Get unverified products failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/products/test-batches")
def get_test_batches(limit: int = Query(default=50, ge=1, le=200), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT id, name, type, 
                       COALESCE(verification_status, 'verified') as verification_status,
                       COALESCE(production_context, 'standard') as production_context,
                       verification_notes
                FROM products
                WHERE COALESCE(production_context, 'standard') IN ('test_batch', 'sample', 'one_off')
                  AND COALESCE(active, true) = true
                ORDER BY id DESC
                LIMIT %s
            """, (limit,))
            products = cur.fetchall()
        return {"count": len(products), "products": products}
    except Exception as e:
        logger.error(f"Get test batches failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# PRODUCT BY ID (AFTER specific /products/* routes)
# ═══════════════════════════════════════════════════════════════

@app.get("/products/{product_id}")
def get_product(product_id: int, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT * FROM products WHERE id = %s", (product_id,))
            product = cur.fetchone()
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")
        return product
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get product failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# INVENTORY ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/inventory/current")
def get_current_inventory(
    product_type: Optional[str] = Query(default=None),
    limit: int = Query(default=100, ge=1, le=500),
    _: bool = Depends(verify_api_key)
):
    try:
        with get_transaction() as cur:
            query = f"""
                SELECT p.id as product_id, p.name as product_name, p.odoo_code, p.type as product_type,
                       p.case_size_lb, p.default_batch_lb,
                       l.id as lot_id, l.lot_code,
                       COALESCE(SUM(tl.quantity_lb), 0) as quantity_on_hand
                FROM products p
                JOIN lots l ON l.product_id = p.id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE COALESCE(p.active, true) = true
            """
            params = []
            if product_type:
                query += " AND p.type = %s"
                params.append(product_type)
            query += " GROUP BY p.id, l.id HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0 ORDER BY p.name, l.lot_code LIMIT %s"
            params.append(limit)
            cur.execute(query, params)
            rows = cur.fetchall()
            inventory = []
            for r in rows:
                item = dict(r)
                oh = float(r['quantity_on_hand'])
                cs = float(r['case_size_lb']) if r.get('case_size_lb') else None
                db = float(r['default_batch_lb']) if r.get('default_batch_lb') else None
                if cs and cs > 0 and r['product_type'] != 'ingredient':
                    item['unit_count'] = round(oh / cs)
                elif db and db > 0 and r['product_type'] == 'batch':
                    item['batch_count'] = round(oh / db, 1)
                inventory.append(item)
        return {"count": len(inventory), "inventory": inventory}
    except Exception as e:
        logger.error(f"Get current inventory failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


def _inventory_detail_for_product(cur, product_id: int) -> dict:
    """Fetch lot-level inventory detail for a single product."""
    results = _inventory_detail_for_products(cur, [product_id])
    return results.get(product_id)


def _inventory_detail_for_products(cur, product_ids: list[int]) -> dict[int, dict]:
    """Fetch lot-level inventory detail for multiple products in bulk (2 queries total)."""
    if not product_ids:
        return {}

    cur.execute(
        """SELECT p.id, p.name, p.odoo_code, p.type, p.uom,
                  p.case_size_lb, p.default_batch_lb
           FROM products p WHERE p.id = ANY(%s)""",
        (product_ids,)
    )
    prods = {row['id']: row for row in cur.fetchall()}
    if not prods:
        return {}

    cur.execute(
        f"""SELECT l.product_id,
                  l.lot_code,
                  COALESCE(SUM(tl.quantity_lb), 0) AS qty_on_hand
           FROM lots l
           LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
           WHERE l.product_id = ANY(%s)
           GROUP BY l.product_id, l.id, l.lot_code
           HAVING COALESCE(SUM(tl.quantity_lb), 0) != 0
           ORDER BY l.product_id, l.lot_code""",
        (product_ids,)
    )
    lot_rows = cur.fetchall()

    # Group lots by product_id
    lots_by_product: dict[int, list] = {pid: [] for pid in prods}
    for lr in lot_rows:
        lots_by_product.setdefault(lr['product_id'], []).append(lr)

    results = {}
    for pid, prod in prods.items():
        lots = []
        total_on_hand = 0.0
        for lr in lots_by_product.get(pid, []):
            qty = float(lr['qty_on_hand'])
            total_on_hand += qty
            lots.append({"lot": lr['lot_code'], "qty_on_hand": round(qty, 2), "unit": prod['uom'] or "lb"})

        total_on_hand = round(total_on_hand, 2)
        result = {
            "product": prod['name'],
            "sku": prod['odoo_code'],
            "lots": lots,
            "total_on_hand": total_on_hand,
        }

        cs = float(prod['case_size_lb']) if prod.get('case_size_lb') else None
        db = float(prod['default_batch_lb']) if prod.get('default_batch_lb') else None
        if cs and cs > 0 and prod['type'] != 'ingredient':
            result['unit_count'] = round(total_on_hand / cs)
        elif db and db > 0 and prod['type'] == 'batch':
            result['batch_count'] = round(total_on_hand / db, 1)

        results[pid] = result

    return results


@app.get("/inventory/lookup")
def inventory_lookup(
    q: str = Query(..., min_length=1),
    limit: int = Query(default=5, ge=1, le=50),
    _: bool = Depends(verify_api_key)
):
    """Search inventory by product name (fuzzy). This is the PRIMARY endpoint for all inventory questions."""
    try:
        with get_transaction() as cur:
            matches = _tiered_product_search(cur, q, limit=limit)
            if not matches:
                return {"query": q, "results": []}
            product_ids = [m['id'] for m in matches]
            details_by_id = _inventory_detail_for_products(cur, product_ids)
            results = []
            for m in matches:
                detail = details_by_id.get(m['id'])
                if detail:
                    detail['match_tier'] = m['match_tier']
                    results.append(detail)
        return {"query": q, "results": results}
    except Exception as e:
        logger.error(f"Inventory lookup failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/inventory/{item_name}")
def get_inventory(item_name: str, _: bool = Depends(verify_api_key)):
    """Get inventory for an exact product name. Prefer /inventory/lookup for general queries."""
    try:
        with get_transaction() as cur:
            # Try exact LIKE match first (original behavior)
            cur.execute(f"""
                SELECT p.id, p.name, p.odoo_code, p.type, p.uom,
                       p.case_size_lb, p.default_batch_lb,
                       COALESCE(SUM(tl.quantity_lb), 0) as total_on_hand
                FROM products p
                LEFT JOIN lots l ON l.product_id = p.id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE LOWER(p.name) LIKE LOWER(%s) OR LOWER(p.odoo_code) LIKE LOWER(%s)
                GROUP BY p.id
            """, (f"%{item_name}%", f"%{item_name}%"))
            rows = cur.fetchall()

            # Fuzzy fallback: if no exact LIKE match, try tiered search
            if not rows:
                fuzzy_matches = _tiered_product_search(cur, item_name, limit=5)
                if not fuzzy_matches:
                    return JSONResponse(status_code=404, content={"error": "not_found", "message": f"No product matching '{item_name}'"})
                if len(fuzzy_matches) == 1:
                    # Single match — return its inventory
                    detail = _inventory_detail_for_product(cur, fuzzy_matches[0]['id'])
                    if detail:
                        return {"count": 1, "inventory": [detail]}
                    return JSONResponse(status_code=404, content={"error": "not_found", "message": f"No product matching '{item_name}'"})
                # Multiple matches — ask caller to disambiguate
                return JSONResponse(status_code=300, content={
                    "error": "multiple_matches",
                    "matches": [{"name": m['name'], "odoo_code": m['odoo_code'], "match_tier": m['match_tier']} for m in fuzzy_matches]
                })

            results = []
            for r in rows:
                item = dict(r)
                oh = float(r['total_on_hand'])
                cs = float(r['case_size_lb']) if r.get('case_size_lb') else None
                db = float(r['default_batch_lb']) if r.get('default_batch_lb') else None
                if cs and cs > 0 and r['type'] != 'ingredient':
                    item['unit_count'] = round(oh / cs)
                elif db and db > 0 and r['type'] == 'batch':
                    item['batch_count'] = round(oh / db, 1)
                results.append(item)
        return {"count": len(results), "inventory": results}
    except Exception as e:
        logger.error(f"Get inventory failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# LOT ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/lots/by-supplier-lot/{supplier_lot_code}")
def get_lots_by_supplier_lot(supplier_lot_code: str, _: bool = Depends(verify_api_key)):
    """Find internal lot(s) by supplier lot code — for recall tracing."""
    try:
        with get_transaction() as cur:
            # Search lots table (direct supplier_lot_code on lot)
            cur.execute(f"""
                SELECT l.id, l.lot_code, l.supplier_lot_code, l.lot_type,
                       l.product_id, p.name AS product_name, p.odoo_code,
                       COALESCE(SUM(tl.quantity_lb), 0) AS quantity_on_hand
                FROM lots l
                JOIN products p ON p.id = l.product_id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE LOWER(l.supplier_lot_code) = LOWER(%s)
                GROUP BY l.id, p.id
            """, (supplier_lot_code,))
            direct_matches = cur.fetchall()

            # Search lot_supplier_codes table (commingled entries)
            cur.execute(f"""
                SELECT DISTINCT l.id, l.lot_code, l.supplier_lot_code, l.lot_type,
                       l.product_id, p.name AS product_name, p.odoo_code,
                       COALESCE(oh.on_hand, 0) AS quantity_on_hand,
                       lsc.supplier_lot_code AS commingled_supplier_lot,
                       lsc.supplier_name, lsc.quantity_lb AS supplier_qty_lb, lsc.notes
                FROM lot_supplier_codes lsc
                JOIN lots l ON l.id = lsc.lot_id
                JOIN products p ON p.id = l.product_id
                LEFT JOIN LATERAL (
                    SELECT COALESCE(SUM(tl2.quantity_lb), 0) AS on_hand
                    FROM {POSTED_LINES} tl2 WHERE tl2.lot_id = l.id
                ) oh ON TRUE
                WHERE LOWER(lsc.supplier_lot_code) = LOWER(%s)
            """, (supplier_lot_code,))
            commingled_matches = cur.fetchall()

            # Combine results, dedup by lot_id
            seen_ids = set()
            results = []
            for lot in direct_matches:
                seen_ids.add(lot['id'])
                results.append({
                    "lot_id": lot['id'],
                    "lot_code": lot['lot_code'],
                    "supplier_lot_code": lot['supplier_lot_code'],
                    "lot_type": lot['lot_type'],
                    "product_name": lot['product_name'],
                    "odoo_code": lot['odoo_code'],
                    "quantity_on_hand": float(lot['quantity_on_hand']),
                    "match_source": "lots.supplier_lot_code"
                })
            for row in commingled_matches:
                entry = {
                    "lot_id": row['id'],
                    "lot_code": row['lot_code'],
                    "supplier_lot_code": row['supplier_lot_code'],
                    "lot_type": row['lot_type'],
                    "product_name": row['product_name'],
                    "odoo_code": row['odoo_code'],
                    "quantity_on_hand": float(row['quantity_on_hand']),
                    "match_source": "lot_supplier_codes",
                    "commingled_detail": {
                        "supplier_lot_code": row['commingled_supplier_lot'],
                        "supplier_name": row['supplier_name'],
                        "quantity_lb": float(row['supplier_qty_lb']) if row['supplier_qty_lb'] else None,
                        "notes": row['notes']
                    }
                }
                if row['id'] not in seen_ids:
                    results.append(entry)
                    seen_ids.add(row['id'])
                else:
                    # Lot already in results from direct match — add commingled detail
                    for r in results:
                        if r['lot_id'] == row['id'] and 'commingled_detail' not in r:
                            r['commingled_detail'] = entry['commingled_detail']
                            break

            if not results:
                raise HTTPException(404, f"No lots found with supplier lot code '{supplier_lot_code}'")

            return {
                "supplier_lot_code_searched": supplier_lot_code,
                "matching_lots": results,
                "total_matches": len(results)
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Search by supplier lot failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/lots/by-code/{lot_code}")
def get_lot_by_code(lot_code: str, product_id: Optional[int] = Query(None), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            query = f"""
                SELECT l.id, l.lot_code, l.product_id, p.name as product_name, p.odoo_code,
                       COALESCE(SUM(tl.quantity_lb), 0) as quantity_on_hand,
                       l.entry_source, l.found_location, l.estimated_age,
                       l.supplier_lot_code, l.lot_type,
                       p.case_size_lb, p.default_batch_lb, p.type as product_type
                FROM lots l
                JOIN products p ON p.id = l.product_id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE LOWER(l.lot_code) = LOWER(%s)
            """
            params = [lot_code]
            if product_id is not None:
                query += " AND l.product_id = %s"
                params.append(product_id)
            query += " GROUP BY l.id, p.id"
            cur.execute(query, params)
            rows = cur.fetchall()

        if not rows:
            raise HTTPException(status_code=404, detail=f"Lot '{lot_code}' not found")
        if len(rows) > 1:
            return JSONResponse(status_code=409, content={
                "error": "ambiguous_lot_code",
                "message": f"Lot code '{lot_code}' matches multiple products",
                "matches": [{"lot_id": r['id'], "product_id": r['product_id'],
                             "product_name": r['product_name'], "entry_source": r['entry_source']} for r in rows]
            })
        lot = rows[0]
        result = dict(lot)
        # Add unit_count based on product type
        oh = float(lot['quantity_on_hand'])
        cs = float(lot['case_size_lb']) if lot.get('case_size_lb') else None
        db = float(lot['default_batch_lb']) if lot.get('default_batch_lb') else None
        if cs and cs > 0 and lot.get('product_type') != 'ingredient':
            result['unit_count'] = round(oh / cs)
        elif db and db > 0 and lot.get('product_type') == 'batch':
            result['batch_count'] = round(oh / db, 1)
        with get_transaction() as cur:
            cur.execute("""
                SELECT supplier_lot_code, supplier_name, quantity_lb, notes
                FROM lot_supplier_codes WHERE lot_id = %s ORDER BY id
            """, (lot['id'],))
            result['supplier_lot_entries'] = [dict(r) for r in cur.fetchall()]
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get lot by code failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/lots/{lot_id}")
def get_lot(lot_id: int, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute(f"""
                SELECT l.*, p.name as product_name, p.odoo_code,
                       p.case_size_lb, p.default_batch_lb, p.type as product_type,
                       COALESCE(SUM(tl.quantity_lb), 0) as quantity_on_hand
                FROM lots l
                JOIN products p ON p.id = l.product_id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE l.id = %s
                GROUP BY l.id, p.id
            """, (lot_id,))
            lot = cur.fetchone()
        if not lot:
            raise HTTPException(status_code=404, detail="Lot not found")
        result = dict(lot)
        oh = float(lot['quantity_on_hand'])
        cs = float(lot['case_size_lb']) if lot.get('case_size_lb') else None
        db = float(lot['default_batch_lb']) if lot.get('default_batch_lb') else None
        if cs and cs > 0 and lot.get('product_type') != 'ingredient':
            result['unit_count'] = round(oh / cs)
        elif db and db > 0 and lot.get('product_type') == 'batch':
            result['batch_count'] = round(oh / db, 1)
        with get_transaction() as cur:
            cur.execute("""
                SELECT supplier_lot_code, supplier_name, quantity_lb, notes
                FROM lot_supplier_codes WHERE lot_id = %s ORDER BY id
            """, (lot_id,))
            result['supplier_lot_entries'] = [dict(r) for r in cur.fetchall()]
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get lot failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


class SupplierLotUpdate(BaseModel):
    supplier_lot_code: str
    notes: Optional[str] = None


class LotRenameRequest(BaseModel):
    new_lot_code: str


@app.patch("/lots/{lot_code}/supplier-lot")
def update_supplier_lot(lot_code: str, req: SupplierLotUpdate, product_id: Optional[int] = Query(None), _: bool = Depends(verify_api_key)):
    """Attach or update the supplier lot cross-reference on an existing lot.

    Use this when a packing slip or physical label shows a supplier lot number
    that differs from (or was missing on) the system lot. Updates lots.supplier_lot_code
    and optionally adds to lot_supplier_codes for commingled lots.
    """
    try:
        with get_transaction() as cur:
            # Find the lot
            query = """
                SELECT l.id, l.lot_code, l.supplier_lot_code, l.lot_type,
                       p.name AS product_name, l.product_id
                FROM lots l
                JOIN products p ON p.id = l.product_id
                WHERE LOWER(l.lot_code) = LOWER(%s)
            """
            params = [lot_code]
            if product_id is not None:
                query += " AND l.product_id = %s"
                params.append(product_id)
            cur.execute(query, params)
            rows = cur.fetchall()
            if not rows:
                raise HTTPException(404, f"Lot '{lot_code}' not found")
            if len(rows) > 1:
                return JSONResponse(status_code=409, content={
                    "error": "ambiguous_lot_code",
                    "message": f"Lot code '{lot_code}' matches multiple products. Provide product_id to disambiguate.",
                    "matches": [{"lot_id": r['id'], "product_id": r['product_id'],
                                 "product_name": r['product_name']} for r in rows]
                })
            lot = rows[0]

            old_supplier_lot = lot['supplier_lot_code']
            new_supplier_lot = req.supplier_lot_code.strip()

            # Update lots.supplier_lot_code
            cur.execute("""
                UPDATE lots SET supplier_lot_code = %s WHERE id = %s
            """, (new_supplier_lot, lot['id']))

            # If commingled lot, also add an entry in lot_supplier_codes
            if lot['lot_type'] == 'commingled':
                cur.execute("""
                    INSERT INTO lot_supplier_codes (lot_id, supplier_lot_code, notes)
                    VALUES (%s, %s, %s)
                """, (lot['id'], new_supplier_lot, req.notes))

            return {
                "lot_code": lot['lot_code'],
                "product_name": lot['product_name'],
                "previous_supplier_lot_code": old_supplier_lot,
                "supplier_lot_code": new_supplier_lot,
                "notes": req.notes,
                "updated": True
            }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Update supplier lot failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/lots/{lot_id}/rename")
def rename_lot(lot_id: int, req: LotRenameRequest, _: bool = Depends(verify_api_key)):
    """Rename a lot's lot_code (e.g. fix an UNKNOWN lot to the real code).

    Validates no duplicate lot_code exists for the same product.
    Only lots.lot_code needs updating — all other tables use integer FKs.
    """
    new_code = req.new_lot_code.strip()
    if not new_code:
        raise HTTPException(400, "new_lot_code must not be empty")

    try:
        with get_transaction() as cur:
            # Fetch the lot
            cur.execute("""
                SELECT l.id, l.lot_code, l.product_id, p.name AS product_name
                FROM lots l
                JOIN products p ON p.id = l.product_id
                WHERE l.id = %s
            """, (lot_id,))
            lot = cur.fetchone()
            if not lot:
                raise HTTPException(404, f"Lot id {lot_id} not found")

            old_code = lot['lot_code']
            product_id = lot['product_id']

            if old_code == new_code:
                return {"lot_id": lot_id, "lot_code": new_code,
                        "product_name": lot['product_name'],
                        "renamed": False, "message": "Already has that lot_code"}

            # Check for conflict
            cur.execute("""
                SELECT id FROM lots
                WHERE product_id = %s AND LOWER(lot_code) = LOWER(%s)
            """, (product_id, new_code))
            conflict = cur.fetchone()
            if conflict:
                raise HTTPException(409,
                    f"Lot code '{new_code}' already exists for product {product_id} "
                    f"(lot id {conflict['id']})")

            # Rename
            cur.execute("""
                UPDATE lots SET lot_code = %s WHERE id = %s
            """, (new_code, lot_id))

            logger.info(f"Lot {lot_id} renamed: '{old_code}' -> '{new_code}' "
                        f"(product_id={product_id})")

            return {
                "lot_id": lot_id,
                "previous_lot_code": old_code,
                "lot_code": new_code,
                "product_id": product_id,
                "product_name": lot['product_name'],
                "renamed": True
            }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Lot rename failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# LOT IDENTITY — Find-or-Create Pattern
# ═══════════════════════════════════════════════════════════════
# Lot Identity Policy: A physical lot must map to exactly one canonical lot_id.
# If lot_code is provided, find-or-create by (product_id, lot_code).
# Only auto-generate if lot_code is omitted.

def find_or_create_lot(cur, product_id: int, lot_code: str, entry_source: str,
                       entry_source_notes: str = None, entry_source_notes_es: str = None,
                       found_location: str = None, estimated_age: str = None) -> tuple:
    """Find existing lot or create a new one. Returns (lot_id, is_new).

    Uses INSERT ... ON CONFLICT DO NOTHING + SELECT to guarantee exactly one lot
    per (product_id, lot_code) pair, leveraging the unique index.
    """
    # Build dynamic INSERT with optional columns
    columns = ["product_id", "lot_code", "entry_source"]
    values = [product_id, lot_code, entry_source]
    placeholders = ["%s", "%s", "%s"]

    if entry_source_notes:
        columns.append("entry_source_notes")
        values.append(entry_source_notes)
        placeholders.append("%s")
    if entry_source_notes_es:
        columns.append("entry_source_notes_es")
        values.append(entry_source_notes_es)
        placeholders.append("%s")
    if found_location:
        columns.append("found_location")
        values.append(found_location)
        placeholders.append("%s")
    if estimated_age:
        columns.append("estimated_age")
        values.append(estimated_age)
        placeholders.append("%s")

    col_str = ", ".join(columns)
    ph_str = ", ".join(placeholders)

    cur.execute(f"""
        INSERT INTO lots ({col_str})
        VALUES ({ph_str})
        ON CONFLICT (product_id, lot_code) DO NOTHING
    """, values)
    is_new = cur.rowcount > 0

    # Fetch the lot (whether just created or already existed)
    cur.execute("SELECT id FROM lots WHERE product_id = %s AND lot_code = %s", (product_id, lot_code))
    lot_id = cur.fetchone()['id']

    if not is_new:
        logger.info(f"Found existing lot {lot_code} (id={lot_id}) for product_id={product_id}")

    return lot_id, is_new


# ═══════════════════════════════════════════════════════════════
# RECEIVE ENDPOINTS
# ═══════════════════════════════════════════════════════════════

def generate_lot_code(cur, shipper_name: str, shipper_code_override: str = None) -> tuple:
    now = get_plant_now()
    date_part = now.strftime("%y-%m-%d")
    
    if shipper_code_override:
        shipper_code = shipper_code_override.upper()[:4]
        auto = False
    else:
        shipper_code = ''.join(c for c in shipper_name.upper() if c.isalpha())[:4]
        auto = True
    
    shipper_code = shipper_code or "UNKN"
    
    cur.execute("""
        SELECT lot_code FROM lots 
        WHERE lot_code LIKE %s 
        ORDER BY lot_code DESC LIMIT 1
    """, (f"{date_part}-{shipper_code}-%",))
    existing = cur.fetchone()
    
    if existing:
        try:
            last_seq = int(existing['lot_code'].split('-')[-1])
            seq = last_seq + 1
        except (ValueError, IndexError):
            seq = 1
    else:
        seq = 1
    
    lot_code = f"{date_part}-{shipper_code}-{seq:03d}"
    return lot_code, shipper_code, auto


@app.post("/receive")
def receive(req: ReceiveRequest, _: bool = Depends(verify_api_key)):
    """Receive inventory. mode=preview returns what will happen; mode=commit executes."""
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at, req.backfill
    )
    # Fallback: use lot_code if supplier_lot_code not provided, then 'N/A'
    supplier_lot = (req.supplier_lot_code or "").strip()
    if not supplier_lot:
        supplier_lot = (req.lot_code or "").strip() or "N/A"
    req.supplier_lot_code = supplier_lot

    if req.mode == "preview":
        try:
            with get_transaction() as cur:
                product = resolve_product_full(cur, req.product_name)

                # Lot Identity Policy: honor physical lot code if provided
                if req.lot_code:
                    lot_code = req.lot_code
                    shipper_code = req.shipper_code_override or ''.join(c for c in req.shipper_name.upper() if c.isalpha())[:4] or "UNKN"
                    auto = False
                    cur.execute("SELECT id FROM lots WHERE product_id = %s AND lot_code = %s", (product['id'], req.lot_code))
                    existing = cur.fetchone()
                else:
                    lot_code, shipper_code, auto = generate_lot_code(cur, req.shipper_name, req.shipper_code_override)
                    existing = None
                total_lb = req.cases * req.case_size_lb

                response = {
                    "mode": "preview",
                    "product_id": product['id'],
                    "product_name": product['name'],
                    "odoo_code": product['odoo_code'],
                    "cases": req.cases,
                    "case_size_lb": req.case_size_lb,
                    "total_lb": total_lb,
                    "shipper_name": req.shipper_name,
                    "shipper_code": shipper_code,
                    "shipper_code_auto": auto,
                    "lot_code": lot_code,
                    "bol_reference": req.bol_reference,
                    "preview_message": f"Ready to receive {req.cases} cases ({total_lb} lb) of {product['name']} as lot {lot_code}"
                }
                if existing:
                    response["lot_exists"] = True
                    response["existing_lot_id"] = existing['id']
                    response["preview_message"] += f" (lot already exists — will add to existing)"
                if req.supplier_lot_entries:
                    response["commingled"] = True
                    response["supplier_lot_entries"] = req.supplier_lot_entries
                # FR-2 lookahead: which open expected receipt (if any) this
                # commit would link to. Informational only.
                er_match = preview_expected_receipt_match(cur, product['id'], req.shipper_name)
                response["expected_receipt_match"] = (
                    {"id": er_match["id"], "expected_qty": er_match["expected_qty"],
                     "remaining": er_match["remaining"], "expected_date": er_match["expected_date"],
                     "reference_number": er_match["reference_number"]}
                    if er_match else None
                )
                return response
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Receive preview failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})
    else:
        # mode == "commit"
        try:
            with get_db_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT pg_advisory_xact_lock(1)")
                    product = resolve_product_full(cur, req.product_name)

                    if req.lot_code:
                        lot_code = req.lot_code
                        shipper_code = req.shipper_code_override or ''.join(c for c in req.shipper_name.upper() if c.isalpha())[:4] or "UNKN"
                    else:
                        lot_code, shipper_code, _ = generate_lot_code(cur, req.shipper_name, req.shipper_code_override)
                    total_lb = req.cases * req.case_size_lb
                    now = get_plant_now()

                    # Determine lot_type
                    lot_type = req.lot_type or ("commingled" if req.supplier_lot_entries else "single_supplier")

                    lot_id, is_new_lot = find_or_create_lot(cur, product['id'], lot_code, 'received')

                    # Update lot with LAT Code Policy v1.1 fields
                    cur.execute("""
                        UPDATE lots SET received_at = COALESCE(received_at, %s),
                                        supplier_lot_code = COALESCE(%s, supplier_lot_code),
                                        lot_type = COALESCE(%s, lot_type)
                        WHERE id = %s
                    """, (now, req.supplier_lot_code, lot_type, lot_id))

                    # FR-2 auto-match: open expected receipt for (product, supplier),
                    # FIFO by expected_date. Supplier resolved by normalised name;
                    # unknown OR INACTIVE supplier, or no open record → post normally,
                    # unlinked. The link is set at INSERT (transactions is append-only).
                    expected_receipt_id = None
                    er_supplier = resolve_supplier(cur, req.shipper_name)
                    if er_supplier and er_supplier["active"]:
                        er_match = find_open_expected_receipt(cur, product['id'], er_supplier['id'], lock=True)
                        if er_match:
                            expected_receipt_id = er_match['id']

                    cur.execute("""
                        INSERT INTO transactions (
                            type, timestamp, bol_reference, shipper_name,
                            shipper_code, cases_received, case_size_lb,
                            expected_receipt_id, occurred_at, created_at_source
                        )
                        VALUES ('receive', %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                    """, (
                        now, req.bol_reference, req.shipper_name, shipper_code,
                        req.cases, req.case_size_lb, expected_receipt_id,
                        occurred_at, created_at_source,
                    ))
                    txn_id = cur.fetchone()['id']

                    cur.execute("""
                        INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb)
                        VALUES (%s, %s, %s, %s)
                    """, (txn_id, product['id'], lot_id, total_lb))

                    # Recompute received (ledger SUM) and auto-close on full/over receipt.
                    expected_receipt_summary = (
                        settle_expected_receipt(cur, expected_receipt_id) if expected_receipt_id else None
                    )

                    # Insert commingled supplier lot entries if provided
                    supplier_entries_saved = []
                    if req.supplier_lot_entries:
                        for entry in req.supplier_lot_entries:
                            cur.execute("""
                                INSERT INTO lot_supplier_codes (lot_id, supplier_lot_code, supplier_name, quantity_lb, notes)
                                VALUES (%s, %s, %s, %s, %s)
                                RETURNING id
                            """, (lot_id, entry.get('supplier_lot_code'), entry.get('supplier_name'),
                                  entry.get('quantity_lb'), entry.get('notes')))
                            supplier_entries_saved.append(cur.fetchone()['id'])

                    date_str, time_str = format_timestamp(now)
                    receipt = f"RECEIVED: {req.cases} cases ({total_lb} lb) {product['name']}\nLot: {lot_code}\nBOL: {req.bol_reference}\n{date_str} {time_str}"

                    lot_verb = "created" if is_new_lot else "found existing"
                    logger.info(f"Receive committed: {lot_code} ({lot_verb}) - {total_lb} lb of {product['name']}")

                    response = {
                        "mode": "commit",
                        "success": True,
                        "transaction_id": txn_id,
                        "confirmation_code": generate_confirmation_code(txn_id),
                        "lot_id": lot_id,
                        "lot_code": lot_code,
                        "lot_is_new": is_new_lot,
                        "lot_type": lot_type,
                        "total_lb": total_lb,
                        "receipt_text": receipt,
                        "message": f"Received {total_lb} lb as lot {lot_code}" + ("" if is_new_lot else " (existing lot)")
                    }
                    if supplier_entries_saved:
                        response["supplier_lot_entries_created"] = len(supplier_entries_saved)
                    response["expected_receipt"] = expected_receipt_summary
                    if expected_receipt_summary:
                        response["message"] += (
                            f"; linked to expected receipt #{expected_receipt_summary['id']}"
                            + (" (now closed)" if expected_receipt_summary["auto_closed"] else
                               f" ({expected_receipt_summary['remaining']:g} lb still expected)")
                        )
                    return response
        except HTTPException:
            raise
        except Exception as e:
            if _is_readonly_error(e): raise
            logger.error(f"Receive commit failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# FR-2: SUPPLIERS + EXPECTED RECEIPTS
# Lightweight expected-receipt tracking — explicitly NOT a PO system.
#
# Invariant: there is no stored "remaining" or "received" balance.
#   remaining = expected_qty - SUM(posted transaction_lines of receipts
#                                  whose transactions.expected_receipt_id
#                                  points at the record), floored at 0.
# Linking happens ONLY at receive-commit INSERT time (transactions is
# append-only). expected_receipts is never read by any inventory
# balance / availability query.
# ═══════════════════════════════════════════════════════════════

# Python mirror of SQL supplier_name_norm() (migration 041). Both sides must
# agree; the SQL function is what the unique index and all lookups use.
def normalize_supplier_name(name: Optional[str]) -> str:
    if not name:
        return ""
    s = name.strip().replace("’", "'").replace("`", "'")
    return re.sub(r"\s+", " ", s).lower()


def resolve_supplier(cur, supplier_name: Optional[str]) -> Optional[dict]:
    """Case/whitespace-insensitive exact match against suppliers. Returns the
    row (id, name, active) or None. Never creates anything."""
    if not supplier_name or not supplier_name.strip():
        return None
    cur.execute(
        """SELECT id, name, active FROM suppliers
           WHERE supplier_name_norm(name) = supplier_name_norm(%s)
           LIMIT 1""",
        (supplier_name,),
    )
    row = cur.fetchone()
    return dict(row) if row else None


def supplier_candidates(cur, supplier_name: str, limit: int = 5) -> list:
    """Up to `limit` closest ACTIVE supplier names (trigram similarity on the
    normalised form, substring hits first). Used for the 422 payload."""
    cur.execute(
        """SELECT id, name,
                  similarity(supplier_name_norm(name), supplier_name_norm(%s)) AS sim,
                  (supplier_name_norm(name) LIKE '%%' || supplier_name_norm(%s) || '%%'
                   OR supplier_name_norm(%s) LIKE '%%' || supplier_name_norm(name) || '%%') AS substr_hit
           FROM suppliers
           WHERE active
           ORDER BY substr_hit DESC, sim DESC, name
           LIMIT %s""",
        (supplier_name, supplier_name, supplier_name, limit * 4),
    )
    rows = cur.fetchall()
    out = []
    for r in rows:
        if r["substr_hit"] or float(r["sim"] or 0) > 0:
            out.append({"supplier_id": r["id"], "name": r["name"], "similarity": round(float(r["sim"] or 0), 3)})
        if len(out) >= limit:
            break
    return out


def caller_source_tag(request: Request, body_tag: Optional[str] = None) -> Optional[str]:
    """Interim attribution until FR-15 (user attribution) exists: a plain-text
    SOURCE tag, never a fake user id.
      * scoped dashboard key authenticated the call → 'dashboard' (body ignored)
      * master key → the caller-supplied tag if any (the office GPT schema
        defaults created_by to 'gpt-sales-admin'), else NULL
    Deliberately NOT the 'legacy-shared-key' operator_id placeholder."""
    key = request.headers.get("X-API-Key") or ""
    if DASHBOARD_API_KEY and key and secrets.compare_digest(key, DASHBOARD_API_KEY):
        return "dashboard"
    tag = re.sub(r"\s+", " ", (body_tag or "").strip())
    return tag[:60] or None


def require_supplier(cur, supplier_name: str) -> dict:
    """Resolve or raise 422 SUPPLIER_NOT_FOUND with up to 5 candidates."""
    supplier = resolve_supplier(cur, supplier_name)
    if supplier and supplier["active"]:
        return supplier
    candidates = supplier_candidates(cur, supplier_name, limit=5)
    if supplier and not supplier["active"]:
        code, msg = "SUPPLIER_INACTIVE", f"Supplier '{supplier['name']}' is inactive."
    else:
        code, msg = "SUPPLIER_NOT_FOUND", f"No supplier matches '{supplier_name}'."
    raise HTTPException(
        status_code=422,
        detail={
            "error_code": code,
            "message": msg + " Pick one of the candidates or create the supplier explicitly (POST /suppliers).",
            "input": supplier_name,
            "candidates": candidates,
            "suggestions": [c["name"] for c in candidates],
        },
    )


# Received quantity for one expected receipt: SUM of posted-only lines on the
# receive transactions linked to it. This is the ONLY place received/remaining
# comes from — it is never stored.
EXPECTED_RECEIPT_RECEIVED_SQL = """
    SELECT COALESCE(SUM(tl.quantity_lb), 0) AS received_qty,
           COUNT(DISTINCT t.id)               AS receipt_count,
           MAX(t.timestamp)                   AS last_received_at
    FROM transactions t
    JOIN ledger_current_transactions ct
      ON ct.id = t.id AND ct.effective_status = 'posted'
    JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
    WHERE t.expected_receipt_id = %s
"""

EXPECTED_RECEIPT_SELECT_SQL = """
    SELECT er.id, er.product_id, er.supplier_id, er.expected_qty, er.expected_date,
           er.reference_number, er.notes, er.status, er.created_at, er.created_by,
           er.updated_at,
           p.name AS product_name, p.odoo_code,
           s.name AS supplier_name,
           COALESCE(rcv.received_qty, 0) AS received_qty,
           COALESCE(rcv.receipt_count, 0) AS receipt_count,
           rcv.last_received_at
    FROM expected_receipts er
    JOIN products p ON p.id = er.product_id
    JOIN suppliers s ON s.id = er.supplier_id
    LEFT JOIN LATERAL (
        SELECT SUM(tl.quantity_lb) AS received_qty,
               COUNT(DISTINCT t.id) AS receipt_count,
               MAX(t.timestamp)     AS last_received_at
        FROM transactions t
        JOIN ledger_current_transactions ct
          ON ct.id = t.id AND ct.effective_status = 'posted'
        JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
        WHERE t.expected_receipt_id = er.id
    ) rcv ON true
"""


def _serialize_expected_receipt(row: dict, today: Optional[date] = None) -> dict:
    today = today or get_plant_now().date()
    expected = float(row["expected_qty"] or 0)
    received = float(row["received_qty"] or 0)
    raw_remaining = expected - received
    exp_date = row["expected_date"]
    is_overdue = bool(row["status"] == "open" and exp_date is not None and exp_date < today)
    return {
        "id": row["id"],
        "product_id": row["product_id"],
        "product_name": row["product_name"],
        "odoo_code": row.get("odoo_code"),
        "supplier_id": row["supplier_id"],
        "supplier_name": row["supplier_name"],
        "expected_qty": expected,
        "received_qty": received,
        "remaining": max(raw_remaining, 0.0),
        "over_receipt_qty": max(-raw_remaining, 0.0),
        "receipt_count": int(row["receipt_count"] or 0),
        "last_received_at": row["last_received_at"].isoformat() if row.get("last_received_at") else None,
        "expected_date": exp_date.isoformat() if exp_date else None,
        "is_overdue": is_overdue,
        "days_overdue": (today - exp_date).days if is_overdue else 0,
        "reference_number": row["reference_number"],
        "notes": row["notes"],
        "status": row["status"],
        "created_at": row["created_at"].isoformat() if row.get("created_at") else None,
        "created_by": row["created_by"],
        "updated_at": row["updated_at"].isoformat() if row.get("updated_at") else None,
    }


def fetch_expected_receipt(cur, expected_receipt_id: int) -> Optional[dict]:
    cur.execute(EXPECTED_RECEIPT_SELECT_SQL + " WHERE er.id = %s", (expected_receipt_id,))
    row = cur.fetchone()
    return _serialize_expected_receipt(dict(row)) if row else None


def find_open_expected_receipt(cur, product_id: int, supplier_id: int, lock: bool = False) -> Optional[dict]:
    """FIFO auto-match: the open expected receipt for (product, supplier) with
    the oldest expected_date (NULL dates last, then lowest id)."""
    cur.execute(
        f"""SELECT id, expected_qty, expected_date, reference_number
            FROM expected_receipts
            WHERE status = 'open' AND product_id = %s AND supplier_id = %s
            ORDER BY expected_date ASC NULLS LAST, id ASC
            LIMIT 1{' FOR UPDATE' if lock else ''}""",
        (product_id, supplier_id),
    )
    row = cur.fetchone()
    return dict(row) if row else None


def preview_expected_receipt_match(cur, product_id: int, shipper_name: str) -> Optional[dict]:
    """Non-locking lookahead used by receive preview. Returns the FIFO match
    (with computed remaining) or None. Never raises."""
    supplier = resolve_supplier(cur, shipper_name)
    if not supplier or not supplier["active"]:
        return None
    match = find_open_expected_receipt(cur, product_id, supplier["id"])
    if not match:
        return None
    return fetch_expected_receipt(cur, match["id"])


def settle_expected_receipt(cur, expected_receipt_id: int) -> dict:
    """After a receipt is linked: recompute received (ledger SUM) and auto-close
    when received >= expected. Over-receipt closes too. Returns the summary."""
    cur.execute(
        "SELECT id, expected_qty, status FROM expected_receipts WHERE id = %s FOR UPDATE",
        (expected_receipt_id,),
    )
    er = cur.fetchone()
    cur.execute(EXPECTED_RECEIPT_RECEIVED_SQL, (expected_receipt_id,))
    rcv = cur.fetchone()
    expected = float(er["expected_qty"])
    received = float(rcv["received_qty"] or 0)
    auto_closed = False
    status = er["status"]
    if status == "open" and received + BALANCE_EPSILON >= expected:
        cur.execute(
            "UPDATE expected_receipts SET status = 'closed', updated_at = clock_timestamp() WHERE id = %s",
            (expected_receipt_id,),
        )
        status = "closed"
        auto_closed = True
    return {
        "id": expected_receipt_id,
        "expected_qty": expected,
        "received_qty": received,
        "remaining": max(expected - received, 0.0),
        "over_receipt_qty": max(received - expected, 0.0),
        "status": status,
        "auto_closed": auto_closed,
    }


# ── Suppliers ──────────────────────────────────────────────────

@app.get("/suppliers")
def list_suppliers(
    q: Optional[str] = Query(None, description="Substring / fuzzy filter"),
    include_inactive: bool = Query(False),
    _: bool = Depends(verify_api_key),
):
    """Curated supplier list (backfilled from ledger shipper names by migration 041)."""
    with get_transaction() as cur:
        if q and q.strip():
            cur.execute(
                """SELECT id, name, active, created_at,
                          similarity(supplier_name_norm(name), supplier_name_norm(%s)) AS sim
                   FROM suppliers
                   WHERE (%s OR active)
                     AND (supplier_name_norm(name) LIKE '%%' || supplier_name_norm(%s) || '%%'
                          OR similarity(supplier_name_norm(name), supplier_name_norm(%s)) > 0.2)
                   ORDER BY sim DESC, name
                   LIMIT 50""",
                (q, include_inactive, q, q),
            )
        else:
            cur.execute(
                """SELECT id, name, active, created_at, NULL::real AS sim
                   FROM suppliers WHERE (%s OR active) ORDER BY name""",
                (include_inactive,),
            )
        rows = cur.fetchall()
    return {
        "suppliers": [
            {"id": r["id"], "name": r["name"], "active": r["active"],
             "created_at": r["created_at"].isoformat() if r["created_at"] else None}
            for r in rows
        ],
        "count": len(rows),
    }


@app.post("/suppliers", status_code=201)
def create_supplier(req: SupplierCreate, _: bool = Depends(verify_api_key)):
    """Explicitly create a supplier. 409 if a normalised-equal name exists."""
    name = re.sub(r"\s+", " ", (req.name or "").strip())
    if not name:
        raise HTTPException(status_code=422, detail={"error_code": "SUPPLIER_NAME_REQUIRED", "message": "name is required"})
    with get_transaction() as cur:
        existing = resolve_supplier(cur, name)
        if existing:
            raise HTTPException(
                status_code=409,
                detail={
                    "error_code": "SUPPLIER_EXISTS",
                    "message": f"Supplier already exists as '{existing['name']}' (id {existing['id']}).",
                    "supplier_id": existing["id"],
                    "name": existing["name"],
                    "active": existing["active"],
                },
            )
        cur.execute(
            "INSERT INTO suppliers (name, active) VALUES (%s, %s) RETURNING id, name, active, created_at",
            (name, req.active),
        )
        row = cur.fetchone()
    return {
        "supplier_id": row["id"], "id": row["id"], "name": row["name"], "active": row["active"],
        "created_at": row["created_at"].isoformat(),
        "message": f"Supplier '{row['name']}' created",
    }


# ── Expected receipts ──────────────────────────────────────────

@app.post("/expected-receipts", status_code=201)
def create_expected_receipt(req: ExpectedReceiptCreate, request: Request, _: bool = Depends(verify_api_key)):
    """Record an expected/incoming delivery (lb). supplier_name must resolve to an
    existing supplier — otherwise 422 with up to 5 candidate names. Never
    auto-creates suppliers. No inventory effect."""
    if req.expected_qty is None or req.expected_qty <= 0:
        raise HTTPException(status_code=422, detail={"error_code": "INVALID_QUANTITY", "message": "expected_qty must be > 0 (lb)"})
    if not req.product_id and not (req.product_name and req.product_name.strip()):
        raise HTTPException(status_code=422, detail={"error_code": "PRODUCT_REQUIRED", "message": "product_id or product_name is required"})

    with get_transaction() as cur:
        if req.product_id:
            cur.execute("SELECT id, name FROM products WHERE id = %s", (req.product_id,))
            prod = cur.fetchone()
            if not prod:
                raise HTTPException(status_code=404, detail={"error_code": "PRODUCT_NOT_FOUND", "message": f"Product id {req.product_id} not found"})
            product_id = prod["id"]
        else:
            product_id, _name = resolve_product_id(cur, req.product_name.strip())

        supplier = require_supplier(cur, req.supplier_name)

        cur.execute(
            """INSERT INTO expected_receipts
                   (product_id, supplier_id, expected_qty, expected_date, reference_number, notes, created_by)
               VALUES (%s, %s, %s, %s, %s, %s, %s)
               RETURNING id""",
            (
                product_id, supplier["id"], req.expected_qty, req.expected_date,
                (req.reference_number or "").strip() or None,
                (req.notes or "").strip() or None,
                caller_source_tag(request, req.created_by),
            ),
        )
        new_id = cur.fetchone()["id"]
        record = fetch_expected_receipt(cur, new_id)

    logger.info(f"Expected receipt {new_id} created: {record['expected_qty']} lb {record['product_name']} from {record['supplier_name']}")
    return {
        "expected_receipt_id": new_id,
        "expected_receipt": record,
        "message": (
            f"Expecting {record['expected_qty']:g} lb of {record['product_name']} from {record['supplier_name']}"
            + (f" on {record['expected_date']}" if record["expected_date"] else "")
        ),
    }


@app.get("/expected-receipts")
def list_expected_receipts(
    status: Optional[str] = Query(None, description="open | closed | cancelled | all (default all)"),
    product_id: Optional[int] = Query(None),
    supplier_id: Optional[int] = Query(None),
    overdue_only: bool = Query(False),
    limit: int = Query(500, ge=1, le=2000),
    _: bool = Depends(verify_api_key),
):
    """List expected receipts with computed remaining (ledger SUM, floored at 0)
    and is_overdue (open AND expected_date < today, plant timezone)."""
    where, params = [], []
    if status and status != "all":
        if status not in ("open", "closed", "cancelled"):
            raise HTTPException(status_code=422, detail={"error_code": "INVALID_STATUS", "message": "status must be open, closed, cancelled or all"})
        where.append("er.status = %s")
        params.append(status)
    if product_id is not None:
        where.append("er.product_id = %s")
        params.append(product_id)
    if supplier_id is not None:
        where.append("er.supplier_id = %s")
        params.append(supplier_id)
    today = get_plant_now().date()
    if overdue_only:
        where.append("er.status = 'open' AND er.expected_date < %s")
        params.append(today)
    sql = EXPECTED_RECEIPT_SELECT_SQL
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += """ ORDER BY (er.status = 'open') DESC, er.expected_date ASC NULLS LAST, er.id ASC
               LIMIT %s"""
    params.append(limit)
    with get_transaction() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()
    items = [_serialize_expected_receipt(dict(r), today) for r in rows]
    return {
        "expected_receipts": items,
        "count": len(items),
        "open_count": sum(1 for i in items if i["status"] == "open"),
        "overdue_count": sum(1 for i in items if i["is_overdue"]),
        "as_of": today.isoformat(),
    }


@app.get("/expected-receipts/{expected_receipt_id}")
def get_expected_receipt(expected_receipt_id: int, _: bool = Depends(verify_api_key)):
    with get_transaction() as cur:
        record = fetch_expected_receipt(cur, expected_receipt_id)
        if not record:
            raise HTTPException(status_code=404, detail={"error_code": "EXPECTED_RECEIPT_NOT_FOUND", "message": f"Expected receipt {expected_receipt_id} not found"})
        cur.execute(
            """SELECT t.id AS transaction_id, t.timestamp, t.bol_reference, t.shipper_name,
                      ct.effective_status,
                      COALESCE(SUM(tl.quantity_lb), 0) AS quantity_lb,
                      MIN(l.lot_code) AS lot_code
               FROM transactions t
               JOIN ledger_current_transactions ct ON ct.id = t.id
               LEFT JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
               LEFT JOIN lots l ON l.id = tl.lot_id
               WHERE t.expected_receipt_id = %s
               GROUP BY t.id, t.timestamp, t.bol_reference, t.shipper_name, ct.effective_status
               ORDER BY t.timestamp, t.id""",
            (expected_receipt_id,),
        )
        receipts = [
            {
                "transaction_id": r["transaction_id"],
                "timestamp": r["timestamp"].isoformat() if r["timestamp"] else None,
                "bol_reference": r["bol_reference"],
                "shipper_name": r["shipper_name"],
                "lot_code": r["lot_code"],
                "quantity_lb": float(r["quantity_lb"] or 0),
                "status": r["effective_status"],
                "counted": r["effective_status"] == "posted",
            }
            for r in cur.fetchall()
        ]
    record["linked_receipts"] = receipts
    return record


@app.patch("/expected-receipts/{expected_receipt_id}")
def update_expected_receipt(expected_receipt_id: int, req: ExpectedReceiptUpdate, _: bool = Depends(verify_api_key)):
    """Edit qty/date/reference/notes while open; or move status to closed /
    cancelled (only from open). Omitted fields untouched; null clears."""
    data = req.dict(exclude_unset=True)
    if not data:
        raise HTTPException(status_code=422, detail={"error_code": "NO_FIELDS", "message": "No fields to update"})
    field_edits = {k: v for k, v in data.items() if k != "status"}
    new_status = data.get("status")

    with get_transaction() as cur:
        cur.execute("SELECT id, status FROM expected_receipts WHERE id = %s FOR UPDATE", (expected_receipt_id,))
        er = cur.fetchone()
        if not er:
            raise HTTPException(status_code=404, detail={"error_code": "EXPECTED_RECEIPT_NOT_FOUND", "message": f"Expected receipt {expected_receipt_id} not found"})
        if er["status"] != "open":
            raise HTTPException(
                status_code=409,
                detail={
                    "error_code": "EXPECTED_RECEIPT_NOT_OPEN",
                    "message": f"Expected receipt {expected_receipt_id} is {er['status']}; only open records can be edited or closed/cancelled.",
                    "status": er["status"],
                },
            )
        sets, params = [], []
        if "expected_qty" in field_edits:
            q = field_edits["expected_qty"]
            if q is None or q <= 0:
                raise HTTPException(status_code=422, detail={"error_code": "INVALID_QUANTITY", "message": "expected_qty must be > 0 (lb)"})
            sets.append("expected_qty = %s"); params.append(q)
        if "expected_date" in field_edits:
            sets.append("expected_date = %s"); params.append(field_edits["expected_date"])
        if "reference_number" in field_edits:
            v = (field_edits["reference_number"] or "").strip() or None
            sets.append("reference_number = %s"); params.append(v)
        if "notes" in field_edits:
            v = (field_edits["notes"] or "").strip() or None
            sets.append("notes = %s"); params.append(v)
        if new_status:
            sets.append("status = %s"); params.append(new_status)
        sets.append("updated_at = clock_timestamp()")
        params.append(expected_receipt_id)
        cur.execute(f"UPDATE expected_receipts SET {', '.join(sets)} WHERE id = %s", params)
        record = fetch_expected_receipt(cur, expected_receipt_id)

    changed = sorted(field_edits.keys()) + (["status"] if new_status else [])
    return {
        "expected_receipt_id": expected_receipt_id,
        "expected_receipt": record,
        "changed_fields": changed,
        "message": f"Expected receipt {expected_receipt_id} " + (f"{new_status}" if new_status else "updated"),
    }


# ═══════════════════════════════════════════════════════════════
# SUPPLIES (dashboard-only) — packaging / consumables inventory view,
# low-stock thresholds, supply-request queue. Migration 043.
# Not in openapi-gpt-v3.yaml on purpose (30-op cap; GPTs don't need these).
# ═══════════════════════════════════════════════════════════════

# The Supplies filter intentionally exposes only supply-relevant categories;
# the unfiltered endpoint still returns every products.type value.
PRODUCT_CATEGORIES = ("ingredient", "packaging", "consumable")

# Every product (active or not, zero inventory or not) with its posted-only
# ledger SUM. LEFT JOINs so a product with no lots / no lines still appears
# with on_hand 0. This is a read of the ledger, never of a stored balance.
SUPPLIES_INVENTORY_SQL = f"""
    SELECT p.id AS product_id, p.name, p.odoo_code, p.type AS category, p.uom,
           p.active, p.case_size_lb, p.low_stock_threshold,
           COALESCE(SUM(tl.quantity_lb), 0) AS on_hand
    FROM products p
    LEFT JOIN lots l ON l.product_id = p.id
    LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
"""


def _serialize_supplies_item(row: dict) -> dict:
    on_hand = float(row["on_hand"] or 0)
    threshold = float(row["low_stock_threshold"]) if row.get("low_stock_threshold") is not None else None
    return {
        "product_id": row["product_id"],
        "name": row["name"],
        "sku": row.get("odoo_code"),
        "category": row["category"],
        "unit": row.get("uom") or "lb",
        "active": bool(row["active"]) if row.get("active") is not None else True,
        "case_size_lb": float(row["case_size_lb"]) if row.get("case_size_lb") is not None else None,
        "on_hand": round(on_hand, 4),
        "low_stock_threshold": threshold,
        # is_low is false whenever no threshold is set (NULL = no alerting)
        "is_low": bool(threshold is not None and on_hand < threshold),
    }


@app.get("/supplies/inventory")
def supplies_inventory(
    category: Optional[str] = Query(None, description="ingredient | packaging | consumable"),
    active_only: bool = Query(False, description="Drop inactive products (default: every product)"),
    low_only: bool = Query(False, description="Only rows with is_low = true"),
    _: bool = Depends(verify_api_key),
):
    """ALL products with on_hand as the posted-only ledger SUM (zero-inventory
    products included), low_stock_threshold and is_low (on_hand < threshold;
    false when threshold is null). Alphabetical by name."""
    where, params = [], []
    if category:
        cat = category.strip().lower()
        if cat not in PRODUCT_CATEGORIES:
            raise HTTPException(
                status_code=422,
                detail={"error_code": "INVALID_CATEGORY",
                        "message": f"category must be one of {', '.join(PRODUCT_CATEGORIES)}",
                        "input": category},
            )
        where.append("p.type = %s")
        params.append(cat)
    if active_only:
        where.append("COALESCE(p.active, true) = true")
    sql = SUPPLIES_INVENTORY_SQL
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " GROUP BY p.id ORDER BY lower(p.name), p.id"
    with get_transaction() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()
    items = [_serialize_supplies_item(dict(r)) for r in rows]
    if low_only:
        items = [i for i in items if i["is_low"]]
    return {
        "items": items,
        "count": len(items),
        "low_count": sum(1 for i in items if i["is_low"]),
        "category": category.strip().lower() if category else None,
    }


@app.get("/supplies/inventory/{product_id}/lots")
def supplies_product_lots(
    product_id: int,
    include_empty: bool = Query(False, description="Also list depleted / negative lots"),
    _: bool = Depends(verify_api_key),
):
    """FIFO lot breakdown for one product: lot code, lot date (the FIFO key),
    remaining per lot as the posted-only ledger SUM, in the exact order the
    allocation paths consume lots (fifo_lot_balances)."""
    with get_transaction() as cur:
        cur.execute(
            "SELECT id AS product_id, name, odoo_code, type AS category, uom, active, case_size_lb, low_stock_threshold "
            "FROM products WHERE id = %s",
            (product_id,),
        )
        prod = cur.fetchone()
        if not prod:
            raise HTTPException(status_code=404, detail={"error_code": "PRODUCT_NOT_FOUND", "message": f"Product id {product_id} not found"})
        lots = fifo_lot_balances(cur, product_id, include_empty=include_empty)
    unit = prod["uom"] or "lb"
    out_lots = []
    for rank, l in enumerate(lots, start=1):
        out_lots.append({
            "fifo_rank": rank,
            "lot_id": l["id"],
            "lot_code": l["lot_code"],
            "lot_date": l["lot_date"].isoformat() if l.get("lot_date") else None,
            "received_at": l["received_at"].isoformat() if l.get("received_at") else None,
            "created_at": l["created_at"].isoformat() if l.get("created_at") else None,
            "entry_source": l.get("entry_source"),
            "supplier_lot_code": l.get("supplier_lot_code"),
            "lot_status": l.get("status") or "active",
            "remaining": round(float(l["available"] or 0), 4),
            "unit": unit,
        })
    total = round(sum(x["remaining"] for x in out_lots), 4)
    threshold = float(prod["low_stock_threshold"]) if prod["low_stock_threshold"] is not None else None
    return {
        "product_id": prod["product_id"],
        "name": prod["name"],
        "sku": prod["odoo_code"],
        "category": prod["category"],
        "unit": unit,
        "active": bool(prod["active"]) if prod["active"] is not None else True,
        "low_stock_threshold": threshold,
        "on_hand": total,
        "is_low": bool(threshold is not None and total < threshold),
        "fifo_order": "COALESCE(received_at, created_at) ASC",
        "include_empty": include_empty,
        "lot_count": len(out_lots),
        "lots": out_lots,
    }


# ── Supply requests ────────────────────────────────────────────

SUPPLY_REQUEST_SELECT_SQL = """
    SELECT sr.id, sr.product_id, sr.item_text, sr.qty, sr.note, sr.requested_by,
           sr.status, sr.created_at, sr.done_at,
           p.name AS product_name, p.odoo_code, p.type AS category, p.uom
    FROM supply_requests sr
    LEFT JOIN products p ON p.id = sr.product_id
"""


def _serialize_supply_request(row: dict) -> dict:
    return {
        "id": row["id"],
        "product_id": row["product_id"],
        "product_name": row.get("product_name"),
        "sku": row.get("odoo_code"),
        "category": row.get("category"),
        "unit": (row.get("uom") or "lb") if row["product_id"] is not None else None,
        "item_text": row["item_text"],
        "display_name": row.get("product_name") if row["product_id"] is not None else row["item_text"],
        "qty": float(row["qty"]) if row["qty"] is not None else None,
        "note": row["note"],
        "requested_by": row["requested_by"],
        "status": row["status"],
        "created_at": row["created_at"].isoformat() if row.get("created_at") else None,
        "done_at": row["done_at"].isoformat() if row.get("done_at") else None,
    }


def fetch_supply_request(cur, supply_request_id: int) -> Optional[dict]:
    cur.execute(SUPPLY_REQUEST_SELECT_SQL + " WHERE sr.id = %s", (supply_request_id,))
    row = cur.fetchone()
    return _serialize_supply_request(dict(row)) if row else None


@app.post("/supply-requests", status_code=201)
def create_supply_request(req: SupplyRequestCreate, _: bool = Depends(verify_api_key)):
    """Create a supply request. Exactly one of product_id / item_text (422
    otherwise). No inventory effect."""
    item_text = re.sub(r"\s+", " ", (req.item_text or "").strip()) or None
    has_product = req.product_id is not None
    if has_product and item_text:
        raise HTTPException(status_code=422, detail={
            "error_code": "SUPPLY_REQUEST_TARGET_AMBIGUOUS",
            "message": "Send either product_id or item_text, not both",
        })
    if not has_product and not item_text:
        raise HTTPException(status_code=422, detail={
            "error_code": "SUPPLY_REQUEST_TARGET_REQUIRED",
            "message": "Either product_id (catalogued product) or item_text (free text) is required",
        })
    requested_by = re.sub(r"\s+", " ", (req.requested_by or "").strip())
    if not requested_by:
        raise HTTPException(status_code=422, detail={"error_code": "REQUESTED_BY_REQUIRED", "message": "requested_by is required"})
    if req.qty is not None and req.qty <= 0:
        raise HTTPException(status_code=422, detail={"error_code": "INVALID_QUANTITY", "message": "qty must be > 0 when provided"})

    with get_transaction() as cur:
        if has_product:
            cur.execute("SELECT id FROM products WHERE id = %s", (req.product_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail={"error_code": "PRODUCT_NOT_FOUND", "message": f"Product id {req.product_id} not found"})
        cur.execute(
            """INSERT INTO supply_requests (product_id, item_text, qty, note, requested_by)
               VALUES (%s, %s, %s, %s, %s) RETURNING id""",
            (req.product_id if has_product else None, item_text, req.qty,
             (req.note or "").strip() or None, requested_by[:120]),
        )
        new_id = cur.fetchone()["id"]
        record = fetch_supply_request(cur, new_id)

    logger.info(f"Supply request {new_id} created by {requested_by}: {record['display_name']} qty={record['qty']}")
    return {
        "supply_request_id": new_id,
        "supply_request": record,
        "message": f"Supply request #{new_id} opened for {record['display_name']}",
    }


@app.get("/supply-requests")
def list_supply_requests(
    status: Optional[str] = Query("all", description="open | done | all (default all)"),
    limit: int = Query(500, ge=1, le=2000),
    _: bool = Depends(verify_api_key),
):
    """List supply requests, newest first."""
    where, params = [], []
    st = (status or "all").strip().lower()
    if st != "all":
        if st not in ("open", "done"):
            raise HTTPException(status_code=422, detail={"error_code": "INVALID_STATUS", "message": "status must be open, done or all"})
        where.append("sr.status = %s")
        params.append(st)
    sql = SUPPLY_REQUEST_SELECT_SQL
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY sr.created_at DESC, sr.id DESC LIMIT %s"
    params.append(limit)
    with get_transaction() as cur:
        cur.execute(sql, params)
        rows = cur.fetchall()
    items = [_serialize_supply_request(dict(r)) for r in rows]
    return {
        "supply_requests": items,
        "count": len(items),
        "open_count": sum(1 for i in items if i["status"] == "open"),
        "done_count": sum(1 for i in items if i["status"] == "done"),
        "status": st,
    }


@app.patch("/supply-requests/{supply_request_id}")
def update_supply_request(supply_request_id: int, req: SupplyRequestUpdate, _: bool = Depends(verify_api_key)):
    """The only transition: open -> done (sets done_at). Anything else 409."""
    with get_transaction() as cur:
        cur.execute("SELECT id, status FROM supply_requests WHERE id = %s FOR UPDATE", (supply_request_id,))
        sr = cur.fetchone()
        if not sr:
            raise HTTPException(status_code=404, detail={"error_code": "SUPPLY_REQUEST_NOT_FOUND", "message": f"Supply request {supply_request_id} not found"})
        if sr["status"] != "open":
            raise HTTPException(status_code=409, detail={
                "error_code": "SUPPLY_REQUEST_NOT_OPEN",
                "message": f"Supply request {supply_request_id} is already {sr['status']}; only open requests can be marked done.",
                "status": sr["status"],
            })
        cur.execute(
            "UPDATE supply_requests SET status = 'done', done_at = clock_timestamp() WHERE id = %s",
            (supply_request_id,),
        )
        record = fetch_supply_request(cur, supply_request_id)
    return {
        "supply_request_id": supply_request_id,
        "supply_request": record,
        "changed_fields": ["status", "done_at"],
        "message": f"Supply request #{supply_request_id} marked done",
    }


# ═══════════════════════════════════════════════════════════════
# SHIP ENDPOINTS
# ═══════════════════════════════════════════════════════════════


def check_open_orders_for_ship(cur, customer_id: int, customer_name: str) -> dict | None:
    """Check if a customer has open sales order lines with remaining quantity.
    Returns a structured payload dict suitable for 409 responses, or None if no open orders.
    Uses customer_id (not fuzzy name LIKE) for reliable matching.
    """
    cur.execute("""
        SELECT so.id AS order_id, so.order_number, so.status, so.requested_ship_date,
               sol.id AS line_id, sol.product_id,
               p.name AS product_name,
               sol.quantity_lb - sol.quantity_shipped_lb AS remaining_lb
        FROM sales_orders so
        JOIN sales_order_lines sol ON sol.sales_order_id = so.id
        JOIN products p ON p.id = sol.product_id
        WHERE so.customer_id = %s
          AND so.status NOT IN ('shipped', 'invoiced', 'cancelled')
          AND sol.line_status NOT IN ('fulfilled', 'cancelled')
          AND sol.quantity_lb - sol.quantity_shipped_lb > 0
        ORDER BY so.requested_ship_date ASC NULLS LAST, so.id ASC
    """, (customer_id,))
    matches = cur.fetchall()

    if not matches:
        return None

    # Group by order
    orders = {}
    for m in matches:
        oid = m['order_id']
        if oid not in orders:
            orders[oid] = {
                "order_id": oid,
                "order_number": m['order_number'],
                "status": m['status'],
                "requested_ship_date": str(m['requested_ship_date']) if m['requested_ship_date'] else None,
                "lines": [],
                "use_instead": {
                    "method": "POST",
                    "url": f"/sales/orders/{oid}/ship",
                    "note": "Set mode=commit in request body"
                }
            }
        orders[oid]["lines"].append({
            "line_id": m['line_id'],
            "product_name": m['product_name'],
            "remaining_lb": float(m['remaining_lb'])
        })

    total_remaining = sum(float(m['remaining_lb']) for m in matches)
    orders_list = list(orders.values())
    order_nums = ", ".join(o['order_number'] for o in orders_list)

    return {
        "error_code": "OPEN_SALES_ORDER_EXISTS",
        "message": (
            f"Standalone ship blocked: '{customer_name}' has {len(orders_list)} open sales order(s) "
            f"({order_nums}) with {total_remaining:,.0f} lb remaining. "
            f"Use the order ship endpoint instead, or set force_standalone=true to bypass."
        ),
        "open_orders": orders_list,
        "note": "No inventory was moved."
    }


@app.post("/ship")
def ship(req: ShipRequest, _: bool = Depends(verify_api_key)):
    """Ship inventory. mode=preview returns allocation plan; mode=commit executes."""
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at, req.backfill
    )
    if req.mode == "preview":
        try:
            with get_transaction() as cur:
                product = resolve_product_full(cur, req.product_name)

                open_orders_warning = None
                try:
                    cust_id, cust_canonical = resolve_customer_id(
                        cur, req.customer_name, auto_create=False, address=req.customer_address
                    )
                except HTTPException as e:
                    if e.status_code == 409:
                        raise
                    cust_id, cust_canonical = None, req.customer_name

                if cust_id and not req.force_standalone:
                    oo_payload = check_open_orders_for_ship(cur, cust_id, cust_canonical)
                    if oo_payload:
                        open_orders_warning = oo_payload

                lots = available_lots_for_product(cur, int(product['id']))
                reservation_summary = _allocation_reservation_summary(
                    cur, int(product['id'])
                )

                if not lots:
                    if open_orders_warning:
                        raise HTTPException(status_code=409, detail={
                            "error_code": "OPEN_SALES_ORDER_EXISTS",
                            "message": f"No inventory for {product['name']}, AND {cust_canonical} has open sales orders. "
                                       f"Use the order ship endpoint to check fulfillment feasibility.",
                            "open_orders": open_orders_warning.get("open_orders", []),
                            "note": "No inventory available for standalone shipping."
                        })
                    raise HTTPException(status_code=400, detail=f"No inventory available for {product['name']}")

                total_available = sum(float(l['on_hand']) for l in lots)
                if total_available < req.quantity_lb:
                    raise HTTPException(status_code=400,
                        detail=f"Insufficient total inventory for {product['name']}. "
                               f"Have {total_available} lb across {len(lots)} lot(s), need {req.quantity_lb} lb")

                # Pin fidelity (owner ruling 2026-08-20): a pinned lot_code
                # restricts the plan to that lot only — never spill to other
                # lots to avoid its reserved pounds.
                plan_lots = lots
                if req.lot_code:
                    selected = next((l for l in lots if l['lot_code'].lower() == req.lot_code.lower()), None)
                    if not selected:
                        raise HTTPException(status_code=404, detail=f"Lot '{req.lot_code}' not found or empty")
                    plan_lots = [selected]

                plan = _takeable_deduction_plan(plan_lots, req.quantity_lb)
                if plan["short_lb"] > BALANCE_EPSILON:
                    raise HTTPException(status_code=400,
                        detail=f"Insufficient total inventory for {product['name']}. "
                               f"Have {plan['total_on_hand_lb']:.4f} lb across "
                               f"{len(plan_lots)} lot(s), need {req.quantity_lb} lb")
                allocations = [
                    {
                        "lot_code": item["lot"]["lot_code"],
                        "lot_id": item["lot"]["lot_id"],
                        "available_lb": float(item["lot"]["on_hand"]),
                        "takeable_lb": float(item["lot"]["takeable"]),
                        "allocated_lb": float(item["quantity_lb"]),
                    }
                    for item in plan["lots"]
                ]
                ship_mode = "single_lot" if len(allocations) == 1 else "multi_lot_fifo"
                if ship_mode == "single_lot":
                    preview_msg = (
                        f"Ready to ship {req.quantity_lb} lb of {product['name']} "
                        f"from lot {allocations[0]['lot_code']}"
                    )
                else:
                    preview_msg = (f"Will ship {req.quantity_lb} lb of {product['name']} "
                                   f"from {len(allocations)} lot(s) (reservation-aware FIFO)")
                allocation_warning = _allocation_observe_warning(
                    "Standalone ship",
                    req.quantity_lb,
                    plan["can_take_lb"],
                    plan["reserved_taken_lb"],
                    reservation_summary,
                    preview=True,
                )

                response = {
                    "mode": "preview",
                    "product_id": product['id'],
                    "product_name": product['name'],
                    "odoo_code": product['odoo_code'],
                    "quantity_lb": req.quantity_lb,
                    "customer_name": cust_canonical,
                    "order_reference": req.order_reference,
                    "ship_mode": ship_mode,
                    "allocations": allocations,
                    "total_available_lb": total_available,
                    "total_takeable_lb": sum(float(l['takeable'] or 0) for l in lots),
                    "can_ship_lb": plan["can_take_lb"],
                    "reserved_others_lb": reservation_summary["reserved_others_lb"],
                    "reserved_by_orders": reservation_summary["reserved_by_orders"],
                    "open_orders_warning": open_orders_warning,
                    "preview_message": preview_msg
                }
                if allocation_warning:
                    response["allocation_warning"] = allocation_warning
                    response["warning"] = allocation_warning["message"]
                return response
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Ship preview failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})
    else:
        # mode == "commit"
        try:
            with get_db_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    product = resolve_product_full(cur, req.product_name)

                    customer_id, canonical_customer = resolve_customer_id(
                        cur, req.customer_name,
                        auto_create=True,
                        force_create=req.force_create_customer,
                        address=req.customer_address
                    )

                    standalone_warning = None
                    oo_payload = check_open_orders_for_ship(cur, customer_id, canonical_customer)
                    if oo_payload:
                        if not req.force_standalone:
                            raise HTTPException(status_code=409, detail=oo_payload)
                        # force_standalone=True: allow but log and tag
                        n_open = len(oo_payload.get("open_orders", []))
                        standalone_warning = f"This shipment was not linked to any sales order. Customer has {n_open} open orders."

                    all_lots = available_lots_for_product(
                        cur,
                        int(product['id']),
                        lock=True,
                        persist_expired=True,
                        released_by=_operator_id(_),
                    )
                    reservation_summary = _allocation_reservation_summary(
                        cur, int(product['id'])
                    )

                    if not all_lots:
                        raise HTTPException(status_code=400, detail=f"No inventory available for {product['name']}")

                    # Pin fidelity (owner ruling 2026-08-20): a pinned lot_code
                    # restricts the plan to that lot only — never spill to
                    # other lots to avoid its reserved pounds. Observe mode
                    # takes the pinned lot's reserved stock instead (warning +
                    # inventory_shipped shrink below).
                    plan_lots = all_lots
                    if req.lot_code:
                        plan_lots = [
                            lot for lot in all_lots
                            if lot['lot_code'].lower() == req.lot_code.lower()
                        ]
                        if not plan_lots:
                            raise HTTPException(
                                status_code=400,
                                detail=f"Lot '{req.lot_code}' has no available inventory",
                            )

                    plan = _takeable_deduction_plan(plan_lots, req.quantity_lb)
                    if plan["short_lb"] > BALANCE_EPSILON:
                        raise HTTPException(
                            status_code=400,
                            detail=f"Insufficient total inventory for {product['name']}. "
                                   f"Have {plan['total_on_hand_lb']:.4f} lb across "
                                   f"{len(plan_lots)} lot(s), need {req.quantity_lb} lb",
                        )
                    allocation_warning = _allocation_observe_warning(
                        "Standalone ship",
                        req.quantity_lb,
                        plan["can_take_lb"],
                        plan["reserved_taken_lb"],
                        reservation_summary,
                    )
                    _enforce_allocation_takeable(
                        "Standalone ship",
                        req.quantity_lb,
                        plan["can_take_lb"],
                        plan["reserved_taken_lb"],
                        reservation_summary,
                        product_id=int(product["id"]),
                    )

                    txn_notes = None
                    if standalone_warning:
                        txn_notes = f"standalone_override=true | {standalone_warning}"
                    now = get_plant_now()
                    cur.execute("""
                        INSERT INTO transactions (
                            type, timestamp, customer_name, order_reference,
                            notes, occurred_at, created_at_source
                        )
                        VALUES ('ship', %s, %s, %s, %s, %s, %s) RETURNING id
                    """, (
                        now, canonical_customer, req.order_reference, txn_notes,
                        occurred_at, created_at_source,
                    ))
                    txn_id = cur.fetchone()['id']

                    shipped_lots = []
                    for item in plan["lots"]:
                        lot = item["lot"]
                        take = float(item["quantity_lb"])
                        cur.execute("""
                            INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb)
                            VALUES (%s, %s, %s, %s)
                        """, (txn_id, product['id'], lot['lot_id'], -take))
                        shipped_lots.append({"lot_code": lot['lot_code'], "shipped_lb": take})

                    ship_mode = "single_lot" if len(shipped_lots) == 1 else "multi_lot_fifo"
                    if ship_mode == "single_lot":
                        only_lot = plan["lots"][0]["lot"]
                        remaining = float(only_lot["on_hand"]) - req.quantity_lb
                        log_msg = (
                            f"Shipped {req.quantity_lb} lb from lot {only_lot['lot_code']}. "
                            f"{remaining} lb remaining in lot."
                        )
                    else:
                        remaining = None
                        log_msg = (
                            f"Shipped {req.quantity_lb} lb from {len(shipped_lots)} lot(s) "
                            "(reservation-aware FIFO)."
                        )

                    # ── Create shipment + shipment_lines (unified shipping model) ──
                    cur.execute("""
                        INSERT INTO shipments (transaction_id, shipped_at, customer_id)
                        VALUES (%s, %s, %s) RETURNING id
                    """, (txn_id, occurred_at or now, customer_id))
                    shipment_id = cur.fetchone()['id']

                    for sl in shipped_lots:
                        cur.execute("""
                            INSERT INTO shipment_lines (shipment_id, transaction_id, product_id, quantity_lb)
                            VALUES (%s, %s, %s, %s)
                        """, (shipment_id, txn_id, product['id'], sl['shipped_lb']))

                    allocations_released = []
                    if plan["reserved_taken_lb"] > BALANCE_EPSILON:
                        allocations_released = _shrink_overallocated_products(
                            cur,
                            [int(product['id'])],
                            _operator_id(_),
                            release_reason="inventory_shipped",
                        )

                    if standalone_warning:
                        logger.warning(f"force_standalone ship for {canonical_customer} who has {len(oo_payload.get('open_orders', []))} open orders: txn {txn_id}")

                    logger.info(f"Ship committed: {req.quantity_lb} lb of {product['name']} to {canonical_customer} ({ship_mode})")

                    response = {
                        "mode": "commit",
                        "success": True,
                        "transaction_id": txn_id,
                        "shipment_id": shipment_id,
                        "confirmation_code": generate_confirmation_code(txn_id),
                        "quantity_shipped": req.quantity_lb,
                        "ship_mode": ship_mode,
                        "lots_used": shipped_lots,
                        "remaining_in_lot": remaining,
                        "can_ship_lb": plan["can_take_lb"],
                        "reserved_others_lb": reservation_summary["reserved_others_lb"],
                        "reserved_by_orders": reservation_summary["reserved_by_orders"],
                        "allocations_released": allocations_released,
                        "customer_name": canonical_customer,
                        "message": log_msg
                    }
                    if standalone_warning:
                        response["standalone_override"] = True
                        response["warning"] = standalone_warning
                    if allocation_warning:
                        response["allocation_warning"] = allocation_warning
                        if "warning" not in response:
                            response["warning"] = allocation_warning["message"]
                    return response
        except HTTPException:
            raise
        except Exception as e:
            if _is_readonly_error(e): raise
            logger.error(f"Ship commit failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# MAKE (PRODUCTION) ENDPOINTS
# ═══════════════════════════════════════════════════════════════

def build_production_warning(product: dict) -> dict | None:
    """Non-empty verification_notes on the product being made (e.g. Kosher
    Ignition on 90025/90026) → warning block the GPT must relay to the
    operator verbatim before committing."""
    note = (product.get('verification_notes') or '').strip()
    if not note:
        return None
    warning = {
        "verification_notes": note,
        "message": f"PRODUCTION WARNING — relay to operator verbatim before proceeding: {note}"
    }
    note_es = (product.get('verification_notes_es') or '').strip()
    if note_es:
        warning["verification_notes_es"] = note_es
    return warning


@app.post("/make")
def make(req: MakeRequest, _: bool = Depends(verify_api_key)):
    """Record batch production. mode=preview returns ingredient check; mode=commit executes."""
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at, req.backfill
    )
    if req.mode == "preview":
        try:
            with get_transaction() as cur:
                product = resolve_product_full(cur, req.product_name)
                batch_size = float(product.get('default_batch_lb') or 0)
                yield_multiplier = float(product.get('yield_multiplier') or 1.0)
                formula_weight_lb = batch_size * req.batches
                total_output = formula_weight_lb * yield_multiplier
                manual_excluded_ids = set(req.excluded_ingredients or [])

                cur.execute("""
                    SELECT bf.ingredient_product_id, p.name as ingredient_name, bf.quantity_lb,
                           COALESCE(bf.exclude_from_inventory, false) as exclude_from_inventory
                    FROM batch_formulas bf
                    JOIN products p ON p.id = bf.ingredient_product_id
                    WHERE bf.product_id = %s
                """, (product['id'],))
                formula = cur.fetchall()

                auto_excluded_ids = set()
                for ing in formula:
                    if ing.get('exclude_from_inventory'):
                        auto_excluded_ids.add(ing['ingredient_product_id'])
                excluded_ids = manual_excluded_ids | auto_excluded_ids

                ingredients_needed = []
                excluded_ingredients = []
                lot_overrides_applied = []
                lot_overrides = req.get_lot_overrides()

                all_ing_ids = [ing['ingredient_product_id'] for ing in formula
                               if ing['ingredient_product_id'] not in excluded_ids]
                ingredient_lots_map = {}
                if all_ing_ids:
                    cur.execute(f"""
                        SELECT l.product_id, l.id, l.lot_code,
                               COALESCE(SUM(tl.quantity_lb), 0) as available
                        FROM lots l
                        LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                        WHERE l.product_id = ANY(%s)
                        GROUP BY l.id
                        HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                        ORDER BY l.product_id, COALESCE(l.received_at, l.created_at) ASC
                    """, (all_ing_ids,))
                    for row in cur.fetchall():
                        pid = row['product_id']
                        if pid not in ingredient_lots_map:
                            ingredient_lots_map[pid] = []
                        ingredient_lots_map[pid].append(dict(row))

                for ing in formula:
                    ing_id = ing['ingredient_product_id']
                    needed = float(ing['quantity_lb']) * req.batches
                    if ing_id in excluded_ids:
                        excluded_ingredients.append({
                            "ingredient_id": ing_id, "ingredient_name": ing['ingredient_name'],
                            "would_need_lb": needed, "excluded": True,
                            "exclusion_type": "auto" if ing_id in auto_excluded_ids else "manual"
                        })
                        continue
                    if lot_overrides and str(ing_id) in lot_overrides:
                        override_code = lot_overrides[str(ing_id)]
                        cur.execute(f"""
                            SELECT l.id, l.lot_code, COALESCE(SUM(tl.quantity_lb), 0) as available
                            FROM lots l LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                            WHERE l.product_id = %s AND LOWER(l.lot_code) = LOWER(%s) GROUP BY l.id
                        """, (ing_id, override_code))
                        override_lot = cur.fetchone()
                        if not override_lot:
                            ingredients_needed.append({
                                "ingredient_id": ing_id, "ingredient_name": ing['ingredient_name'],
                                "needed_lb": needed, "available_lb": 0, "sufficient": False,
                                "override_lot": override_code,
                                "override_error": f"Lot '{override_code}' not found for this ingredient"
                            })
                            continue
                        avail = float(override_lot['available'])
                        lot_overrides_applied.append({
                            "ingredient_id": ing_id, "ingredient_name": ing['ingredient_name'],
                            "lot_code": override_lot['lot_code'], "needed_lb": needed,
                            "available_lb": avail, "sufficient": avail >= needed
                        })
                        ingredients_needed.append({
                            "ingredient_id": ing_id, "ingredient_name": ing['ingredient_name'],
                            "needed_lb": needed, "available_lb": avail, "sufficient": avail >= needed,
                            "override_lot": override_lot['lot_code']
                        })
                    else:
                        available_lots = ingredient_lots_map.get(ing_id, [])
                        total_avail = sum(float(lot['available']) for lot in available_lots)
                        lot_details = [{"lot_code": lot['lot_code'], "available_lb": float(lot['available'])} for lot in available_lots]
                        ingredients_needed.append({
                            "ingredient_id": ing_id, "ingredient_name": ing['ingredient_name'],
                            "needed_lb": needed, "available_lb": total_avail, "sufficient": total_avail >= needed,
                            "lot_count": len(available_lots), "lots": lot_details
                        })

                all_sufficient = all(i['sufficient'] for i in ingredients_needed)
                if req.lot_code:
                    lot_code = req.lot_code
                else:
                    now = get_plant_now()
                    date_part = now.strftime("%y-%m%d")
                    cur.execute("SELECT lot_code FROM lots WHERE lot_code LIKE %s ORDER BY lot_code DESC LIMIT 1", (f"B{date_part}-%",))
                    existing = cur.fetchone()
                    if existing:
                        try: seq = int(existing['lot_code'].split('-')[-1]) + 1
                        except (ValueError, IndexError): seq = 1
                    else: seq = 1
                    lot_code = f"B{date_part}-{seq:03d}"

                siblings = get_sibling_skus(cur, product['id'])
                yield_note = f" (estimated yield with {yield_multiplier}x multiplier; actual weight may differ)" if yield_multiplier != 1.0 else ""

                response = {
                    "mode": "preview",
                    "product_id": product['id'], "product_name": product['name'],
                    "batches": req.batches, "batch_size_lb": batch_size,
                    "yield_multiplier": yield_multiplier, "formula_weight_lb": formula_weight_lb,
                    "estimated_yield_lb": total_output, "total_output_lb": total_output,
                    "lot_code": lot_code, "ingredients": ingredients_needed,
                    "all_ingredients_available": all_sufficient,
                    "preview_message": f"Ready to make {req.batches} batch(es) of {product['name']} ({total_output} lb){yield_note}"
                }
                production_warning = build_production_warning(product)
                if production_warning:
                    response["production_warning"] = production_warning
                    response["preview_message"] += f" ⚠ {production_warning['verification_notes']}"
                if siblings:
                    sibling_names = [s['name'] for s in siblings]
                    response["sibling_skus"] = siblings
                    response["sku_confirmation_required"] = True
                    response["sku_warning"] = (
                        f"This batch source has {len(siblings) + 1} finished-good SKUs with the same formula. "
                        f"You selected '{product['name']}'. Other options: {sibling_names}. "
                        f"Confirm this is the correct output SKU before committing."
                    )
                if lot_overrides_applied:
                    response["lot_overrides"] = lot_overrides_applied
                    response["preview_message"] += f" (with {len(lot_overrides_applied)} lot override(s))"
                if excluded_ingredients:
                    auto_count = sum(1 for e in excluded_ingredients if e.get('exclusion_type') == 'auto')
                    manual_count = len(excluded_ingredients) - auto_count
                    response["excluded_ingredients"] = excluded_ingredients
                    parts = []
                    if auto_count: parts.append(f"{auto_count} auto-excluded")
                    if manual_count: parts.append(f"{manual_count} manually excluded")
                    response["preview_message"] += f" ({', '.join(parts)} ingredient(s))"
                return response
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Make preview failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})
    else:
        # mode == "commit"
        try:
            with get_db_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    product = resolve_product_full(cur, req.product_name)

                    siblings = get_sibling_skus(cur, product['id'])
                    if siblings and not req.confirmed_sku:
                        sibling_names = [s['name'] for s in siblings]
                        raise HTTPException(status_code=400, detail=(
                            f"SKU confirmation required. '{product['name']}' shares a batch formula with: "
                            f"{sibling_names}. Set confirmed_sku=true to confirm this is the correct "
                            f"output SKU. Never assume — ask the operator which SKU they are packing."
                        ))

                    batch_size = float(product.get('default_batch_lb') or 0)
                    yield_multiplier = float(product.get('yield_multiplier') or 1.0)
                    formula_weight_lb = float(to_decimal(batch_size) * to_decimal(req.batches))
                    total_output = float(to_decimal(formula_weight_lb) * to_decimal(yield_multiplier))
                    if total_output <= 0:
                        raise HTTPException(400, f"Make rejected: output quantity is 0 lb. Product '{product['name']}' has batch_size={batch_size}, batches={req.batches}.")
                    now = get_plant_now()
                    manual_excluded_ids = set(req.excluded_ingredients or [])
                    auto_excluded_ids = set()

                    if req.lot_code:
                        lot_code = req.lot_code
                    else:
                        date_part = now.strftime("%y-%m%d")
                        cur.execute("SELECT lot_code FROM lots WHERE lot_code LIKE %s ORDER BY lot_code DESC LIMIT 1", (f"B{date_part}-%",))
                        existing = cur.fetchone()
                        if existing:
                            try: seq = int(existing['lot_code'].split('-')[-1]) + 1
                            except (ValueError, IndexError): seq = 1
                        else: seq = 1
                        lot_code = f"B{date_part}-{seq:03d}"

                    output_lot_id, is_new_lot = find_or_create_lot(cur, product['id'], lot_code, 'production_output')

                    cur.execute("""
                        SELECT bf.ingredient_product_id, bf.quantity_lb,
                               COALESCE(bf.exclude_from_inventory, false) as exclude_from_inventory
                        FROM batch_formulas bf WHERE bf.product_id = %s
                    """, (product['id'],))
                    formula = cur.fetchall()

                    auto_excluded_ids = set()
                    for ing in formula:
                        if ing.get('exclude_from_inventory'):
                            auto_excluded_ids.add(ing['ingredient_product_id'])
                    excluded_ids = manual_excluded_ids | auto_excluded_ids

                    exclusion_note = ""
                    if manual_excluded_ids:
                        exclusion_note += f" (manually excluded IDs: {sorted(manual_excluded_ids)})"
                    if auto_excluded_ids:
                        exclusion_note += f" (auto-excluded IDs: {sorted(auto_excluded_ids)})"

                    cur.execute("""
                        INSERT INTO transactions (
                            type, timestamp, notes, occurred_at, created_at_source
                        )
                        VALUES ('make', %s, %s, %s, %s) RETURNING id
                    """, (
                        now, f"{req.batches} batch(es) of {product['name']}{exclusion_note}",
                        occurred_at, created_at_source,
                    ))
                    txn_id = cur.fetchone()['id']

                    cur.execute("""
                        INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb)
                        VALUES (%s, %s, %s, %s)
                    """, (txn_id, product['id'], output_lot_id, total_output))

                    consumed_by_ingredient = {}
                    excluded_from_run = []
                    lot_overrides = req.get_lot_overrides()

                    all_formula_ids = [ing['ingredient_product_id'] for ing in formula]
                    ing_names = {}
                    if all_formula_ids:
                        cur.execute("SELECT id, name FROM products WHERE id = ANY(%s)", (all_formula_ids,))
                        for row in cur.fetchall():
                            ing_names[row['id']] = row['name']

                    for ing in formula:
                        ing_id = ing['ingredient_product_id']
                        needed = float(to_decimal(ing['quantity_lb']) * to_decimal(req.batches))
                        ing_name = ing_names.get(ing_id, f"ID {ing_id}")
                        if ing_id in excluded_ids:
                            excluded_from_run.append({
                                "ingredient_id": ing_id, "ingredient_name": ing_name,
                                "skipped_lb": needed,
                                "exclusion_type": "auto" if ing_id in auto_excluded_ids else "manual"
                            })
                            continue
                        if ing_id not in consumed_by_ingredient:
                            consumed_by_ingredient[ing_id] = {
                                "ingredient_id": ing_id, "ingredient_name": ing_name,
                                "total_consumed_lb": 0.0, "lots": []
                            }
                        override_lot = None
                        if lot_overrides and str(ing_id) in lot_overrides:
                            override_code = lot_overrides[str(ing_id)]
                            cur.execute("SELECT l.id, l.lot_code FROM lots l WHERE l.product_id = %s AND LOWER(l.lot_code) = LOWER(%s)", (ing_id, override_code))
                            override_lot = cur.fetchone()
                        if override_lot:
                            cur.execute("SELECT id FROM lots WHERE id = %s FOR UPDATE", (override_lot['id'],))
                            available = validate_lot_deduction(cur, override_lot['id'], override_lot['lot_code'], needed)
                            cur.execute("INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, ing_id, override_lot['id'], -needed))
                            cur.execute("INSERT INTO ingredient_lot_consumption (transaction_id, ingredient_product_id, ingredient_lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, ing_id, override_lot['id'], needed))
                            consumed_by_ingredient[ing_id]["total_consumed_lb"] += needed
                            consumed_by_ingredient[ing_id]["lots"].append({"lot_code": override_lot['lot_code'], "consumed_lb": needed, "override": True})
                        else:
                            cur.execute(f"""
                                SELECT l.id, l.lot_code, COALESCE(SUM(tl.quantity_lb), 0) as available
                                FROM lots l LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                                WHERE l.product_id = %s GROUP BY l.id
                                HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0 ORDER BY COALESCE(l.received_at, l.created_at) ASC
                            """, (ing_id,))
                            candidate_lots = cur.fetchall()
                            if not candidate_lots:
                                raise HTTPException(status_code=400, detail=f"No inventory available for ingredient ID {ing_id}")
                            lot_ids = [lot['id'] for lot in candidate_lots]
                            cur.execute("SELECT id FROM lots WHERE id = ANY(%s) ORDER BY id ASC FOR UPDATE", (lot_ids,))
                            cur.execute(f"""
                                SELECT l.id, l.lot_code, COALESCE(SUM(tl.quantity_lb), 0) as available
                                FROM lots l LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                                WHERE l.id = ANY(%s) GROUP BY l.id
                                HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0 ORDER BY COALESCE(l.received_at, l.created_at) ASC
                            """, (lot_ids,))
                            lots = cur.fetchall()
                            remaining = needed
                            for lot in lots:
                                if remaining <= BALANCE_EPSILON: break
                                avail = float(lot['available'])
                                if avail < BALANCE_EPSILON: continue
                                take = min(avail, remaining)
                                cur.execute("INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, ing_id, lot['id'], -take))
                                cur.execute("INSERT INTO ingredient_lot_consumption (transaction_id, ingredient_product_id, ingredient_lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, ing_id, lot['id'], take))
                                consumed_by_ingredient[ing_id]["total_consumed_lb"] += take
                                consumed_by_ingredient[ing_id]["lots"].append({"lot_code": lot['lot_code'], "consumed_lb": take})
                                remaining -= take
                            if remaining > BALANCE_EPSILON:
                                raise HTTPException(status_code=400, detail=f"Insufficient inventory for ingredient ID {ing_id}. Missing {remaining:.2f} lb")

                    consumed_flat = []
                    for group in consumed_by_ingredient.values():
                        for lot_entry in group["lots"]:
                            consumed_flat.append({"ingredient_id": group["ingredient_id"], "ingredient_name": group["ingredient_name"], **lot_entry})

                    logger.info(f"Make committed: {lot_code} - {total_output} lb of {product['name']}")

                    response = {
                        "mode": "commit",
                        "success": True, "transaction_id": txn_id,
                        "confirmation_code": generate_confirmation_code(txn_id),
                        "lot_id": output_lot_id, "lot_code": lot_code,
                        "yield_multiplier": yield_multiplier, "formula_weight_lb": formula_weight_lb,
                        "estimated_yield_lb": total_output, "output_lb": total_output,
                        "ingredients_consumed": consumed_flat,
                        "ingredients_consumed_grouped": list(consumed_by_ingredient.values()),
                        "message": f"Produced {total_output} lb as lot {lot_code}"
                    }
                    production_warning = build_production_warning(product)
                    if production_warning:
                        response["production_warning"] = production_warning
                    if siblings:
                        response["confirmed_sku"] = True
                        response["sibling_skus"] = [s['name'] for s in siblings]
                    if excluded_from_run:
                        auto_count = sum(1 for e in excluded_from_run if e.get('exclusion_type') == 'auto')
                        manual_count = len(excluded_from_run) - auto_count
                        response["excluded_ingredients"] = excluded_from_run
                        parts = []
                        if auto_count: parts.append(f"{auto_count} auto-excluded")
                        if manual_count: parts.append(f"{manual_count} manually excluded")
                        response["message"] += f" ({', '.join(parts)} ingredient(s))"

                    # Fix 2: Auto-prompt /pack after /make — query FG products that
                    # can be packed from this batch product via parent_batch_product_id.
                    # This tells the GPT/operator which /pack calls to make next.
                    cur.execute("""
                        SELECT id, name, case_size_lb
                        FROM products
                        WHERE parent_batch_product_id = %s
                          AND type != 'ingredient'
                        ORDER BY name
                    """, (product['id'],))
                    fg_products = cur.fetchall()
                    if fg_products:
                        response["pack_needed"] = {
                            "batch_lot_code": lot_code,
                            "batch_product_name": product['name'],
                            "batch_on_hand_lb": total_output,
                            "finished_goods": [
                                {
                                    "product_id": fg['id'],
                                    "name": fg['name'],
                                    "case_size_lb": float(fg['case_size_lb']) if fg['case_size_lb'] else None
                                }
                                for fg in fg_products
                            ],
                            "message": (
                                f"Run /pack to convert {product['name']} lot {lot_code} "
                                f"into finished goods: {', '.join(fg['name'] for fg in fg_products)}"
                            )
                        }

                    response["daily_production_summary"] = get_daily_production_summary(cur)
                    return response
        except HTTPException:
            raise
        except Exception as e:
            if _is_readonly_error(e): raise
            logger.error(f"Make commit failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# PACK (BATCH → FINISHED GOOD) ENDPOINTS
# ═══════════════════════════════════════════════════════════════

def resolve_pack_add_ins(cur, source: dict, target: dict, total_lb: float) -> dict | None:
    """Detect add-in ingredients needed when packing from a base batch into an FG
    whose intermediate batch BOM bridges them. Returns add-in info dict or None."""
    cur.execute("SELECT parent_batch_product_id FROM products WHERE id = %s", (target['id'],))
    row = cur.fetchone()
    parent_id = row['parent_batch_product_id'] if row else None
    if parent_id is None or parent_id == source['id']:
        return None  # no mismatch — normal pack

    # Look up intermediate batch BOM
    cur.execute("""
        SELECT bf.ingredient_product_id, p.name as ingredient_name, bf.quantity_lb,
               COALESCE(bf.exclude_from_inventory, false) as exclude_from_inventory
        FROM batch_formulas bf
        JOIN products p ON p.id = bf.ingredient_product_id
        WHERE bf.product_id = %s
    """, (parent_id,))
    formula = cur.fetchall()
    if not formula:
        # No BOM on intermediate — fall back to old mismatch warning
        cur.execute("SELECT name, odoo_code FROM products WHERE id = %s", (parent_id,))
        expected = cur.fetchone()
        if not expected:
            return None
        return {
            "warning": f"Source batch mismatch: Target FG '{target['name']}' is normally packed from '{expected['name']}' ({expected['odoo_code']}), not from '{source['name']}' ({source.get('odoo_code', 'N/A')}). No BOM found for intermediate product.",
            "warning_es": f"Lote fuente no coincide: el producto final '{target['name']}' normalmente se empaca desde '{expected['name']}' ({expected['odoo_code']}), no desde '{source['name']}' ({source.get('odoo_code', 'N/A')}). No se encontró fórmula para el producto intermedio.",
        }

    # Find the base ingredient (source batch) in the BOM to get the ratio
    base_qty = None
    add_in_formulas = []
    for ing in formula:
        if ing['ingredient_product_id'] == source['id']:
            base_qty = float(ing['quantity_lb'])
        elif not ing['exclude_from_inventory']:
            add_in_formulas.append(ing)

    if base_qty is None or base_qty <= 0:
        # Source batch not found in intermediate BOM — genuine mismatch
        cur.execute("SELECT name, odoo_code FROM products WHERE id = %s", (parent_id,))
        expected = cur.fetchone()
        if not expected:
            return None
        return {
            "warning": f"Source batch mismatch: Target FG '{target['name']}' is normally packed from '{expected['name']}' ({expected['odoo_code']}), not from '{source['name']}' ({source.get('odoo_code', 'N/A')}). Source batch not found in intermediate BOM.",
            "warning_es": f"Lote fuente no coincide: el producto final '{target['name']}' normalmente se empaca desde '{expected['name']}' ({expected['odoo_code']}), no desde '{source['name']}' ({source.get('odoo_code', 'N/A')}). Lote fuente no encontrado en la fórmula intermedia.",
        }

    if not add_in_formulas:
        return None  # no add-ins — straight repack

    ratio = total_lb / base_qty
    add_in_ingredients = []
    all_add_in_ids = [ing['ingredient_product_id'] for ing in add_in_formulas]

    # Batch-fetch available lots for all add-in ingredients
    if all_add_in_ids:
        cur.execute(f"""
            SELECT l.product_id, l.id, l.lot_code,
                   COALESCE(SUM(tl.quantity_lb), 0) as available
            FROM lots l LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
            WHERE l.product_id = ANY(%s) GROUP BY l.id
            HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
            ORDER BY l.product_id, COALESCE(l.received_at, l.created_at) ASC
        """, (all_add_in_ids,))
        add_in_lots_map = {}
        for r in cur.fetchall():
            pid = r['product_id']
            if pid not in add_in_lots_map:
                add_in_lots_map[pid] = []
            add_in_lots_map[pid].append(dict(r))

    for ing in add_in_formulas:
        ing_id = ing['ingredient_product_id']
        needed = round(float(ing['quantity_lb']) * ratio, 2)
        lots = add_in_lots_map.get(ing_id, [])
        total_avail = sum(float(lot['available']) for lot in lots)
        lot_details = [{"lot_code": lot['lot_code'], "available_lb": round(float(lot['available']), 2)} for lot in lots]
        add_in_ingredients.append({
            "ingredient_id": ing_id,
            "ingredient_name": ing['ingredient_name'],
            "needed_lb": needed,
            "available_lb": round(total_avail, 2),
            "sufficient": total_avail >= needed - BALANCE_EPSILON,
            "lots": lot_details,
        })

    all_sufficient = all(ai['sufficient'] for ai in add_in_ingredients)
    add_in_names = " + ".join(ai['ingredient_name'] for ai in add_in_ingredients)
    result = {
        "add_in_ingredients": add_in_ingredients,
        "all_add_ins_sufficient": all_sufficient,
        "add_in_note": f"Add-in ingredients will be deducted automatically ({add_in_names} added at packing hopper)",
        "_add_in_formulas": add_in_formulas,  # internal: used by commit
        "_base_qty": base_qty,  # internal: used by commit
    }
    if not all_sufficient:
        short = [ai for ai in add_in_ingredients if not ai['sufficient']]
        short_names = ", ".join(f"{ai['ingredient_name']} (need {ai['needed_lb']} lb, have {ai['available_lb']} lb)" for ai in short)
        result["warning"] = f"Insufficient add-in ingredients: {short_names}"
        result["warning_es"] = f"Ingredientes adicionales insuficientes: {short_names}"
    return result


@app.post("/pack")
def pack(req: PackRequest, _: bool = Depends(verify_api_key)):
    """Pack batch into finished goods. mode=preview returns allocation plan; mode=commit executes."""
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at, req.backfill
    )
    if req.mode == "preview":
        try:
            with get_transaction() as cur:
                source = resolve_product_full(cur, req.source_product)
                target = resolve_product_full(cur, req.target_product)
                case_weight = req.case_weight_lb
                if case_weight is None:
                    case_weight = float(target.get('case_size_lb') or 0)
                if case_weight <= 0:
                    raise HTTPException(400, f"Case weight required. Product '{target['name']}' has no case_size_lb set. Provide case_weight_lb parameter to override.")
                total_lb = req.cases * case_weight

                available_lots = available_lots_for_product(cur, int(source['id']))
                reservation_summary = _allocation_reservation_summary(
                    cur, int(source['id'])
                )
                total_available = sum(float(lot['on_hand']) for lot in available_lots)

                if req.lot_allocations:
                    allocations = []
                    can_pack = 0.0
                    reserved_taken = 0.0
                    for alloc in req.lot_allocations:
                        matched = next((l for l in available_lots if l['lot_code'].lower() == alloc.lot_code.lower()), None)
                        if not matched:
                            allocations.append({"lot_code": alloc.lot_code, "available_lb": 0, "allocated_lb": alloc.quantity_lb, "sufficient": False, "error": f"Lot '{alloc.lot_code}' not found or has no inventory for {source['name']}"})
                            continue
                        physical = float(matched['on_hand'])
                        takeable = float(matched['takeable'])
                        can_pack += min(float(alloc.quantity_lb), takeable)
                        reserved_taken += max(
                            0.0,
                            min(float(alloc.quantity_lb), physical)
                            - min(float(alloc.quantity_lb), takeable),
                        )
                        allocations.append({"lot_id": matched['lot_id'], "lot_code": matched['lot_code'], "available_lb": physical, "takeable_lb": takeable, "allocated_lb": alloc.quantity_lb, "sufficient": physical >= alloc.quantity_lb})
                    alloc_total = sum(a['allocated_lb'] for a in allocations)
                    if abs(alloc_total - total_lb) > 0.01:
                        return JSONResponse(status_code=400, content={"error": f"Lot allocations sum to {alloc_total} lb but {total_lb} lb needed ({req.cases} cases x {case_weight} lb)"})
                else:
                    plan = _takeable_deduction_plan(available_lots, total_lb)
                    can_pack = plan["can_take_lb"]
                    reserved_taken = plan["reserved_taken_lb"]
                    allocations = [
                        {
                            "lot_id": item["lot"]["lot_id"],
                            "lot_code": item["lot"]["lot_code"],
                            "available_lb": float(item["lot"]["on_hand"]),
                            "takeable_lb": float(item["lot"]["takeable"]),
                            "allocated_lb": float(item["quantity_lb"]),
                            "sufficient": plan["short_lb"] <= BALANCE_EPSILON,
                        }
                        for item in plan["lots"]
                    ]

                all_sufficient = all(a.get('sufficient', False) for a in allocations)
                if req.target_lot_code:
                    output_lot_code = req.target_lot_code
                elif allocations and allocations[0].get('lot_code'):
                    output_lot_code = allocations[0]['lot_code']
                else:
                    output_lot_code = "UNKNOWN"

                allocation_warning = _allocation_observe_warning(
                    "Pack",
                    total_lb,
                    can_pack,
                    reserved_taken,
                    reservation_summary,
                    preview=True,
                )
                result = {
                    "mode": "preview",
                    "source_product_id": source['id'], "source_product_name": source['name'],
                    "target_product_id": target['id'], "target_product_name": target['name'],
                    "cases": req.cases, "case_weight_lb": case_weight, "total_lb": total_lb,
                    "output_lot_code": output_lot_code, "allocations": allocations,
                    "all_lots_sufficient": all_sufficient, "total_batch_available_lb": total_available,
                    "total_takeable_lb": sum(float(lot['takeable']) for lot in available_lots),
                    "can_pack_lb": can_pack,
                    "reserved_others_lb": reservation_summary["reserved_others_lb"],
                    "reserved_by_orders": reservation_summary["reserved_by_orders"],
                    "source_lot_count": len(available_lots),
                    "source_lots": [{"lot_code": lot['lot_code'], "available_lb": float(lot['on_hand']), "takeable_lb": float(lot['takeable'])} for lot in available_lots],
                    "preview_message": f"Ready to pack {req.cases} cases ({total_lb} lb) of {target['name']} from {source['name']} ({len(available_lots)} batch lot(s))"
                }
                if allocation_warning:
                    result["allocation_warning"] = allocation_warning
                    result["warning"] = allocation_warning["message"]
                add_in_info = resolve_pack_add_ins(cur, source, target, total_lb)
                if add_in_info:
                    # Strip internal keys before returning
                    public_info = {k: v for k, v in add_in_info.items() if not k.startswith('_')}
                    result.update(public_info)
                return result
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Pack preview failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})
    else:
        # mode == "commit"
        try:
            with get_db_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    source = resolve_product_full(cur, req.source_product)
                    target = resolve_product_full(cur, req.target_product)
                    case_weight = req.case_weight_lb
                    if case_weight is None:
                        case_weight = float(target.get('case_size_lb') or 0)
                    if case_weight <= 0:
                        raise HTTPException(400, f"Case weight required for '{target['name']}'. Provide case_weight_lb parameter to override.")
                    total_lb = req.cases * case_weight
                    now = get_plant_now()

                    lots = available_lots_for_product(
                        cur,
                        int(source['id']),
                        lock=True,
                        persist_expired=True,
                        released_by=_operator_id(_),
                    )
                    reservation_summary = _allocation_reservation_summary(
                        cur, int(source['id'])
                    )
                    if not lots:
                        raise HTTPException(400, f"No batch inventory available for {source['name']}")

                    lots_by_code = {lot['lot_code'].lower(): lot for lot in lots}

                    if req.lot_allocations:
                        alloc_plan = []
                        can_pack = 0.0
                        reserved_taken = 0.0
                        for alloc in req.lot_allocations:
                            lot = lots_by_code.get(alloc.lot_code.lower())
                            if not lot:
                                raise HTTPException(400, f"Lot '{alloc.lot_code}' not found or empty for {source['name']}")
                            validate_lot_deduction(cur, lot['lot_id'], lot['lot_code'], alloc.quantity_lb)
                            takeable = float(lot['takeable'])
                            physical = float(lot['on_hand'])
                            # Clamp by physical for symmetry with the preview
                            # arithmetic. validate_lot_deduction already
                            # guarantees quantity <= physical here, but the
                            # clamp keeps commit == preview by construction
                            # rather than by reliance on that guard.
                            can_pack += min(float(alloc.quantity_lb), takeable)
                            reserved_taken += max(
                                0.0,
                                min(float(alloc.quantity_lb), physical)
                                - min(float(alloc.quantity_lb), takeable),
                            )
                            alloc_plan.append((lot, alloc.quantity_lb))
                        alloc_total = sum(qty for _, qty in alloc_plan)
                        if abs(alloc_total - total_lb) > 0.01:
                            raise HTTPException(400, f"Allocations sum to {alloc_total} lb, need {total_lb} lb ({req.cases} cases x {case_weight} lb)")
                    else:
                        plan = _takeable_deduction_plan(lots, total_lb)
                        if plan["short_lb"] > BALANCE_EPSILON:
                            raise HTTPException(400, f"Insufficient batch inventory. Have {plan['total_on_hand_lb']:.4f} lb, need {total_lb} lb")
                        can_pack = plan["can_take_lb"]
                        reserved_taken = plan["reserved_taken_lb"]
                        alloc_plan = [
                            (item["lot"], float(item["quantity_lb"]))
                            for item in plan["lots"]
                        ]

                    allocation_warning = _allocation_observe_warning(
                        "Pack",
                        total_lb,
                        can_pack,
                        reserved_taken,
                        reservation_summary,
                    )
                    _enforce_allocation_takeable(
                        "Pack",
                        total_lb,
                        can_pack,
                        reserved_taken,
                        reservation_summary,
                        product_id=int(source["id"]),
                    )

                    if req.target_lot_code:
                        output_lot_code = req.target_lot_code
                    else:
                        output_lot_code = alloc_plan[0][0]['lot_code']

                    output_lot_id, is_new_lot = find_or_create_lot(cur, target['id'], output_lot_code, 'pack_output')
                    source_lot_summary = ", ".join(f"{lot['lot_code']} ({qty} lb)" for lot, qty in alloc_plan)
                    cur.execute("""
                        INSERT INTO transactions (
                            type, timestamp, notes, occurred_at, created_at_source
                        )
                        VALUES ('pack', %s, %s, %s, %s) RETURNING id
                    """, (
                        now,
                        f"Pack {req.cases} cases of {target['name']} from {source['name']} lots: {source_lot_summary}",
                        occurred_at, created_at_source,
                    ))
                    txn_id = cur.fetchone()['id']

                    cur.execute("INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, target['id'], output_lot_id, total_lb))

                    consumed = []
                    for lot, qty in alloc_plan:
                        cur.execute("INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, source['id'], lot['lot_id'], -qty))
                        cur.execute("INSERT INTO ingredient_lot_consumption (transaction_id, ingredient_product_id, ingredient_lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, source['id'], lot['lot_id'], qty))
                        consumed.append({"lot_code": lot['lot_code'], "consumed_lb": qty})

                    # --- Add-in ingredient deduction ---
                    add_in_info = resolve_pack_add_ins(cur, source, target, total_lb)
                    add_in_consumed = []
                    if add_in_info and 'add_in_ingredients' in add_in_info:
                        # Check all add-ins are sufficient before deducting
                        if not add_in_info.get('all_add_ins_sufficient'):
                            short = [ai for ai in add_in_info['add_in_ingredients'] if not ai['sufficient']]
                            short_msg = "; ".join(f"{ai['ingredient_name']}: have {ai['available_lb']} lb, need {ai['needed_lb']} lb" for ai in short)
                            raise HTTPException(400, f"Insufficient inventory for add-in ingredient(s): {short_msg}")

                        ratio = total_lb / add_in_info['_base_qty']
                        for ing in add_in_info['_add_in_formulas']:
                            ing_id = ing['ingredient_product_id']
                            needed = round(float(ing['quantity_lb']) * ratio, 2)
                            # FIFO deduction with locking
                            cur.execute(f"""
                                SELECT l.id, l.lot_code, COALESCE(SUM(tl.quantity_lb), 0) as available
                                FROM lots l LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                                WHERE l.product_id = %s GROUP BY l.id
                                HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                                ORDER BY COALESCE(l.received_at, l.created_at) ASC
                            """, (ing_id,))
                            candidate_lots = cur.fetchall()
                            lot_ids = [lot['id'] for lot in candidate_lots]
                            cur.execute("SELECT id FROM lots WHERE id = ANY(%s) ORDER BY id ASC FOR UPDATE", (lot_ids,))
                            cur.execute(f"""
                                SELECT l.id, l.lot_code, COALESCE(SUM(tl.quantity_lb), 0) as available
                                FROM lots l LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                                WHERE l.id = ANY(%s) GROUP BY l.id
                                HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                                ORDER BY COALESCE(l.received_at, l.created_at) ASC
                            """, (lot_ids,))
                            ing_lots = cur.fetchall()
                            remaining = needed
                            cur.execute("SELECT name FROM products WHERE id = %s", (ing_id,))
                            ing_name = cur.fetchone()['name']
                            for lot in ing_lots:
                                if remaining <= BALANCE_EPSILON:
                                    break
                                avail = float(lot['available'])
                                if avail < BALANCE_EPSILON:
                                    continue
                                take = min(avail, remaining)
                                cur.execute("INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, ing_id, lot['id'], -take))
                                cur.execute("INSERT INTO ingredient_lot_consumption (transaction_id, ingredient_product_id, ingredient_lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, ing_id, lot['id'], take))
                                add_in_consumed.append({"ingredient_name": ing_name, "lot_code": lot['lot_code'], "consumed_lb": round(take, 2)})
                                remaining -= take
                            if remaining > BALANCE_EPSILON:
                                raise HTTPException(400, f"Insufficient inventory for add-in ingredient {ing_name}: need {needed} lb, could only allocate {needed - remaining:.2f} lb")

                    logger.info(f"Pack committed: {output_lot_code} - {total_lb} lb of {target['name']} from {source['name']}")

                    response = {
                        "mode": "commit",
                        "success": True, "transaction_id": txn_id,
                        "confirmation_code": generate_confirmation_code(txn_id),
                        "output_lot_id": output_lot_id, "output_lot_code": output_lot_code,
                        "target_product_name": target['name'], "source_product_name": source['name'],
                        "cases": req.cases, "case_weight_lb": case_weight, "total_lb": total_lb,
                        "can_pack_lb": can_pack,
                        "reserved_others_lb": reservation_summary["reserved_others_lb"],
                        "reserved_by_orders": reservation_summary["reserved_by_orders"],
                        "batch_lots_consumed": consumed,
                        "message": f"Packed {req.cases} cases ({total_lb} lb) of {target['name']} as lot {output_lot_code}"
                    }
                    if allocation_warning:
                        response["allocation_warning"] = allocation_warning
                        response["warning"] = allocation_warning["message"]
                    if add_in_consumed:
                        response["add_in_ingredients_consumed"] = add_in_consumed
                        add_in_names = ", ".join(set(ai['ingredient_name'] for ai in add_in_consumed))
                        response["add_in_note"] = f"Add-in ingredients deducted: {add_in_names}"
                        response["message"] += f" (with add-ins: {add_in_names})"
                    elif add_in_info and add_in_info.get('warning'):
                        response["warning"] = add_in_info["warning"]
                        if add_in_info.get('warning_es'):
                            response["warning_es"] = add_in_info["warning_es"]
                    response["daily_production_summary"] = get_daily_production_summary(cur)
                    return response
        except HTTPException:
            raise
        except Exception as e:
            if _is_readonly_error(e): raise
            logger.error(f"Pack commit failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# ADJUST ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.post("/adjust")
def adjust(req: AdjustRequest, _: bool = Depends(verify_api_key)):
    """Adjust inventory. mode=preview returns balance check; mode=commit executes."""
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at, req.backfill
    )
    validate_bilingual(req.reason, req.reason_es, "reason")
    if req.mode == "preview":
        try:
            with get_transaction() as cur:
                cur.execute(f"""
                    SELECT p.id as product_id, p.name, p.odoo_code,
                           COALESCE(p.label_type, 'house') as label_type,
                           l.id as lot_id, l.lot_code,
                           COALESCE(SUM(tl.quantity_lb), 0) as quantity_on_hand
                    FROM products p
                    JOIN lots l ON l.product_id = p.id
                    LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                    WHERE (LOWER(p.name) LIKE LOWER(%s) OR LOWER(p.odoo_code) LIKE LOWER(%s))
                      AND LOWER(l.lot_code) = LOWER(%s)
                    GROUP BY p.id, l.id
                """, (f"%{req.product_name}%", f"%{req.product_name}%", req.lot_code))
                result = cur.fetchone()
                if not result:
                    return JSONResponse(status_code=404, content={"error": f"Product/lot combination not found for '{req.product_name}' / '{req.lot_code}'"})
                warning = check_private_label_merge(result['name'], result['label_type'], req.reason, req.adjustment_lb)
                if warning:
                    return JSONResponse(status_code=403, content={"blocked": True, "warning": warning, "product_name": result['name'], "label_type": result['label_type']})

                quantity_on_hand = float(result['quantity_on_hand'])
                new_balance = quantity_on_hand + req.adjustment_lb
                response = {
                    "mode": "preview",
                    "product_id": result['product_id'], "product_name": result['name'],
                    "odoo_code": result['odoo_code'], "label_type": result['label_type'],
                    "lot_code": result['lot_code'], "current_quantity_lb": quantity_on_hand,
                    "adjustment_lb": req.adjustment_lb, "new_balance_lb": new_balance,
                    "reason": req.reason,
                    "preview_message": f"Will adjust {result['name']} lot {result['lot_code']} by {req.adjustment_lb} lb ({quantity_on_hand} → {new_balance} lb)"
                }
                if req.reason_es:
                    response["reason_es"] = req.reason_es
                if new_balance < 0:
                    response["balance_warning"] = f"This will result in negative inventory ({new_balance} lb)"
                return response
        except Exception as e:
            logger.error(f"Adjust preview failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})
    else:
        # mode == "commit"
        try:
            with get_db_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    product = resolve_product_full(cur, req.product_name)
                    cur.execute("SELECT id as lot_id, lot_code FROM lots WHERE product_id = %s AND LOWER(lot_code) = LOWER(%s)", (product['id'], req.lot_code))
                    lot = cur.fetchone()
                    if not lot:
                        raise HTTPException(404, f"Lot '{req.lot_code}' not found for product '{product['name']}'")
                    cur.execute("SELECT COALESCE(label_type, 'house') as label_type FROM products WHERE id = %s", (product['id'],))
                    lt_row = cur.fetchone()
                    result = {**product, 'product_id': product['id'], 'lot_id': lot['lot_id'], 'lot_code': lot['lot_code'],
                              'label_type': lt_row['label_type'] if lt_row else 'house'}
                    warning = check_private_label_merge(result['name'], result['label_type'], req.reason, req.adjustment_lb)
                    if warning:
                        return JSONResponse(status_code=403, content={"blocked": True, "warning": warning, "product_name": result['name'], "label_type": result['label_type']})

                    now = get_plant_now()
                    cur.execute("""
                        INSERT INTO transactions (
                            type, timestamp, adjust_reason, adjust_reason_es,
                            notes, occurred_at, created_at_source
                        )
                        VALUES ('adjust', %s, %s, %s, %s, %s, %s) RETURNING id
                    """, (
                        now, req.reason, req.reason_es,
                        f"Adjustment: {req.adjustment_lb} lb",
                        occurred_at, created_at_source,
                    ))
                    txn_id = cur.fetchone()['id']
                    cur.execute("INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, result['product_id'], result['lot_id'], req.adjustment_lb))

                    new_balance = lot_on_hand(cur, result['lot_id'])
                    logger.info(f"Adjust committed: {req.adjustment_lb} lb to lot {result['lot_code']} (balance: {new_balance} lb)")

                    response = {
                        "mode": "commit",
                        "success": True, "transaction_id": txn_id,
                        "confirmation_code": generate_confirmation_code(txn_id),
                        "product_id": result['product_id'], "product_name": result['name'],
                        "lot_code": result['lot_code'], "adjustment_lb": req.adjustment_lb,
                        "new_balance_lb": new_balance, "reason": req.reason,
                        "message": f"Adjusted lot {result['lot_code']} by {req.adjustment_lb} lb (new balance: {new_balance} lb)"
                    }
                    if req.reason_es:
                        response["reason_es"] = req.reason_es
                    return response
        except HTTPException:
            raise
        except Exception as e:
            if _is_readonly_error(e): raise
            logger.error(f"Adjust commit failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# VOID TRANSACTION ENDPOINT
# ═══════════════════════════════════════════════════════════════

class VoidRequest(BaseModel):
    reason: str


class LedgerCorrectionRequest(BaseModel):
    event_type: Literal["amend", "void", "restore"] = "amend"
    reason: str
    replacement_values: Optional[Dict[str, Any]] = None


def _operator_id(auth_context: Any) -> str:
    """Phase-1 compatibility shim for the existing shared credential."""
    if isinstance(auth_context, dict) and auth_context.get("operator_id"):
        return str(auth_context["operator_id"])
    if isinstance(auth_context, str) and auth_context.strip():
        return auth_context.strip()
    return "legacy-shared-key"


_TRANSACTION_AMENDABLE_FIELDS = {
    "occurred_at", "business_date", "notes", "bol_reference",
    "shipper_name", "shipper_code", "cases_received", "case_size_lb",
    "customer_name", "order_reference", "adjust_reason", "adjust_reason_es",
}


def _append_transaction_correction(
    cur,
    transaction_id: int,
    event_type: str,
    reason: str,
    replacement_values: Optional[Dict[str, Any]],
    operator_id: str,
):
    reason = (reason or "").strip()
    if not reason:
        raise HTTPException(422, "A non-empty correction reason is required")

    cur.execute("SELECT id, type FROM transactions WHERE id = %s FOR UPDATE", (transaction_id,))
    transaction = cur.fetchone()
    if not transaction:
        raise HTTPException(404, f"Transaction #{transaction_id} not found")

    cur.execute(
        """SELECT effective_record, effective_status
           FROM ledger_current_transactions WHERE id = %s""",
        (transaction_id,),
    )
    current = cur.fetchone()
    previous = dict(current["effective_record"] or {})
    current_status = current["effective_status"]

    if event_type == "void" and current_status == "voided":
        raise HTTPException(400, f"Transaction #{transaction_id} is already voided")
    if event_type == "restore" and current_status != "voided":
        raise HTTPException(400, f"Transaction #{transaction_id} is not voided")

    restore_product_ids = []
    prepared_restore = None
    if event_type == "restore" and transaction["type"] == "ship":
        restore_product_ids = _preflight_restore_ship_stock(cur, transaction_id)
        prepared_restore = _prepare_restore_ship_allocations(
            cur, transaction_id, operator_id
        )
        # Preserve the addendum's stock-first, own-coverage-second error
        # precedence. PR 5's foreign-stock gate is additive after those guards
        # and still precedes the correction INSERT and inventory_restored shrink.
        _preflight_restore_ship_takeable(cur, transaction_id)

    replacement = dict(previous)
    if event_type == "amend":
        changes = replacement_values or {}
        forbidden = sorted(set(changes) - _TRANSACTION_AMENDABLE_FIELDS)
        if forbidden:
            raise HTTPException(
                422,
                f"Immutable or unsupported correction fields: {', '.join(forbidden)}",
            )
        if not changes:
            raise HTTPException(422, "replacement_values must include at least one field")
        replacement.update(changes)
    elif event_type == "void":
        replacement["status"] = "voided"
    else:
        replacement["status"] = "posted"

    cur.execute(
        """INSERT INTO ledger_corrections
               (target_table, target_id, event_type, previous_values,
                replacement_values, reason, operator_id)
           VALUES ('transactions', %s, %s, %s::jsonb, %s::jsonb, %s, %s)
           RETURNING id, created_at, operator_id""",
        (
            transaction_id,
            event_type,
            json.dumps(previous, default=str),
            json.dumps(replacement, default=str),
            reason,
            operator_id,
        ),
    )
    event = cur.fetchone()
    correction_id = event["id"]
    result = {
        "correction_id": str(correction_id),
        "created_at": event["created_at"],
        "operator_id": event["operator_id"],
        "event_type": event_type,
        "previous_values": previous,
        "replacement_values": replacement,
    }
    if event_type == "void":
        if transaction["type"] == "ship":
            allocations_restored = _void_ship_allocations(
                cur, transaction_id, operator_id
            )
            _record_void_allocation_reactivations(
                cur, transaction_id, correction_id, allocations_restored
            )
            result["allocations_restored"] = allocations_restored
        else:
            cur.execute(
                "SELECT DISTINCT product_id FROM ledger_current_transaction_lines "
                "WHERE transaction_id = %s ORDER BY product_id",
                (transaction_id,),
            )
            result["allocations_released"] = _shrink_overallocated_products(
                cur,
                [int(row["product_id"]) for row in cur.fetchall()],
                operator_id,
            )
    elif event_type == "restore" and transaction["type"] == "ship":
        result["allocations_reshipped"] = _restore_ship_allocations(
            cur, transaction_id, operator_id, prepared_restore
        )
        _shrink_overallocated_products(
            cur,
            restore_product_ids,
            operator_id,
            release_reason="inventory_restored",
        )
        unknown_line_ids = prepared_restore["unknown_line_ids"]
        if unknown_line_ids:
            result["allocation_reactivation_unknown"] = True
            result["allocation_reactivation_unknown_line_ids"] = unknown_line_ids
    return result


def _append_transaction_line_correction(
    cur,
    line_id: int,
    replacement_values: Dict[str, Any],
    reason: str,
    operator_id: str,
):
    allowed = {"product_id", "lot_id", "quantity_lb"}
    forbidden = set(replacement_values) - allowed
    if forbidden:
        raise HTTPException(422, f"Unsupported line correction fields: {sorted(forbidden)}")
    cur.execute("SELECT id FROM transaction_lines WHERE id = %s FOR UPDATE", (line_id,))
    if not cur.fetchone():
        raise HTTPException(404, f"Transaction line #{line_id} not found")
    cur.execute(
        "SELECT effective_record FROM ledger_current_transaction_lines WHERE id = %s",
        (line_id,),
    )
    previous = dict(cur.fetchone()["effective_record"])
    replacement = dict(previous)
    replacement.update(replacement_values)
    cur.execute(
        """INSERT INTO ledger_corrections
               (target_table, target_id, event_type, previous_values,
                replacement_values, reason, operator_id)
           VALUES ('transaction_lines', %s, 'amend', %s::jsonb, %s::jsonb, %s, %s)
           RETURNING id""",
        (
            line_id,
            json.dumps(previous, default=str),
            json.dumps(replacement, default=str),
            reason,
            operator_id,
        ),
    )
    return str(cur.fetchone()["id"])


@app.post("/void/{transaction_id}")
def void_transaction(transaction_id: int, req: VoidRequest, _: bool = Depends(verify_api_key)):
    """Append a void correction while preserving the original transaction.

    All balance math reads lines through the effective append-only state, so
    the original's lines drop out immediately without mutating or hiding it.

    Response keeps the historical shape: reversal_transaction_id and
    reversal_lines are retained for backward compatibility but are always
    null/empty now that voiding no longer posts reversals.
    """
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                event = _append_transaction_correction(
                    cur,
                    transaction_id,
                    "void",
                    req.reason,
                    None,
                    _operator_id(_),
                )

                logger.info(
                    f"Voided transaction #{transaction_id} through correction "
                    f"{event['correction_id']}"
                )
                return {
                    "success": True,
                    "voided_transaction_id": transaction_id,
                    "correction_id": event["correction_id"],
                    "created_at": event["created_at"],
                    "reversal_transaction_id": None,
                    "reversal_lines": [],
                    "allocations_restored": event.get("allocations_restored", []),
                    "allocations_released": event.get("allocations_released", []),
                    "message": (
                        f"Transaction #{transaction_id} voided by append-only correction; "
                        "the original is preserved and its lines are excluded from balances"
                    ),
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Void transaction failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/records/transactions/{transaction_id}/corrections")
def correct_transaction(
    transaction_id: int,
    req: LedgerCorrectionRequest,
    _: bool = Depends(verify_api_key),
):
    try:
        with get_transaction() as cur:
            event = _append_transaction_correction(
                cur,
                transaction_id,
                req.event_type,
                req.reason,
                req.replacement_values,
                _operator_id(_),
            )
            return {"transaction_id": transaction_id, **event}
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e):
            raise
        logger.error(f"Transaction correction failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# TRACE ENDPOINTS
# ═══════════════════════════════════════════════════════════════

# Canonical entry_source values written by lot creation endpoints
INGREDIENT_ENTRY_SOURCES = {'received', 'found_inventory', 'adjusted', None}
OUTPUT_ENTRY_SOURCES = {'production_output', 'pack_output'}

@app.get("/trace/batch/{lot_code}")
def trace_batch(lot_code: str, product_id: Optional[int] = Query(None), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            # First: look up the lot to determine its type
            query = """
                SELECT l.id, l.lot_code, l.entry_source, l.supplier_lot_code,
                       p.name as product_name, p.id as product_id
                FROM lots l
                JOIN products p ON p.id = l.product_id
                WHERE LOWER(l.lot_code) = LOWER(%s)
            """
            params = [lot_code]
            if product_id is not None:
                query += " AND l.product_id = %s"
                params.append(product_id)
            cur.execute(query, params)
            rows = cur.fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail=f"Lot '{lot_code}' not found")
            if len(rows) > 1:
                return JSONResponse(status_code=409, content={
                    "error": "ambiguous_lot_code",
                    "message": f"Lot code '{lot_code}' matches multiple products",
                    "matches": [{"lot_id": r['id'], "product_id": r['product_id'],
                                 "product_name": r['product_name'], "entry_source": r['entry_source']} for r in rows]
                })
            lot_row = rows[0]

            # Ingredient lot: entry_source is received, adjusted, found_inventory (not production/pack output)
            if lot_row['entry_source'] in INGREDIENT_ENTRY_SOURCES:
                return _trace_ingredient_backward(cur, lot_row)

            # Finished goods lot: entry_source is make or pack — existing logic
            cur.execute("""
                SELECT t.id as transaction_id, t.timestamp, l.lot_code, p.name as product_name,
                       tl.quantity_lb as output_lb
                FROM ledger_current_transactions t
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN lots l ON l.id = tl.lot_id
                JOIN products p ON p.id = tl.product_id
                WHERE t.type IN ('make', 'pack')
                  AND t.effective_status = 'posted'
                  AND l.id = %s
                  AND tl.quantity_lb > 0
            """, (lot_row['id'],))
            batch = cur.fetchone()

            if not batch:
                # Lot exists but no production transaction found — treat as ingredient
                return _trace_ingredient_backward(cur, lot_row)

            cur.execute("""
                SELECT l.id as lot_id, p.id as product_id,
                       p.name as ingredient_name, l.lot_code as ingredient_lot,
                       l.supplier_lot_code, ilc.quantity_lb as quantity_consumed
                FROM ingredient_lot_consumption ilc
                JOIN products p ON p.id = ilc.ingredient_product_id
                JOIN lots l ON l.id = ilc.ingredient_lot_id
                WHERE ilc.transaction_id = %s
            """, (batch['transaction_id'],))
            ingredients = cur.fetchall()

            date_str, time_str = format_timestamp(batch['timestamp'])

            # Customer shipments for this batch lot
            cur.execute("""
                SELECT t.id as transaction_id, t.timestamp as shipped_at,
                       ABS(tl.quantity_lb) as quantity_lb,
                       t.customer_name, t.order_reference,
                       so.id as sales_order_id,
                       so.order_number,
                       s.id as shipment_id
                FROM ledger_current_transaction_lines tl
                JOIN ledger_current_transactions t ON t.id = tl.transaction_id
                LEFT JOIN sales_order_shipments sos ON sos.transaction_id = t.id
                LEFT JOIN sales_order_lines sol ON sol.id = sos.sales_order_line_id
                LEFT JOIN sales_orders so ON so.id = sol.sales_order_id
                LEFT JOIN shipment_lines sl ON sl.transaction_id = t.id
                LEFT JOIN shipments s ON s.id = sl.shipment_id
                WHERE tl.lot_id = %s
                  AND t.type = 'ship'
                  AND tl.quantity_lb < 0
                  AND t.effective_status = 'posted'
                ORDER BY t.timestamp
            """, (lot_row['id'],))
            shipment_rows = cur.fetchall()

            customer_shipments = []
            for sh in shipment_rows:
                sh_date, sh_time = format_timestamp(sh['shipped_at'])
                customer_shipments.append({
                    "transaction_id": sh['transaction_id'],
                    "shipped_at": sh['shipped_at'].isoformat() if sh['shipped_at'] else None,
                    "shipped_date": sh_date,
                    "shipped_time": sh_time,
                    "quantity_lb": float(sh['quantity_lb']),
                    "customer_name": sh['customer_name'],
                    "order_reference": sh['order_reference'],
                    "sales_order_id": sh['sales_order_id'],
                    "order_number": sh['order_number'],
                    "shipment_id": sh['shipment_id']
                })

            total_shipped = sum(s['quantity_lb'] for s in customer_shipments)

            # Current on-hand quantity (posted-only)
            on_hand = lot_on_hand(cur, lot_row['id'])

            # Fetch case_size_lb for unit counts
            cur.execute("SELECT case_size_lb, default_batch_lb, type FROM products WHERE id = %s", (lot_row['product_id'],))
            prod_info = cur.fetchone()
            cs_lb = float(prod_info['case_size_lb']) if prod_info and prod_info['case_size_lb'] else None
            db_lb = float(prod_info['default_batch_lb']) if prod_info and prod_info['default_batch_lb'] else None
            output_lb = float(batch['output_lb'])
            result = {
                "trace_type": "batch",
                "lot_id": lot_row['id'],
                "product_id": lot_row['product_id'],
                "batch_lot_code": batch['lot_code'],
                "product_name": batch['product_name'],
                "output_lb": output_lb,
                "produced_date": date_str,
                "produced_time": time_str,
                "ingredients": [
                    {
                        "lot_id": ing['lot_id'],
                        "product_id": ing['product_id'],
                        "ingredient_name": ing['ingredient_name'],
                        "lot_code": ing['ingredient_lot'],
                        "supplier_lot_code": ing['supplier_lot_code'],
                        "quantity_lb": float(ing['quantity_consumed'])
                    } for ing in ingredients
                ],
                "customer_shipments": customer_shipments,
                "total_shipped_lb": total_shipped,
                "on_hand_lb": on_hand
            }
            if cs_lb and cs_lb > 0:
                result["case_size_lb"] = cs_lb
                result["output_units"] = round(output_lb / cs_lb)
                result["on_hand_units"] = round(on_hand / cs_lb)
                result["total_shipped_units"] = round(total_shipped / cs_lb)
            elif db_lb and db_lb > 0 and prod_info.get('type') == 'batch':
                result["default_batch_lb"] = db_lb
                result["output_batches"] = round(output_lb / db_lb, 1)
                result["on_hand_batches"] = round(on_hand / db_lb, 1)
            return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Trace batch failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


def _trace_ingredient_backward(cur, lot_row):
    """Backward trace for an ingredient lot: supplier origin + downstream batches."""
    lot_code = lot_row['lot_code']

    # 1. Find receive transaction (supplier origin)
    supplier_info = None
    cur.execute("""
        SELECT t.id as transaction_id, t.timestamp, t.shipper_name, t.bol_reference,
               tl.quantity_lb
        FROM ledger_current_transactions t
        JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
        JOIN lots l ON l.id = tl.lot_id
        WHERE t.type = 'receive' AND t.effective_status = 'posted'
          AND l.id = %s
        ORDER BY t.timestamp DESC LIMIT 1
    """, (lot_row['id'],))
    recv = cur.fetchone()
    if recv:
        recv_date, recv_time = format_timestamp(recv['timestamp'])
        supplier_info = {
            "supplier_name": recv['shipper_name'] or 'Unknown supplier',
            "bol_reference": recv['bol_reference'],
            "received_date": recv_date,
            "received_time": recv_time,
            "quantity_lb": float(recv['quantity_lb']),
            "transaction_id": recv['transaction_id']
        }

    # 2. Find downstream batches that consumed this ingredient lot
    cur.execute("""
        SELECT DISTINCT bl.id as lot_id, bp.id as product_id,
               bl.lot_code as batch_lot, bp.name as batch_product,
               ilc.quantity_lb as quantity_consumed, t.timestamp
        FROM ingredient_lot_consumption ilc
        JOIN ledger_current_transactions t ON t.id = ilc.transaction_id
        JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id AND tl.quantity_lb > 0
        JOIN lots bl ON bl.id = tl.lot_id
        JOIN products bp ON bp.id = bl.product_id
        WHERE ilc.ingredient_lot_id = %s
          AND t.effective_status = 'posted'
        ORDER BY t.timestamp
    """, (lot_row['id'],))
    downstream = cur.fetchall()

    batches_out = []
    for b in downstream:
        b_date, b_time = format_timestamp(b['timestamp'])
        batches_out.append({
            "lot_id": b['lot_id'],
            "product_id": b['product_id'],
            "batch_lot_code": b['batch_lot'],
            "batch_product": b['batch_product'],
            "quantity_consumed": float(b['quantity_consumed']),
            "produced_date": b_date,
            "produced_time": b_time
        })

    # 3. Find direct shipments (ship transactions that deducted from this lot)
    cur.execute("""
        SELECT t.id as transaction_id, t.timestamp as shipped_at,
               ABS(tl.quantity_lb) as quantity_lb,
               t.customer_name, t.order_reference,
               so.id as sales_order_id,
               so.order_number,
               s.id as shipment_id
        FROM ledger_current_transaction_lines tl
        JOIN ledger_current_transactions t ON t.id = tl.transaction_id
        LEFT JOIN sales_order_shipments sos ON sos.transaction_id = t.id
        LEFT JOIN sales_order_lines sol ON sol.id = sos.sales_order_line_id
        LEFT JOIN sales_orders so ON so.id = sol.sales_order_id
        LEFT JOIN shipment_lines sl ON sl.transaction_id = t.id
        LEFT JOIN shipments s ON s.id = sl.shipment_id
        WHERE tl.lot_id = %s
          AND t.type = 'ship'
          AND tl.quantity_lb < 0
          AND t.effective_status = 'posted'
        ORDER BY t.timestamp
    """, (lot_row['id'],))
    shipments = cur.fetchall()

    shipments_out = []
    for sh in shipments:
        sh_date, sh_time = format_timestamp(sh['shipped_at'])
        shipments_out.append({
            "transaction_id": sh['transaction_id'],
            "shipped_at": sh['shipped_at'].isoformat() if sh['shipped_at'] else None,
            "shipped_date": sh_date,
            "shipped_time": sh_time,
            "quantity_lb": float(sh['quantity_lb']),
            "customer_name": sh['customer_name'],
            "order_reference": sh['order_reference'],
            "sales_order_id": sh['sales_order_id'],
            "order_number": sh['order_number'],
            "shipment_id": sh['shipment_id']
        })

    total_shipped = sum(s['quantity_lb'] for s in shipments_out)

    # 4. Current on-hand quantity (posted-only)
    on_hand = lot_on_hand(cur, lot_row['id'])

    return {
        "trace_type": "ingredient",
        "lot_id": lot_row['id'],
        "product_id": lot_row['product_id'],
        "lot_code": lot_code,
        "product_name": lot_row['product_name'],
        "supplier_lot_code": lot_row['supplier_lot_code'],
        "entry_source": lot_row['entry_source'],
        "supplier": supplier_info,
        "downstream_batches": batches_out,
        "direct_shipments": shipments_out,
        "total_shipped_lb": total_shipped,
        "on_hand_lb": on_hand
    }


@app.get("/trace/ingredient/{lot_code}")
def trace_ingredient(lot_code: str, product_id: Optional[int] = Query(None), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            query = """
                SELECT l.id, l.lot_code, l.supplier_lot_code, l.entry_source,
                       p.name as ingredient_name, l.product_id
                FROM lots l
                JOIN products p ON p.id = l.product_id
                WHERE LOWER(l.lot_code) = LOWER(%s)
            """
            params = [lot_code]
            if product_id is not None:
                query += " AND l.product_id = %s"
                params.append(product_id)
            cur.execute(query, params)
            rows = cur.fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail=f"Lot '{lot_code}' not found")
            if len(rows) > 1:
                return JSONResponse(status_code=409, content={
                    "error": "ambiguous_lot_code",
                    "message": f"Lot code '{lot_code}' matches multiple products",
                    "matches": [{"lot_id": r['id'], "product_id": r['product_id'],
                                 "product_name": r['ingredient_name'], "entry_source": r['entry_source']} for r in rows]
                })
            lot = rows[0]

            # Output lot (pack/make): don't reject, provide full picture
            is_output_lot = lot['entry_source'] in OUTPUT_ENTRY_SOURCES

            if is_output_lot:
                # Find upstream ingredients: get the make/pack transaction that created this lot,
                # then pull ingredient_lot_consumption rows for that transaction
                cur.execute("""
                    SELECT t.id as transaction_id
                    FROM ledger_current_transactions t
                    JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                    WHERE t.type IN ('make', 'pack')
                      AND t.effective_status = 'posted'
                      AND tl.lot_id = %s
                      AND tl.quantity_lb > 0
                """, (lot['id'],))
                prod_txn = cur.fetchone()

                upstream_ingredients = []
                if prod_txn:
                    cur.execute("""
                        SELECT l.id as lot_id, p.id as product_id,
                               p.name as ingredient_name, l.lot_code,
                               ilc.quantity_lb, l.supplier_lot_code
                        FROM ingredient_lot_consumption ilc
                        JOIN products p ON p.id = ilc.ingredient_product_id
                        JOIN lots l ON l.id = ilc.ingredient_lot_id
                        WHERE ilc.transaction_id = %s
                    """, (prod_txn['transaction_id'],))
                    upstream_ingredients = [
                        {
                            "lot_id": r['lot_id'],
                            "product_id": r['product_id'],
                            "ingredient_name": r['ingredient_name'],
                            "lot_code": r['lot_code'],
                            "quantity_lb": float(r['quantity_lb']),
                            "supplier_lot_code": r['supplier_lot_code']
                        } for r in cur.fetchall()
                    ]

            else:
                upstream_ingredients = None

            # Production/pack consumption, keyed by lot id because lot codes can collide.
            cur.execute("""
                SELECT DISTINCT bl.id as lot_id, bp.id as product_id,
                       bl.lot_code as batch_lot, bp.name as batch_product,
                       ilc.quantity_lb as quantity_consumed
                FROM ingredient_lot_consumption ilc
                JOIN ledger_current_transactions t ON t.id = ilc.transaction_id
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id AND tl.quantity_lb > 0
                JOIN lots bl ON bl.id = tl.lot_id
                JOIN products bp ON bp.id = bl.product_id
                WHERE ilc.ingredient_lot_id = %s
                  AND t.effective_status = 'posted'
            """, (lot['id'],))
            batches = cur.fetchall()

            # Direct shipments (ship transactions that deducted from this lot)
            cur.execute("""
                SELECT t.id as transaction_id, t.timestamp as shipped_at,
                       ABS(tl.quantity_lb) as quantity_lb,
                       t.customer_name, t.order_reference,
                       so.id as sales_order_id,
                       so.order_number,
                       s.id as shipment_id
                FROM ledger_current_transaction_lines tl
                JOIN ledger_current_transactions t ON t.id = tl.transaction_id
                LEFT JOIN sales_order_shipments sos ON sos.transaction_id = t.id
                LEFT JOIN sales_order_lines sol ON sol.id = sos.sales_order_line_id
                LEFT JOIN sales_orders so ON so.id = sol.sales_order_id
                LEFT JOIN shipment_lines sl ON sl.transaction_id = t.id
                LEFT JOIN shipments s ON s.id = sl.shipment_id
                WHERE tl.lot_id = %s
                  AND t.type = 'ship'
                  AND tl.quantity_lb < 0
                  AND t.effective_status = 'posted'
                ORDER BY t.timestamp
            """, (lot['id'],))
            shipments = cur.fetchall()

            direct_shipments = []
            for sh in shipments:
                sh_date, sh_time = format_timestamp(sh['shipped_at'])
                direct_shipments.append({
                    "transaction_id": sh['transaction_id'],
                    "shipped_at": sh['shipped_at'].isoformat() if sh['shipped_at'] else None,
                    "shipped_date": sh_date,
                    "shipped_time": sh_time,
                    "quantity_lb": float(sh['quantity_lb']),
                    "customer_name": sh['customer_name'],
                    "order_reference": sh['order_reference'],
                    "sales_order_id": sh['sales_order_id'],
                    "order_number": sh['order_number'],
                    "shipment_id": sh['shipment_id']
                })

            total_shipped = sum(s['quantity_lb'] for s in direct_shipments)

            # Current on-hand quantity (posted-only)
            on_hand = lot_on_hand(cur, lot['id'])

            result = {
                "lot_id": lot['id'],
                "product_id": lot['product_id'],
                "ingredient_lot_code": lot['lot_code'],
                "supplier_lot_code": lot['supplier_lot_code'],
                "ingredient_name": lot['ingredient_name'],
                "used_in_batches": [
                    {
                        "lot_id": b['lot_id'],
                        "product_id": b['product_id'],
                        "batch_lot_code": b['batch_lot'],
                        "batch_product": b['batch_product'],
                        "quantity_used": float(b['quantity_consumed'])
                    } for b in batches
                ],
                "direct_shipments": direct_shipments,
                "total_shipped_lb": total_shipped,
                "on_hand_lb": on_hand
            }

            if is_output_lot:
                result["lot_origin"] = lot['entry_source']
                result["origin_note"] = (
                    f"This lot was created via /{lot['entry_source'].replace('_output', '')}. "
                    f"For full upstream ingredients, see /trace/batch/{lot['lot_code']}"
                )
                result["upstream_ingredients"] = upstream_ingredients

            return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Trace ingredient failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/trace/supplier-lot/{supplier_lot_code}")
def trace_supplier_lot(supplier_lot_code: str, product_id: Optional[int] = Query(None), _: bool = Depends(verify_api_key)):
    """FDA recall-ready endpoint: given a supplier lot code, find all internal lots and trace full exposure."""
    try:
        with get_transaction() as cur:
            # Find all internal lots matching this supplier lot code
            # Check both lots.supplier_lot_code and lot_supplier_codes table (commingled receipts)
            query = """
                SELECT DISTINCT l.id as lot_id, l.lot_code, l.supplier_lot_code,
                       l.entry_source, p.id as product_id, p.name as product_name
                FROM lots l
                JOIN products p ON p.id = l.product_id
                WHERE LOWER(l.supplier_lot_code) = LOWER(%s)
            """
            params = [supplier_lot_code]
            if product_id is not None:
                query += " AND l.product_id = %s"
                params.append(product_id)
            query += """
                UNION

                SELECT DISTINCT l.id as lot_id, l.lot_code, l.supplier_lot_code,
                       l.entry_source, p.id as product_id, p.name as product_name
                FROM lot_supplier_codes lsc
                JOIN lots l ON l.id = lsc.lot_id
                JOIN products p ON p.id = l.product_id
                WHERE LOWER(lsc.supplier_lot_code) = LOWER(%s)
            """
            params.append(supplier_lot_code)
            if product_id is not None:
                query += " AND l.product_id = %s"
                params.append(product_id)
            query += " ORDER BY lot_code"
            cur.execute(query, params)
            matched_lots = cur.fetchall()

            if not matched_lots:
                raise HTTPException(status_code=404,
                    detail=f"No internal lots found for supplier lot code '{supplier_lot_code}'")

            total_received = 0.0
            total_in_production = 0.0
            total_shipped = 0.0
            total_on_hand = 0.0

            results = []
            for lot in matched_lots:
                lot_id = lot['lot_id']

                # On-hand quantity (posted-only)
                on_hand = lot_on_hand(cur, lot_id)

                # Total received (positive qty from receive transactions)
                cur.execute("""
                    SELECT COALESCE(SUM(tl.quantity_lb), 0) as total_received
                    FROM ledger_current_transaction_lines tl
                    JOIN ledger_current_transactions t ON t.id = tl.transaction_id
                    WHERE tl.lot_id = %s AND t.type = 'receive'
                      AND tl.quantity_lb > 0
                      AND t.effective_status = 'posted'
                """, (lot_id,))
                received = float(cur.fetchone()['total_received'])

                # Production consumption (ingredient_lot_consumption)
                cur.execute("""
                    SELECT bl.id as lot_id, bp.id as product_id,
                           bl.lot_code as batch_lot_code, bp.name as batch_product,
                           ilc.quantity_lb as quantity_used_lb
                    FROM ingredient_lot_consumption ilc
                    JOIN ledger_current_transactions t ON t.id = ilc.transaction_id
                    JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id AND tl.quantity_lb > 0
                    JOIN lots bl ON bl.id = tl.lot_id
                    JOIN products bp ON bp.id = bl.product_id
                    WHERE ilc.ingredient_lot_id = %s
                      AND t.effective_status = 'posted'
                    ORDER BY t.timestamp
                """, (lot_id,))
                production = cur.fetchall()
                production_usage = [{
                    "lot_id": p['lot_id'],
                    "product_id": p['product_id'],
                    "batch_lot_code": p['batch_lot_code'],
                    "batch_product": p['batch_product'],
                    "quantity_used_lb": float(p['quantity_used_lb'])
                } for p in production]
                lot_production_total = sum(p['quantity_used_lb'] for p in production_usage)

                # Direct customer shipments
                cur.execute("""
                    SELECT t.id as transaction_id, t.timestamp as shipped_at,
                           ABS(tl.quantity_lb) as quantity_lb,
                           t.customer_name, t.order_reference,
                           so.order_number
                    FROM ledger_current_transaction_lines tl
                    JOIN ledger_current_transactions t ON t.id = tl.transaction_id
                    LEFT JOIN sales_order_shipments sos ON sos.transaction_id = t.id
                    LEFT JOIN sales_order_lines sol ON sol.id = sos.sales_order_line_id
                    LEFT JOIN sales_orders so ON so.id = sol.sales_order_id
                    WHERE tl.lot_id = %s
                      AND t.type = 'ship'
                      AND tl.quantity_lb < 0
                      AND t.effective_status = 'posted'
                    ORDER BY t.timestamp
                """, (lot_id,))
                shipments = cur.fetchall()
                customer_shipments = []
                for sh in shipments:
                    sh_date, sh_time = format_timestamp(sh['shipped_at'])
                    customer_shipments.append({
                        "transaction_id": sh['transaction_id'],
                        "shipped_at": sh['shipped_at'].isoformat() if sh['shipped_at'] else None,
                        "shipped_date": sh_date,
                        "shipped_time": sh_time,
                        "quantity_lb": float(sh['quantity_lb']),
                        "customer_name": sh['customer_name'],
                        "order_number": sh['order_number']
                    })
                lot_shipped_total = sum(s['quantity_lb'] for s in customer_shipments)

                results.append({
                    "lot_id": lot['lot_id'],
                    "lot_code": lot['lot_code'],
                    "product_name": lot['product_name'],
                    "product_id": lot['product_id'],
                    "on_hand_lb": on_hand,
                    "total_received_lb": received,
                    "production_usage": production_usage,
                    "customer_shipments": customer_shipments
                })

                total_received += received
                total_in_production += lot_production_total
                total_shipped += lot_shipped_total
                total_on_hand += on_hand

            return {
                "supplier_lot_code": supplier_lot_code,
                "matched_internal_lots": results,
                "total_exposure_summary": {
                    "total_received_lb": total_received,
                    "total_in_production_lb": total_in_production,
                    "total_shipped_to_customers_lb": total_shipped,
                    "total_on_hand_lb": total_on_hand
                }
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Trace supplier lot failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# TRANSACTION HISTORY ENDPOINTS
# ═══════════════════════════════════════════════════════════════

class CertificationCreate(BaseModel):
    business_date: date
    certified_at: datetime
    source_type: Literal["manual", "whatsapp_export"] = "manual"
    source_message_id: Optional[str] = None
    notes: Optional[str] = None


class CertificationCorrection(BaseModel):
    certified_at: datetime
    reason: str
    source_message_id: Optional[str] = None
    notes: Optional[str] = None


def _effective_certification(cur, business_date: date):
    cur.execute(
        "SELECT * FROM current_certifications WHERE business_date = %s",
        (business_date,),
    )
    return cur.fetchone()


@app.post("/records/certifications")
def create_certification(req: CertificationCreate, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute(
                """INSERT INTO certifications
                       (business_date, certified_at, operator_id, source_type,
                        source_message_id, notes)
                   VALUES (%s, %s, %s, %s, %s, %s)
                   RETURNING id, business_date, certified_at, operator_id,
                             source_type, source_message_id, notes, created_at""",
                (
                    req.business_date,
                    req.certified_at,
                    _operator_id(_),
                    req.source_type,
                    req.source_message_id,
                    req.notes,
                ),
            )
            row = dict(cur.fetchone())
            row["certification_id"] = str(row.pop("id"))
            return row
    except psycopg2.errors.UniqueViolation:
        raise HTTPException(
            409,
            f"A certification already exists for {req.business_date}; use the correction endpoint",
        )
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e):
            raise
        logger.error(f"Certification create failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/records/certifications/{certification_id}/corrections")
def correct_certification(
    certification_id: uuid.UUID,
    req: CertificationCorrection,
    _: bool = Depends(verify_api_key),
):
    reason = req.reason.strip()
    if not reason:
        raise HTTPException(422, "A non-empty correction reason is required")
    try:
        with get_transaction() as cur:
            cur.execute(
                "SELECT * FROM certifications WHERE id = %s FOR UPDATE",
                (str(certification_id),),
            )
            original = cur.fetchone()
            if not original:
                raise HTTPException(404, "Certification not found")
            current = _effective_certification(cur, original["business_date"])
            if str(current["certification_id"]) != str(certification_id):
                raise HTTPException(409, "Only the current certification can be corrected")
            cur.execute(
                """INSERT INTO certifications
                       (business_date, certified_at, operator_id, source_type,
                        source_message_id, notes, supersedes_certification_id,
                        correction_reason)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                   RETURNING id, business_date, certified_at, operator_id,
                             source_type, source_message_id, notes,
                             supersedes_certification_id, correction_reason, created_at""",
                (
                    original["business_date"],
                    req.certified_at,
                    _operator_id(_),
                    original["source_type"],
                    req.source_message_id or original["source_message_id"],
                    req.notes if req.notes is not None else original["notes"],
                    str(certification_id),
                    reason,
                ),
            )
            row = dict(cur.fetchone())
            row["certification_id"] = str(row.pop("id"))
            row["supersedes_certification_id"] = str(row["supersedes_certification_id"])
            return row
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e):
            raise
        logger.error(f"Certification correction failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/records/certifications/{business_date}")
def get_certification(business_date: date, _: bool = Depends(verify_api_key)):
    with get_transaction() as cur:
        current = _effective_certification(cur, business_date)
        if not current:
            raise HTTPException(404, f"No certification for {business_date}")
        cur.execute(
            """SELECT id, business_date, certified_at, operator_id, source_type,
                      source_message_id, notes, supersedes_certification_id,
                      correction_reason, created_at
               FROM certifications WHERE business_date = %s
               ORDER BY created_at, id""",
            (business_date,),
        )
        chain = [dict(row) for row in cur.fetchall()]
        for row in chain:
            row["id"] = str(row["id"])
            if row["supersedes_certification_id"]:
                row["supersedes_certification_id"] = str(row["supersedes_certification_id"])
        result = dict(current)
        result["certification_id"] = str(result["certification_id"])
        result["chain"] = chain
        return result


def _late_records(cur, target_date: date):
    certification = _effective_certification(cur, target_date)
    if not certification:
        raise HTTPException(409, f"Business day {target_date} has not been certified")
    cutoff = certification["certified_at"]

    cur.execute(
        """SELECT ct.id AS transaction_id, ct.type, ct.business_date,
                  ct.occurred_at, ct.created_at, ct.operator_id,
                  ct.effective_status, ct.latest_correction_id,
                  (ct.created_at > %s) AS late,
                  GREATEST(EXTRACT(EPOCH FROM (ct.created_at - %s)) / 60.0, 0) AS minutes_after_cutoff
           FROM ledger_current_transactions ct
           WHERE ct.business_date = %s
           ORDER BY ct.created_at, ct.id""",
        (cutoff, cutoff, target_date),
    )
    entries = [dict(row) for row in cur.fetchall()]
    for row in entries:
        if row["latest_correction_id"]:
            row["latest_correction_id"] = str(row["latest_correction_id"])
        row["minutes_after_cutoff"] = float(row["minutes_after_cutoff"] or 0)

    cur.execute(
        """SELECT c.id AS correction_id,
                  COALESCE(t.id, line_t.id) AS transaction_id,
                  c.target_table, c.target_id,
                  c.event_type, c.reason, c.operator_id, c.created_at,
                  (c.created_at > %s) AS late,
                  GREATEST(EXTRACT(EPOCH FROM (c.created_at - %s)) / 60.0, 0) AS minutes_after_cutoff
           FROM ledger_corrections c
           LEFT JOIN transactions t
             ON c.target_table = 'transactions' AND t.id = c.target_id
           LEFT JOIN transaction_lines tl
             ON c.target_table = 'transaction_lines' AND tl.id = c.target_id
           LEFT JOIN transactions line_t ON line_t.id = tl.transaction_id
           WHERE COALESCE(t.business_date, line_t.business_date) = %s
           ORDER BY c.created_at, c.id""",
        (cutoff, cutoff, target_date),
    )
    corrections = [dict(row) for row in cur.fetchall()]
    for row in corrections:
        row["correction_id"] = str(row["correction_id"])
        row["minutes_after_cutoff"] = float(row["minutes_after_cutoff"] or 0)

    return {
        "business_date": target_date,
        "certification": {
            "certification_id": str(certification["certification_id"]),
            "certified_at": cutoff,
            "operator_id": certification["operator_id"],
        },
        "entries": entries,
        "late_entries": [row for row in entries if row["late"]],
        "corrections": corrections,
        "late_corrections": [row for row in corrections if row["late"]],
    }


@app.get("/records/late")
def get_late_records(
    business_date: date = Query(...),
    _: bool = Depends(verify_api_key),
):
    with get_transaction() as cur:
        return _late_records(cur, business_date)


@app.get("/records/late.csv")
def export_late_records_csv(
    business_date: date = Query(...),
    _: bool = Depends(verify_api_key),
):
    with get_transaction() as cur:
        data = _late_records(cur, business_date)

    output = io.StringIO()
    columns = [
        "record_kind", "record_id", "transaction_id", "transaction_type",
        "business_date", "occurred_at", "created_at", "operator_id",
        "effective_status", "correction_type", "reason", "late",
        "minutes_after_cutoff", "certified_at",
    ]
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    cutoff = data["certification"]["certified_at"]
    for row in data["entries"]:
        writer.writerow({
            "record_kind": "original",
            "record_id": row["transaction_id"],
            "transaction_id": row["transaction_id"],
            "transaction_type": row["type"],
            "business_date": row["business_date"],
            "occurred_at": row["occurred_at"],
            "created_at": row["created_at"],
            "operator_id": row["operator_id"],
            "effective_status": row["effective_status"],
            "late": row["late"],
            "minutes_after_cutoff": row["minutes_after_cutoff"],
            "certified_at": cutoff,
        })
    for row in data["corrections"]:
        writer.writerow({
            "record_kind": "correction",
            "record_id": row["correction_id"],
            "transaction_id": row["transaction_id"],
            "created_at": row["created_at"],
            "operator_id": row["operator_id"],
            "correction_type": row["event_type"],
            "reason": row["reason"],
            "late": row["late"],
            "minutes_after_cutoff": row["minutes_after_cutoff"],
            "certified_at": cutoff,
        })
    headers = {
        "Content-Disposition": f'attachment; filename="late-entries-{business_date}.csv"'
    }
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers=headers)

@app.get("/transactions/history")
def get_transaction_history(
    limit: int = Query(default=20, ge=1, le=1000),
    transaction_type: Optional[str] = Query(default=None),
    product_name: Optional[str] = Query(default=None),
    since: Optional[date] = Query(default=None, description="Start date (YYYY-MM-DD), inclusive"),
    until: Optional[date] = Query(default=None, description="End date (YYYY-MM-DD), inclusive"),
    _: bool = Depends(verify_api_key)
):
    try:
        with get_transaction() as cur:
            query = """
                SELECT ct.id, ct.type, ct.timestamp, ct.occurred_at,
                       ct.business_date, ct.created_at, ct.operator_id,
                       ct.effective_status AS status,
                       ct.latest_correction_id, ct.latest_correction_type,
                       ct.latest_correction_created_at,
                       ct.effective_record ->> 'bol_reference' AS bol_reference,
                       ct.effective_record ->> 'shipper_name' AS shipper_name,
                       ct.effective_record ->> 'customer_name' AS customer_name,
                       ct.effective_record ->> 'order_reference' AS order_reference,
                       ct.effective_record ->> 'adjust_reason' AS adjust_reason,
                       ct.effective_record ->> 'notes' AS notes,
                       lines.lines,
                       corrections.correction_chain
                FROM ledger_current_transactions ct
                LEFT JOIN LATERAL (
                    SELECT json_agg(json_build_object(
                           'product_name', p.name,
                           'product_id', tl.product_id,
                           'lot_code', l.lot_code,
                           'lot_id', tl.lot_id,
                           'quantity_lb', tl.quantity_lb,
                           'case_size_lb', p.case_size_lb,
                           'product_type', p.type
                       ) ORDER BY tl.id) AS lines
                    FROM ledger_current_transaction_lines tl
                    LEFT JOIN products p ON p.id = tl.product_id
                    LEFT JOIN lots l ON l.id = tl.lot_id
                    WHERE tl.transaction_id = ct.id
                ) lines ON true
                LEFT JOIN LATERAL (
                    SELECT json_agg(json_build_object(
                        'correction_id', c.id,
                        'event_type', c.event_type,
                        'reason', c.reason,
                        'operator_id', c.operator_id,
                        'created_at', c.created_at,
                        'previous_values', c.previous_values,
                        'replacement_values', c.replacement_values
                    ) ORDER BY c.created_at, c.id) AS correction_chain
                    FROM ledger_corrections c
                    WHERE (c.target_table = 'transactions' AND c.target_id = ct.id)
                       OR (c.target_table = 'transaction_lines' AND EXISTS (
                           SELECT 1 FROM transaction_lines target_line
                           WHERE target_line.id = c.target_id
                             AND target_line.transaction_id = ct.id
                       ))
                ) corrections ON true
                WHERE true
            """
            params = []

            if transaction_type:
                query += " AND ct.type = %s"
                params.append(transaction_type)

            if product_name:
                query += " AND EXISTS (SELECT 1 FROM ledger_current_transaction_lines ftl JOIN products fp ON fp.id = ftl.product_id WHERE ftl.transaction_id = ct.id AND fp.name ILIKE %s)"
                params.append(f"%{product_name}%")

            if since:
                query += " AND ct.business_date >= %s"
                params.append(since)

            if until:
                query += " AND ct.business_date <= %s"
                params.append(until)

            query += " ORDER BY ct.created_at DESC, ct.id DESC LIMIT %s"
            params.append(limit)
            
            cur.execute(query, params)
            transactions = cur.fetchall()
            
            for txn in transactions:
                date_str, time_str = format_timestamp(txn['occurred_at'])
                txn['date'] = date_str
                txn['time'] = time_str
                created_date, created_time = format_timestamp(txn['created_at'])
                txn['created_date'] = created_date
                txn['created_time'] = created_time
                if txn.get('latest_correction_id'):
                    txn['latest_correction_id'] = str(txn['latest_correction_id'])
                # Enrich lines with unit counts
                if txn.get('lines'):
                    for ln in txn['lines']:
                        cs = float(ln['case_size_lb']) if ln.get('case_size_lb') else None
                        qty = abs(float(ln['quantity_lb'] or 0))
                        ln['unit_count'] = round(qty / cs) if cs and cs > 0 and ln.get('product_type') != 'ingredient' else None
            
            return {"count": len(transactions), "transactions": transactions}
    except Exception as e:
        logger.error(f"Get transaction history failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/ledger/recent")
def get_recent_ledger_events(
    limit: int = Query(default=20, ge=1, le=50),
    _: bool = Depends(verify_api_key),
):
    """Return the global, append-only recent-ledger feed for the dashboard.

    Feed rows are either original transactions or append-only correction events.
    They are ordered by each row's own database-entered timestamp, rather than
    by business date or by the original transaction timestamp.  The effective
    status and effective lines always come from the ledger-current views, so a
    voided original remains visible while its balance impact is correctly
    excluded by POSTED_LINES elsewhere.

    Direction is intentionally derived here, never from a browser quantity sign:
    receive -> received; make -> produced; pack -> packed; ship -> shipped;
    adjust -> adjusted_in when its effective line total is non-negative,
    otherwise adjusted_out.  Pack is a transformation and therefore has its
    own explicit direction even when its lines contain both signs.
    """
    try:
        with get_transaction() as cur:
            # One read-only SELECT: line_data supplies effective product/lot/
            # quantity/unit data for originals and the target metadata needed
            # to render correction events.  No client-side ledger inference is
            # required or permitted.
            cur.execute("""
                WITH line_data AS (
                    SELECT ctl.transaction_id,
                           COALESCE(SUM(ctl.quantity_lb), 0) AS net_quantity,
                           json_agg(json_build_object(
                               'product_name', p.name,
                               'quantity', ctl.quantity_lb,
                               'unit', COALESCE(p.uom, 'lb'),
                               'lot_code', l.lot_code
                           ) ORDER BY ctl.id) AS lines
                    FROM ledger_current_transaction_lines ctl
                    LEFT JOIN products p ON p.id = ctl.product_id
                    LEFT JOIN lots l ON l.id = ctl.lot_id
                    GROUP BY ctl.transaction_id
                ),
                feed AS (
                    SELECT
                        'transaction'::text AS event_kind,
                        'TX-' || ct.id::text AS event_id,
                        ct.id AS transaction_id,
                        ct.type AS transaction_type,
                        ct.type AS event_type,
                        ct.business_date,
                        ct.occurred_at,
                        ct.created_at AS entered_at,
                        ct.created_at_source,
                        ct.effective_status,
                        ct.status AS raw_status,
                        NULL::uuid AS correction_id,
                        NULL::text AS correction_event_type,
                        NULL::text AS correction_reason,
                        NULL::text AS correction_target_table,
                        NULL::bigint AS correction_target_id,
                        CASE ct.type
                            WHEN 'receive' THEN 'received'
                            WHEN 'make' THEN 'produced'
                            WHEN 'pack' THEN 'packed'
                            WHEN 'ship' THEN 'shipped'
                            WHEN 'adjust' THEN CASE
                                WHEN COALESCE(ld.net_quantity, 0) >= 0 THEN 'adjusted_in'
                                ELSE 'adjusted_out'
                            END
                            ELSE ct.type
                        END AS direction,
                        COALESCE(ld.lines, '[]'::json) AS lines
                    FROM ledger_current_transactions ct
                    LEFT JOIN line_data ld ON ld.transaction_id = ct.id

                    UNION ALL

                    SELECT
                        'correction'::text AS event_kind,
                        'COR-' || c.id::text AS event_id,
                        target.id AS transaction_id,
                        target.type AS transaction_type,
                        c.event_type AS event_type,
                        target.business_date,
                        target.occurred_at,
                        c.created_at AS entered_at,
                        target.created_at_source,
                        target.effective_status,
                        target.status AS raw_status,
                        c.id AS correction_id,
                        c.event_type AS correction_event_type,
                        c.reason AS correction_reason,
                        c.target_table AS correction_target_table,
                        c.target_id AS correction_target_id,
                        CASE target.type
                            WHEN 'receive' THEN 'received'
                            WHEN 'make' THEN 'produced'
                            WHEN 'pack' THEN 'packed'
                            WHEN 'ship' THEN 'shipped'
                            WHEN 'adjust' THEN CASE
                                WHEN COALESCE(ld.net_quantity, 0) >= 0 THEN 'adjusted_in'
                                ELSE 'adjusted_out'
                            END
                            ELSE target.type
                        END AS direction,
                        '[]'::json AS lines
                    FROM ledger_corrections c
                    JOIN ledger_current_transactions target
                      ON target.id = CASE
                          WHEN c.target_table = 'transactions' THEN c.target_id
                          WHEN c.target_table = 'transaction_lines' THEN (
                              SELECT tl.transaction_id
                              FROM transaction_lines tl
                              WHERE tl.id = c.target_id
                          )
                      END
                    LEFT JOIN line_data ld ON ld.transaction_id = target.id
                )
                SELECT *
                FROM feed
                ORDER BY entered_at DESC, event_id DESC
                LIMIT %s
            """, (limit,))
            rows = cur.fetchall()

        events = []
        for row in rows:
            event = {
                "event_id": row["event_id"],
                "event_kind": row["event_kind"],
                "transaction_id": row["transaction_id"],
                "transaction_type": row["transaction_type"],
                "event_type": row["event_type"],
                "business_date": row["business_date"].isoformat() if row["business_date"] else None,
                "occurred_at": row["occurred_at"],
                "entered_at": row["entered_at"],
                "created_at_source": row["created_at_source"],
                "effective_status": row["effective_status"],
                "raw_status": row["raw_status"],
                "direction": row["direction"],
                "lines": row["lines"] or [],
                "correction": None,
            }
            if row["event_kind"] == "correction":
                event["correction"] = {
                    "id": str(row["correction_id"]),
                    "event_type": row["correction_event_type"],
                    "reason": row["correction_reason"],
                    "target_table": row["correction_target_table"],
                    "target_id": row["correction_target_id"],
                }
            events.append(event)
        return {"count": len(events), "events": events}
    except Exception as e:
        logger.error(f"Get recent ledger events failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# QUICK-CREATE PRODUCT ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.post("/products/quick-create")
def quick_create_product(req: QuickCreateProductRequest, _: bool = Depends(verify_api_key)):
    validate_bilingual(req.notes, req.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT id, name FROM products WHERE LOWER(name) = LOWER(%s)", (req.product_name,))
                existing = cur.fetchone()
                if existing:
                    return JSONResponse(status_code=409, content={
                        "error": f"Product '{req.product_name}' already exists",
                        "existing_product_id": existing['id']
                    })

                verification_notes = f"Quick-created. Name confidence: {req.name_confidence}."
                if req.notes:
                    verification_notes += f" {req.notes}"
                verification_notes_es = req.notes_es

                cur.execute("""
                    INSERT INTO products (name, type, uom, storage_type, verification_status, verification_notes, verification_notes_es, created_via, active)
                    VALUES (%s, %s, %s, %s, 'unverified', %s, %s, 'quick_create', true)
                    RETURNING id, name, type, uom, verification_status
                """, (req.product_name, req.product_type, req.uom, req.storage_type, verification_notes, verification_notes_es))
                product = cur.fetchone()

                try:
                    cur.execute("""
                        INSERT INTO product_verification_history (product_id, from_status, to_status, action, action_notes, action_notes_es, performed_by)
                        VALUES (%s, NULL, 'unverified', 'created', %s, %s, %s)
                    """, (product['id'], f"Quick-created during receive. {verification_notes}", verification_notes_es, req.performed_by))
                except Exception:
                    pass

                logger.info(f"Quick-created product: {product['name']} (ID: {product['id']})")

                return {
                    "success": True,
                    "product_id": product['id'],
                    "product_name": product['name'],
                    "product_type": product['type'],
                    "verification_status": product['verification_status'],
                    "message": f"Created '{product['name']}' - flagged for verification"
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Quick-create product failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/products/quick-create-batch")
def quick_create_batch_product(req: QuickCreateBatchProductRequest, _: bool = Depends(verify_api_key)):
    validate_bilingual(req.notes, req.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT id, name FROM products WHERE LOWER(name) = LOWER(%s)", (req.product_name,))
                existing = cur.fetchone()
                if existing:
                    return JSONResponse(status_code=409, content={
                        "error": f"Product '{req.product_name}' already exists",
                        "existing_product_id": existing['id']
                    })

                verification_notes = f"Quick-created for production. Category: {req.category}. Context: {req.production_context}. Name confidence: {req.name_confidence}."
                if req.notes:
                    verification_notes += f" {req.notes}"
                verification_notes_es = req.notes_es

                cur.execute("""
                    INSERT INTO products (name, type, uom, verification_status, verification_notes, verification_notes_es, production_context, created_via, active)
                    VALUES (%s, 'batch', 'lb', 'unverified', %s, %s, %s, 'quick_create_batch', true)
                    RETURNING id, name, type, verification_status
                """, (req.product_name, verification_notes, verification_notes_es, req.production_context))
                product = cur.fetchone()

                logger.info(f"Quick-created batch product: {product['name']} (ID: {product['id']})")

                return {
                    "success": True,
                    "product_id": product['id'],
                    "product_name": product['name'],
                    "product_type": product['type'],
                    "production_context": req.production_context,
                    "verification_status": product['verification_status'],
                    "message": f"Created batch '{product['name']}' - flagged for verification"
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Quick-create batch product failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# LOT REASSIGNMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.post("/lots/{lot_id}/reassign")
def reassign_lot(lot_id: int, req: LotReassignmentRequest, _: bool = Depends(verify_api_key)):
    validate_bilingual(req.reason_notes, req.reason_notes_es, "reason_notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Postgres forbids FOR UPDATE with GROUP BY (same bug fixed in
                # the ship paths in 89e15ae): lock the lot row first, then
                # aggregate on-hand in a separate query.
                cur.execute("""
                    SELECT l.id, l.lot_code, l.product_id, p.name as product_name,
                           COALESCE(p.label_type, 'house') as label_type
                    FROM lots l
                    JOIN products p ON p.id = l.product_id
                    WHERE l.id = %s
                    FOR UPDATE OF l
                """, (lot_id,))
                lot = cur.fetchone()

                if not lot:
                    raise HTTPException(status_code=404, detail=f"Lot ID {lot_id} not found")

                cur.execute(f"""
                    SELECT COALESCE(SUM(tl.quantity_lb), 0) as quantity_on_hand
                    FROM {POSTED_LINES} tl
                    WHERE tl.lot_id = %s
                """, (lot_id,))
                lot['quantity_on_hand'] = cur.fetchone()['quantity_on_hand']

                if lot['product_id'] == req.to_product_id:
                    return JSONResponse(status_code=400, content={
                        "error": f"Lot is already assigned to {lot['product_name']}"
                    })

                cur.execute("""
                    SELECT id, name, COALESCE(label_type, 'house') as label_type
                    FROM products WHERE id = %s
                """, (req.to_product_id,))
                to_product = cur.fetchone()

                if not to_product:
                    raise HTTPException(status_code=404, detail=f"Target product ID {req.to_product_id} not found")

                # SKU protection: block product_merge if either side is private-label
                if req.reason_code == 'product_merge':
                    if lot['label_type'] == 'private_label':
                        return JSONResponse(status_code=403, content={
                            "blocked": True,
                            "warning": (
                                f"BLOCKED: Cannot merge lots from private-label SKU '{lot['product_name']}'. "
                                f"Private-label products are identity-protected and cannot be merged or consolidated. "
                                f"If this is a correction, use reason_code 'incorrect_receive' or 'data_entry_error' instead."
                            ),
                            "source_product": lot['product_name'],
                            "source_label_type": lot['label_type']
                        })
                    if to_product['label_type'] == 'private_label':
                        return JSONResponse(status_code=403, content={
                            "blocked": True,
                            "warning": (
                                f"BLOCKED: Cannot merge lots into private-label SKU '{to_product['name']}'. "
                                f"Private-label products are identity-protected and cannot be merged or consolidated. "
                                f"If this is a correction, use reason_code 'incorrect_receive' or 'data_entry_error' instead."
                            ),
                            "target_product": to_product['name'],
                            "target_label_type": to_product['label_type']
                        })
                
                cur.execute("""
                    SELECT COUNT(*) as count FROM ingredient_lot_consumption WHERE ingredient_lot_id = %s
                """, (lot_id,))
                usage = cur.fetchone()
                
                cur.execute(
                    "SELECT id FROM ledger_current_transaction_lines WHERE lot_id = %s ORDER BY id",
                    (lot_id,),
                )
                affected_line_ids = [row["id"] for row in cur.fetchall()]
                cur.execute("UPDATE lots SET product_id = %s WHERE id = %s", (req.to_product_id, lot_id))
                line_correction_ids = [
                    _append_transaction_line_correction(
                        cur,
                        line_id,
                        {"product_id": req.to_product_id},
                        req.reason_notes or req.reason_code,
                        _operator_id(_),
                    )
                    for line_id in affected_line_ids
                ]
                
                reassignment_id = None
                try:
                    cur.execute("""
                        INSERT INTO lot_reassignments
                        (lot_id, lot_code, from_product_id, from_product_name, to_product_id, to_product_name,
                         quantity_affected, uom, reason_code, reason_notes, reason_notes_es, reassigned_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                    """, (lot_id, lot['lot_code'], lot['product_id'], lot['product_name'],
                          req.to_product_id, to_product['name'], float(lot['quantity_on_hand']), 'lb',
                          req.reason_code, req.reason_notes, req.reason_notes_es, req.performed_by))
                    _hist = cur.fetchone()
                    reassignment_id = _hist['id'] if _hist else None
                except Exception as e:
                    logger.warning(f"Failed to record lot reassignment history: {e}")

                logger.info(f"Reassigned lot {lot['lot_code']} from {lot['product_name']} to {to_product['name']}")

                response = {
                    "success": True,
                    "lot_id": lot_id,
                    "reassignment_id": reassignment_id,
                    "line_correction_ids": line_correction_ids,
                    "lot_code": lot['lot_code'],
                    "from_product": lot['product_name'],
                    "to_product": to_product['name'],
                    "reason_code": req.reason_code,
                    "production_usage_updated": usage['count'] if usage else 0,
                    "message": f"Reassigned lot {lot['lot_code']} to {to_product['name']}"
                }
                if req.reason_notes:
                    response["reason_notes"] = req.reason_notes
                if req.reason_notes_es:
                    response["reason_notes_es"] = req.reason_notes_es
                return response
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Lot reassignment failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# FOUND INVENTORY ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.post("/inventory/found")
def add_found_inventory(req: AddFoundInventoryRequest, _: bool = Depends(verify_api_key)):
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at, req.backfill
    )
    validate_bilingual(req.notes, req.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT id, name FROM products WHERE id = %s", (req.product_id,))
                product = cur.fetchone()
                
                if not product:
                    raise HTTPException(status_code=404, detail=f"Product ID {req.product_id} not found")
                
                cur.execute("SELECT pg_advisory_xact_lock(2)")

                now = get_plant_now()

                # Lot Identity Policy: honor physical lot code if provided
                if req.lot_code:
                    lot_code = req.lot_code
                else:
                    date_part = now.strftime("%y-%m-%d")
                    cur.execute("""
                        SELECT lot_code FROM lots WHERE lot_code LIKE %s ORDER BY lot_code DESC LIMIT 1
                    """, (f"{date_part}-FOUND-%",))
                    existing = cur.fetchone()
                    if existing:
                        try:
                            last_seq = int(existing['lot_code'].split('-')[-1])
                            seq = last_seq + 1
                        except (ValueError, IndexError):
                            seq = 1
                    else:
                        seq = 1
                    lot_code = f"{date_part}-FOUND-{seq:03d}"

                lot_id, is_new_lot = find_or_create_lot(
                    cur, req.product_id, lot_code, 'found_inventory',
                    entry_source_notes=req.notes, entry_source_notes_es=req.notes_es,
                    found_location=req.found_location, estimated_age=req.estimated_age
                )

                cur.execute("""
                    INSERT INTO transactions (
                        type, timestamp, notes, occurred_at, created_at_source
                    )
                    VALUES ('adjust', %s, %s, %s, %s)
                    RETURNING id
                """, (
                    now, f"Found inventory: {req.reason_code}",
                    occurred_at, created_at_source,
                ))
                txn_id = cur.fetchone()['id']

                cur.execute("""
                    INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb)
                    VALUES (%s, %s, %s, %s)
                """, (txn_id, req.product_id, lot_id, req.quantity))

                try:
                    cur.execute("""
                        INSERT INTO inventory_adjustments
                        (lot_id, product_id, adjustment_type, quantity_before, quantity_adjustment, quantity_after,
                         uom, reason_code, reason_notes, reason_notes_es, found_location, estimated_age, suspected_supplier, adjusted_by)
                        VALUES (%s, %s, 'found', 0, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (lot_id, req.product_id, req.quantity, req.quantity, req.uom,
                          req.reason_code, req.notes, req.notes_es, req.found_location, req.estimated_age,
                          req.suspected_supplier, req.performed_by))
                except Exception as e:
                    logger.warning(f"Failed to record inventory adjustment: {e}")
                
                logger.info(f"Added found inventory: {lot_code} - {req.quantity} {req.uom} of {product['name']}")
                
                return {
                    "success": True,
                    "lot_id": lot_id,
                    "lot_code": lot_code,
                    "product_name": product['name'],
                    "quantity": req.quantity,
                    "uom": req.uom,
                    "entry_source": "found_inventory",
                    "message": f"Added {req.quantity} {req.uom} of {product['name']} as lot {lot_code}"
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Add found inventory failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/inventory/found-with-new-product")
def add_found_inventory_with_new_product(req: AddFoundInventoryWithNewProductRequest, _: bool = Depends(verify_api_key)):
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at, req.backfill
    )
    validate_bilingual(req.notes, req.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT id, name FROM products WHERE LOWER(name) = LOWER(%s)", (req.product_name,))
                existing = cur.fetchone()
                if existing:
                    return JSONResponse(status_code=409, content={
                        "error": f"Product '{req.product_name}' already exists",
                        "existing_product_id": existing['id'],
                        "suggestion": "Use /inventory/found with the existing product_id"
                    })
                
                verification_notes = f"Quick-created during inventory count. {req.notes or ''}"
                verification_notes_es = req.notes_es
                cur.execute("""
                    INSERT INTO products (name, type, uom, storage_type, verification_status, verification_notes, verification_notes_es, created_via, active)
                    VALUES (%s, %s, %s, %s, 'unverified', %s, %s, 'quick_create_found_inventory', true)
                    RETURNING id, name
                """, (req.product_name, req.product_type, req.uom, req.storage_type, verification_notes, verification_notes_es))
                product = cur.fetchone()
                
                cur.execute("SELECT pg_advisory_xact_lock(2)")

                now = get_plant_now()

                # Lot Identity Policy: honor physical lot code if provided
                if req.lot_code:
                    lot_code = req.lot_code
                else:
                    date_part = now.strftime("%y-%m-%d")
                    cur.execute("SELECT lot_code FROM lots WHERE lot_code LIKE %s ORDER BY lot_code DESC LIMIT 1", (f"{date_part}-FOUND-%",))
                    existing_lot = cur.fetchone()
                    seq = (int(existing_lot['lot_code'].split('-')[-1]) + 1) if existing_lot else 1
                    lot_code = f"{date_part}-FOUND-{seq:03d}"

                lot_id, is_new_lot = find_or_create_lot(
                    cur, product['id'], lot_code, 'found_inventory',
                    entry_source_notes=req.notes, entry_source_notes_es=req.notes_es,
                    found_location=req.found_location, estimated_age=req.estimated_age
                )
                
                cur.execute("""
                    INSERT INTO transactions (
                        type, timestamp, notes, occurred_at, created_at_source
                    )
                    VALUES ('adjust', %s, %s, %s, %s)
                    RETURNING id
                """, (
                    now, f"Found inventory with new product: {req.reason_code}",
                    occurred_at, created_at_source,
                ))
                txn_id = cur.fetchone()['id']
                
                cur.execute("""
                    INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb)
                    VALUES (%s, %s, %s, %s)
                """, (txn_id, product['id'], lot_id, req.quantity))
                
                logger.info(f"Created product and found inventory: {product['name']} - {lot_code}")
                
                return {
                    "success": True,
                    "product_id": product['id'],
                    "product_name": product['name'],
                    "verification_status": "unverified",
                    "lot_id": lot_id,
                    "lot_code": lot_code,
                    "quantity": req.quantity,
                    "message": f"Created '{product['name']}' and added {req.quantity} {req.uom} as lot {lot_code}"
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Add found inventory with new product failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/inventory/found/queue")
def get_found_inventory_queue(limit: int = Query(default=50, ge=1, le=200), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute(f"""
                SELECT l.id as lot_id, l.lot_code, p.id as product_id, p.name as product_name,
                       COALESCE(SUM(tl.quantity_lb), 0) as quantity_on_hand,
                       l.entry_source, l.found_location, l.estimated_age, l.entry_source_notes
                FROM lots l
                JOIN products p ON p.id = l.product_id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE l.entry_source = 'found_inventory'
                GROUP BY l.id, p.id
                HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                ORDER BY l.id DESC
                LIMIT %s
            """, (limit,))
            lots = cur.fetchall()
        return {"count": len(lots), "found_inventory": lots}
    except Exception as e:
        logger.error(f"Get found inventory queue failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# PRODUCT VERIFICATION ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.post("/products/{product_id}/verify")
def verify_product(product_id: int, req: VerifyProductRequest, _: bool = Depends(verify_api_key)):
    validate_bilingual(req.notes, req.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT id, name, verification_status FROM products WHERE id = %s FOR UPDATE", (product_id,))
                product = cur.fetchone()

                if not product:
                    raise HTTPException(status_code=404, detail=f"Product ID {product_id} not found")

                old_status = product.get('verification_status', 'unverified')

                if req.action == 'verify':
                    new_status = 'verified'
                    new_name = req.verified_name or product['name']
                    cur.execute("""
                        UPDATE products SET verification_status = %s, name = %s, verification_notes = %s, verification_notes_es = %s
                        WHERE id = %s
                    """, (new_status, new_name, req.notes, req.notes_es, product_id))

                elif req.action == 'reject':
                    new_status = 'rejected'
                    cur.execute("""
                        UPDATE products SET verification_status = %s, active = false, verification_notes = %s, verification_notes_es = %s
                        WHERE id = %s
                    """, (new_status, req.notes, req.notes_es, product_id))

                elif req.action == 'archive':
                    new_status = 'archived'
                    cur.execute("""
                        UPDATE products SET verification_status = %s, active = false, verification_notes = %s, verification_notes_es = %s
                        WHERE id = %s
                    """, (new_status, req.notes, req.notes_es, product_id))
                else:
                    raise HTTPException(status_code=400, detail=f"Invalid action: {req.action}")

                try:
                    cur.execute("""
                        INSERT INTO product_verification_history (product_id, from_status, to_status, action, action_notes, action_notes_es, performed_by)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (product_id, old_status, new_status, req.action, req.notes, req.notes_es, req.performed_by))
                except Exception:
                    pass
                
                logger.info(f"Product {product_id} {req.action}: {old_status} -> {new_status}")
                
                return {
                    "success": True,
                    "product_id": product_id,
                    "action": req.action,
                    "from_status": old_status,
                    "to_status": new_status,
                    "message": f"Product {req.action}d successfully"
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Verify product failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# BOM ENDPOINTS
# ═══════════════════════════════════════════════════════════════

@app.get("/bom/products")
def list_bom_products(
    product_type: Optional[str] = Query(default=None),
    q: Optional[str] = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    _: bool = Depends(verify_api_key)
):
    try:
        with get_transaction() as cur:
            query = "SELECT id, name, odoo_code, type, uom, default_batch_lb, case_size_lb FROM products WHERE COALESCE(active, true) = true"
            params = []
            
            if product_type:
                query += " AND type = %s"
                params.append(product_type)
            
            if q:
                query += " AND (LOWER(name) LIKE LOWER(%s) OR LOWER(odoo_code) LIKE LOWER(%s))"
                params.extend([f"%{q}%", f"%{q}%"])
            
            query += " ORDER BY name LIMIT %s"
            params.append(limit)
            
            cur.execute(query, params)
            products = cur.fetchall()
        return {"count": len(products), "products": products}
    except Exception as e:
        logger.error(f"List BOM products failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/bom/batches/{batch_id}/formula")
def get_batch_formula(batch_id: int, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT id, name, default_batch_lb, verification_notes, verification_notes_es FROM products WHERE id = %s", (batch_id,))
            batch = cur.fetchone()

            if not batch:
                raise HTTPException(status_code=404, detail=f"Batch product ID {batch_id} not found")

            cur.execute("""
                SELECT bf.ingredient_product_id, p.name as ingredient_name, p.odoo_code, bf.quantity_lb,
                       COALESCE(bf.exclude_from_inventory, false) as exclude_from_inventory
                FROM batch_formulas bf
                JOIN products p ON p.id = bf.ingredient_product_id
                WHERE bf.product_id = %s
                ORDER BY bf.quantity_lb DESC
            """, (batch_id,))
            ingredients = cur.fetchall()

            response = {
                "batch_id": batch['id'],
                "batch_name": batch['name'],
                "batch_weight_lb": float(batch['default_batch_lb']) if batch['default_batch_lb'] else None,
                "ingredients": ingredients
            }
            production_warning = build_production_warning(batch)
            if production_warning:
                response["production_warning"] = production_warning
                response["verification_notes"] = production_warning["verification_notes"]
                if "verification_notes_es" in production_warning:
                    response["verification_notes_es"] = production_warning["verification_notes_es"]
            return response
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get batch formula failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# REASON CODES ENDPOINT
# ═══════════════════════════════════════════════════════════════

@app.get("/reason-codes")
def get_reason_codes(_: bool = Depends(verify_api_key)):
    return {
        "found_inventory_reasons": [
            {"code": "found_during_count", "description": "Discovered during physical inventory count"},
            {"code": "found_back_stock", "description": "Found in back stock or secondary location"},
            {"code": "predates_system", "description": "Inventory that existed before system go-live"},
            {"code": "unreceived_delivery", "description": "Delivery that was never formally received"}
        ],
        "lot_reassignment_reasons": [
            {"code": "incorrect_receive", "description": "Originally received against wrong product"},
            {"code": "product_merge", "description": "Products being merged/consolidated"},
            {"code": "data_entry_error", "description": "Simple data entry mistake"},
            {"code": "supplier_relabel", "description": "Supplier changed product labeling"},
            {"code": "other", "description": "Other reason (specify in notes)"}
        ],
        "adjustment_reasons": [
            {"code": "damage", "description": "Product damaged"},
            {"code": "spoilage", "description": "Product spoiled or expired"},
            {"code": "count_correction", "description": "Correction from physical count"},
            {"code": "sample", "description": "Used for samples"},
            {"code": "hydration_yield", "description": "Hydration/processing yield correction"},
            {"code": "other", "description": "Other reason (specify in notes)"}
        ]
    }


# ═══════════════════════════════════════════════════════════════
# CUSTOMER ENDPOINTS (v2.3.0)
# ═══════════════════════════════════════════════════════════════

@app.get("/customers")
def list_customers(active_only: bool = True, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            if active_only:
                cur.execute("SELECT id, name, contact_name, email, phone, active FROM customers WHERE active = true ORDER BY name")
            else:
                cur.execute("SELECT id, name, contact_name, email, phone, active FROM customers ORDER BY name")
            return {"customers": cur.fetchall()}
    except Exception as e:
        logger.error(f"List customers failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/customers/search")
def search_customers(q: str = Query(..., min_length=1), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute(
                """SELECT DISTINCT c.id, c.name, c.contact_name, c.phone, c.email
                   FROM customers c
                   LEFT JOIN customer_aliases ca ON ca.customer_id = c.id
                   WHERE c.active = true
                     AND (LOWER(c.name) LIKE LOWER(%s) OR LOWER(ca.alias) LIKE LOWER(%s))
                   ORDER BY c.name""",
                (f"%{q}%", f"%{q}%")
            )
            rows = cur.fetchall()
        return {"results": rows}
    except Exception as e:
        logger.error(f"Search customers failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/customers")
def create_customer(req: CustomerCreate, _: bool = Depends(verify_api_key)):
    validate_bilingual(req.notes, req.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """INSERT INTO customers (name, contact_name, email, phone, address, notes, notes_es)
                       VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING id, name""",
                    (req.name, req.contact_name, req.email, req.phone, req.address, req.notes, req.notes_es)
                )
                row = cur.fetchone()
                logger.info(f"Created customer: {row['name']} (ID: {row['id']})")
                return {"customer_id": row['id'], "name": row['name'], "message": f"Customer '{row['name']}' created"}
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        if "unique" in str(e).lower():
            raise HTTPException(409, f"Customer '{req.name}' already exists")
        logger.error(f"Create customer failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/customers/{customer_id}")
def update_customer(customer_id: int, req: CustomerUpdate, _: bool = Depends(verify_api_key)):
    try:
        def _work(conn):
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Separate aliases from customer table fields
                aliases = req.aliases
                updates = req.dict(exclude_none=True, exclude={'aliases'})

                if not updates and aliases is None:
                    raise HTTPException(400, "No fields to update")

                # Update customer table fields
                if updates:
                    set_clause = ", ".join(f"{k} = %s" for k in updates)
                    values = list(updates.values()) + [customer_id]
                    cur.execute(
                        f"UPDATE customers SET {set_clause} WHERE id = %s RETURNING id, name",
                        values
                    )
                    row = cur.fetchone()
                    if not row:
                        raise HTTPException(404, "Customer not found")
                else:
                    cur.execute("SELECT id, name FROM customers WHERE id = %s", (customer_id,))
                    row = cur.fetchone()
                    if not row:
                        raise HTTPException(404, "Customer not found")

                # Handle aliases: clear and replace
                if aliases is not None:
                    cur.execute("DELETE FROM customer_aliases WHERE customer_id = %s", (customer_id,))
                    for alias in aliases:
                        alias_stripped = alias.strip()
                        if alias_stripped:
                            cur.execute(
                                "INSERT INTO customer_aliases (customer_id, alias) VALUES (%s, %s)",
                                (customer_id, alias_stripped)
                            )

                # Fetch current aliases for response
                cur.execute(
                    "SELECT alias FROM customer_aliases WHERE customer_id = %s ORDER BY alias",
                    (customer_id,)
                )
                current_aliases = [r['alias'] for r in cur.fetchall()]

                return {
                    "customer_id": row['id'],
                    "name": row['name'],
                    "aliases": current_aliases,
                    "message": "Customer updated"
                }
        return run_idempotent_write_with_readonly_retry("PATCH /customers/{customer_id}", _work)
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        if "idx_customer_aliases_lower_alias" in str(e):
            raise HTTPException(409, f"One of the aliases is already in use by another customer")
        logger.error(f"Update customer failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# SALES ORDER ENDPOINTS (v2.3.0)
# ═══════════════════════════════════════════════════════════════

# Status state machine — all valid transitions (includes auto-transitions from shipOrderCommit)
VALID_TRANSITIONS = {
    'new':            ['confirmed', 'cancelled'],
    'confirmed':      ['in_production', 'cancelled'],
    'in_production':  ['ready', 'cancelled'],
    'ready':          ['in_production', 'shipped', 'partial_ship', 'cancelled'],
    'partial_ship':   ['shipped', 'cancelled'],
    'shipped':        ['invoiced'],
    'invoiced':       [],   # terminal
    'cancelled':      [],   # terminal
}

# Manual transitions — subset excluding shipped/partial_ship (those are auto-only via shipOrderCommit)
MANUAL_TRANSITIONS = {
    'new':            ['confirmed', 'cancelled'],
    'confirmed':      ['in_production', 'cancelled'],
    'in_production':  ['ready', 'cancelled'],
    'ready':          ['in_production', 'cancelled'],
    'partial_ship':   ['cancelled'],
    'shipped':        ['invoiced'],
    'invoiced':       [],
    'cancelled':      [],
}

@app.post("/sales/orders")
def create_sales_order(req: OrderCreate, _: bool = Depends(verify_api_key)):
    validate_bilingual(req.notes, req.notes_es, "notes")
    for line in req.lines:
        validate_bilingual(line.notes, line.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                customer_id, customer_name = resolve_customer_id(
                    cur, req.customer_name, address=req.customer_address
                )

                cur.execute(
                    """INSERT INTO sales_orders (customer_id, requested_ship_date, notes, notes_es, order_number, status)
                       VALUES (%s, %s, %s, %s, '', 'confirmed')
                       RETURNING id, order_number""",
                    (customer_id, req.requested_ship_date, req.notes, req.notes_es)
                )
                row = cur.fetchone()
                order_id, order_number = row['id'], row['order_number']

                line_results = []
                total_lb = 0
                warnings = []
                for line in req.lines:
                    product_id, prod_name = resolve_product_id(cur, line.product_name)

                    # Detect service items (Pallets, freight, etc.) — zero weight is valid
                    cur.execute(
                        "SELECT case_size_lb, COALESCE(is_service, false) AS is_service FROM products WHERE id = %s",
                        (product_id,)
                    )
                    prod_row = cur.fetchone()
                    is_service = prod_row and prod_row['is_service']

                    if is_service:
                        # Service items get zero weight, skip case-weight logic
                        line.quantity_lb = line.quantity_lb if line.quantity_lb else 0
                        effective_case_weight = None
                        used_unit = line.unit or 'each'
                    else:
                        # Fix #2: Auto-lookup case weight from product if not provided
                        effective_case_weight = line.case_weight_lb
                        used_unit = line.unit or 'lb'
                        if used_unit in ('cases', 'bags', 'boxes') and effective_case_weight is None:
                            if prod_row and prod_row.get('case_size_lb'):
                                effective_case_weight = float(prod_row['case_size_lb'])
                            else:
                                raise HTTPException(
                                    status_code=400,
                                    detail={
                                        "error_code": "CASE_WEIGHT_REQUIRED",
                                        "message": f"case_weight_lb is required for '{prod_name}' when ordering in {used_unit}. No default case weight is set for this product.",
                                        "input": prod_name,
                                        "suggestions": [],
                                    }
                                )
                            # Recalculate quantity_lb with looked-up weight
                            line.quantity_lb = line.quantity * effective_case_weight

                        # Fix #1: Warn if unit was not explicitly provided and quantity was given
                        if line.quantity is not None and line.unit is None:
                            warnings.append(
                                f"⚠️ '{prod_name}': No unit specified for quantity {line.quantity:,.0f} — "
                                f"defaulting to lb. Did you mean cases?"
                            )

                    cur.execute(
                        """INSERT INTO sales_order_lines (sales_order_id, product_id, quantity_lb, unit_price, notes, notes_es)
                           VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
                        (order_id, product_id, line.quantity_lb, line.unit_price, line.notes, line.notes_es)
                    )
                    line_id = cur.fetchone()['id']
                    total_lb += line.quantity_lb

                    # Fix #3: Quantity sanity check — compare to customer's average order size
                    if not is_service:
                        cur.execute("""
                            SELECT AVG(sol.quantity_lb) as avg_qty
                            FROM sales_order_lines sol
                            JOIN sales_orders so ON so.id = sol.sales_order_id
                            WHERE so.customer_id = %s AND sol.product_id = %s
                              AND sol.id != %s
                              AND sol.line_status != 'cancelled'
                        """, (customer_id, product_id, line_id))
                        avg_row = cur.fetchone()
                        if avg_row and avg_row['avg_qty'] and line.quantity_lb < float(avg_row['avg_qty']) * 0.25:
                            warnings.append(
                                f"⚠️ '{prod_name}': {line.quantity_lb:,.0f} lb is unusually low for {customer_name}. "
                                f"Their average order is {float(avg_row['avg_qty']):,.0f} lb. Double-check the quantity."
                            )

                    line_results.append({
                        "line_id": line_id,
                        "product": prod_name,
                        "quantity_lb": line.quantity_lb,
                        "original_quantity": line.quantity,
                        "original_unit": used_unit,
                        "case_weight_lb": effective_case_weight,
                        "unit_price": line.unit_price
                    })

                logger.info(f"Created sales order {order_number} for {customer_name} with {len(line_results)} lines")
                return {
                    "order_id": order_id,
                    "order_number": order_number,
                    "customer": customer_name,
                    "requested_ship_date": req.requested_ship_date,
                    "status": "confirmed",
                    "total_lb": total_lb,
                    "lines": line_results,
                    "warnings": warnings if warnings else None,
                    "message": f"Order {order_number} created with {len(line_results)} line(s)"
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Create sales order failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


_ORDERS_MATRIX_PAN_YIELD = {
    # Source: CNS Production Source of Truth v1.0, 2026-07-02.
    "70050": 322.6, "10300": 322.6, "70048": 380.12, "70052": 370.12,
    "70012": 322.6, "1614": 322.6, "70073": 426, "10001": 360,
    "10020": 360, "10002": 360, "67470": 360, "67473": 360,
    "67476": 360, "893": 360, "10010": 300, "10029": 300,
    "31012": None,
}
_ORDERS_MATRIX_BATCH_NUMBER = {
    "70050": "90002",
    "10300": "90002",
    "70052": "90024",
}

_MATRIX_COLORS = {
    "left": "2F5496", "granola": "C55A11", "coconut": "2E7D5B",
    "graham": "7B6A50", "other": "2F5496",
}
_MATRIX_BANDS = {
    "granola": ("FDF3EC", "FAE5D5"),
    "coconut": ("EDF7F1", "DCEEE4"),
    "graham": ("F5F2EC", "EAE4D8"),
    "other": ("FFFFFF", "EFEFEF"),
}
_MATRIX_SKU_ORDER = {
    sku: rank for rank, sku in enumerate(
        ("70050", "10300", "70048", "70052", "70012", "1614", "70073",
         "10001", "10020", "10002", "67470", "67473", "67476", "893",
         "10010", "10029", "31012")
    )
}


def _matrix_family(product_name: str) -> str:
    name = (product_name or "").lower()
    if "granola" in name:
        return "granola"
    if "coconut" in name or re.search(r"\bcoco\b", name):
        return "coconut"
    if "graham" in name:
        return "graham"
    return "other"


def _matrix_product_sort(product: dict):
    name = product["product_name"].lower()
    family = product["family"]
    family_rank = {"granola": 0, "coconut": 1, "graham": 3, "other": 4}[family]
    coconut_rank = 0
    vendor_rank = 0
    if family == "coconut":
        coconut_rank = 1 if "toast" in name else 0
        vendor_rank = 0 if "cns" in name else 1 if "unipro" in name else 2 if "cq" in name else 3
    return (family_rank + coconut_rank, vendor_rank, _MATRIX_SKU_ORDER.get(product["sku"], 999), name, product["sku"])


def _matrix_short_header(product_name: str) -> str:
    header = (product_name or "").replace("–", "-").replace("—", "-")
    header = re.sub(r"\bCoconut\b", "Coco", header, flags=re.IGNORECASE)
    header = re.sub(r"\bSweetened\b", "Swt", header, flags=re.IGNORECASE)
    header = re.sub(r"\s*-\s*", " ", header)
    header = re.sub(r"\s+Case\b", "", header, flags=re.IGNORECASE)
    header = re.sub(r"\b(\d+(?:\.\d+)?)\s*LB\b", r"\1#", header, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", header).strip()


_PER_LB_HEADER_RE = re.compile(r"per.?lb$", re.IGNORECASE)


def _matrix_column_header(product: dict) -> str:
    """Column header for one product column.

    Per-lb bulk SKUs are marked "(lb)" so a Cases-sheet reader does not mistake
    pounds for cases — unless the product name already ends in "per/lb", where
    the suffix would just stutter.
    """
    header = _matrix_short_header(product["product_name"])
    if product["per_lb"] and not _PER_LB_HEADER_RE.search(header):
        header += " (lb)"
    return header


def _matrix_apply_numeric(cell, number_format='#,##0;(#,##0);"—"'):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = number_format


def _matrix_quantity_note(
    sku: str, qty_cases: float, lb_per_case: float, sheet_name: str, per_lb: bool = False
) -> Optional[str]:
    """Build the production-planning note for one nonzero matrix quantity.

    Returns None when there is nothing truthful to say: a per-lb bulk SKU has no
    cases to convert, so without a pan yield on file it gets no note at all
    rather than misleading cases-to-pans text.
    """
    pounds = qty_cases * lb_per_case
    if per_lb:
        pan_yield = _ORDERS_MATRIX_PAN_YIELD.get(sku)
        if pan_yield is None:
            return None
        lead = f"{pounds:,.0f} lb (sold per lb)"
    else:
        qty_text = f"{qty_cases:,g}"
        if sheet_name == "Cases":
            lead = f"{qty_text} cases = {pounds:,.0f} lb"
        else:
            lead = f"{pounds:,.0f} lb ({qty_text} cases)"

        if sku == "31012":
            return f"{lead} — repack from 50 lb bulk, no pans"

        pan_yield = _ORDERS_MATRIX_PAN_YIELD.get(sku)
        if pan_yield is None:
            return f"{lead} — no pan yield on file"

    pans = pounds / pan_yield
    pans_text = "<0.1" if pans > 0 and f"{pans:,.1f}" == "0.0" else f"{pans:,.1f}"
    note = f"{lead} ≈ {pans_text} pans @ {pan_yield:g} lb/pan"
    batch_number = _ORDERS_MATRIX_BATCH_NUMBER.get(sku)
    if batch_number:
        note += f" (Batch {batch_number})"
    if sku == "70073":
        note += " (finished wt; oven load 341 lb + 85 lb premix)"
    return note


def _matrix_comment(text: str) -> Comment:
    comment = Comment(text, "Factory Ledger")
    comment.width = 260
    comment.height = 80
    return comment


def _build_orders_matrix_workbook(lines: List[dict], export_date: date) -> Workbook:
    products_by_sku = {}
    grouped = {}
    for line in lines:
        sku = str(line["sku"])
        qty = float(line["qty"])
        product = products_by_sku.setdefault(sku, {
            "sku": sku,
            "product_name": line["product_name"],
            "lb_per_case": line["lb_per_case"],
            "family": _matrix_family(line["product_name"]),
            "per_lb": bool(line.get("per_lb")),
            "has_fractional_qty": False,
        })
        product["has_fractional_qty"] = product["has_fractional_qty"] or not math.isclose(
            qty, round(qty), abs_tol=1e-9
        )
        key = (line["due_date"], line["customer"], line["order_id"])
        order_quantities = grouped.setdefault(key, {})
        order_quantities[sku] = order_quantities.get(sku, 0) + qty

    products = sorted(products_by_sku.values(), key=_matrix_product_sort)
    rows = sorted(grouped.items(), key=lambda item: (item[0][0] or date.max, item[0][1].lower(), str(item[0][2])))
    workbook = Workbook()
    workbook.remove(workbook.active)

    thin_bottom = Side(style="thin", color="D0D0D0")
    medium_top = Side(style="medium", color="404040")
    source_text = "CNS Production Source of Truth v1.0, 2026-07-02"

    for sheet_name in ("Cases", "Pounds"):
        ws = workbook.create_sheet(sheet_name)
        total_col = 5 + len(products)
        headers = ["Due date", "Weekday", "Customer", "Order ID"] + [
            _matrix_column_header(p) for p in products
        ] + ["Row total"]
        ws.append(headers)
        ws.row_dimensions[1].height = 34
        for col, value in enumerate(headers, 1):
            family = products[col - 5]["family"] if 5 <= col < total_col else "left"
            cell = ws.cell(1, col, value)
            cell.font = Font(name="Arial", size=7, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_MATRIX_COLORS[family])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 11
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["D"].hidden = True
        for col in range(5, total_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 11

        week_indices = {}
        previous_week = None
        for row_offset, ((due_date, customer, order_id), quantities) in enumerate(rows, 2):
            week_key = due_date - timedelta(days=due_date.weekday()) if due_date else None
            if week_key not in week_indices:
                week_indices[week_key] = len(week_indices)
            week_band = week_indices[week_key] % 2
            values = [due_date, due_date.strftime("%A") if due_date else "", customer, order_id]
            ws.append(values)
            ws.cell(row_offset, 1).number_format = "mm/dd/yyyy"

            for col in range(1, total_col + 1):
                cell = ws.cell(row_offset, col)
                cell.font = Font(name="Arial", size=6 if col == 3 else 10)
                cell.border = Border(bottom=thin_bottom)
                cell.alignment = Alignment(vertical="center")
                if col <= 4:
                    cell.fill = PatternFill("solid", fgColor="FFFFFF" if week_band == 0 else "DCE6F1")

            for product_index, product in enumerate(products, 5):
                qty = quantities.get(product["sku"])
                cell = ws.cell(row_offset, product_index)
                if qty is not None:
                    cell.value = qty if sheet_name == "Cases" else qty * product["lb_per_case"]
                    if not math.isclose(qty, 0, abs_tol=1e-9):
                        note = _matrix_quantity_note(
                            product["sku"], qty, product["lb_per_case"], sheet_name,
                            per_lb=product["per_lb"],
                        )
                        if note:
                            cell.comment = _matrix_comment(note)
                cell.fill = PatternFill("solid", fgColor=_MATRIX_BANDS[product["family"]][row_offset % 2])
                _matrix_apply_numeric(
                    cell,
                    '#,##0.#;(#,##0.#);"—"' if product["has_fractional_qty"] else '#,##0;(#,##0);"—"',
                )

            total_cell = ws.cell(row_offset, total_col)
            total_cell.value = f"=SUM(E{row_offset}:{get_column_letter(total_col - 1)}{row_offset})"
            total_cell.fill = PatternFill("solid", fgColor=("FFFFFF", "EFEFEF")[row_offset % 2])
            _matrix_apply_numeric(total_cell)

            if previous_week is not None and week_key != previous_week:
                for col in range(1, total_col + 1):
                    old = ws.cell(row_offset, col).border
                    ws.cell(row_offset, col).border = Border(top=medium_top, bottom=old.bottom)
            previous_week = week_key

            if due_date and due_date < export_date:
                for col in range(1, 5):
                    ws.cell(row_offset, col).font = Font(
                        name="Arial", size=6 if col == 3 else 10, bold=True, color="C00000"
                    )

        data_end = 1 + len(rows)
        total_label = "TOTAL CASES" if sheet_name == "Cases" else "TOTAL LB"
        summary_specs = [total_label, "Lb per case", "TOTAL POUNDS", "Lb per batch/pan (input)", "BATCHES / PANS"]
        summary_start = data_end + 1
        for offset, label in enumerate(summary_specs):
            row_num = summary_start + offset
            ws.cell(row_num, 1, label)
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
            ws.cell(row_num, 1).font = Font(name="Arial", size=10, bold=True)
            ws.cell(row_num, 1).fill = PatternFill("solid", fgColor="D9D2C4")
            ws.cell(row_num, 1).alignment = Alignment(vertical="center")
            for col in range(2, 5):
                ws.cell(row_num, col).fill = PatternFill("solid", fgColor="D9D2C4")
            for product_index, product in enumerate(products, 5):
                cell = ws.cell(row_num, product_index)
                cell.fill = PatternFill("solid", fgColor=_MATRIX_COLORS[product["family"]])
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                _matrix_apply_numeric(cell)

        total_row, lb_case_row, pounds_row, input_row, batches_row = range(summary_start, summary_start + 5)
        first_product = "E"
        last_product = get_column_letter(total_col - 1)
        for product_index, product in enumerate(products, 5):
            letter = get_column_letter(product_index)
            ws.cell(total_row, product_index, f"=SUM({letter}2:{letter}{data_end})")
            ws.cell(lb_case_row, product_index, product["lb_per_case"])
            ws.cell(pounds_row, product_index, f"={letter}{total_row}*{letter}{lb_case_row}" if sheet_name == "Cases" else f"={letter}{total_row}")

            input_cell = ws.cell(input_row, product_index)
            if product["sku"] in _ORDERS_MATRIX_PAN_YIELD:
                pan_yield = _ORDERS_MATRIX_PAN_YIELD[product["sku"]]
                input_cell.value = "n/a" if pan_yield is None else pan_yield
                note = source_text
                if product["sku"] == "70073":
                    note += ". Finished batch weight = 341 lb oven base + 85 lb premix; never schedule 426 lb through the oven."
                elif product["sku"] == "31012":
                    note += ". Repack routing; batch/pan input is not applicable."
                input_cell.comment = _matrix_comment(note)
            else:
                input_cell.fill = PatternFill("solid", fgColor="FFF2CC")
                input_cell.font = Font(name="Arial", size=10, bold=True, color="0000FF")
            batches_cell = ws.cell(
                batches_row, product_index,
                f'=IF(N({letter}{input_row})>0,{letter}{pounds_row}/{letter}{input_row},"—")',
            )
            pan_yield = _ORDERS_MATRIX_PAN_YIELD.get(product["sku"])
            if pan_yield is not None:
                total_lb = sum(
                    quantities.get(product["sku"], 0) * product["lb_per_case"]
                    for _, quantities in rows
                )
                batches_cell.comment = _matrix_comment(
                    f"{total_lb:,.0f} lb ÷ {pan_yield:g} lb/pan"
                )
            elif product["sku"] == "31012":
                batches_cell.comment = _matrix_comment("Repack from 50 lb bulk; no pans")
            else:
                batches_cell.comment = _matrix_comment("No pan yield on file")
            _matrix_apply_numeric(batches_cell, '#,##0.0;(#,##0.0);"—"')
            if product["has_fractional_qty"]:
                for row_num in (total_row, lb_case_row, pounds_row, input_row):
                    _matrix_apply_numeric(ws.cell(row_num, product_index), '#,##0.#;(#,##0.#);"—"')

        for row_num in (total_row, pounds_row):
            ws.cell(row_num, total_col, f"=SUM({first_product}{row_num}:{last_product}{row_num})")
        ws.cell(lb_case_row, total_col, "")
        ws.cell(input_row, total_col, "")
        ws.cell(batches_row, total_col, f'=IF(COUNT({first_product}{batches_row}:{last_product}{batches_row})>0,SUM({first_product}{batches_row}:{last_product}{batches_row}),"—")')
        for row_num in range(summary_start, summary_start + 5):
            total_cell = ws.cell(row_num, total_col)
            total_cell.fill = PatternFill("solid", fgColor=_MATRIX_COLORS["left"])
            total_cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
            _matrix_apply_numeric(total_cell, '#,##0.0;(#,##0.0);"—"' if row_num == batches_row else '#,##0;(#,##0);"—"')

        for col in range(1, total_col + 1):
            cell = ws.cell(summary_start, col)
            cell.border = Border(top=medium_top)
        ws.freeze_panes = "E2"
        ws.auto_filter.ref = f"A1:{get_column_letter(total_col)}{data_end}"

    return workbook


@app.get("/export/orders-matrix.xlsx", include_in_schema=False)
def export_orders_matrix(_: bool = Depends(verify_api_key)):
    """Dashboard-only styled matrix of open, production-relevant sales-order lines."""
    with get_transaction() as cur:
        cur.execute(
            """
            SELECT c.name AS customer, so.order_number AS order_id,
                   so.requested_ship_date AS due_date, p.odoo_code AS sku,
                   p.name AS product_name,
                   COALESCE(sol.quantity_lb / NULLIF(p.case_size_lb, 0), sol.quantity_lb) AS qty,
                   p.case_size_lb AS lb_per_case
            FROM sales_orders so
            JOIN customers c ON c.id = so.customer_id
            JOIN sales_order_lines sol ON sol.sales_order_id = so.id
            JOIN products p ON p.id = sol.product_id
            WHERE so.status NOT IN ('shipped', 'invoiced', 'cancelled')
              AND p.type = 'finished'
              AND NOT COALESCE(p.is_service, false)
              AND NOT COALESCE(p.no_production, false)
              AND NULLIF(TRIM(p.odoo_code), '') IS NOT NULL
            ORDER BY so.requested_ship_date ASC NULLS LAST, c.name, so.order_number, sol.id
            """
        )
        raw_lines = [dict(row) for row in cur.fetchall()]

    offending = []
    lines = []
    for line in raw_lines:
        lb_per_case = line.get("lb_per_case")
        # Bulk per-lb finished goods (70004/70013/70016) carry no case size by
        # design — NULL case_size_lb is the marker for "sold by the pound", so
        # their quantity is already pounds and one "case" is one pound. Only
        # genuinely invalid case data (<= 0) still fails the export.
        per_lb = lb_per_case is None
        if per_lb:
            lb_per_case = 1.0
        if lb_per_case <= 0 or line["qty"] is None:
            offending.append(str(line["sku"]))
            continue
        line["lb_per_case"] = float(lb_per_case)
        line["per_lb"] = per_lb
        lines.append(line)
    if offending:
        raise HTTPException(
            status_code=422,
            detail={
                "error_code": "INVALID_ORDER_EXPORT_CASE_SIZE",
                "message": "Cannot export orders matrix: missing or invalid case size",
                "offending_skus": sorted(set(offending)),
            },
        )

    export_date = datetime.now(ZoneInfo("America/New_York")).date()
    workbook = _build_orders_matrix_workbook(lines, export_date)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"CNS_Open_Orders_Matrix_{export_date.isoformat()}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_READINESS_BLOCKER_ORDER = {
    code: index for index, code in enumerate((
        "shortage",
        "unallocated",
        "partial_allocation",
        "unstaged",
        "missing_lot_dates",
        "not_floor_ready",
        "fulfillment_diverged",
        "no_ship_date",
        "inbound_cover",
        "service_only",
    ))
}


# Shared by all three readiness GETs. requested_orders is the complete page/set,
# so shipped/on-hand/allocation inputs are aggregated once rather than queried
# per order or line. Expired auto allocations are ignored by formula only: this
# SELECT deliberately contains no UPDATE.
SALES_ORDER_READINESS_SQL = """
    WITH requested_orders AS (
        SELECT unnest(%s::integer[]) AS sales_order_id
    ),
    line_base AS (
        SELECT sol.id AS line_id, sol.sales_order_id, sol.product_id,
               sol.quantity_lb AS ordered_lb,
               sol.quantity_shipped_lb AS shipped_recorded_lb,
               sol.line_status, p.name AS product, p.odoo_code AS sku,
               COALESCE(p.is_service, false) AS is_service
        FROM requested_orders ro
        JOIN sales_order_lines sol ON sol.sales_order_id = ro.sales_order_id
        JOIN products p ON p.id = sol.product_id
        WHERE sol.line_status <> 'cancelled'
    ),
    relevant_products AS (
        SELECT DISTINCT product_id
        FROM line_base
        WHERE NOT is_service
    ),
    posted AS (
        SELECT tl.transaction_id, tl.lot_id, tl.product_id, tl.quantity_lb,
               ct.type AS transaction_type
        FROM ledger_current_transaction_lines tl
        JOIN ledger_current_transactions ct ON ct.id = tl.transaction_id
        JOIN relevant_products rp ON rp.product_id = tl.product_id
        WHERE ct.effective_status = 'posted'
    ),
    lot_balances AS (
        SELECT l.id AS lot_id, l.product_id, l.lot_code, l.entry_source,
               l.received_at, l.created_at,
               COALESCE(SUM(posted.quantity_lb), 0) AS on_hand_lb
        FROM relevant_products rp
        JOIN lots l ON l.product_id = rp.product_id
        LEFT JOIN posted ON posted.lot_id = l.id
        GROUP BY l.id, l.product_id, l.lot_code, l.entry_source,
                 l.received_at, l.created_at
    ),
    on_hand_sku AS (
        SELECT product_id, COALESCE(SUM(on_hand_lb), 0) AS on_hand_lb
        FROM lot_balances
        GROUP BY product_id
    ),
    shipped_eff AS (
        SELECT sos.sales_order_line_id AS line_id,
               SUM(ABS(posted.quantity_lb)) AS shipped_effective_lb
        FROM line_base lb
        JOIN sales_order_shipments sos ON sos.sales_order_line_id = lb.line_id
        JOIN posted ON posted.transaction_id = sos.transaction_id
                   AND posted.product_id = lb.product_id
                   AND posted.transaction_type = 'ship'
        WHERE NOT lb.is_service
        GROUP BY sos.sales_order_line_id
    ),
    live_alloc AS (
        SELECT soa.sales_order_line_id AS line_id, soa.product_id, soa.lot_id,
               soa.quantity_lb
        FROM sales_order_allocations soa
        JOIN relevant_products rp ON rp.product_id = soa.product_id
        WHERE soa.status = 'active'
          AND (soa.expires_at IS NULL OR soa.expires_at > now())
    ),
    alloc_by_line AS (
        SELECT line_id, product_id,
               SUM(quantity_lb) AS allocated_lb,
               SUM(quantity_lb) FILTER (WHERE lot_id IS NULL) AS allocated_sku_lb,
               SUM(quantity_lb) FILTER (WHERE lot_id IS NOT NULL) AS allocated_lot_lb
        FROM live_alloc
        GROUP BY line_id, product_id
    ),
    alloc_by_product AS (
        SELECT product_id,
               SUM(quantity_lb) AS allocated_product_lb,
               SUM(quantity_lb) FILTER (WHERE lot_id IS NULL) AS allocated_product_sku_lb
        FROM live_alloc
        GROUP BY product_id
    ),
    alloc_by_lot AS (
        SELECT product_id, lot_id, SUM(quantity_lb) AS allocated_lot_lb
        FROM live_alloc
        WHERE lot_id IS NOT NULL
        GROUP BY product_id, lot_id
    ),
    line_lot_payload AS (
        SELECT line_id,
               jsonb_agg(jsonb_build_object(
                   'lot_id', lot_id,
                   'quantity_lb', quantity_lb
               ) ORDER BY lot_id) AS line_lot_allocations
        FROM live_alloc
        WHERE lot_id IS NOT NULL
        GROUP BY line_id
    ),
    inbound AS (
        SELECT er.product_id, SUM(er.expected_qty) AS inbound_open_lb
        FROM expected_receipts er
        JOIN relevant_products rp ON rp.product_id = er.product_id
        WHERE er.status = 'open'
        GROUP BY er.product_id
    ),
    lot_payload AS (
        SELECT lb.product_id,
               jsonb_agg(jsonb_build_object(
                   'lot_id', lb.lot_id,
                   'lot_code', lb.lot_code,
                   'entry_source', lb.entry_source,
                   'received_at', lb.received_at,
                   'on_hand_lb', lb.on_hand_lb,
                   'allocated_lot_lb', COALESCE(abl.allocated_lot_lb, 0)
               ) ORDER BY COALESCE(lb.received_at, lb.created_at) ASC, lb.lot_id ASC) AS lots
        FROM lot_balances lb
        LEFT JOIN alloc_by_lot abl
               ON abl.product_id = lb.product_id AND abl.lot_id = lb.lot_id
        GROUP BY lb.product_id
    )
    SELECT lb.*,
           COALESCE(se.shipped_effective_lb, 0) AS shipped_effective_lb,
           COALESCE(oh.on_hand_lb, 0) AS on_hand_lb,
           COALESCE(abl.allocated_lb, 0) AS allocated_lb,
           COALESCE(abl.allocated_sku_lb, 0) AS allocated_sku_lb,
           COALESCE(abl.allocated_lot_lb, 0) AS allocated_lot_lb,
           COALESCE(abp.allocated_product_lb, 0) AS allocated_product_lb,
           COALESCE(abp.allocated_product_sku_lb, 0) AS allocated_product_sku_lb,
           COALESCE(i.inbound_open_lb, 0) AS inbound_open_lb,
           COALESCE(lp.lots, '[]'::jsonb) AS lots,
           COALESCE(llp.line_lot_allocations, '[]'::jsonb) AS line_lot_allocations
    FROM line_base lb
    LEFT JOIN shipped_eff se ON se.line_id = lb.line_id
    LEFT JOIN on_hand_sku oh ON oh.product_id = lb.product_id
    LEFT JOIN alloc_by_line abl
           ON abl.line_id = lb.line_id AND abl.product_id = lb.product_id
    LEFT JOIN alloc_by_product abp ON abp.product_id = lb.product_id
    LEFT JOIN inbound i ON i.product_id = lb.product_id
    LEFT JOIN lot_payload lp ON lp.product_id = lb.product_id
    LEFT JOIN line_lot_payload llp ON llp.line_id = lb.line_id
    ORDER BY lb.sales_order_id, lb.line_id
"""


def _factory_ready_required() -> bool:
    return (os.getenv("FACTORY_READY_REQUIRED", "true").strip().lower()
            not in {"0", "false", "no", "off"})


def _blocker(code: str, severity: str, detail: Optional[str] = None) -> dict:
    item = {"code": code, "severity": severity}
    if detail:
        item["detail"] = detail
    return item


def _lot_is_incomplete(lot: dict) -> bool:
    lot_code = str(lot.get("lot_code") or "")
    return (
        (lot_code.upper().startswith("STAGED-") or lot.get("entry_source") == "found_inventory")
        and lot.get("received_at") is None
    )


def _line_readiness(row: dict) -> dict:
    """Apply the PR-2 readiness formula to one physical line input row."""
    ordered = float(row["ordered_lb"] or 0)
    shipped_recorded = float(row["shipped_recorded_lb"] or 0)
    shipped_effective = float(row["shipped_effective_lb"] or 0)
    remaining = max(0.0, ordered - shipped_effective)
    on_hand = float(row["on_hand_lb"] or 0)
    allocated = float(row["allocated_lb"] or 0)
    allocated_sku = float(row["allocated_sku_lb"] or 0)
    allocated_lot = float(row["allocated_lot_lb"] or 0)
    allocated_product = float(row["allocated_product_lb"] or 0)
    allocated_others = max(0.0, allocated_product - allocated)
    available = on_hand - allocated_product
    coverable = max(0.0, on_hand - allocated_others)
    shortage = max(0.0, remaining - coverable)
    unallocated_need = max(0.0, remaining - allocated)
    inbound_open = float(row["inbound_open_lb"] or 0)
    diverged = abs(shipped_recorded - shipped_effective) > BALANCE_EPSILON

    line_lot_allocations = {
        int(item["lot_id"]): float(item["quantity_lb"] or 0)
        for item in (row.get("line_lot_allocations") or [])
    }
    incomplete_pins = [
        lot for lot in (row.get("lots") or [])
        if int(lot["lot_id"]) in line_lot_allocations
        and _lot_is_incomplete(lot)
    ]

    # Determine whether FIFO coverage of the portion not already pinned to a
    # lot would need an incomplete unpinned lot. Foreign SKU allocations claim
    # the same unpinned pool first, in deterministic FIFO order.
    need_from_unpinned = max(0.0, remaining - allocated_lot)
    foreign_sku = max(
        0.0,
        float(row["allocated_product_sku_lb"] or 0) - allocated_sku,
    )
    unstaged_lots = []
    for lot in (row.get("lots") or []):
        free = max(
            0.0,
            float(lot.get("on_hand_lb") or 0) - float(lot.get("allocated_lot_lb") or 0),
        )
        if foreign_sku > BALANCE_EPSILON:
            shadowed = min(foreign_sku, free)
            foreign_sku -= shadowed
            free -= shadowed
        if free <= BALANCE_EPSILON or need_from_unpinned <= BALANCE_EPSILON:
            continue
        take = min(need_from_unpinned, free)
        if (
            take > BALANCE_EPSILON
            and _lot_is_incomplete(lot)
            and int(lot["lot_id"]) not in line_lot_allocations
        ):
            unstaged_lots.append(str(lot.get("lot_code") or lot["lot_id"]))
        need_from_unpinned -= take

    blockers = []
    if shortage > BALANCE_EPSILON:
        blockers.append(_blocker("shortage", "block", f"Short {shortage:.4f} lb of posted cover"))
    if remaining > BALANCE_EPSILON and allocated <= BALANCE_EPSILON:
        blockers.append(_blocker("unallocated", "block", f"{remaining:.4f} lb remains with no allocation"))
    elif (allocated > BALANCE_EPSILON
          and allocated < remaining - BALANCE_EPSILON):
        blockers.append(_blocker("partial_allocation", "block", f"{unallocated_need:.4f} lb remains unallocated"))
    if unstaged_lots:
        blockers.append(_blocker("unstaged", "block", "Incomplete FIFO stock must be lot-pinned: " + ", ".join(unstaged_lots)))
    if incomplete_pins:
        blockers.append(_blocker("missing_lot_dates", "block", "Allocated lots need received_at: " + ", ".join(str(lot.get("lot_code") or lot["lot_id"]) for lot in incomplete_pins)))
    if diverged:
        blockers.append(_blocker("fulfillment_diverged", "block", "Recorded and effective shipped pounds differ"))
    if remaining > BALANCE_EPSILON and inbound_open > BALANCE_EPSILON:
        blockers.append(_blocker("inbound_cover", "warn", f"{inbound_open:.4f} lb is expected inbound and is not on-hand"))

    inventory_ready = (
        remaining <= BALANCE_EPSILON
        or (allocated + BALANCE_EPSILON >= remaining and shortage <= BALANCE_EPSILON)
    )
    return {
        "ordered_lb": ordered,
        "shipped_recorded_lb": shipped_recorded,
        "shipped_effective_lb": shipped_effective,
        "remaining_lb": remaining,
        "on_hand_lb": on_hand,
        "allocated_lb": allocated,
        "allocated_sku_lb": allocated_sku,
        "allocated_lot_lb": allocated_lot,
        "available_lb": available,
        "coverable_lb": coverable,
        "shortage_lb": shortage,
        "unallocated_need_lb": unallocated_need,
        "inbound_open_lb": inbound_open,
        "inventory_ready": inventory_ready,
        "fulfillment_diverged": diverged,
        "blockers": blockers,
    }


def _load_sales_order_readiness(cur, order_rows: list) -> tuple[dict, dict]:
    """Return ({order_id: readiness}, {line_id: readiness}) without writes."""
    if not order_rows:
        return {}, {}
    order_ids = [int(row["id"]) for row in order_rows]
    cur.execute(SALES_ORDER_READINESS_SQL, (order_ids,))
    inputs = cur.fetchall()

    order_meta = {int(row["id"]): row for row in order_rows}
    line_results = {}
    grouped = {order_id: [] for order_id in order_ids}
    for row in inputs:
        if row["is_service"]:
            readiness = {
                "inventory_ready": True,
                "fulfillment_diverged": False,
                "blockers": [],
            }
        else:
            readiness = _line_readiness(row)
            grouped[int(row["sales_order_id"])].append(readiness)
        line_results[int(row["line_id"])] = {
            "sales_order_id": int(row["sales_order_id"]),
            "product": row["product"],
            "sku": row["sku"],
            "is_service": bool(row["is_service"]),
            "ordered_lb": float(row["ordered_lb"] or 0),
            "shipped_recorded_lb": float(row["shipped_recorded_lb"] or 0),
            "readiness": readiness,
        }

    order_results = {}
    for order_id in order_ids:
        physical = grouped[order_id]
        ordered = sum(line["ordered_lb"] for line in physical)
        shipped_recorded = sum(line["shipped_recorded_lb"] for line in physical)
        shipped_effective = sum(line["shipped_effective_lb"] for line in physical)
        remaining = sum(line["remaining_lb"] for line in physical)
        allocated = sum(line["allocated_lb"] for line in physical)
        shortage = sum(line["shortage_lb"] for line in physical)
        inventory_ready = all(line["inventory_ready"] for line in physical)

        blockers_by_code = {}
        for line in physical:
            for blocker in line["blockers"]:
                blockers_by_code.setdefault(
                    blocker["code"],
                    {"code": blocker["code"], "severity": blocker["severity"]},
                )

        meta = order_meta[order_id]
        floor_ready = bool(meta.get("ready") or meta.get("floor_ready"))
        if not floor_ready:
            blockers_by_code["not_floor_ready"] = {
                "code": "not_floor_ready",
                "severity": "block" if _factory_ready_required() else "warn",
            }
        if meta.get("requested_ship_date") is None:
            blockers_by_code["no_ship_date"] = {"code": "no_ship_date", "severity": "warn"}
        if not any(line["remaining_lb"] > BALANCE_EPSILON for line in physical):
            blockers_by_code["service_only"] = {"code": "service_only", "severity": "info"}

        blockers = sorted(
            blockers_by_code.values(),
            key=lambda item: _READINESS_BLOCKER_ORDER[item["code"]],
        )
        fulfillment_diverged = any(line["fulfillment_diverged"] for line in physical)
        dispatch_ready = inventory_ready and not any(
            blocker["severity"] == "block" for blocker in blockers
        )
        order_results[order_id] = {
            "ordered_lb": ordered,
            "shipped_recorded_lb": shipped_recorded,
            "shipped_effective_lb": shipped_effective,
            "remaining_effective_lb": remaining,
            "allocated_lb": allocated,
            "shortage_lb": shortage,
            "inventory_ready": inventory_ready,
            "dispatch_ready": dispatch_ready,
            "floor_ready": floor_ready,
            "fulfillment_diverged": fulfillment_diverged,
            "blockers": blockers,
        }
    return order_results, line_results


@app.get("/sales/orders")
def list_sales_orders(
    status: Optional[str] = None,
    customer: Optional[str] = None,
    overdue_only: bool = False,
    limit: int = Query(default=50, ge=1, le=200),
    _: bool = Depends(verify_api_key)
):
    try:
        with get_transaction() as cur:
            query = """
                SELECT so.id, so.order_number, c.name AS customer,
                       so.order_date, so.requested_ship_date, so.status,
                       COALESCE(sof.ready, false) AS ready,
                       sof.ready_at, sof.ready_by, sof.note AS ready_note,
                       COUNT(sol.id) AS line_count,
                       COALESCE(SUM(sol.quantity_lb) FILTER (WHERE NOT COALESCE(p.is_service, false)), 0) AS total_lb,
                       COALESCE(SUM(sol.quantity_shipped_lb) FILTER (WHERE NOT COALESCE(p.is_service, false)), 0) AS shipped_lb,
                       COALESCE(SUM(sol.quantity_lb / NULLIF(p.case_size_lb, 0)) FILTER (WHERE NOT COALESCE(p.is_service, false) AND p.case_size_lb IS NOT NULL AND p.case_size_lb > 0), 0) AS total_units,
                       COALESCE(SUM(sol.quantity_shipped_lb / NULLIF(p.case_size_lb, 0)) FILTER (WHERE NOT COALESCE(p.is_service, false) AND p.case_size_lb IS NOT NULL AND p.case_size_lb > 0), 0) AS shipped_units,
                       COALESCE((
                           SELECT jsonb_agg(jsonb_build_object(
                               'line_id', pallet_sol.id,
                               'product', pallet_p.name,
                               'sku', pallet_p.odoo_code,
                               'uom', COALESCE(pallet_p.uom, 'lb'),
                               'case_size_lb', pallet_p.case_size_lb,
                               'unit_count', CASE WHEN pallet_p.case_size_lb > 0 THEN ROUND(pallet_sol.quantity_lb / pallet_p.case_size_lb) END,
                               'shipped_units', CASE WHEN pallet_p.case_size_lb > 0 THEN ROUND(pallet_sol.quantity_shipped_lb / pallet_p.case_size_lb) END,
                               'remaining_units', CASE WHEN pallet_p.case_size_lb > 0 THEN ROUND((pallet_sol.quantity_lb - pallet_sol.quantity_shipped_lb) / pallet_p.case_size_lb) END,
                               'is_non_weight', COALESCE(pallet_p.is_service, false)
                           ) ORDER BY pallet_sol.id)
                           FROM sales_order_lines pallet_sol
                           JOIN products pallet_p ON pallet_p.id = pallet_sol.product_id
                           WHERE pallet_sol.sales_order_id = so.id
                       ), '[]'::jsonb) AS pallet_lines
                FROM sales_orders so
                JOIN customers c ON c.id = so.customer_id
                LEFT JOIN sales_order_lines sol ON sol.sales_order_id = so.id
                LEFT JOIN products p ON p.id = sol.product_id
                LEFT JOIN sales_order_flags sof ON sof.so_number = so.order_number
                LEFT JOIN customer_aliases ca ON ca.customer_id = c.id
                WHERE 1=1
            """
            params = []
            if status:
                if status == 'open':
                    query += " AND so.status NOT IN ('shipped', 'invoiced', 'cancelled')"
                else:
                    query += " AND so.status = %s"
                    params.append(status)
            if customer:
                query += " AND (LOWER(c.name) LIKE LOWER(%s) OR LOWER(ca.alias) LIKE LOWER(%s))"
                params.append(f"%{customer}%")
                params.append(f"%{customer}%")
            if overdue_only:
                query += " AND so.requested_ship_date < CURRENT_DATE AND so.status NOT IN ('shipped', 'invoiced', 'cancelled')"

            query += " GROUP BY so.id, c.name, sof.ready, sof.ready_at, sof.ready_by, sof.note ORDER BY so.requested_ship_date ASC NULLS LAST LIMIT %s"
            params.append(limit)
            cur.execute(query, params)
            rows = cur.fetchall()
            readiness_by_order, _ = _load_sales_order_readiness(cur, rows)

            orders = []
            for r in rows:
                total = float(r['total_lb'] or 0)
                shipped = float(r['shipped_lb'] or 0)
                ship_date = r['requested_ship_date']
                is_open = r['status'] not in ('shipped', 'invoiced', 'cancelled')

                # Proactive warnings for open orders
                order_warnings = []
                if is_open:
                    if ship_date is None:
                        order_warnings.append("⚠️ No ship date set")
                    if 'test' in r['customer'].lower():
                        order_warnings.append("⚠️ Possible test order")
                    if total == 0 and r['line_count'] == 0:
                        order_warnings.append("⚠️ Empty order — no line items")

                total_units = round(float(r['total_units'] or 0))
                shipped_units = round(float(r['shipped_units'] or 0))
                order = {
                    "order_id": r['id'],
                    "order_number": r['order_number'],
                    "customer": r['customer'],
                    "order_date": str(r['order_date']),
                    "requested_ship_date": str(ship_date) if ship_date else None,
                    "status": r['status'],
                    "line_count": r['line_count'],
                    "total_lb": total,
                    "shipped_lb": shipped,
                    "remaining_lb": total - shipped,
                    "total_units": total_units,
                    "shipped_units": shipped_units,
                    "remaining_units": total_units - shipped_units,
                    "pallet_lines": r['pallet_lines'],
                    "ready": bool(r['ready']),
                    "ready_at": r['ready_at'].isoformat() if r['ready_at'] else None,
                    "ready_by": r['ready_by'],
                    "note": r['ready_note'],
                    "overdue": ship_date is not None and ship_date < date.today() and is_open
                }
                readiness = readiness_by_order[r['id']]
                order.update({
                    "inventory_ready": readiness["inventory_ready"],
                    "dispatch_ready": readiness["dispatch_ready"],
                    "fulfillment_diverged": readiness["fulfillment_diverged"],
                    "shortage_lb": readiness["shortage_lb"],
                    "allocated_lb": readiness["allocated_lb"],
                    "remaining_effective_lb": readiness["remaining_effective_lb"],
                    "blockers": readiness["blockers"],
                })
                if order_warnings:
                    order["warnings"] = order_warnings
                orders.append(order)
            return {"orders": orders, "count": len(orders)}
    except Exception as e:
        logger.error(f"List sales orders failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


def _sales_order_flag_row_to_dict(row):
    d = dict(row)
    if d.get("ready_at"):
        d["ready_at"] = d["ready_at"].isoformat()
    if d.get("updated_at"):
        d["updated_at"] = d["updated_at"].isoformat()
    return d


@app.post("/sales-orders/{so_number}/ready")
def set_sales_order_ready_flag(so_number: str, req: SalesOrderReadyFlagRequest, _: bool = Depends(verify_api_key)):
    """Upsert the dashboard-only Factory Ready annotation for a sales order."""
    try:
        with get_transaction() as cur:
            cur.execute(
                """
                SELECT order_number, status
                FROM sales_orders
                WHERE order_number = %s
                """,
                (so_number,)
            )
            order = cur.fetchone()
            if not order:
                return JSONResponse(status_code=404, content={"error": "Sales order not found"})
            if order["status"] in ("shipped", "invoiced", "cancelled"):
                return JSONResponse(status_code=400, content={"error": "Factory Ready can only be set on open sales orders"})

            ready_by = (req.by or "floor").strip() or "floor"
            note = req.note.strip() if isinstance(req.note, str) else req.note

            cur.execute(
                """
                INSERT INTO sales_order_flags (so_number, ready, ready_at, ready_by, note, updated_at)
                VALUES (%s, %s, CASE WHEN %s THEN NOW() ELSE NULL END, %s, %s, NOW())
                ON CONFLICT (so_number) DO UPDATE SET
                    ready = EXCLUDED.ready,
                    ready_at = CASE
                        WHEN EXCLUDED.ready THEN COALESCE(sales_order_flags.ready_at, EXCLUDED.ready_at)
                        ELSE NULL
                    END,
                    ready_by = EXCLUDED.ready_by,
                    note = EXCLUDED.note,
                    updated_at = NOW()
                RETURNING so_number, ready, ready_at, ready_by, note, updated_at
                """,
                (order["order_number"], req.ready, req.ready, ready_by, note)
            )
            return _sales_order_flag_row_to_dict(cur.fetchone())
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Sales order ready flag update failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/sales/orders/fulfillment-check")
def fulfillment_check(
    customer_name: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    order_id: Optional[int] = Query(default=None),
    _: bool = Depends(verify_api_key)
):
    """Read-only dispatch queue for open and fulfillment-diverged orders."""
    OPEN_STATUSES = ('confirmed', 'in_production', 'ready', 'partial_ship')

    if status and status not in OPEN_STATUSES:
        raise HTTPException(400,
            f"Invalid status filter '{status}'. Must be one of: {list(OPEN_STATUSES)}"
        )

    try:
        with get_transaction() as cur:
            query = """
                WITH matching_orders AS (
                    SELECT so.id, so.order_number, so.status, so.requested_ship_date,
                           c.name AS customer, COALESCE(sof.ready, false) AS ready
                    FROM sales_orders so
                    JOIN customers c ON c.id = so.customer_id
                    LEFT JOIN sales_order_flags sof ON sof.so_number = so.order_number
                    WHERE 1=1
            """
            params: list = []

            if order_id is not None:
                query += " AND so.id = %s"
                params.append(order_id)

            if customer_name:
                query += """ AND (
                    LOWER(c.name) LIKE LOWER(%s)
                    OR EXISTS (
                        SELECT 1 FROM customer_aliases ca
                        WHERE ca.customer_id = c.id
                          AND LOWER(ca.alias) LIKE LOWER(%s)
                    )
                )"""
                params.append(f"%{customer_name}%")
                params.append(f"%{customer_name}%")

            if status:
                query += " AND so.status = %s"
                params.append(status)

            query += """
                ),
                physical_lines AS (
                    SELECT sol.id AS line_id, sol.sales_order_id, sol.product_id,
                           sol.quantity_shipped_lb AS shipped_recorded_lb
                    FROM matching_orders mo
                    JOIN sales_order_lines sol ON sol.sales_order_id = mo.id
                    JOIN products p ON p.id = sol.product_id
                    WHERE sol.line_status <> 'cancelled'
                      AND NOT COALESCE(p.is_service, false)
                ),
                shipped_eff AS (
                    SELECT pl.line_id,
                           SUM(ABS(tl.quantity_lb)) AS shipped_effective_lb
                    FROM physical_lines pl
                    JOIN sales_order_shipments sos ON sos.sales_order_line_id = pl.line_id
                    JOIN ledger_current_transactions ct ON ct.id = sos.transaction_id
                    JOIN ledger_current_transaction_lines tl
                      ON tl.transaction_id = sos.transaction_id
                     AND tl.product_id = pl.product_id
                    WHERE ct.effective_status = 'posted'
                      AND ct.type = 'ship'
                    GROUP BY pl.line_id
                ),
                divergent_orders AS (
                    SELECT pl.sales_order_id
                    FROM physical_lines pl
                    LEFT JOIN shipped_eff se ON se.line_id = pl.line_id
                    GROUP BY pl.sales_order_id
                    HAVING BOOL_OR(
                        ABS(pl.shipped_recorded_lb - COALESCE(se.shipped_effective_lb, 0)) > %s
                    )
                )
                SELECT mo.*
                FROM matching_orders mo
                LEFT JOIN divergent_orders d ON d.sales_order_id = mo.id
                WHERE mo.status = ANY(%s) OR d.sales_order_id IS NOT NULL
                ORDER BY mo.requested_ship_date ASC NULLS LAST, mo.id ASC
            """
            params.extend([BALANCE_EPSILON, list(OPEN_STATUSES)])
            cur.execute(query, tuple(params))
            orders = cur.fetchall()
            readiness_by_order, readiness_by_line = _load_sales_order_readiness(cur, orders)

            results = []
            summary = {"total_orders_checked": 0, "fulfillable": 0, "partially_fulfillable": 0, "blocked": 0}
            lines_by_order = {int(order["id"]): [] for order in orders}
            for line_id, payload in readiness_by_line.items():
                lines_by_order[payload["sales_order_id"]].append((line_id, payload))

            for order in orders:
                order_lines = []
                line_payloads = lines_by_order[order['id']]
                for line_id, payload in line_payloads:
                    ready = payload["readiness"]
                    remaining = ready.get(
                        "remaining_lb",
                        max(0.0, payload["ordered_lb"] - payload["shipped_recorded_lb"]),
                    )
                    order_lines.append({
                        "line_id": line_id,
                        "product": payload["product"],
                        "sku": payload["sku"],
                        "ordered_lb": payload["ordered_lb"],
                        "shipped_lb": payload["shipped_recorded_lb"],
                        "remaining_lb": remaining,
                        "on_hand_lb": ready.get("on_hand_lb", 0.0),
                        "can_fulfill": ready["inventory_ready"],
                        "shortfall_lb": ready.get("shortage_lb", 0.0),
                        "readiness": ready,
                    })

                readiness = readiness_by_order[order['id']]
                physical_states = [
                    payload["readiness"] for _, payload in line_payloads
                    if not payload["is_service"]
                ]
                ready_lines = sum(line["inventory_ready"] for line in physical_states)

                # Classify for summary
                summary["total_orders_checked"] += 1
                if readiness["inventory_ready"]:
                    summary["fulfillable"] += 1
                elif ready_lines > 0:
                    summary["partially_fulfillable"] += 1
                else:
                    summary["blocked"] += 1

                result = {
                    "order_id": order['id'],
                    "order_number": order['order_number'],
                    "customer": order['customer'],
                    "status": order['status'],
                    "requested_ship_date": str(order['requested_ship_date']) if order['requested_ship_date'] else None,
                    "fulfillable": readiness["inventory_ready"],
                    "lines": order_lines,
                    "total_remaining_lb": readiness["remaining_effective_lb"],
                    "total_on_hand_lb": sum(line.get("on_hand_lb", 0.0) for line in physical_states),
                    "total_shortfall_lb": readiness["shortage_lb"],
                }
                result.update(readiness)
                results.append(result)

            # Sort: fulfillable first within each date group
            results.sort(key=lambda o: (
                o['requested_ship_date'] or '9999-12-31',
                0 if o['dispatch_ready'] else 1
            ))

            return {
                "summary": summary,
                "orders": results
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Fulfillment check failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/sales/orders/{order_id}")
def get_sales_order(order_id: int = Depends(resolve_order_id), _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute(
                """SELECT so.id, so.order_number, c.name AS customer, so.order_date,
                          so.requested_ship_date, so.status, so.notes, so.notes_es, so.created_at,
                          COALESCE(sof.ready, false) AS ready
                   FROM sales_orders so
                   JOIN customers c ON c.id = so.customer_id
                   LEFT JOIN sales_order_flags sof ON sof.so_number = so.order_number
                   WHERE so.id = %s""",
                (order_id,)
            )
            row = cur.fetchone()
            if not row:
                raise HTTPException(
                    status_code=404,
                    detail={
                        "error_code": "ORDER_NOT_FOUND",
                        "message": f"Order #{order_id} not found",
                        "input": str(order_id),
                        "suggestions": [],
                    }
                )

            readiness_by_order, readiness_by_line = _load_sales_order_readiness(cur, [row])
            date_str, time_str = format_timestamp(row['created_at'])
            order = {
                "order_id": row['id'],
                "order_number": row['order_number'],
                "customer": row['customer'],
                "order_date": str(row['order_date']),
                "requested_ship_date": str(row['requested_ship_date']) if row['requested_ship_date'] else None,
                "status": row['status'],
                "notes": row['notes'],
                "created_date": date_str,
                "created_time": time_str
            }
            if row.get('notes_es'):
                order["notes_es"] = row['notes_es']

            cur.execute(
                """SELECT sol.id, p.name, p.odoo_code, p.uom, sol.quantity_lb, sol.quantity_shipped_lb,
                          sol.unit_price, sol.line_status, sol.notes, sol.notes_es,
                          p.case_size_lb, COALESCE(p.is_service, false) AS is_service,
                          COALESCE(p.no_production, false) AS no_production
                   FROM sales_order_lines sol
                   JOIN products p ON p.id = sol.product_id
                   WHERE sol.sales_order_id = %s
                   ORDER BY sol.id""",
                (order_id,)
            )
            lines = []
            total_ordered = 0
            total_shipped = 0
            total_value = 0
            total_ordered_units = 0
            total_shipped_units = 0
            for r in cur.fetchall():
                qty = float(r['quantity_lb'])
                shipped = float(r['quantity_shipped_lb'])
                price = float(r['unit_price']) if r['unit_price'] else None
                case_size = float(r['case_size_lb']) if r['case_size_lb'] else None
                cases = round(qty / case_size) if case_size else None

                # Exclude service/charge lines (pallets, freight, etc.) from weight totals
                is_service = r['is_service']
                if not is_service:
                    total_ordered += qty
                    total_shipped += shipped
                    if case_size:
                        total_ordered_units += round(qty / case_size)
                        total_shipped_units += round(shipped / case_size)

                # line_value = cases * price_per_case (not lb * price)
                line_value = None
                if price and cases:
                    line_value = round(cases * price, 2)
                    total_value += line_value
                elif price:
                    # Fallback for products without case_size_lb: treat unit_price as price/lb
                    line_value = round(qty * price, 2)
                    total_value += line_value

                # Detect non-weight items (pallets, freight, surcharges, etc.)
                # Primary: DB flag; fallback: keyword matching
                product_name_lower = r['name'].lower()
                non_weight_keywords = ('pallet', 'freight', 'delivery', 'surcharge', 'charge', 'fee')
                is_non_weight = is_service or any(kw in product_name_lower for kw in non_weight_keywords)

                if is_non_weight:
                    price_basis = "per_unit"
                elif case_size:
                    price_basis = "per_case"
                else:
                    price_basis = "per_lb"

                line_units = round(qty / case_size) if case_size else None
                shipped_units_line = round(shipped / case_size) if case_size else None
                remaining_units_line = (line_units - shipped_units_line) if line_units is not None and shipped_units_line is not None else None
                line_data = {
                    "line_id": r['id'],
                    "product": r['name'],
                    "sku": r['odoo_code'],
                    "uom": r['uom'] or "lb",
                    "no_production": bool(r['no_production']),
                    "quantity_lb": qty,
                    "cases": cases,
                    "case_size_lb": case_size,
                    "unit_count": line_units,
                    "quantity_shipped_lb": shipped,
                    "shipped_units": shipped_units_line,
                    "remaining_lb": qty - shipped,
                    "remaining_units": remaining_units_line,
                    "case_price": price,
                    "price_basis": price_basis,
                    "line_value": line_value,
                    "line_status": r['line_status'],
                    "notes": r['notes']
                }
                if is_non_weight:
                    line_data["is_non_weight"] = True
                    line_data["unit_quantity"] = int(qty) if qty == int(qty) else qty
                if r.get('notes_es'):
                    line_data["notes_es"] = r['notes_es']
                line_readiness = readiness_by_line.get(r['id'])
                if line_readiness is not None:
                    line_data["readiness"] = line_readiness["readiness"]
                lines.append(line_data)

            cur.execute(
                """SELECT sos.id, sol.id AS line_id, p.name AS product,
                          sos.quantity_lb, sos.shipped_at, t.id AS transaction_id
                   FROM sales_order_shipments sos
                   JOIN sales_order_lines sol ON sol.id = sos.sales_order_line_id
                   JOIN products p ON p.id = sol.product_id
                   JOIN ledger_current_transactions t ON t.id = sos.transaction_id
                   WHERE sol.sales_order_id = %s
                     AND t.effective_status = 'posted'
                   ORDER BY sos.shipped_at DESC""",
                (order_id,)
            )
            shipments = []
            for r in cur.fetchall():
                s_date, s_time = format_timestamp(r['shipped_at'])
                shipments.append({
                    "shipment_id": r['id'],
                    "line_id": r['line_id'],
                    "product": r['product'],
                    "quantity_lb": float(r['quantity_lb']),
                    "shipped_date": s_date,
                    "shipped_time": s_time,
                    "transaction_id": r['transaction_id']
                })

            order["lines"] = lines
            order["shipments"] = shipments
            order["totals"] = {
                "total_ordered_lb": total_ordered,
                "total_shipped_lb": total_shipped,
                "remaining_lb": total_ordered - total_shipped,
                "total_ordered_units": total_ordered_units,
                "total_shipped_units": total_shipped_units,
                "total_remaining_units": total_ordered_units - total_shipped_units,
                "total_value": round(total_value, 2) if total_value > 0 else None
            }
            order.update(readiness_by_order[order_id])
            return order
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get sales order failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


def _order_readiness_after_write(cur, order_id: int) -> tuple[dict, dict]:
    cur.execute(
        """SELECT so.id, so.order_number, so.requested_ship_date,
                  COALESCE(sof.ready, false) AS ready
             FROM sales_orders so
             LEFT JOIN sales_order_flags sof ON sof.so_number = so.order_number
            WHERE so.id = %s""",
        (order_id,),
    )
    order = cur.fetchone()
    if not order:
        _allocation_error(
            "ORDER_NOT_FOUND",
            f"Order #{order_id} not found",
            status_code=404,
            order_id=order_id,
        )
    order_readiness, line_readiness = _load_sales_order_readiness(cur, [order])
    return order_readiness[order_id], line_readiness


@app.post("/sales/orders/{order_id}/allocations")
def create_sales_order_allocation(
    req: SalesOrderAllocationCreate,
    request: Request,
    order_id: int = Depends(resolve_order_id),
    _: bool = Depends(verify_api_key),
):
    try:
        with get_transaction() as cur:
            line = _load_allocatable_line(cur, order_id, req.line_id)
            product_id = int(line["product_id"])
            created_by = caller_source_tag(request)
            _lock_allocation_product(cur, product_id)
            expired = _expire_auto_fifo_allocations(cur, product_id, created_by)

            created = []
            if req.mode == "manual":
                if req.quantity_lb is None:
                    _allocation_error(
                        "ALLOCATION_QUANTITY_REQUIRED",
                        "quantity_lb is required for manual allocation",
                        status_code=422,
                    )
                if req.expires_at is not None:
                    _allocation_error(
                        "MANUAL_ALLOCATION_CANNOT_EXPIRE",
                        "expires_at is only valid for auto_fifo allocations",
                        status_code=422,
                    )
                quantity = float(req.quantity_lb)
                source = req.source or "manual"
                _validate_allocation_addition(
                    cur, line, quantity, lot_id=req.lot_id
                )
                if req.lot_id is not None:
                    cur.execute(
                        "SELECT lot_code, entry_source FROM lots WHERE id = %s",
                        (req.lot_id,),
                    )
                    pinned_lot = cur.fetchone()
                    if (
                        str(pinned_lot["lot_code"] or "").upper().startswith("STAGED-")
                        or pinned_lot["entry_source"] == "found_inventory"
                    ):
                        source = "staged_lot"
                created.append(_upsert_live_allocation(
                    cur,
                    line=line,
                    lot_id=req.lot_id,
                    quantity_lb=quantity,
                    source=source,
                    expires_at=None,
                    note=req.note,
                    created_by=created_by,
                ))
            else:
                if req.lot_id is not None or req.source is not None:
                    _allocation_error(
                        "INVALID_AUTO_FIFO_REQUEST",
                        "auto_fifo chooses its own lots; omit lot_id and source",
                        status_code=422,
                    )
                expires_at = req.expires_at or (datetime.now(timezone.utc) + timedelta(hours=48))
                if expires_at.tzinfo is None or expires_at.utcoffset() is None:
                    _allocation_error(
                        "INVALID_ALLOCATION_EXPIRY",
                        "expires_at must include a timezone offset or Z",
                        status_code=422,
                    )
                if expires_at <= datetime.now(timezone.utc):
                    _allocation_error(
                        "INVALID_ALLOCATION_EXPIRY",
                        "expires_at must be in the future",
                        status_code=422,
                    )
                _, line_allocated = _allocation_totals(cur, product_id, req.line_id)
                remaining_effective = max(
                    0.0,
                    float(line["quantity_lb"])
                    - _line_shipped_effective(cur, req.line_id, product_id),
                )
                unallocated_need = max(0.0, remaining_effective - line_allocated)
                quantity = unallocated_need if req.quantity_lb is None else float(req.quantity_lb)
                _validate_allocation_addition(cur, line, quantity)
                to_allocate = quantity
                for lot in available_lots_for_product(cur, product_id, req.line_id):
                    if to_allocate <= BALANCE_EPSILON:
                        break
                    take = min(to_allocate, float(lot["takeable"] or 0))
                    if take <= BALANCE_EPSILON:
                        continue
                    created.append(_upsert_live_allocation(
                        cur,
                        line=line,
                        lot_id=int(lot["lot_id"]),
                        quantity_lb=take,
                        source="auto_fifo",
                        expires_at=expires_at,
                        note=req.note,
                        created_by=created_by,
                    ))
                    to_allocate -= take
                if to_allocate > BALANCE_EPSILON:
                    _allocation_error(
                        "OVER_ALLOCATION",
                        f"Only {quantity - to_allocate:g} of {quantity:g} lb could be allocated FIFO",
                        requested_lb=quantity,
                        allocated_lb=quantity - to_allocate,
                        coverable_lb=quantity - to_allocate,
                    )

            order_readiness, lines = _order_readiness_after_write(cur, order_id)
            logger.info(
                "Allocation create/upsert: order=%s line=%s product=%s mode=%s qty=%s rows=%s",
                line["order_number"], req.line_id, product_id, req.mode,
                sum(float(row["quantity_lb"]) for row in created),
                [int(row["id"]) for row in created],
            )
            return {
                "order_id": order_id,
                "order_number": line["order_number"],
                "line_id": req.line_id,
                "mode": req.mode,
                "allocations": created,
                "expired_allocations_released": expired,
                "line_readiness": lines[req.line_id]["readiness"],
                "order_readiness": order_readiness,
            }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e):
            raise
        logger.error(f"Create allocation failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/sales/orders/{order_id}/allocations")
def get_sales_order_allocations(
    order_id: int = Depends(resolve_order_id),
    _: bool = Depends(verify_api_key),
):
    try:
        with get_transaction() as cur:
            order_readiness, lines = _order_readiness_after_write(cur, order_id)
            cur.execute(
                """SELECT soa.*, p.name AS product_name, p.odoo_code AS sku,
                          l.lot_code,
                          CASE
                            WHEN soa.status = 'active'
                             AND soa.expires_at IS NOT NULL
                             AND soa.expires_at <= clock_timestamp()
                            THEN 'released'
                            ELSE soa.status
                          END AS effective_status
                     FROM sales_order_allocations soa
                     JOIN products p ON p.id = soa.product_id
                     LEFT JOIN lots l ON l.id = soa.lot_id
                    WHERE soa.sales_order_id = %s
                    ORDER BY soa.created_at, soa.id""",
                (order_id,),
            )
            allocations = [dict(row) for row in cur.fetchall()]
            return {
                "order_id": order_id,
                "allocations": allocations,
                "lines": [
                    {"line_id": line_id, **payload}
                    for line_id, payload in sorted(lines.items())
                ],
                "order_readiness": order_readiness,
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get allocations failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/sales/orders/{order_id}/allocations/{allocation_id}/release")
def release_sales_order_allocation(
    allocation_id: int,
    request: Request,
    order_id: int = Depends(resolve_order_id),
    _: bool = Depends(verify_api_key),
):
    try:
        with get_transaction() as cur:
            cur.execute(
                """SELECT soa.*, so.order_number
                     FROM sales_order_allocations soa
                     JOIN sales_orders so ON so.id = soa.sales_order_id
                    WHERE soa.id = %s AND soa.sales_order_id = %s""",
                (allocation_id, order_id),
            )
            allocation = cur.fetchone()
            if not allocation:
                _allocation_error(
                    "ALLOCATION_NOT_FOUND",
                    f"Allocation #{allocation_id} not found on order #{order_id}",
                    status_code=404,
                    allocation_id=allocation_id,
                    order_id=order_id,
                )
            released_by = caller_source_tag(request)
            _lock_allocation_product(cur, int(allocation["product_id"]))
            expired = _expire_auto_fifo_allocations(
                cur, int(allocation["product_id"]), released_by
            )
            released = _release_active_allocations(
                cur,
                allocation_id=allocation_id,
                order_id=order_id,
                reason="manual_release",
                released_by=released_by,
            )
            if not released:
                cur.execute(
                    "SELECT status, release_reason, released_at FROM sales_order_allocations WHERE id=%s",
                    (allocation_id,),
                )
                state = cur.fetchone()
                if state and state["status"] == "released":
                    released = [{"id": allocation_id, **dict(state)}]
                else:
                    _allocation_error(
                        "ALLOCATION_NOT_ACTIVE",
                        f"Allocation #{allocation_id} is not active",
                        allocation_id=allocation_id,
                        status=state["status"] if state else None,
                    )
            order_readiness, lines = _order_readiness_after_write(cur, order_id)
            logger.info(
                "Allocation release: order=%s allocation=%s reason=manual_release",
                allocation["order_number"], allocation_id,
            )
            return {
                "order_id": order_id,
                "allocation_id": allocation_id,
                "released": released,
                "expired_allocations_released": expired,
                "line_readiness": lines[int(allocation["sales_order_line_id"])]["readiness"],
                "order_readiness": order_readiness,
            }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e):
            raise
        logger.error(f"Release allocation failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/lots/{lot_id}/received-at")
def update_lot_received_at(
    lot_id: int,
    payload: Dict[str, Any],
    _: bool = Depends(verify_api_key),
):
    if "received_at" not in payload or payload.get("received_at") in (None, ""):
        _allocation_error(
            "RECEIVED_AT_REQUIRED",
            "received_at is required and cannot be null or empty",
            status_code=422,
            lot_id=lot_id,
        )
    value = payload.get("received_at")
    if not isinstance(value, str):
        _allocation_error(
            "INVALID_RECEIVED_AT",
            "received_at must be an ISO-8601 timestamp with a timezone offset or Z",
            status_code=422,
            lot_id=lot_id,
        )
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except (TypeError, ValueError):
        _allocation_error(
            "INVALID_RECEIVED_AT",
            "received_at must be an ISO-8601 timestamp with a timezone offset or Z",
            status_code=422,
            lot_id=lot_id,
        )
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        _allocation_error(
            "INVALID_RECEIVED_AT",
            "received_at must include a timezone offset or Z",
            status_code=422,
            lot_id=lot_id,
        )
    if parsed > datetime.now(timezone.utc):
        _allocation_error(
            "RECEIVED_AT_IN_FUTURE",
            "received_at cannot be in the future",
            status_code=422,
            lot_id=lot_id,
        )
    try:
        with get_transaction() as cur:
            cur.execute(
                """UPDATE lots SET received_at = %s
                     WHERE id = %s
                 RETURNING id, lot_code, product_id, received_at, entry_source""",
                (parsed, lot_id),
            )
            lot = cur.fetchone()
            if not lot:
                _allocation_error(
                    "LOT_NOT_FOUND",
                    f"Lot #{lot_id} not found",
                    status_code=404,
                    lot_id=lot_id,
                )
            result = dict(lot)
            result["lot_id"] = result.pop("id")
            result["lot_is_incomplete"] = _lot_is_incomplete(result)
            return result
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e):
            raise
        logger.error(f"Update lot received_at failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/sales/orders/{order_id}/status")
def update_order_status(order_id: int = Depends(resolve_order_id), req: OrderStatusUpdate = ..., _: bool = Depends(verify_api_key)):
    all_statuses = list(VALID_TRANSITIONS.keys())
    if req.status not in all_statuses:
        raise HTTPException(400, f"Invalid status. Must be one of: {all_statuses}")

    # Block manual setting of shipped/partial_ship — those are auto-only via shipOrderCommit
    if req.status in ('shipped', 'partial_ship'):
        raise HTTPException(400,
            f"'{req.status}' status is set automatically when an order is shipped. "
            f"Use the ship endpoint instead."
        )

    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # Get current status first
                cur.execute(
                    "SELECT order_number, status FROM sales_orders WHERE id = %s",
                    (order_id,)
                )
                row = cur.fetchone()
                if not row:
                    raise HTTPException(404, f"Order #{order_id} not found")

                current = row['status']
                allowed = MANUAL_TRANSITIONS.get(current, [])

                if req.status not in allowed:
                    if not allowed:
                        raise HTTPException(400,
                            f"Order {row['order_number']} is '{current}' — this is a terminal status. "
                            f"No further status changes are allowed."
                        )
                    raise HTTPException(400,
                        f"Invalid status transition: '{current}' → '{req.status}'. "
                        f"Allowed transitions from '{current}': {allowed}."
                    )

                cur.execute(
                    "UPDATE sales_orders SET status = %s WHERE id = %s RETURNING order_number, status",
                    (req.status, order_id)
                )
                updated = cur.fetchone()
                released_allocations = []
                if req.status == 'cancelled':
                    cur.execute(
                        """SELECT DISTINCT product_id
                             FROM sales_order_allocations
                            WHERE sales_order_id = %s AND status = 'active'
                            ORDER BY product_id""",
                        (order_id,),
                    )
                    product_ids = [int(item['product_id']) for item in cur.fetchall()]
                    for product_id in product_ids:
                        _lock_allocation_product(cur, product_id)
                        _expire_auto_fifo_allocations(cur, product_id, _operator_id(_))
                    released_allocations = _release_active_allocations(
                        cur,
                        order_id=order_id,
                        reason='order_cancelled',
                        released_by=_operator_id(_),
                    )
                logger.info(f"Order {updated['order_number']} status: {current} → {req.status}")
                return {
                    "order_id": order_id,
                    "order_number": updated['order_number'],
                    "previous_status": current,
                    "status": updated['status'],
                    "allocations_released": released_allocations,
                    "message": f"Order {updated['order_number']}: {current} → {req.status}"
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Update order status failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/sales/orders/{order_id}")
def update_order_header(order_id: int = Depends(resolve_order_id), req: OrderHeaderUpdate = ..., _: bool = Depends(verify_api_key)):
    """Update order header fields (ship date, notes, customer). Only allowed when status is 'new' or 'confirmed'."""
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    "SELECT id, order_number, status, customer_id, requested_ship_date, notes, notes_es FROM sales_orders WHERE id = %s",
                    (order_id,)
                )
                order = cur.fetchone()
                if not order:
                    raise HTTPException(
                        status_code=404,
                        detail={
                            "error_code": "ORDER_NOT_FOUND",
                            "message": f"Order #{order_id} not found",
                            "input": order_id,
                            "suggestions": [],
                        }
                    )

                if order['status'] not in ('new', 'confirmed'):
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "error_code": "ORDER_HEADER_LOCKED",
                            "message": f"Order {order['order_number']} is '{order['status']}' — header edits only allowed when status is 'new' or 'confirmed'.",
                            "input": str(order_id),
                            "suggestions": [],
                        }
                    )

                updates = {}
                if req.requested_ship_date is not None:
                    updates['requested_ship_date'] = req.requested_ship_date if req.requested_ship_date else None
                if req.notes is not None:
                    updates['notes'] = req.notes if req.notes else None
                if req.notes_es is not None:
                    updates['notes_es'] = req.notes_es if req.notes_es else None
                if req.customer_id is not None:
                    # Verify customer exists
                    cur.execute("SELECT id, name FROM customers WHERE id = %s", (req.customer_id,))
                    cust = cur.fetchone()
                    if not cust:
                        raise HTTPException(
                            status_code=404,
                            detail={
                                "error_code": "CUSTOMER_NOT_FOUND",
                                "message": f"Customer ID {req.customer_id} not found",
                                "input": str(req.customer_id),
                                "suggestions": [],
                            }
                        )
                    updates['customer_id'] = req.customer_id

                if not updates:
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "error_code": "NO_FIELDS_TO_UPDATE",
                            "message": "No fields to update",
                            "input": "",
                            "suggestions": [],
                        }
                    )

                set_clause = ", ".join(f"{k} = %s" for k in updates)
                values = list(updates.values()) + [order_id]
                cur.execute(
                    f"UPDATE sales_orders SET {set_clause} WHERE id = %s RETURNING id, order_number, status, customer_id, requested_ship_date, notes, notes_es",
                    values
                )
                updated = cur.fetchone()

                # Get customer name for response
                cur.execute("SELECT name FROM customers WHERE id = %s", (updated['customer_id'],))
                customer_name = cur.fetchone()['name']

                changes = list(updates.keys())
                logger.info(f"Order {updated['order_number']} header updated: {changes}")
                return {
                    "order_id": updated['id'],
                    "order_number": updated['order_number'],
                    "status": updated['status'],
                    "customer_id": updated['customer_id'],
                    "customer_name": customer_name,
                    "requested_ship_date": str(updated['requested_ship_date']) if updated['requested_ship_date'] else None,
                    "notes": updated['notes'],
                    "notes_es": updated['notes_es'],
                    "fields_updated": changes,
                    "message": f"Order {updated['order_number']} updated: {', '.join(changes)}"
                }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Update order header failed: {e}")
        if _is_readonly_error(e):
            diagnostics = _capture_readonly_diagnostics()
            logger.error(
                "READONLY_TRIPWIRE: "
                + json.dumps({"error": str(e), "diagnostics": diagnostics}, default=str)
            )
            return JSONResponse(
                status_code=500,
                content={"error": str(e), "diagnostics": diagnostics},
            )
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/sales/orders/{order_id}/lines")
def add_order_lines(order_id: int = Depends(resolve_order_id), req: AddOrderLines = ..., _: bool = Depends(verify_api_key)):
    for line in req.lines:
        validate_bilingual(line.notes, line.notes_es, "notes")
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT order_number, status FROM sales_orders WHERE id = %s", (order_id,))
                row = cur.fetchone()
                if not row:
                    raise HTTPException(
                        status_code=404,
                        detail={
                            "error_code": "ORDER_NOT_FOUND",
                            "message": f"Order #{order_id} not found",
                            "input": str(order_id),
                            "suggestions": [],
                        }
                    )
                if row['status'] in ('shipped', 'invoiced', 'cancelled'):
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "error_code": "ORDER_LINES_LOCKED",
                            "message": f"Cannot add lines to {row['status']} order",
                            "input": str(order_id),
                            "suggestions": [],
                        }
                    )

                results = []
                warnings = []
                for line in req.lines:
                    product_id, prod_name = resolve_product_id(cur, line.product_name)

                    # Detect service items (Pallets, freight, etc.) — zero weight is valid
                    cur.execute(
                        "SELECT case_size_lb, COALESCE(is_service, false) AS is_service FROM products WHERE id = %s",
                        (product_id,)
                    )
                    prod_row = cur.fetchone()
                    is_service = prod_row and prod_row['is_service']

                    if is_service:
                        # Service items get zero weight, skip case-weight logic
                        line.quantity_lb = line.quantity_lb if line.quantity_lb else 0
                        effective_case_weight = None
                        used_unit = line.unit or 'each'
                    else:
                        # Fix #2: Auto-lookup case weight from product if not provided
                        effective_case_weight = line.case_weight_lb
                        used_unit = line.unit or 'lb'
                        if used_unit in ('cases', 'bags', 'boxes') and effective_case_weight is None:
                            if prod_row and prod_row.get('case_size_lb'):
                                effective_case_weight = float(prod_row['case_size_lb'])
                            else:
                                raise HTTPException(
                                    status_code=400,
                                    detail={
                                        "error_code": "CASE_WEIGHT_REQUIRED",
                                        "message": f"case_weight_lb is required for '{prod_name}' when ordering in {used_unit}. No default case weight is set for this product.",
                                        "input": prod_name,
                                        "suggestions": [],
                                    }
                                )
                            line.quantity_lb = line.quantity * effective_case_weight

                        # Fix #1: Warn if unit was not explicitly provided
                        if line.quantity is not None and line.unit is None:
                            warnings.append(
                                f"⚠️ '{prod_name}': No unit specified for quantity {line.quantity:,.0f} — "
                                f"defaulting to lb. Did you mean cases?"
                            )

                    cur.execute(
                        """INSERT INTO sales_order_lines (sales_order_id, product_id, quantity_lb, unit_price, notes, notes_es)
                           VALUES (%s, %s, %s, %s, %s, %s) RETURNING id""",
                        (order_id, product_id, line.quantity_lb, line.unit_price, line.notes, line.notes_es)
                    )
                    line_id = cur.fetchone()['id']
                    results.append({
                        "line_id": line_id,
                        "product": prod_name,
                        "quantity_lb": line.quantity_lb,
                        "original_quantity": line.quantity,
                        "original_unit": used_unit,
                        "case_weight_lb": effective_case_weight
                    })

                response = {"order_id": order_id, "order_number": row['order_number'], "lines_added": results, "message": f"Added {len(results)} line(s) to {row['order_number']}"}
                if warnings:
                    response["warnings"] = warnings
                return response
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Add order lines failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/sales/orders/{order_id}/lines/{line_id}/cancel")
def cancel_order_line(order_id: int = Depends(resolve_order_id), line_id: int = Path(...), _: bool = Depends(verify_api_key)):
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """UPDATE sales_order_lines SET line_status = 'cancelled'
                       WHERE id = %s AND sales_order_id = %s AND line_status != 'fulfilled'
                       RETURNING id, product_id""",
                    (line_id, order_id)
                )
                row = cur.fetchone()
                if not row:
                    raise HTTPException(404, "Line not found or already fulfilled")
                _lock_allocation_product(cur, int(row['product_id']))
                _expire_auto_fifo_allocations(cur, int(row['product_id']), _operator_id(_))
                released = _release_active_allocations(
                    cur,
                    line_id=line_id,
                    reason='line_cancelled',
                    released_by=_operator_id(_),
                )
                return {"order_id": order_id, "line_id": line_id, "line_status": "cancelled",
                        "allocations_released": released, "message": "Line cancelled"}
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Cancel order line failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.patch("/sales/orders/{order_id}/lines/{line_id}/update")
def update_order_line(
    order_id: int = Depends(resolve_order_id),
    line_id: int = Path(...),
    quantity_lb: Optional[float] = Query(default=None),
    unit_price: Optional[float] = Query(default=None),
    _: bool = Depends(verify_api_key)
):
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(
                    """SELECT id, product_id, quantity_lb, unit_price, line_status
                         FROM sales_order_lines
                        WHERE id = %s AND sales_order_id = %s
                          AND line_status NOT IN ('fulfilled', 'cancelled')
                        FOR UPDATE""",
                    (line_id, order_id),
                )
                existing = cur.fetchone()
                if not existing:
                    raise HTTPException(
                        status_code=404,
                        detail={
                            "error_code": "LINE_NOT_FOUND",
                            "message": "Line not found or already fulfilled/cancelled",
                            "input": str(line_id),
                            "suggestions": [],
                        }
                    )
                fields = []
                values = []
                allocations_released = []
                if quantity_lb is not None:
                    if quantity_lb <= 0:
                        _allocation_error(
                            "INVALID_LINE_QUANTITY",
                            "quantity_lb must be greater than zero",
                            status_code=422,
                            line_id=line_id,
                            quantity_lb=quantity_lb,
                        )
                    shipped_effective = _line_shipped_effective(
                        cur, line_id, int(existing['product_id'])
                    )
                    if quantity_lb + BALANCE_EPSILON < shipped_effective:
                        _allocation_error(
                            "QTY_BELOW_SHIPPED_EFFECTIVE",
                            f"Line #{line_id} cannot be reduced below {shipped_effective:.4f} lb already shipped",
                            status_code=422,
                            line_id=line_id,
                            requested_lb=quantity_lb,
                            shipped_effective_lb=shipped_effective,
                        )
                    fields.append("quantity_lb = %s")
                    values.append(quantity_lb)
                if unit_price is not None:
                    fields.append("unit_price = %s")
                    values.append(unit_price)
                if not fields:
                    raise HTTPException(
                        status_code=400,
                        detail={
                            "error_code": "NO_FIELDS_TO_UPDATE",
                            "message": "Nothing to update",
                            "input": "",
                            "suggestions": [],
                        }
                    )
                values.extend([line_id, order_id])
                cur.execute(
                    f"""UPDATE sales_order_lines SET {', '.join(fields)}
                        WHERE id = %s AND sales_order_id = %s AND line_status NOT IN ('fulfilled', 'cancelled')
                        RETURNING id, quantity_lb, unit_price, product_id""",
                    values
                )
                row = cur.fetchone()
                if quantity_lb is not None:
                    product_id = int(row['product_id'])
                    _lock_allocation_product(cur, product_id)
                    _expire_auto_fifo_allocations(cur, product_id, _operator_id(_))
                    remaining_effective = max(
                        0.0,
                        float(quantity_lb) - _line_shipped_effective(cur, line_id, product_id),
                    )
                    cur.execute(
                        """SELECT * FROM sales_order_allocations
                             WHERE sales_order_line_id = %s AND status = 'active'
                               AND (expires_at IS NULL OR expires_at > clock_timestamp())
                             ORDER BY created_at DESC, id DESC FOR UPDATE""",
                        (line_id,),
                    )
                    active_rows = cur.fetchall()
                    allocated = sum(float(item['quantity_lb']) for item in active_rows)
                    excess = max(0.0, allocated - remaining_effective)
                    if excess > BALANCE_EPSILON:
                        allocations_released = _shrink_active_allocations(
                            cur,
                            active_rows,
                            excess,
                            'line_quantity_reduced',
                            _operator_id(_),
                        )
                # Fetch case_size_lb for unit count
                cur.execute("SELECT case_size_lb FROM products WHERE id = %s", (row['product_id'],))
                prow = cur.fetchone()
                cs = float(prow['case_size_lb']) if prow and prow['case_size_lb'] else None
                qty = float(row['quantity_lb'])
                return {"line_id": row['id'], "quantity_lb": qty, "unit_price": float(row['unit_price']) if row['unit_price'] else None,
                        "case_size_lb": cs, "unit_count": round(qty / cs) if cs else None,
                        "allocations_released": allocations_released}
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Update order line failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# SHIP AGAINST ORDER ENDPOINTS (v2.3.0)
# ═══════════════════════════════════════════════════════════════

@app.post("/sales/orders/{order_id}/ship")
def ship_order(order_id: int = Depends(resolve_order_id), req: Optional[ShipOrderRequest] = None, _: bool = Depends(verify_api_key)):
    """Ship against a sales order. mode=preview returns feasibility; mode=commit executes and creates shipment record."""
    occurred_at, created_at_source = validate_inventory_occurred_at(
        req.occurred_at if req else None,
        req.backfill if req else False,
    )
    mode = "preview" if req is None else req.mode
    if mode == "preview":
        try:
            with get_transaction() as cur:
                cur.execute("SELECT so.order_number, so.status, c.name FROM sales_orders so JOIN customers c ON c.id = so.customer_id WHERE so.id = %s", (order_id,))
                order_row = cur.fetchone()
                if not order_row:
                    raise HTTPException(404, f"Order #{order_id} not found")
                if order_row['status'] == 'new':
                    raise HTTPException(400, f"Cannot ship order {order_row['order_number']} — status is 'new'. Confirm the order first.")
                if order_row['status'] in ('invoiced', 'cancelled'):
                    raise HTTPException(400, f"Cannot ship {order_row['status']} order")
                ship_all = (req is None) or (req.ship_all)
                cur.execute("""SELECT sol.id, p.id AS product_id, p.name, sol.quantity_lb, sol.quantity_shipped_lb,
                                      COALESCE(p.is_service, false) AS is_service
                               FROM sales_order_lines sol JOIN products p ON p.id = sol.product_id
                               WHERE sol.sales_order_id = %s AND sol.line_status NOT IN ('fulfilled', 'cancelled') ORDER BY sol.id""", (order_id,))
                lines = cur.fetchall()
                preview = []
                warnings = []
                for line in lines:
                    remaining = float(line['quantity_lb']) - float(line['quantity_shipped_lb'])
                    if remaining <= 0: continue
                    if ship_all:
                        ship_qty = remaining
                    elif req and req.lines:
                        match = next((rl for rl in req.lines if rl.line_id == line['id']), None)
                        if not match: continue
                        ship_qty = match.quantity_lb
                    else:
                        ship_qty = remaining
                    # F05-04: service lines auto-fulfill at commit — skip the stock lookup
                    # so preview doesn't misleadingly report "no stock" for a pallet charge.
                    if line['is_service']:
                        preview.append({"line_id": line['id'], "product": line['name'], "ordered_lb": float(line['quantity_lb']),
                                        "already_shipped_lb": float(line['quantity_shipped_lb']), "remaining_lb": remaining,
                                        "requested_ship_lb": ship_qty, "can_ship_lb": ship_qty, "on_hand_lb": None,
                                        "short": 0, "is_service": True})
                        continue
                    plan = _sales_order_ship_plan(
                        cur,
                        int(line['product_id']),
                        int(line['id']),
                        ship_qty,
                        lock=False,
                        persist_expired=False,
                    )
                    reservation_summary = _allocation_reservation_summary(
                        cur, int(line['product_id']), int(line['id'])
                    )
                    on_hand = _product_on_hand(cur, int(line['product_id']))
                    can_ship = float(plan['actual_ship_lb'])
                    if can_ship < ship_qty:
                        warnings.append(f"{line['name']}: only {can_ship:.1f} lb currently takeable, need {ship_qty:.1f} lb")
                    line_preview = {"line_id": line['id'], "product": line['name'], "ordered_lb": float(line['quantity_lb']),
                                    "already_shipped_lb": float(line['quantity_shipped_lb']), "remaining_lb": remaining,
                                    "requested_ship_lb": ship_qty, "can_ship_lb": can_ship, "on_hand_lb": on_hand,
                                    "reserved_others_lb": reservation_summary["reserved_others_lb"],
                                    "reserved_by_orders": reservation_summary["reserved_by_orders"],
                                    "short": max(0, ship_qty - can_ship)}
                    if _allocations_enforced():
                        reserved_taken = max(0.0, min(float(ship_qty), on_hand) - can_ship)
                        allocation_warning = _allocation_observe_warning(
                            "Sales-order ship",
                            ship_qty,
                            can_ship,
                            reserved_taken,
                            reservation_summary,
                            preview=True,
                        )
                        if allocation_warning:
                            line_preview["allocation_warning"] = allocation_warning
                    preview.append(line_preview)
                return {"mode": "preview", "order_number": order_row['order_number'], "customer": order_row['name'],
                        "status": order_row['status'], "lines": preview, "warnings": warnings,
                        "message": "Preview only — set mode=commit to execute"}
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Ship order preview failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})
    else:
        # mode == "commit"
        try:
            with get_db_connection() as conn:
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT so.id, so.order_number, so.status, so.customer_id, c.name FROM sales_orders so JOIN customers c ON c.id = so.customer_id WHERE so.id = %s", (order_id,))
                    order_row = cur.fetchone()
                    if not order_row:
                        raise HTTPException(404, f"Order #{order_id} not found")
                    if order_row['status'] == 'new':
                        raise HTTPException(400, f"Cannot ship order {order_row['order_number']} — status is 'new'. Confirm the order first.")
                    if order_row['status'] in ('invoiced', 'cancelled'):
                        raise HTTPException(400, f"Cannot ship {order_row['status']} order")
                    ship_all = (req is None) or (req.ship_all)

                    if ship_all:
                        cur.execute("""SELECT sol.id, sol.product_id, sol.quantity_lb, sol.quantity_shipped_lb, p.name,
                                              COALESCE(p.is_service, false) AS is_service
                                       FROM sales_order_lines sol JOIN products p ON p.id = sol.product_id
                                       WHERE sol.sales_order_id = %s AND sol.line_status NOT IN ('fulfilled', 'cancelled') ORDER BY sol.id""", (order_id,))
                        lines_to_ship = [{"line_id": r['id'], "product_id": r['product_id'],
                                          "quantity_lb": float(r['quantity_lb']) - float(r['quantity_shipped_lb']),
                                          "product_name": r['name'],
                                          "is_service": r['is_service']}
                                         for r in cur.fetchall()
                                         if float(r['quantity_lb']) - float(r['quantity_shipped_lb']) > 0]
                    else:
                        lines_to_ship = []
                        for rl in (req.lines or []):
                            cur.execute("""SELECT sol.id, sol.product_id, sol.quantity_lb, sol.quantity_shipped_lb, p.name,
                                                  COALESCE(p.is_service, false) AS is_service
                                           FROM sales_order_lines sol JOIN products p ON p.id = sol.product_id
                                           WHERE sol.id = %s AND sol.sales_order_id = %s""", (rl.line_id, order_id))
                            r = cur.fetchone()
                            if not r:
                                raise HTTPException(404, f"Line #{rl.line_id} not found on order #{order_id}")
                            remaining = float(r['quantity_lb']) - float(r['quantity_shipped_lb'])
                            if remaining <= 0:
                                raise HTTPException(status_code=409, detail={"error_code": "LINE_ALREADY_FULFILLED", "message": f"Line #{rl.line_id} ({r['name']}) is already fully shipped.", "line_id": rl.line_id, "product": r['name'], "ordered_lb": float(r['quantity_lb']), "shipped_lb": float(r['quantity_shipped_lb']), "remaining_lb": 0})
                            if rl.quantity_lb > remaining:
                                raise HTTPException(status_code=422, detail={"error_code": "QTY_EXCEEDS_REMAINING", "message": f"Line #{rl.line_id} ({r['name']}): requested {rl.quantity_lb} lb but only {remaining} lb remaining.", "line_id": rl.line_id, "product": r['name'], "requested_lb": rl.quantity_lb, "remaining_lb": remaining, "suggestion": f"Retry with quantity_lb={remaining}"})
                            lines_to_ship.append({"line_id": r['id'], "product_id": r['product_id'], "quantity_lb": rl.quantity_lb, "product_name": r['name'], "is_service": r['is_service']})

                    if not lines_to_ship:
                        raise HTTPException(status_code=409, detail={"error_code": "ORDER_ALREADY_FULFILLED", "message": f"Order {order_row['order_number']} has no remaining lines to ship.", "order_id": order_id, "order_number": order_row['order_number'], "status": order_row['status']})

                    # PR 5 hard-gate preflight.  Run before the shipment header or
                    # any line/service updates.  This is deliberately absent from
                    # the flag-off path, whose PR 4 response/write behavior is a
                    # regression contract.
                    if _allocations_enforced():
                        for item in lines_to_ship:
                            if item["is_service"]:
                                continue
                            plan = _sales_order_ship_plan(
                                cur,
                                int(item["product_id"]),
                                int(item["line_id"]),
                                float(item["quantity_lb"]),
                                released_by=_operator_id(_),
                                lock=True,
                                persist_expired=False,
                            )
                            can_ship = float(plan["actual_ship_lb"])
                            # Match the ship plan's physical basis exactly:
                            # only positive-balance FIFO lots are consumable.
                            # A negative audit lot must not offset those lots
                            # and hide a foreign-reservation steal.
                            on_hand = sum(
                                float(lot["available"] or 0)
                                for lot in fifo_lot_balances(
                                    cur, int(item["product_id"]), include_empty=False
                                )
                            )
                            reserved_taken = max(
                                0.0,
                                min(float(item["quantity_lb"]), on_hand) - can_ship,
                            )
                            reservation_summary = _allocation_reservation_summary(
                                cur, int(item["product_id"]), int(item["line_id"])
                            )
                            _enforce_allocation_takeable(
                                "Sales-order ship",
                                float(item["quantity_lb"]),
                                can_ship,
                                reserved_taken,
                                reservation_summary,
                                order_id=int(order_id),
                                order_number=order_row["order_number"],
                                sales_order_line_id=int(item["line_id"]),
                                product_id=int(item["product_id"]),
                            )

                    now = get_plant_now()

                    # Create shipment record (v3.0.0)
                    cur.execute("""
                        INSERT INTO shipments (sales_order_id, shipped_at, customer_id)
                        VALUES (%s, %s, %s) RETURNING id
                    """, (order_id, occurred_at or now, order_row['customer_id']))
                    shipment_id = cur.fetchone()['id']

                    results = []
                    all_fully_shipped = True

                    for item in lines_to_ship:
                        qty_to_ship = item["quantity_lb"]

                        # F05-04: service lines (pallet charges, freight, etc.) auto-fulfill
                        # with no inventory lookup — they represent billing, not stock movement.
                        if item["is_service"]:
                            cur.execute(
                                "UPDATE sales_order_lines SET quantity_shipped_lb = quantity_shipped_lb + %s, line_status = 'fulfilled' WHERE id = %s",
                                (qty_to_ship, item["line_id"]),
                            )
                            results.append({
                                "line_id": item["line_id"],
                                "product": item["product_name"],
                                "requested_lb": qty_to_ship,
                                "shipped_lb": qty_to_ship,
                                "status": "fulfilled",
                                "line_status": "fulfilled",
                                "is_service": True,
                            })
                            continue

                        plan = _sales_order_ship_plan(
                            cur,
                            int(item["product_id"]),
                            int(item["line_id"]),
                            qty_to_ship,
                            released_by=_operator_id(_),
                        )
                        actual_ship = float(plan["actual_ship_lb"])
                        if actual_ship <= BALANCE_EPSILON:
                            results.append({"line_id": item["line_id"], "product": item["product_name"], "requested_lb": qty_to_ship, "shipped_lb": 0, "status": "no_stock"})
                            all_fully_shipped = False
                            continue

                        cur.execute("""
                            INSERT INTO transactions (
                                type, timestamp, customer_name, notes,
                                occurred_at, created_at_source
                            )
                            VALUES ('ship', %s, %s, %s, %s, %s) RETURNING id
                        """, (
                            now, order_row['name'],
                            f"Sales order {order_row['order_number']} — {item['product_name']}",
                            occurred_at, created_at_source,
                        ))
                        txn_id = cur.fetchone()['id']
                        lots_used = []
                        for lot in plan["lots"]:
                            take = float(lot["quantity_lb"])
                            cur.execute("INSERT INTO transaction_lines (transaction_id, product_id, lot_id, quantity_lb) VALUES (%s, %s, %s, %s)", (txn_id, item["product_id"], lot['lot_id'], -take))
                            lots_used.append({"lot_code": lot['lot_code'], "quantity_lb": take})

                        allocation_changes = _consume_sales_order_allocations(
                            cur, plan, txn_id
                        )

                        cur.execute("UPDATE sales_order_lines SET quantity_shipped_lb = quantity_shipped_lb + %s WHERE id = %s RETURNING quantity_lb, quantity_shipped_lb", (actual_ship, item["line_id"]))
                        updated = cur.fetchone()
                        ordered = float(updated['quantity_lb'])
                        new_shipped = float(updated['quantity_shipped_lb'])
                        if new_shipped >= ordered:
                            new_line_status = 'fulfilled'
                        elif new_shipped > 0:
                            new_line_status = 'partial'
                            all_fully_shipped = False
                        else:
                            new_line_status = 'pending'
                            all_fully_shipped = False
                        cur.execute("UPDATE sales_order_lines SET line_status = %s WHERE id = %s", (new_line_status, item["line_id"]))

                        # Record in sales_order_shipments
                        cur.execute("INSERT INTO sales_order_shipments (sales_order_line_id, transaction_id, quantity_lb) VALUES (%s, %s, %s)", (item["line_id"], txn_id, actual_ship))

                        # Record in shipment_lines (v3.0.0)
                        cur.execute("""
                            INSERT INTO shipment_lines (shipment_id, transaction_id, sales_order_line_id, product_id, quantity_lb)
                            VALUES (%s, %s, %s, %s, %s)
                        """, (shipment_id, txn_id, item["line_id"], item["product_id"], actual_ship))

                        # Fetch case_size_lb for unit count
                        cur.execute("SELECT case_size_lb FROM products WHERE id = %s", (item["product_id"],))
                        cs_row = cur.fetchone()
                        cs_lb = float(cs_row['case_size_lb']) if cs_row and cs_row['case_size_lb'] else None
                        results.append({"line_id": item["line_id"], "product": item["product_name"],
                                        "requested_lb": qty_to_ship, "shipped_lb": actual_ship,
                                        "shipped_units": round(actual_ship / cs_lb) if cs_lb else None,
                                        "short_lb": max(0, qty_to_ship - actual_ship), "lots_used": lots_used,
                                        "allocation_changes": allocation_changes,
                                        "transaction_id": txn_id, "confirmation_code": generate_confirmation_code(txn_id),
                                        "line_status": new_line_status})
                        if actual_ship < qty_to_ship:
                            all_fully_shipped = False

                    # Block zero-shipment: if no PHYSICAL line items shipped any quantity, roll back.
                    # Service lines (pallet charges, etc.) don't count — they auto-fulfill without
                    # inventory movement and shouldn't bypass this guard on their own.
                    any_actually_shipped = any(
                        r.get("shipped_lb", 0) > 0 and not r.get("is_service") for r in results
                    )
                    if not any_actually_shipped:
                        cur.execute("DELETE FROM shipment_lines WHERE shipment_id = %s", (shipment_id,))
                        cur.execute("DELETE FROM shipments WHERE id = %s", (shipment_id,))
                        raise HTTPException(status_code=409, detail={
                            "error_code": "ZERO_SHIPMENT",
                            "message": "No items could be shipped — all products have zero available stock",
                            "order_id": order_id,
                            "order_number": order_row['order_number'],
                            "lines_attempted": len(results)
                        })

                    new_order_status = 'shipped' if all_fully_shipped else 'partial_ship'
                    cur.execute("UPDATE sales_orders SET status = %s WHERE id = %s", (new_order_status, order_id))
                    logger.info(f"Ship order {order_row['order_number']}: {'fully' if all_fully_shipped else 'partially'} shipped")
                    return {"mode": "commit", "order_number": order_row['order_number'], "customer": order_row['name'],
                            "order_status": new_order_status, "shipment_id": shipment_id, "lines_shipped": results,
                            "message": f"Order {order_row['order_number']} {'fully' if all_fully_shipped else 'partially'} shipped"}
        except HTTPException:
            raise
        except Exception as e:
            if _is_readonly_error(e): raise
            logger.error(f"Ship order commit failed: {e}")
            return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# SPLIT-PATH ALIASES — GPT schema uses /endpoint/preview and
# /endpoint/commit instead of mode in the body.  These thin
# wrappers keep the old schema working.
# ═══════════════════════════════════════════════════════════════

@app.post("/receive/preview", include_in_schema=False)
def receive_preview(req: ReceiveRequest, _: bool = Depends(verify_api_key)):
    req.mode = "preview"
    return receive(req, _)

@app.post("/receive/commit", include_in_schema=False)
def receive_commit(req: ReceiveRequest, _: bool = Depends(verify_api_key)):
    req.mode = "commit"
    return receive(req, _)

@app.post("/ship/preview", include_in_schema=False)
def ship_preview(req: ShipRequest, _: bool = Depends(verify_api_key)):
    req.mode = "preview"
    return ship(req, _)

@app.post("/ship/commit", include_in_schema=False)
def ship_commit(req: ShipRequest, _: bool = Depends(verify_api_key)):
    req.mode = "commit"
    return ship(req, _)

@app.post("/make/preview", include_in_schema=False)
def make_preview(req: MakeRequest, _: bool = Depends(verify_api_key)):
    req.mode = "preview"
    return make(req, _)

@app.post("/make/commit", include_in_schema=False)
def make_commit(req: MakeRequest, _: bool = Depends(verify_api_key)):
    req.mode = "commit"
    return make(req, _)

@app.post("/pack/preview", include_in_schema=False)
def pack_preview(req: PackRequest, _: bool = Depends(verify_api_key)):
    req.mode = "preview"
    return pack(req, _)

@app.post("/pack/commit", include_in_schema=False)
def pack_commit(req: PackRequest, _: bool = Depends(verify_api_key)):
    req.mode = "commit"
    return pack(req, _)

@app.post("/adjust/preview", include_in_schema=False)
def adjust_preview(req: AdjustRequest, _: bool = Depends(verify_api_key)):
    req.mode = "preview"
    return adjust(req, _)

@app.post("/adjust/commit", include_in_schema=False)
def adjust_commit(req: AdjustRequest, _: bool = Depends(verify_api_key)):
    req.mode = "commit"
    return adjust(req, _)

@app.post("/sales/orders/{order_id}/ship/preview", include_in_schema=False)
def ship_order_preview(order_id: int = Depends(resolve_order_id), req: Optional[ShipOrderRequest] = None, _: bool = Depends(verify_api_key)):
    if req is None:
        req = ShipOrderRequest()
    req.mode = "preview"
    return ship_order(order_id, req, _)

@app.post("/sales/orders/{order_id}/ship/commit", operation_id="commitShipOrder")
def commit_ship_order(req: CommitShipOrderRequest, order_id: int = Depends(resolve_order_id), _: bool = Depends(verify_api_key)):
    """Always commit a sales-order shipment through the shared ship_order service."""
    commit_req = ShipOrderRequest(
        mode="commit",
        ship_all=req.ship_all,
        lines=req.lines,
        occurred_at=req.occurred_at,
        backfill=req.backfill,
    )
    return ship_order(order_id, commit_req, _)


# ═══════════════════════════════════════════════════════════════
# PACKING SLIP PDF (v3.0.0)
# ═══════════════════════════════════════════════════════════════

@app.get("/sales/orders/{order_id}/packing-slip")
def generate_packing_slip(order_id: int = Depends(resolve_order_id), _: bool = Depends(verify_api_key_flexible)):
    """Generate a printable packing slip PDF for a sales order."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

    try:
        with get_transaction() as cur:
            # ── Fetch order header + customer ──
            cur.execute("""
                SELECT so.id, so.order_number, so.order_date, so.requested_ship_date,
                       so.status, so.notes,
                       c.name AS customer_name, c.address AS customer_address
                FROM sales_orders so
                JOIN customers c ON c.id = so.customer_id
                WHERE so.id = %s
            """, (order_id,))
            order = cur.fetchone()
            if not order:
                raise HTTPException(404, f"Order #{order_id} not found")
            if order['status'] == 'cancelled':
                raise HTTPException(400, f"Order {order['order_number']} is cancelled — cannot generate packing slip")

            # ── Fetch order lines ──
            cur.execute("""
                SELECT sol.id, sol.product_id, p.name AS product_name,
                       sol.quantity_lb, sol.quantity_shipped_lb,
                       p.case_size_lb, sol.line_status
                FROM sales_order_lines sol
                JOIN products p ON p.id = sol.product_id
                WHERE sol.sales_order_id = %s AND sol.line_status != 'cancelled'
                ORDER BY sol.id
            """, (order_id,))
            lines = cur.fetchall()

            # ── Lot allocation: actual shipment data or FIFO preview ──
            line_allocations = []

            # Check if this order has committed shipment records
            cur.execute("""
                SELECT s.id
                FROM shipments s
                JOIN shipment_lines sl ON sl.shipment_id = s.id
                JOIN ledger_current_transactions t ON t.id = sl.transaction_id
                WHERE s.sales_order_id = %s
                  AND t.effective_status = 'posted'
                LIMIT 1
            """, (order_id,))
            has_shipments = cur.fetchone() is not None

            if has_shipments:
                # ── Post-shipment: pull actual lot allocations from shipment records ──
                cur.execute("""
                    SELECT p.name AS product_name, p.case_size_lb,
                           l.lot_code, l.supplier_lot_code,
                           ABS(tl.quantity_lb) AS quantity_lb
                    FROM shipment_lines sl
                    JOIN ledger_current_transactions t ON t.id = sl.transaction_id
                        AND t.effective_status = 'posted'
                    JOIN ledger_current_transaction_lines tl ON tl.transaction_id = sl.transaction_id
                    JOIN lots l ON l.id = tl.lot_id
                    JOIN products p ON p.id = tl.product_id
                    WHERE sl.shipment_id IN (
                        SELECT s.id FROM shipments s WHERE s.sales_order_id = %s
                    )
                    AND tl.quantity_lb != 0
                    ORDER BY p.name, l.created_at ASC
                """, (order_id,))
                shipped_lots = cur.fetchall()

                for row in shipped_lots:
                    product_name = row['product_name']
                    case_size = float(row['case_size_lb']) if row['case_size_lb'] else None
                    qty = float(row['quantity_lb'])

                    product_lower = product_name.lower()
                    non_weight_keywords = ('pallet', 'freight', 'delivery', 'surcharge', 'charge', 'fee')
                    is_non_weight = any(kw in product_lower for kw in non_weight_keywords)

                    if is_non_weight:
                        qty_display = str(int(qty)) if qty == int(qty) else str(qty)
                    elif case_size and case_size > 0:
                        cases = round(qty / case_size)
                        qty_display = f"{qty:g} lb \u00b7 {cases} units"
                    else:
                        qty_display = f"{qty:g} lb"

                    line_allocations.append({
                        "product_name": product_name,
                        "lot_code": row['lot_code'],
                        "supplier_lot_code": row['supplier_lot_code'],
                        "qty_display": qty_display
                    })

                # Include non-weight order lines that may not have shipment lot records
                shipped_product_names = {r['product_name'] for r in shipped_lots}
                for line in lines:
                    product_name = line['product_name']
                    product_lower = product_name.lower()
                    non_weight_keywords = ('pallet', 'freight', 'delivery', 'surcharge', 'charge', 'fee')
                    is_non_weight = any(kw in product_lower for kw in non_weight_keywords)
                    if is_non_weight and product_name not in shipped_product_names:
                        qty_lb = float(line['quantity_lb'])
                        qty_display = str(int(qty_lb)) if qty_lb == int(qty_lb) else str(qty_lb)
                        line_allocations.append({
                            "product_name": product_name,
                            "lot_code": "N/A",
                            "qty_display": qty_display
                        })

            else:
                # ── Pre-shipment: FIFO lot allocation preview ──
                for line in lines:
                    product_name = line['product_name']
                    qty_lb = float(line['quantity_lb'])
                    remaining_lb = qty_lb - float(line['quantity_shipped_lb'])
                    case_size = float(line['case_size_lb']) if line['case_size_lb'] else None

                    # Detect non-weight items
                    product_lower = product_name.lower()
                    non_weight_keywords = ('pallet', 'freight', 'delivery', 'surcharge', 'charge', 'fee')
                    is_non_weight = any(kw in product_lower for kw in non_weight_keywords)

                    if is_non_weight:
                        qty_display = str(int(qty_lb)) if qty_lb == int(qty_lb) else str(qty_lb)
                        line_allocations.append({
                            "product_name": product_name,
                            "lot_code": "N/A",
                            "qty_display": qty_display
                        })
                        continue

                    # Query FIFO lots
                    cur.execute(f"""
                        SELECT l.id, l.lot_code, l.supplier_lot_code,
                               COALESCE(SUM(tl.quantity_lb), 0) AS available
                        FROM lots l
                        LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                        WHERE l.product_id = %s
                        GROUP BY l.id
                        HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                        ORDER BY COALESCE(l.received_at, l.created_at) ASC
                    """, (line['product_id'],))
                    lots = cur.fetchall()

                    remaining_need = remaining_lb if remaining_lb > 0 else qty_lb
                    allocated = False
                    for lot in lots:
                        if remaining_need <= 0:
                            break
                        avail = float(lot['available'])
                        take = min(avail, remaining_need)
                        qty_display_val = take
                        if case_size and case_size > 0:
                            cases = round(take / case_size)
                            qty_display = f"{take:g} lb \u00b7 {cases} units"
                        else:
                            qty_display = f"{take:g} lb"
                        line_allocations.append({
                            "product_name": product_name,
                            "lot_code": lot['lot_code'],
                            "supplier_lot_code": lot['supplier_lot_code'],
                            "qty_display": qty_display
                        })
                        remaining_need -= take
                        allocated = True

                    if remaining_need > 0:
                        # Shortfall — show INSUFFICIENT row
                        if case_size and case_size > 0:
                            short_cases = math.ceil(remaining_need / case_size)
                            qty_display = f"{remaining_need:g} lb \u00b7 {short_cases} units"
                        else:
                            qty_display = f"{remaining_need:g} lb"
                        insufficient_entry = {
                            "product_name": product_name,
                            "lot_code": "INSUFFICIENT",
                            "qty_display": qty_display
                        }
                        # Check if unpacked batch inventory exists for this FG product.
                        # If /make was run but /pack was not, batch inventory sits idle
                        # while the FG SKU shows zero on-hand.
                        cur.execute(f"""
                            SELECT p2.name AS batch_name,
                                   COALESCE(SUM(tl.quantity_lb), 0) AS batch_available
                            FROM products p
                            JOIN products p2 ON p2.id = p.parent_batch_product_id
                            JOIN lots l ON l.product_id = p2.id
                            JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                            WHERE p.id = %s AND p.parent_batch_product_id IS NOT NULL
                            GROUP BY p2.name
                            HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                        """, (line['product_id'],))
                        batch_row = cur.fetchone()
                        if batch_row:
                            batch_avail = float(batch_row['batch_available'])
                            insufficient_entry["batch_hint"] = (
                                f"Note: {batch_avail:g} lb of {batch_row['batch_name']} is available "
                                f"— run /pack to convert to finished goods."
                            )
                        line_allocations.append(insufficient_entry)
                    elif not allocated:
                        # No lots at all
                        if case_size and case_size > 0:
                            cases = round(qty_lb / case_size)
                            qty_display = f"{qty_lb:g} lb \u00b7 {cases} units"
                        else:
                            qty_display = f"{qty_lb:g} lb"
                        insufficient_entry = {
                            "product_name": product_name,
                            "lot_code": "INSUFFICIENT",
                            "qty_display": qty_display
                        }
                        # Same batch-inventory cross-reference for zero-lot case
                        cur.execute(f"""
                            SELECT p2.name AS batch_name,
                                   COALESCE(SUM(tl.quantity_lb), 0) AS batch_available
                            FROM products p
                            JOIN products p2 ON p2.id = p.parent_batch_product_id
                            JOIN lots l ON l.product_id = p2.id
                            JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                            WHERE p.id = %s AND p.parent_batch_product_id IS NOT NULL
                            GROUP BY p2.name
                            HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                        """, (line['product_id'],))
                        batch_row = cur.fetchone()
                        if batch_row:
                            batch_avail = float(batch_row['batch_available'])
                            insufficient_entry["batch_hint"] = (
                                f"Note: {batch_avail:g} lb of {batch_row['batch_name']} is available "
                                f"— run /pack to convert to finished goods."
                            )
                        line_allocations.append(insufficient_entry)

            # ── Build PDF ──
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter,
                                    topMargin=0.5*inch, bottomMargin=0.5*inch,
                                    leftMargin=0.6*inch, rightMargin=0.6*inch)

            styles = getSampleStyleSheet()
            elements = []

            # Colors
            dark_gray = HexColor('#333333')
            med_gray = HexColor('#666666')
            light_gray = HexColor('#999999')
            header_bg = HexColor('#4a4a4a')
            white = HexColor('#FFFFFF')
            row_alt = HexColor('#F5F5F5')

            # Custom styles
            style_company = ParagraphStyle('Company', parent=styles['Normal'],
                                           fontSize=12, leading=14, textColor=dark_gray,
                                           fontName='Helvetica-Bold')
            style_company_detail = ParagraphStyle('CompanyDetail', parent=styles['Normal'],
                                                   fontSize=8, leading=10, textColor=med_gray)
            style_title = ParagraphStyle('SlipTitle', parent=styles['Normal'],
                                          fontSize=18, leading=22, textColor=dark_gray,
                                          fontName='Helvetica-Bold', spaceAfter=6)
            style_label = ParagraphStyle('Label', parent=styles['Normal'],
                                          fontSize=7, leading=9, textColor=light_gray,
                                          fontName='Helvetica-Bold')
            style_value = ParagraphStyle('Value', parent=styles['Normal'],
                                          fontSize=9, leading=11, textColor=dark_gray)
            style_value_bold = ParagraphStyle('ValueBold', parent=styles['Normal'],
                                              fontSize=9, leading=11, textColor=dark_gray,
                                              fontName='Helvetica-Bold')
            style_small_gray = ParagraphStyle('SmallGray', parent=styles['Normal'],
                                              fontSize=7, leading=9, textColor=light_gray)
            style_footer = ParagraphStyle('Footer', parent=styles['Normal'],
                                           fontSize=7, leading=9, textColor=light_gray)
            style_section_label = ParagraphStyle('SectionLabel', parent=styles['Normal'],
                                                  fontSize=10, leading=12, textColor=dark_gray,
                                                  fontName='Helvetica-Bold', spaceBefore=12,
                                                  spaceAfter=4)
            style_sig_line = ParagraphStyle('SigLine', parent=styles['Normal'],
                                            fontSize=9, leading=18, textColor=dark_gray)

            # ── HEADER: Company info ──
            company_block = [
                Paragraph("CNS Confectionery Products LLC", style_company),
                Paragraph("33 Hook Road", style_company_detail),
                Paragraph("Bayonne, NJ 07002 US", style_company_detail),
                Paragraph("(201) 823-1400", style_company_detail),
                Paragraph("miriam@cnscoinc.com", style_company_detail),
            ]
            for p in company_block:
                elements.append(p)

            elements.append(Spacer(1, 0.25*inch))

            # ── TITLE ──
            elements.append(Paragraph("Packing Slip", style_title))
            elements.append(Spacer(1, 0.15*inch))

            # ── INFO ROW (4 columns) ──
            # Customer address
            cust_name = order['customer_name'] or ""
            cust_addr = order['customer_address'] or "Address on file"
            addr_lines = cust_addr.split('\n') if cust_addr else ["Address on file"]
            bill_to_text = f"<b>{cust_name}</b><br/>" + "<br/>".join(addr_lines)
            ship_to_text = bill_to_text  # Mirror for now

            ship_date = str(order['requested_ship_date']) if order['requested_ship_date'] else "TBD"
            order_date = str(order['order_date']) if order['order_date'] else ""
            order_number = order['order_number']

            style_info = ParagraphStyle('InfoCell', parent=styles['Normal'],
                                         fontSize=8, leading=10, textColor=dark_gray)

            col1 = [Paragraph("BILL TO", style_label), Paragraph(bill_to_text, style_info)]
            col2 = [Paragraph("SHIP TO", style_label), Paragraph(ship_to_text, style_info)]
            col3 = [
                Paragraph("SHIP DATE", style_label), Paragraph(ship_date, style_info),
                Spacer(1, 4),
                Paragraph("SHIP VIA", style_label), Paragraph("Customer Pick Up", style_info)
            ]
            col4 = [
                Paragraph("SO #", style_label), Paragraph(f"<b>{order_number}</b>", style_info),
                Spacer(1, 4),
                Paragraph("DATE", style_label), Paragraph(order_date, style_info)
            ]

            # Render the 4-column info as a table
            from reportlab.platypus import KeepTogether
            info_table_data = [[col1, col2, col3, col4]]
            info_table = Table(info_table_data, colWidths=[2.0*inch, 2.0*inch, 1.5*inch, 1.5*inch])
            info_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 0),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(info_table)
            elements.append(Spacer(1, 0.15*inch))

            # ── PURCHASE ORDER ──
            po_ref = order['notes'] if order['notes'] else ""
            if po_ref:
                elements.append(Paragraph("PURCHASE ORDER", style_label))
                elements.append(Paragraph(po_ref, style_value))
                elements.append(Spacer(1, 0.1*inch))

            # ── ITEMS TABLE ──
            style_table_header = ParagraphStyle('TableHeader', parent=styles['Normal'],
                                                 fontSize=8, leading=10, textColor=white,
                                                 fontName='Helvetica-Bold')
            style_table_cell = ParagraphStyle('TableCell', parent=styles['Normal'],
                                               fontSize=8, leading=10, textColor=dark_gray)
            style_table_cell_bold = ParagraphStyle('TableCellBold', parent=styles['Normal'],
                                                    fontSize=8, leading=10, textColor=dark_gray,
                                                    fontName='Helvetica-Bold')

            table_data = [[
                Paragraph("DATE", style_table_header),
                Paragraph("ACTIVITY", style_table_header),
                Paragraph("LOT #", style_table_header),
                Paragraph("QTY", style_table_header),
            ]]

            style_supplier_ref = ParagraphStyle('SupplierRef', parent=styles['Normal'],
                                                    fontSize=6, leading=8, textColor=light_gray)

            style_batch_hint = ParagraphStyle('BatchHint', parent=styles['Normal'],
                                                    fontSize=7, leading=9, textColor=HexColor('#CC6600'),
                                                    fontName='Helvetica-Oblique')

            order_date_str = str(order['order_date']) if order['order_date'] else ""
            for alloc in line_allocations:
                lot_style = style_table_cell_bold if alloc['lot_code'] == 'INSUFFICIENT' else style_table_cell
                # Build lot cell with optional supplier lot reference
                lot_text = alloc['lot_code']
                supplier_code = alloc.get('supplier_lot_code')
                if supplier_code and alloc['lot_code'] not in ('N/A', 'INSUFFICIENT'):
                    lot_cell = [Paragraph(lot_text, lot_style),
                                Paragraph(f"(Supplier: {supplier_code})", style_supplier_ref)]
                else:
                    lot_cell = Paragraph(lot_text, lot_style)
                table_data.append([
                    Paragraph(order_date_str, style_table_cell),
                    Paragraph(alloc['product_name'], style_table_cell),
                    lot_cell,
                    Paragraph(alloc['qty_display'], style_table_cell),
                ])
                # If this INSUFFICIENT line has a batch hint, add a note row
                if alloc.get('batch_hint'):
                    table_data.append([
                        Paragraph("", style_table_cell),
                        Paragraph(alloc['batch_hint'], style_batch_hint),
                        Paragraph("", style_table_cell),
                        Paragraph("", style_table_cell),
                    ])

            items_table = Table(table_data, colWidths=[1.0*inch, 3.2*inch, 1.5*inch, 1.3*inch])

            # Build table style with alternating rows
            table_style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), header_bg),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LINEBELOW', (0, 0), (-1, 0), 0.5, dark_gray),
                ('LINEBELOW', (0, -1), (-1, -1), 0.5, light_gray),
            ]
            # Alternating row backgrounds
            for i in range(1, len(table_data)):
                if i % 2 == 0:
                    table_style_cmds.append(('BACKGROUND', (0, i), (-1, i), row_alt))

            items_table.setStyle(TableStyle(table_style_cmds))
            elements.append(items_table)

            # ── Dashed separator ──
            elements.append(Spacer(1, 0.2*inch))
            from reportlab.platypus import HRFlowable
            elements.append(HRFlowable(width="100%", thickness=0.5, color=light_gray,
                                        dash=[3, 3], spaceBefore=0, spaceAfter=0))
            elements.append(Spacer(1, 0.2*inch))

            # ── WAREHOUSE CONFIRMATION ──
            elements.append(Paragraph("WAREHOUSE CONFIRMATION", style_section_label))
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Paragraph("Picked By: ___________________________________", style_sig_line))
            elements.append(Paragraph("Verified By: ___________________________________", style_sig_line))
            elements.append(Paragraph("Date: ___________________________________", style_sig_line))
            elements.append(Spacer(1, 0.2*inch))

            # ── Traceability note ──
            elements.append(Paragraph(
                "All lot numbers are system-assigned internal codes from Factory Ledger. "
                "Do not substitute with supplier lot numbers.",
                style_small_gray
            ))
            elements.append(Spacer(1, 0.3*inch))

            # ── Footer ──
            now_et = get_plant_now()
            footer_ts = now_et.strftime("%Y-%m-%d %I:%M %p ET")
            footer_table_data = [[
                Paragraph(f"Generated by Factory Ledger | {footer_ts}", style_footer),
                Paragraph("Page 1 of 1", ParagraphStyle('FooterRight', parent=style_footer, alignment=TA_RIGHT))
            ]]
            footer_table = Table(footer_table_data, colWidths=[5.0*inch, 2.0*inch])
            footer_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]))
            elements.append(footer_table)

            # ── Render ──
            doc.build(elements)
            buffer.seek(0)

            return StreamingResponse(
                buffer,
                media_type="application/pdf",
                headers={
                    "Content-Disposition": f'inline; filename="packing_slip_{order_number}.pdf"'
                }
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Packing slip generation failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# SALES DASHBOARD (v2.3.0)
# ═══════════════════════════════════════════════════════════════

@app.get("/sales/dashboard")
def sales_dashboard(_: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            # Status counts
            cur.execute(
                """SELECT status, COUNT(*) as cnt FROM sales_orders
                   WHERE status NOT IN ('invoiced', 'cancelled')
                   GROUP BY status ORDER BY status"""
            )
            status_counts = {r['status']: r['cnt'] for r in cur.fetchall()}

            # Overdue
            cur.execute(
                """SELECT so.order_number, c.name AS customer, so.requested_ship_date,
                          SUM(sol.quantity_lb - sol.quantity_shipped_lb) FILTER (WHERE NOT COALESCE(p.is_service, false)) AS remaining_lb
                   FROM sales_orders so
                   JOIN customers c ON c.id = so.customer_id
                   JOIN sales_order_lines sol ON sol.sales_order_id = so.id
                   JOIN products p ON p.id = sol.product_id
                   WHERE so.requested_ship_date < CURRENT_DATE
                     AND so.status NOT IN ('shipped', 'invoiced', 'cancelled')
                   GROUP BY so.id, c.name
                   HAVING SUM(sol.quantity_lb - sol.quantity_shipped_lb) FILTER (WHERE NOT COALESCE(p.is_service, false)) > 0
                   ORDER BY so.requested_ship_date ASC"""
            )
            overdue = [
                {"order_number": r['order_number'], "customer": r['customer'],
                 "requested_ship_date": str(r['requested_ship_date']), "remaining_lb": float(r['remaining_lb'])}
                for r in cur.fetchall()
            ]

            # Due this week
            cur.execute(
                """SELECT so.order_number, c.name AS customer, so.requested_ship_date,
                          SUM(sol.quantity_lb - sol.quantity_shipped_lb) FILTER (WHERE NOT COALESCE(p.is_service, false)) AS remaining_lb
                   FROM sales_orders so
                   JOIN customers c ON c.id = so.customer_id
                   JOIN sales_order_lines sol ON sol.sales_order_id = so.id
                   JOIN products p ON p.id = sol.product_id
                   WHERE so.requested_ship_date BETWEEN CURRENT_DATE AND CURRENT_DATE + INTERVAL '7 days'
                     AND so.status NOT IN ('shipped', 'invoiced', 'cancelled')
                   GROUP BY so.id, c.name
                   HAVING SUM(sol.quantity_lb - sol.quantity_shipped_lb) FILTER (WHERE NOT COALESCE(p.is_service, false)) > 0
                   ORDER BY so.requested_ship_date ASC"""
            )
            due_this_week = [
                {"order_number": r['order_number'], "customer": r['customer'],
                 "requested_ship_date": str(r['requested_ship_date']), "remaining_lb": float(r['remaining_lb'])}
                for r in cur.fetchall()
            ]

            # Recent shipments
            cur.execute(
                """SELECT so.order_number, c.name AS customer, SUM(sos.quantity_lb) AS shipped_lb,
                          MAX(sos.shipped_at) AS last_shipped
                   FROM sales_order_shipments sos
                   JOIN sales_order_lines sol ON sol.id = sos.sales_order_line_id
                   JOIN sales_orders so ON so.id = sol.sales_order_id
                   JOIN customers c ON c.id = so.customer_id
                   JOIN ledger_current_transactions t ON t.id = sos.transaction_id
                   WHERE sos.shipped_at > now() - INTERVAL '7 days'
                     AND t.effective_status = 'posted'
                   GROUP BY so.id, c.name
                   ORDER BY last_shipped DESC"""
            )
            recent_shipments = []
            for r in cur.fetchall():
                s_date, s_time = format_timestamp(r['last_shipped'])
                recent_shipments.append({
                    "order_number": r['order_number'], "customer": r['customer'],
                    "shipped_lb": float(r['shipped_lb']),
                    "last_shipped_date": s_date, "last_shipped_time": s_time
                })

            now_date, now_time = format_timestamp(get_plant_now())
            return {
                "status_summary": status_counts,
                "overdue_orders": overdue,
                "overdue_count": len(overdue),
                "due_this_week": due_this_week,
                "due_this_week_count": len(due_this_week),
                "recent_shipments_7d": recent_shipments,
                "as_of_date": now_date,
                "as_of_time": now_time
            }
    except Exception as e:
        logger.error(f"Sales dashboard failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# WEB DASHBOARD API (no auth — read-only, same-origin)
# ═══════════════════════════════════════════════════════════════

_DASHBOARD_CONFIG_PATH = pathlib.Path(__file__).parent / "dashboard" / "dashboard_config.json"

def _load_dashboard_config():
    try:
        with open(_DASHBOARD_CONFIG_PATH) as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Failed to load dashboard config: {e}")
        return None


_BATCH_PANEL_FAMILIES = ("coconut", "granola", "graham", "chips", "sprinkles")


def _production_family_from_name(name: str) -> str:
    """Classify a batch product into a production family from its name.

    products.product_category is unused (null on every live batch row), so name
    matching is the durable signal. Chocolate-chip granola is granola, not chips.
    """
    n = (name or "").lower()
    if "coconut" in n:
        return "coconut"
    if "graham" in n:
        return "graham"
    if "sprinkle" in n:
        return "sprinkles"
    if "granola" in n:
        return "granola"
    if "chip" in n:
        return "chips"
    return "other"


def _floor_unit_count(qty, size):
    """Whole cases/units from a weight. Floor so available cases are never overstated."""
    if qty is None or not size or float(size) <= 0:
        return None
    return int(math.floor(float(qty) / float(size) + 1e-9))


def _made_unit_size_lbs(base_batch_lb, yield_multiplier):
    """Finished output weight of one physical batch/pan, including yield gain/loss."""
    if not base_batch_lb or float(base_batch_lb) <= 0:
        return None
    y = float(yield_multiplier) if yield_multiplier is not None else 1.0
    if y <= 0:
        y = 1.0
    return float(base_batch_lb) * y


@app.get("/dashboard/api/production")
def dashboard_api_production(
    days: int = Query(default=5, ge=1, le=31),
    month: Optional[str] = Query(default=None)
):
    """Rolling production calendar — batches made + finished goods packed.
    Excludes ship/receive/adjust. Only 'make' and 'pack' transactions."""
    try:
        with get_transaction() as cur:
            # Timestamps are stored in ET via get_plant_now(), so use DATE()
            # directly — no timezone conversion needed.
            if month:
                # Full month view: e.g. month=2026-02
                try:
                    parts = month.split("-")
                    y, m = int(parts[0]), int(parts[1])
                    start_date = f"{y}-{m:02d}-01"
                    last_day = calendar.monthrange(y, m)[1]
                    end_date = f"{y}-{m:02d}-{last_day}"
                except (ValueError, IndexError):
                    raise HTTPException(400, "month must be YYYY-MM format")
                date_filter = "DATE((t.timestamp AT TIME ZONE 'UTC') AT TIME ZONE 'America/New_York') BETWEEN %s AND %s"
                params = [start_date, end_date]
            else:
                now_et = get_plant_now()
                start = (now_et - timedelta(days=days - 1)).strftime("%Y-%m-%d")
                date_filter = "DATE((t.timestamp AT TIME ZONE 'UTC') AT TIME ZONE 'America/New_York') >= %s"
                params = [start]

            cur.execute(f"""
                SELECT DATE((t.timestamp AT TIME ZONE 'UTC') AT TIME ZONE 'America/New_York') as prod_date,
                       t.type as transaction_type,
                       p.name as product_name, p.odoo_code as sku,
                       p.type as product_type, p.pack_format,
                       p.default_batch_lb, p.yield_multiplier, p.case_size_lb,
                       SUM(tl.quantity_lb) FILTER (WHERE tl.quantity_lb > 0) as total_lbs,
                       COUNT(DISTINCT t.id) as txn_count
                FROM ledger_current_transactions t
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN products p ON p.id = tl.product_id
                WHERE t.type IN ('make', 'pack')
                  AND t.effective_status = 'posted'
                  AND tl.quantity_lb > 0
                  AND {date_filter}
                GROUP BY prod_date, p.id, t.type
                ORDER BY prod_date DESC, p.name
            """, params)
            rows = cur.fetchall()

        # Group by day
        days_map = {}
        for r in rows:
            d = str(r['prod_date'])
            if d not in days_map:
                dt = r['prod_date']
                day_name = dt.strftime("%A") if hasattr(dt, 'strftime') else d
                days_map[d] = {"date": d, "day_name": day_name, "batches": [], "finished_goods": []}
            total_lbs = float(r['total_lbs'] or 0)
            entry = {
                "product_name": r['product_name'],
                "sku": r['sku'],
                "total_lbs": total_lbs,
                "product_type": r['product_type'],
                "transaction_type": r['transaction_type'],
                "pack_format": r['pack_format']
            }
            if r['transaction_type'] == 'make':
                batch_size = float(r['default_batch_lb']) if r['default_batch_lb'] else None
                yield_multiplier = float(r['yield_multiplier']) if r['yield_multiplier'] is not None else 1.0
                # Make lines store finished-output weight. One physical pan/batch is the
                # base formula weight after yield gain or loss; for hydrated coconut,
                # that means default_batch_lb * yield_multiplier per finished pan.
                made_unit_size = batch_size * yield_multiplier if batch_size and yield_multiplier > 0 else None
                entry["standard_batch_size_lbs"] = batch_size
                entry["yield_multiplier"] = yield_multiplier
                entry["made_unit_size_lbs"] = made_unit_size
                entry["batch_count"] = round(total_lbs / made_unit_size) if made_unit_size else None
                days_map[d]["batches"].append(entry)
            else:
                cs = float(r['case_size_lb']) if r['case_size_lb'] else None
                entry["case_size_lb"] = cs
                entry["unit_count"] = _floor_unit_count(total_lbs, cs)
                days_map[d]["finished_goods"].append(entry)

        return {"days": list(days_map.values())}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Dashboard production API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/inventory/finished-goods")
def dashboard_api_finished_goods():
    """On-hand inventory for finished goods, grouped by panel with lot breakdown."""
    config = _load_dashboard_config()
    if not config:
        return JSONResponse(status_code=500, content={"error": "Dashboard config not found"})
    try:
        panels = list(config.get("finished_goods_panels", []))
        coconut = config.get("coconut_panel")
        if coconut:
            panels.append(coconut)

        # Collect all SKU names
        all_skus = []
        for panel in panels:
            all_skus.extend(panel.get("skus", []))

        with get_transaction() as cur:
            # Get on-hand per product
            cur.execute(f"""
                SELECT p.id, p.name, COALESCE(p.case_size_lb, p.case_size_lb) as case_size_lb,
                       COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                FROM products p
                LEFT JOIN lots l ON l.product_id = p.id
                  AND COALESCE(l.status, 'active') = 'active'
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE COALESCE(p.active, true) = true
                  AND LOWER(p.name) = ANY(SELECT LOWER(unnest(%s::text[])))
                GROUP BY p.id
            """, (all_skus,))
            product_rows = {r['name'].lower(): dict(r) for r in cur.fetchall()}

            # Get lot breakdown for all matched products
            matched_ids = [r['id'] for r in product_rows.values()]
            lot_map = {}
            if matched_ids:
                cur.execute(f"""
                    SELECT l.product_id, l.lot_code,
                           COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs,
                           l.id as lot_id
                    FROM lots l
                    LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                    WHERE l.product_id = ANY(%s)
                      AND COALESCE(l.status, 'active') = 'active'
                    GROUP BY l.id
                    HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                    ORDER BY COALESCE(l.received_at, l.created_at) ASC
                """, (matched_ids,))
                for lr in cur.fetchall():
                    pid = lr['product_id']
                    if pid not in lot_map:
                        lot_map[pid] = []
                    lot_map[pid].append({
                        "lot_code": lr['lot_code'],
                        "on_hand_lbs": float(lr['on_hand_lbs']),
                        "product_id": pid
                    })

        # Enrich lot rows with unit counts using product case_size_lb
        for pid, lots in lot_map.items():
            prow = next((v for v in product_rows.values() if v['id'] == pid), None)
            cs = float(prow['case_size_lb']) if prow and prow.get('case_size_lb') else None
            for lot in lots:
                lot['case_size_lb'] = cs
                lot['unit_count'] = _floor_unit_count(lot['on_hand_lbs'], cs)

        result_panels = []
        for panel in panels:
            panel_data = {
                "id": panel.get("id", ""),
                "title": panel.get("title", ""),
                "case_weight_lb": panel.get("case_weight_lb"),
                "products": [],
                "missing_skus": []
            }
            for sku in panel.get("skus", []):
                prow = product_rows.get(sku.lower())
                if prow:
                    pid = prow['id']
                    on_hand = float(prow['on_hand_lbs'])
                    case_wt = panel.get("case_weight_lb")
                    if case_wt is None and prow.get('case_size_lb'):
                        case_wt = float(prow['case_size_lb'])
                    product_entry = {
                        "product_name": prow['name'],
                        "on_hand_lbs": on_hand,
                        "case_weight_lb": case_wt,
                        "lots": lot_map.get(pid, [])
                    }
                    panel_data["products"].append(product_entry)
                else:
                    panel_data["missing_skus"].append(sku)
            # Sort by on_hand descending
            panel_data["products"].sort(key=lambda x: x["on_hand_lbs"], reverse=True)
            result_panels.append(panel_data)

        return {"panels": result_panels}
    except Exception as e:
        logger.error(f"Dashboard finished goods API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/inventory/batches")
def dashboard_api_batches():
    """Batch inventory on-hand for all active production families.

    Includes coconut, granola, graham, chips, and sprinkles (classified from
    product name). Estimated batch/pan counts use default_batch_lb *
    yield_multiplier so hydrated coconut is not overstated. On-hand remains a
    posted ledger SUM.
    """
    config = _load_dashboard_config()
    if not config:
        return JSONResponse(status_code=500, content={"error": "Dashboard config not found"})
    try:
        batch_skus = config.get("batch_skus", [])
        sku_names = [b["name"] for b in batch_skus]
        # Optional size override only when the product has no default_batch_lb.
        config_batch_sizes = {b["name"].lower(): b.get("standard_batch_size_lbs") for b in batch_skus}

        with get_transaction() as cur:
            cur.execute(f"""
                SELECT p.id, p.name, p.default_batch_lb, p.yield_multiplier,
                       COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                FROM products p
                LEFT JOIN lots l ON l.product_id = p.id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE COALESCE(p.active, true) = true
                  AND p.type = 'batch'
                GROUP BY p.id
                ORDER BY COALESCE(SUM(tl.quantity_lb), 0) DESC
            """)
            product_rows = [
                r for r in cur.fetchall()
                if _production_family_from_name(r["name"]) in _BATCH_PANEL_FAMILIES
            ]

            matched_ids = [r['id'] for r in product_rows]
            lot_map = {}
            if matched_ids:
                cur.execute(f"""
                    SELECT l.product_id, l.lot_code,
                           COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                    FROM lots l
                    LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                    WHERE l.product_id = ANY(%s)
                    GROUP BY l.id
                    HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                    ORDER BY COALESCE(l.received_at, l.created_at) ASC
                """, (matched_ids,))
                for lr in cur.fetchall():
                    pid = lr['product_id']
                    if pid not in lot_map:
                        lot_map[pid] = []
                    lot_map[pid].append({
                        "lot_code": lr['lot_code'],
                        "on_hand_lbs": float(lr['on_hand_lbs']),
                        "product_id": pid
                    })

        family_rank = {name: i for i, name in enumerate(_BATCH_PANEL_FAMILIES)}
        found_names = {r['name'].lower() for r in product_rows}
        missing_skus = [n for n in sku_names if n.lower() not in found_names]

        batches = []
        made_unit_by_pid = {}
        for r in product_rows:
            on_hand = float(r['on_hand_lbs'])
            batch_size = float(r['default_batch_lb']) if r['default_batch_lb'] else None
            if batch_size is None:
                cfg = config_batch_sizes.get(r['name'].lower())
                if cfg:
                    batch_size = float(cfg)
            yield_multiplier = float(r['yield_multiplier']) if r['yield_multiplier'] is not None else 1.0
            made_unit = _made_unit_size_lbs(batch_size, yield_multiplier)
            family = _production_family_from_name(r['name'])
            made_unit_by_pid[r['id']] = (batch_size, made_unit)
            batches.append({
                "product_name": r['name'],
                "on_hand_lbs": on_hand,
                "standard_batch_size_lbs": batch_size,
                "yield_multiplier": yield_multiplier,
                "made_unit_size_lbs": made_unit,
                "production_family": family,
                "batch_count": round(on_hand / made_unit, 1) if made_unit else None,
                "lots": lot_map.get(r['id'], [])
            })

        for pid, lots in lot_map.items():
            batch_size, made_unit = made_unit_by_pid.get(pid, (None, None))
            for lot in lots:
                lot['default_batch_lb'] = batch_size
                lot['batch_count'] = round(lot['on_hand_lbs'] / made_unit, 1) if made_unit else None

        batches.sort(key=lambda b: (family_rank.get(b["production_family"], 99), -b["on_hand_lbs"], b["product_name"]))
        return {"batches": batches, "missing_skus": missing_skus}
    except Exception as e:
        logger.error(f"Dashboard batches API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/inventory/ingredients")
def dashboard_api_ingredients(category: Optional[str] = Query(default=None)):
    """Ingredient/raw material on-hand grouped by category."""
    config = _load_dashboard_config()
    if not config:
        return JSONResponse(status_code=500, content={"error": "Dashboard config not found"})
    try:
        categories = config.get("ingredient_categories", [])
        if category:
            categories = [c for c in categories if c["id"] == category]

        all_names = []
        for cat in categories:
            all_names.extend(cat.get("items", []))

        with get_transaction() as cur:
            cur.execute(f"""
                SELECT p.id, p.name, p.uom,
                       COALESCE(SUM(tl.quantity_lb), 0) as on_hand
                FROM products p
                LEFT JOIN lots l ON l.product_id = p.id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE COALESCE(p.active, true) = true
                  AND LOWER(p.name) = ANY(SELECT LOWER(unnest(%s::text[])))
                GROUP BY p.id
            """, (all_names,))
            rows = cur.fetchall()
            product_map = {
                r['name'].lower(): {
                    "id": r['id'],
                    "name": r['name'],
                    "on_hand": float(r['on_hand']),
                    "uom": (r['uom'] or None),
                }
                for r in rows
            }

            # Fetch lot-level breakdown for all matched products
            matched_ids = [r['id'] for r in rows]
            lot_map = {}
            if matched_ids:
                cur.execute(f"""
                    SELECT l.product_id, l.lot_code,
                           COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                    FROM lots l
                    LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                    WHERE l.product_id = ANY(%s)
                    GROUP BY l.id
                    HAVING COALESCE(SUM(tl.quantity_lb), 0) > 0
                    ORDER BY COALESCE(l.received_at, l.created_at) ASC
                """, (matched_ids,))
                for lr in cur.fetchall():
                    pid = lr['product_id']
                    if pid not in lot_map:
                        lot_map[pid] = []
                    lot_map[pid].append({
                        "lot_code": lr['lot_code'],
                        "on_hand_lbs": float(lr['on_hand_lbs']),
                        "product_id": pid
                    })

        result = []
        for cat in categories:
            items = []
            missing = []
            for item_name in cat.get("items", []):
                pdata = product_map.get(item_name.lower())
                if pdata:
                    uom = pdata["uom"] or cat.get("unit") or "lb"
                    lots = lot_map.get(pdata["id"], [])
                    for lot in lots:
                        lot["uom"] = uom
                    items.append({
                        "name": pdata["name"],
                        "on_hand": pdata["on_hand"],
                        "uom": uom,
                        "lots": lots
                    })
                else:
                    missing.append(item_name)
            # Preserve config ordering (don't sort by on_hand)
            result.append({
                "id": cat["id"],
                "title": cat["title"],
                "unit": cat.get("unit", "lb"),
                "items": items,
                "missing_skus": missing,
                "total_skus_expected": len(cat.get("items", []))
            })

        return {"categories": result}
    except Exception as e:
        logger.error(f"Dashboard ingredients API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/activity/shipments")
def dashboard_api_shipments(limit: int = Query(default=100, ge=1, le=500)):
    """Shipping log — most recent first."""
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT t.id, t.timestamp, t.created_at, t.created_at_source,
                       t.customer_name, t.order_reference, t.notes,
                       json_agg(json_build_object(
                           'product_name', p.name,
                           'product_id', p.id,
                           'lot_code', l.lot_code,
                           'quantity_lb', tl.quantity_lb,
                           'case_size_lb', p.case_size_lb,
                           'product_type', p.type,
                           'is_service', COALESCE(p.is_service, false)
                       ) ORDER BY p.name) as lines
                FROM transactions t
                JOIN ledger_current_transactions ct ON ct.id = t.id
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN products p ON p.id = tl.product_id
                LEFT JOIN lots l ON l.id = tl.lot_id
                WHERE t.type = 'ship'
                  AND ct.effective_status = 'posted'
                  AND LOWER(COALESCE(t.customer_name, '')) != 'internal packaging'
                GROUP BY t.id
                ORDER BY t.timestamp DESC
                LIMIT %s
            """, (limit,))
            shipments = cur.fetchall()

        result = []
        for s in shipments:
            d, tm = format_timestamp(s['timestamp'])
            created_date, created_time = format_timestamp(s['created_at'])
            total_lbs = 0
            total_units = 0
            enriched_lines = []
            for ln in (s['lines'] or []):
                qty = abs(float(ln['quantity_lb'] or 0))
                total_lbs += qty
                cs = float(ln['case_size_lb']) if ln.get('case_size_lb') else None
                uc = _floor_unit_count(qty, cs)
                if uc is not None:
                    total_units += uc
                ln['unit_count'] = uc
                enriched_lines.append(ln)
            result.append({
                "transaction_id": s['id'],
                "date": d,
                "time": tm,
                "created_at": s['created_at'],
                "created_date": created_date,
                "created_time": created_time,
                "created_at_source": s['created_at_source'],
                "customer_name": s['customer_name'],
                "order_reference": s['order_reference'],
                "total_lbs": total_lbs,
                "total_units": total_units if total_units > 0 else None,
                "lines": enriched_lines,
                "notes": s['notes']
            })
        return {"shipments": result}
    except Exception as e:
        logger.error(f"Dashboard shipments API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/activity/receipts")
def dashboard_api_receipts(limit: int = Query(default=100, ge=1, le=500)):
    """Receiving log — most recent first."""
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT t.id, t.timestamp, t.created_at, t.created_at_source,
                       t.shipper_name, t.bol_reference, t.notes,
                       t.cases_received, t.case_size_lb,
                       json_agg(json_build_object(
                           'product_name', p.name,
                           'product_id', p.id,
                           'lot_code', l.lot_code,
                           'quantity_lb', tl.quantity_lb,
                           'case_size_lb', p.case_size_lb
                       ) ORDER BY p.name) as lines
                FROM transactions t
                JOIN ledger_current_transactions ct ON ct.id = t.id
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN products p ON p.id = tl.product_id
                LEFT JOIN lots l ON l.id = tl.lot_id
                WHERE t.type = 'receive'
                  AND ct.effective_status = 'posted'
                GROUP BY t.id
                ORDER BY t.timestamp DESC
                LIMIT %s
            """, (limit,))
            receipts = cur.fetchall()

        result = []
        for r in receipts:
            d, tm = format_timestamp(r['timestamp'])
            created_date, created_time = format_timestamp(r['created_at'])
            total_lbs = 0
            enriched_lines = []
            for ln in (r['lines'] or []):
                qty = float(ln['quantity_lb'] or 0)
                total_lbs += qty
                cs = float(ln['case_size_lb']) if ln.get('case_size_lb') else None
                ln['unit_count'] = _floor_unit_count(qty, cs)
                enriched_lines.append(ln)
            result.append({
                "transaction_id": r['id'],
                "date": d,
                "time": tm,
                "created_at": r['created_at'],
                "created_date": created_date,
                "created_time": created_time,
                "created_at_source": r['created_at_source'],
                "shipper_name": r['shipper_name'],
                "bol_reference": r['bol_reference'],
                "total_lbs": total_lbs,
                "cases_received": r['cases_received'],
                "case_size_lb": float(r['case_size_lb']) if r['case_size_lb'] else None,
                "lines": enriched_lines,
                "notes": r['notes']
            })
        return {"receipts": result}
    except Exception as e:
        logger.error(f"Dashboard receipts API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/activity/daily-entries")
def dashboard_api_daily_entries(
    target_date: date = Query(..., alias="date", description="Day to list (YYYY-MM-DD)"),
    date_mode: str = Query(default="event", description="'event' filters by business_date; 'entered' filters by the plant-local calendar day of created_at"),
):
    """Every posted ledger entry for one day, with the database entry timestamp.

    For scoring timely data entry: an entry is 'late' when created_at falls on a
    later America/New_York calendar day than the event date (business_date).
    business_date already encodes the plant-day convention (naive UTC timestamp
    -> America/New_York, migration 039), so event-date filtering uses it directly.
    Late flags are only computed when created_at_source='database' — backfilled
    legacy rows (migration_backfill_039 / legacy_unverified) carry the migration
    run time, not the real entry time, and are reported as unreliable instead.
    """
    if date_mode not in ("event", "entered"):
        raise HTTPException(422, "date_mode must be 'event' or 'entered'")
    try:
        with get_transaction() as cur:
            if date_mode == "entered":
                date_filter = "(ct.created_at AT TIME ZONE 'America/New_York')::date = %s"
            else:
                date_filter = "ct.business_date = %s"
            cur.execute(f"""
                SELECT ct.id, ct.type, ct.business_date, ct.occurred_at,
                       ct.created_at, ct.created_at_source, ct.operator_id,
                       json_agg(json_build_object(
                           'product_name', p.name,
                           'product_id', p.id,
                           'sku', p.odoo_code,
                           'lot_code', l.lot_code,
                           'quantity_lb', tl.quantity_lb
                       ) ORDER BY tl.id) AS lines
                FROM ledger_current_transactions ct
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = ct.id
                LEFT JOIN products p ON p.id = tl.product_id
                LEFT JOIN lots l ON l.id = tl.lot_id
                WHERE ct.effective_status = 'posted'
                  AND {date_filter}
                GROUP BY ct.id, ct.type, ct.business_date, ct.occurred_at,
                         ct.created_at, ct.created_at_source, ct.operator_id
                ORDER BY ct.created_at, ct.id
            """, (target_date,))
            rows = cur.fetchall()

        entries = []
        for t in rows:
            d, tm = format_timestamp(t['occurred_at'])
            created_date, created_time = format_timestamp(t['created_at'])
            entry_time_reliable = t['created_at_source'] == 'database'
            created_at = t['created_at']
            if created_at.tzinfo is None:
                created_at = created_at.replace(tzinfo=timezone.utc)
            days_late = (created_at.astimezone(PLANT_TIMEZONE).date() - t['business_date']).days
            late_entry = entry_time_reliable and days_late > 0
            entries.append({
                "transaction_id": t['id'],
                "type": t['type'],
                "event_date": t['business_date'].isoformat(),
                "date": d,
                "time": tm,
                "created_at": t['created_at'],
                "created_date": created_date,
                "created_time": created_time,
                "created_at_source": t['created_at_source'],
                "entry_time_reliable": entry_time_reliable,
                "late_entry": late_entry,
                "days_late": days_late if entry_time_reliable else None,
                "operator_id": t['operator_id'],
                "lines": t['lines'] or [],
            })
        return {
            "date": target_date.isoformat(),
            "date_mode": date_mode,
            "count": len(entries),
            "entries": entries,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Dashboard daily entries API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/lot/{lot_code}")
def dashboard_api_lot_detail(lot_code: str, product_id: Optional[int] = Query(default=None)):
    """Lot detail with full transaction timeline."""
    try:
        with get_transaction() as cur:
            # Lot info — filter by product_id when provided (lot codes can be shared)
            query = f"""
                SELECT l.id, l.lot_code, l.product_id, p.name as product_name,
                       l.entry_source, p.case_size_lb as product_case_size_lb,
                       COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                FROM lots l
                JOIN products p ON p.id = l.product_id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE LOWER(l.lot_code) = LOWER(%s)
            """
            params = [lot_code]
            if product_id is not None:
                query += " AND l.product_id = %s"
                params.append(product_id)
            query += " GROUP BY l.id, p.id"
            cur.execute(query, params)
            rows = cur.fetchall()
            if not rows:
                raise HTTPException(404, f"Lot '{lot_code}' not found")
            if len(rows) > 1:
                return JSONResponse(status_code=409, content={
                    "error": "ambiguous_lot_code",
                    "message": f"Lot code '{lot_code}' matches multiple products",
                    "matches": [{"lot_id": r['id'], "product_id": r['product_id'],
                                 "product_name": r['product_name'], "entry_source": r['entry_source']} for r in rows]
                })
            lot = rows[0]

            # First transaction to get original quantity
            cur.execute("""
                SELECT tl.quantity_lb
                FROM ledger_current_transaction_lines tl
                JOIN ledger_current_transactions t ON t.id = tl.transaction_id
                WHERE tl.lot_id = %s
                  AND t.effective_status = 'posted'
                ORDER BY t.timestamp ASC
                LIMIT 1
            """, (lot['id'],))
            first_txn = cur.fetchone()
            original_qty = float(first_txn['quantity_lb']) if first_txn else 0

            # Full timeline
            cur.execute("""
                SELECT t.id as transaction_id, t.type, t.timestamp,
                       t.created_at, t.created_at_source,
                       tl.quantity_lb,
                       t.customer_name, t.shipper_name, t.order_reference,
                       t.bol_reference, t.adjust_reason, t.notes,
                       t.cases_received, t.case_size_lb
                FROM ledger_current_transaction_lines tl
                JOIN ledger_current_transactions t ON t.id = tl.transaction_id
                WHERE tl.lot_id = %s
                  AND t.effective_status = 'posted'
                ORDER BY t.timestamp ASC
            """, (lot['id'],))
            timeline_rows = cur.fetchall()

        product_case_size = float(lot['product_case_size_lb']) if lot['product_case_size_lb'] else None

        timeline = []
        for tr in timeline_rows:
            d, tm = format_timestamp(tr['timestamp'])
            created_date, created_time = format_timestamp(tr['created_at'])
            qty_lb = float(tr['quantity_lb'])
            # For receives, use actual cases_received; otherwise derive from product case_size_lb
            if tr['cases_received'] is not None:
                cases = int(tr['cases_received'])
            elif product_case_size:
                cases = round(abs(qty_lb) / product_case_size)
            else:
                cases = None
            timeline.append({
                "transaction_id": tr['transaction_id'],
                "type": tr['type'],
                "date": d,
                "time": tm,
                "created_at": tr['created_at'],
                "created_date": created_date,
                "created_time": created_time,
                "created_at_source": tr['created_at_source'],
                "quantity_lb": qty_lb,
                "cases": cases,
                "customer_name": tr['customer_name'],
                "shipper_name": tr['shipper_name'],
                "order_reference": tr['order_reference'],
                "bol_reference": tr['bol_reference'],
                "adjust_reason": tr['adjust_reason'],
                "notes": tr['notes']
            })

        # Derive header case counts
        on_hand_lbs = float(lot['on_hand_lbs'])
        original_cases = round(original_qty / product_case_size) if product_case_size else None
        on_hand_cases = round(on_hand_lbs / product_case_size) if product_case_size else None

        return {
            "lot_code": lot['lot_code'],
            "product_name": lot['product_name'],
            "entry_source": lot['entry_source'],
            "original_quantity_lbs": original_qty,
            "original_cases": original_cases,
            "on_hand_lbs": on_hand_lbs,
            "on_hand_cases": on_hand_cases,
            "case_size_lb": product_case_size,
            "timeline": timeline
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Dashboard lot detail API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/product/{product_id}/lots")
def dashboard_api_product_lots(product_id: int):
    """Get all lots for a product with on-hand quantities."""
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT p.name, p.type, p.odoo_code, p.case_size_lb, p.default_batch_lb
                FROM products p WHERE p.id = %s
            """, (product_id,))
            product = cur.fetchone()
            if not product:
                raise HTTPException(404, f"Product {product_id} not found")

            cs = float(product['case_size_lb']) if product['case_size_lb'] else None
            db = float(product['default_batch_lb']) if product['default_batch_lb'] else None

            cur.execute(f"""
                SELECT l.lot_code, l.entry_source,
                       COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                FROM lots l
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE l.product_id = %s
                GROUP BY l.id
                ORDER BY l.id DESC
            """, (product_id,))
            lots = []
            for r in cur.fetchall():
                oh = float(r["on_hand_lbs"])
                lot_entry = {"lot_code": r["lot_code"], "entry_source": r["entry_source"],
                             "on_hand_lbs": oh}
                if cs:
                    lot_entry["case_size_lb"] = cs
                    lot_entry["unit_count"] = _floor_unit_count(oh, cs)
                elif db:
                    lot_entry["default_batch_lb"] = db
                    lot_entry["batch_count"] = round(oh / db, 1) if db > 0 else None
                lots.append(lot_entry)

        return {
            "product_name": product["name"],
            "product_type": product["type"],
            "odoo_code": product["odoo_code"],
            "case_size_lb": cs,
            "default_batch_lb": db,
            "lots": lots
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Dashboard product lots API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/dashboard/api/search")
def dashboard_api_search(q: str = Query(min_length=1)):
    """Global search across products, lots, orders, and customers."""
    try:
        term = f"%{q}%"
        results = {}
        with get_transaction() as cur:
            # Products
            cur.execute(f"""
                SELECT p.id as product_id, p.name, p.type, p.odoo_code,
                       COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                FROM products p
                LEFT JOIN lots l ON l.product_id = p.id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE COALESCE(p.active, true) = true
                  AND (p.name ILIKE %s OR p.odoo_code ILIKE %s)
                GROUP BY p.id
                ORDER BY p.name
                LIMIT 20
            """, (term, term))
            results["products"] = [dict(r) for r in cur.fetchall()]
            for p in results["products"]:
                p["on_hand_lbs"] = float(p["on_hand_lbs"])

            # Lots
            cur.execute(f"""
                SELECT l.lot_code, l.product_id, p.name as product_name,
                       COALESCE(SUM(tl.quantity_lb), 0) as on_hand_lbs
                FROM lots l
                JOIN products p ON p.id = l.product_id
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                WHERE l.lot_code ILIKE %s
                GROUP BY l.id, p.id
                ORDER BY l.id DESC
                LIMIT 20
            """, (term,))
            results["lots"] = [dict(r) for r in cur.fetchall()]
            for lt in results["lots"]:
                lt["on_hand_lbs"] = float(lt["on_hand_lbs"])

            # Sales orders (also search by customer alias)
            cur.execute("""
                SELECT DISTINCT so.id as order_id, so.order_number, c.name as customer, so.status,
                       so.order_date
                FROM sales_orders so
                JOIN customers c ON c.id = so.customer_id
                LEFT JOIN customer_aliases ca ON ca.customer_id = c.id
                WHERE so.order_number ILIKE %s OR c.name ILIKE %s OR ca.alias ILIKE %s
                ORDER BY so.id DESC
                LIMIT 20
            """, (term, term, term))
            results["orders"] = [
                {**dict(r), "order_date": str(r["order_date"]) if r["order_date"] else None}
                for r in cur.fetchall()
            ]

            # Customers (also search by alias)
            cur.execute("""
                SELECT DISTINCT c.name, c.contact_name, c.email, c.phone
                FROM customers c
                LEFT JOIN customer_aliases ca ON ca.customer_id = c.id
                WHERE (c.name ILIKE %s OR ca.alias ILIKE %s) AND c.active = true
                ORDER BY c.name
                LIMIT 20
            """, (term, term))
            results["customers"] = [dict(r) for r in cur.fetchall()]

        return results
    except Exception as e:
        logger.error(f"Dashboard search API failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# DASHBOARD API — Notes / To-Dos / Reminders (NO AUTH)
# ═══════════════════════════════════════════════════════════════

def _note_row_to_dict(row):
    """Convert a notes DB row to a JSON-safe dict."""
    d = dict(row)
    for key in ("created_at", "updated_at"):
        if d.get(key):
            date_str, time_str = format_timestamp(d[key])
            d[key] = f"{date_str} {time_str}"
    if d.get("due_date"):
        d["due_date"] = str(d["due_date"])
    return d


@app.get("/dashboard/api/notes")
def dashboard_api_notes(
    category: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    entity_type: Optional[str] = Query(default=None),
    entity_id: Optional[str] = Query(default=None),
):
    """List notes/todos/reminders with optional filters. NO AUTH."""
    try:
        with get_transaction() as cur:
            clauses = []
            params = []
            if category:
                clauses.append("category = %s")
                params.append(category)
            if status:
                clauses.append("status = %s")
                params.append(status)
            if entity_type:
                clauses.append("entity_type = %s")
                params.append(entity_type)
            if entity_id:
                clauses.append("entity_id = %s")
                params.append(entity_id)

            where = ""
            if clauses:
                where = "WHERE " + " AND ".join(clauses)

            cur.execute(f"""
                SELECT * FROM notes
                {where}
                ORDER BY
                    CASE WHEN status = 'open' THEN 0 ELSE 1 END,
                    CASE priority WHEN 'high' THEN 0 WHEN 'normal' THEN 1 ELSE 2 END,
                    due_date ASC NULLS LAST,
                    created_at DESC
            """, params)
            rows = [_note_row_to_dict(r) for r in cur.fetchall()]
            return {"notes": rows}
    except Exception as e:
        logger.error(f"Dashboard notes list failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/dashboard/api/notes")
def dashboard_api_notes_create(req: NoteCreate, _: bool = Depends(verify_api_key)):
    """Create a note/todo/reminder."""
    try:
        with get_transaction() as cur:
            due = None
            if req.due_date:
                due = date.fromisoformat(req.due_date)

            cur.execute("""
                INSERT INTO notes (category, title, body, priority, due_date, entity_type, entity_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                RETURNING *
            """, (req.category, req.title, req.body or "", req.priority, due,
                  req.entity_type, req.entity_id))
            row = cur.fetchone()
            return _note_row_to_dict(row)
    except ValueError as e:
        return JSONResponse(status_code=400, content={"error": str(e)})
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Dashboard notes create failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/dashboard/api/notes/{note_id}")
def dashboard_api_notes_update(note_id: int, req: NoteUpdate, _: bool = Depends(verify_api_key)):
    """Update a note/todo/reminder."""
    try:
        with get_transaction() as cur:
            # Build SET clause dynamically from provided fields
            sets = []
            params = []
            data = req.dict(exclude_unset=True)
            if not data:
                return JSONResponse(status_code=400, content={"error": "No fields to update"})

            for field, value in data.items():
                if field == "due_date":
                    if value == "" or value is None:
                        sets.append("due_date = NULL")
                    else:
                        sets.append("due_date = %s")
                        params.append(date.fromisoformat(value))
                else:
                    sets.append(f"{field} = %s")
                    params.append(value)

            sets.append("updated_at = NOW()")
            params.append(note_id)

            cur.execute(f"""
                UPDATE notes SET {', '.join(sets)}
                WHERE id = %s
                RETURNING *
            """, params)
            row = cur.fetchone()
            if not row:
                return JSONResponse(status_code=404, content={"error": "Note not found"})
            return _note_row_to_dict(row)
    except ValueError as e:
        return JSONResponse(status_code=400, content={"error": str(e)})
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Dashboard notes update failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/dashboard/api/notes/{note_id}")
def dashboard_api_notes_delete(note_id: int, _: bool = Depends(verify_api_key)):
    """Delete a note/todo/reminder."""
    try:
        with get_transaction() as cur:
            cur.execute("DELETE FROM notes WHERE id = %s RETURNING id", (note_id,))
            row = cur.fetchone()
            if not row:
                return JSONResponse(status_code=404, content={"error": "Note not found"})
            return {"deleted": True, "id": note_id}
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Dashboard notes delete failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/dashboard/api/notes/{note_id}/toggle")
def dashboard_api_notes_toggle(note_id: int, _: bool = Depends(verify_api_key)):
    """Toggle a note's status between open and done."""
    try:
        with get_transaction() as cur:
            cur.execute("SELECT status FROM notes WHERE id = %s", (note_id,))
            row = cur.fetchone()
            if not row:
                return JSONResponse(status_code=404, content={"error": "Note not found"})
            new_status = "done" if row["status"] == "open" else "open"
            cur.execute("""
                UPDATE notes SET status = %s, updated_at = NOW()
                WHERE id = %s RETURNING *
            """, (new_status, note_id))
            return _note_row_to_dict(cur.fetchone())
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Dashboard notes toggle failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# ADMIN PRODUCT & BOM MANAGEMENT
# ═══════════════════════════════════════════════════════════════

class ProductUpdate(BaseModel):
    odoo_code: Optional[str] = None
    case_size_lb: Optional[float] = None
    default_batch_lb: Optional[float] = None
    yield_multiplier: Optional[float] = None
    active: Optional[bool] = None


@app.put("/admin/products/{product_id}")
def admin_update_product(product_id: int, req: ProductUpdate, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT id, name FROM products WHERE id = %s", (product_id,))
            product = cur.fetchone()
            if not product:
                raise HTTPException(status_code=404, detail=f"Product ID {product_id} not found")

            updates = []
            params = []
            if req.odoo_code is not None:
                updates.append("odoo_code = %s")
                params.append(req.odoo_code)
            if req.case_size_lb is not None:
                updates.append("case_size_lb = %s")
                params.append(req.case_size_lb)
            if req.default_batch_lb is not None:
                updates.append("default_batch_lb = %s")
                params.append(req.default_batch_lb)
            if req.yield_multiplier is not None:
                updates.append("yield_multiplier = %s")
                params.append(req.yield_multiplier)
            if req.active is not None:
                updates.append("active = %s")
                params.append(req.active)

            if not updates:
                return {"updated": False, "message": "No fields to update"}

            params.append(product_id)
            cur.execute(f"UPDATE products SET {', '.join(updates)} WHERE id = %s", params)

        return {
            "updated": True,
            "product_id": product_id,
            "product_name": product['name'],
            "changes": {
                k: v for k, v in {
                    "odoo_code": req.odoo_code,
                    "case_size_lb": req.case_size_lb,
                    "default_batch_lb": req.default_batch_lb,
                    "yield_multiplier": req.yield_multiplier,
                    "active": req.active
                }.items() if v is not None
            }
        }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Admin product update failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


class BomLineCreate(BaseModel):
    ingredient_product_id: int
    quantity_lb: float
    exclude_from_inventory: Optional[bool] = False

class BomLineUpdate(BaseModel):
    quantity_lb: Optional[float] = None
    exclude_from_inventory: Optional[bool] = None


@app.get("/admin/bom/search")
def admin_bom_search(
    product_name: str = Query(...),
    _: bool = Depends(verify_api_key)
):
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT id, name, odoo_code, type, default_batch_lb
                FROM products
                WHERE COALESCE(active, true) = true
                  AND LOWER(name) LIKE LOWER(%s)
                ORDER BY name
                LIMIT 20
            """, (f"%{product_name}%",))
            products = cur.fetchall()
        return {"count": len(products), "products": products}
    except Exception as e:
        logger.error(f"Admin BOM search failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/admin/bom/{product_id}/lines")
def admin_bom_lines(product_id: int, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT id, name FROM products WHERE id = %s", (product_id,))
            product = cur.fetchone()
            if not product:
                raise HTTPException(status_code=404, detail=f"Product ID {product_id} not found")

            cur.execute("""
                SELECT bf.id AS line_id, bf.ingredient_product_id, p.name AS ingredient_name,
                       bf.quantity_lb, COALESCE(bf.exclude_from_inventory, false) AS exclude_from_inventory
                FROM batch_formulas bf
                JOIN products p ON p.id = bf.ingredient_product_id
                WHERE bf.product_id = %s
                ORDER BY bf.quantity_lb DESC
            """, (product_id,))
            lines = cur.fetchall()

        return {
            "product_id": product['id'],
            "product_name": product['name'],
            "line_count": len(lines),
            "lines": lines
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Admin BOM lines failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/admin/bom/{product_id}/lines")
def admin_bom_add_line(product_id: int, req: BomLineCreate, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("SELECT id, name FROM products WHERE id = %s", (product_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail=f"Product ID {product_id} not found")

            cur.execute("SELECT id, name FROM products WHERE id = %s", (req.ingredient_product_id,))
            ingredient = cur.fetchone()
            if not ingredient:
                raise HTTPException(status_code=404, detail=f"Ingredient product ID {req.ingredient_product_id} not found")

            cur.execute("""
                INSERT INTO batch_formulas (product_id, ingredient_product_id, quantity_lb, exclude_from_inventory)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (product_id, req.ingredient_product_id, req.quantity_lb, req.exclude_from_inventory))
            new_line = cur.fetchone()

        return {
            "created": True,
            "line_id": new_line['id'],
            "ingredient_name": ingredient['name'],
            "quantity_lb": req.quantity_lb,
            "exclude_from_inventory": req.exclude_from_inventory
        }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Admin BOM add line failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.put("/admin/bom/lines/{line_id}")
def admin_bom_update_line(line_id: int, req: BomLineUpdate, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT bf.id, bf.quantity_lb, COALESCE(bf.exclude_from_inventory, false) AS exclude_from_inventory,
                       p.name AS ingredient_name
                FROM batch_formulas bf
                JOIN products p ON p.id = bf.ingredient_product_id
                WHERE bf.id = %s
            """, (line_id,))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail=f"BOM line ID {line_id} not found")

            updates = []
            params = []
            if req.quantity_lb is not None:
                updates.append("quantity_lb = %s")
                params.append(req.quantity_lb)
            if req.exclude_from_inventory is not None:
                updates.append("exclude_from_inventory = %s")
                params.append(req.exclude_from_inventory)

            if not updates:
                return {"updated": False, "message": "No fields to update"}

            params.append(line_id)
            cur.execute(f"UPDATE batch_formulas SET {', '.join(updates)} WHERE id = %s", params)

        return {
            "updated": True,
            "line_id": line_id,
            "ingredient_name": existing['ingredient_name'],
            "previous_quantity_lb": float(existing['quantity_lb']),
            "new_quantity_lb": req.quantity_lb if req.quantity_lb is not None else float(existing['quantity_lb']),
            "exclude_from_inventory": req.exclude_from_inventory if req.exclude_from_inventory is not None else existing['exclude_from_inventory']
        }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Admin BOM update line failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/admin/bom/lines/{line_id}")
def admin_bom_delete_line(line_id: int, _: bool = Depends(verify_api_key)):
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT bf.id, bf.product_id, bf.ingredient_product_id, bf.quantity_lb,
                       p.name AS ingredient_name, prod.name AS product_name
                FROM batch_formulas bf
                JOIN products p ON p.id = bf.ingredient_product_id
                JOIN products prod ON prod.id = bf.product_id
                WHERE bf.id = %s
            """, (line_id,))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail=f"BOM line ID {line_id} not found")

            cur.execute("DELETE FROM batch_formulas WHERE id = %s", (line_id,))

        return {
            "deleted": True,
            "line_id": line_id,
            "product_name": existing['product_name'],
            "ingredient_name": existing['ingredient_name'],
            "quantity_lb": float(existing['quantity_lb'])
        }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Admin BOM delete line failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# ADMIN: FG → BATCH PRODUCT MAPPING (product_bom)
# ═══════════════════════════════════════════════════════════════

class ProductBomCreate(BaseModel):
    finished_product_id: int
    component_product_id: int
    quantity: Optional[float] = 1.0
    uom: Optional[str] = "unit"


@app.get("/admin/product-bom")
def admin_list_product_bom(
    fg_only: bool = Query(False, description="Only show batch product components"),
    _: bool = Depends(verify_api_key)
):
    """List all FG → component mappings from product_bom."""
    try:
        with get_transaction() as cur:
            query = """
                SELECT pb.id, pb.finished_product_id, fg.name AS finished_good_name,
                       pb.component_product_id, cp.name AS component_name, cp.type AS component_type,
                       pb.quantity, pb.uom
                FROM product_bom pb
                JOIN products fg ON fg.id = pb.finished_product_id
                JOIN products cp ON cp.id = pb.component_product_id
            """
            if fg_only:
                query += " WHERE cp.type = 'batch'"
            query += " ORDER BY fg.name, cp.type"
            cur.execute(query)
            rows = cur.fetchall()
        return {"count": len(rows), "mappings": [dict(r) for r in rows]}
    except Exception as e:
        logger.error(f"Admin product-bom list failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/admin/product-bom")
def admin_create_product_bom(req: ProductBomCreate, _: bool = Depends(verify_api_key)):
    """Add a component to a finished good's product_bom."""
    try:
        with get_transaction() as cur:
            cur.execute("SELECT id, name FROM products WHERE id = %s", (req.finished_product_id,))
            fg = cur.fetchone()
            if not fg:
                raise HTTPException(404, f"Finished good ID {req.finished_product_id} not found")

            cur.execute("SELECT id, name FROM products WHERE id = %s", (req.component_product_id,))
            cp = cur.fetchone()
            if not cp:
                raise HTTPException(404, f"Component product ID {req.component_product_id} not found")

            cur.execute("""
                INSERT INTO product_bom (finished_product_id, component_product_id, quantity, uom)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (req.finished_product_id, req.component_product_id, req.quantity, req.uom))
            row = cur.fetchone()

        return {"created": True, "id": row['id'], "finished_good": fg['name'], "component": cp['name'],
                "quantity": req.quantity, "uom": req.uom}
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Admin product-bom create failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.delete("/admin/product-bom/{mapping_id}")
def admin_delete_product_bom(mapping_id: int, _: bool = Depends(verify_api_key)):
    """Remove a component from a finished good's product_bom."""
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT pb.id, fg.name AS finished_good_name, cp.name AS component_name
                FROM product_bom pb
                JOIN products fg ON fg.id = pb.finished_product_id
                JOIN products cp ON cp.id = pb.component_product_id
                WHERE pb.id = %s
            """, (mapping_id,))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(404, f"Mapping ID {mapping_id} not found")

            cur.execute("DELETE FROM product_bom WHERE id = %s", (mapping_id,))

        return {"deleted": True, "finished_good": existing['finished_good_name'], "component": existing['component_name']}
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Admin product-bom delete failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# LOT TRACEABILITY — Duplicate Scanner & Merge
# ═══════════════════════════════════════════════════════════════

@app.get("/admin/lots/duplicates")
def scan_lot_duplicates(_: bool = Depends(verify_api_key)):
    """Scan for duplicate (product_id, lot_code) pairs across the lots table.
    Returns grouped results for review before merging."""
    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT p.name AS product_name, p.id AS product_id, l.lot_code,
                       COUNT(*) AS duplicate_count,
                       ARRAY_AGG(l.id ORDER BY l.created_at) AS lot_ids,
                       ARRAY_AGG(l.created_at ORDER BY l.created_at) AS created_dates
                FROM lots l
                JOIN products p ON p.id = l.product_id
                WHERE l.lot_code IS NOT NULL
                  AND COALESCE(l.status, 'active') = 'active'
                GROUP BY p.id, p.name, l.lot_code
                HAVING COUNT(*) > 1
                ORDER BY COUNT(*) DESC
            """)
            rows = cur.fetchall()

            groups = []
            for r in rows:
                groups.append({
                    "product_name": r['product_name'],
                    "product_id": r['product_id'],
                    "lot_code": r['lot_code'],
                    "duplicate_count": r['duplicate_count'],
                    "lot_ids": r['lot_ids'],
                    "created_dates": [str(d) for d in r['created_dates']]
                })

            return {
                "duplicate_groups": groups,
                "total_groups": len(groups)
            }
    except Exception as e:
        logger.error(f"Scan lot duplicates failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


class LotMergeRequest(BaseModel):
    source_lot_id: int
    target_lot_id: int
    reason: str


@app.post("/admin/lots/merge")
def merge_lots(req: LotMergeRequest, _: bool = Depends(verify_api_key)):
    """Merge source lot into target lot. Moves all transaction_lines and
    ingredient_lot_consumption references, marks source as merged.
    Both lots must belong to the same product."""
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                # 1. Validate both lots exist
                cur.execute("SELECT id, product_id, lot_code, status FROM lots WHERE id = %s", (req.source_lot_id,))
                source = cur.fetchone()
                if not source:
                    raise HTTPException(404, f"Source lot ID {req.source_lot_id} not found")

                cur.execute("SELECT id, product_id, lot_code, status FROM lots WHERE id = %s", (req.target_lot_id,))
                target = cur.fetchone()
                if not target:
                    raise HTTPException(404, f"Target lot ID {req.target_lot_id} not found")

                # 2. Validate neither is already merged
                if source.get('status') == 'merged':
                    raise HTTPException(400,
                        f"Source lot {source['lot_code']} (id={req.source_lot_id}) is already merged. "
                        f"Cannot merge an already-merged lot."
                    )
                if target.get('status') == 'merged':
                    raise HTTPException(400,
                        f"Target lot {target['lot_code']} (id={req.target_lot_id}) is already merged. "
                        f"Cannot merge into an already-merged lot."
                    )

                # 3. Validate same product
                if source['product_id'] != target['product_id']:
                    raise HTTPException(400,
                        f"Cannot merge lots from different products. "
                        f"Source lot {source['lot_code']} is product_id={source['product_id']}, "
                        f"target lot {target['lot_code']} is product_id={target['product_id']}."
                    )

                # 4. Lock every lot and active allocation for the product in
                # the allocation protocol's deterministic order, then coalesce
                # source pins onto the survivor before ledger references move.
                allocation_moves = _coalesce_lot_allocations(
                    cur,
                    int(source['product_id']),
                    req.source_lot_id,
                    req.target_lot_id,
                )

                rows_moved = {}

                # 5. Move transaction lines through append-only effective-state
                # corrections; raw historical line rows never change.
                cur.execute(
                    "SELECT id FROM ledger_current_transaction_lines WHERE lot_id = %s ORDER BY id",
                    (req.source_lot_id,),
                )
                line_ids = [row["id"] for row in cur.fetchall()]
                line_correction_ids = [
                    _append_transaction_line_correction(
                        cur,
                        line_id,
                        {"lot_id": req.target_lot_id},
                        req.reason,
                        _operator_id(_),
                    )
                    for line_id in line_ids
                ]
                rows_moved["transaction_lines"] = len(line_correction_ids)

                # Move ingredient_lot_consumption
                cur.execute(
                    "UPDATE ingredient_lot_consumption SET ingredient_lot_id = %s WHERE ingredient_lot_id = %s",
                    (req.target_lot_id, req.source_lot_id)
                )
                rows_moved["ingredient_lot_consumption"] = cur.rowcount

                # 6. Mark source lot as merged
                now = get_plant_now()
                cur.execute("""
                    UPDATE lots
                    SET status = 'merged',
                        merged_into_lot_id = %s,
                        merged_at = %s,
                        merge_reason = %s
                    WHERE id = %s
                """, (req.target_lot_id, now, req.reason, req.source_lot_id))

                # 7. Recalculate target lot balance from ledger (posted-only)
                computed_balance = lot_on_hand(cur, req.target_lot_id)

                total_rows = sum(rows_moved.values())
                logger.info(
                    f"Lot merge: {source['lot_code']} (id={req.source_lot_id}) → "
                    f"{target['lot_code']} (id={req.target_lot_id}). "
                    f"Moved {total_rows} rows. Balance: {computed_balance} lb. "
                    f"Reason: {req.reason}"
                )

                return {
                    "merged": True,
                    "source_lot_id": req.source_lot_id,
                    "source_lot_code": source['lot_code'],
                    "target_lot_id": req.target_lot_id,
                    "target_lot_code": target['lot_code'],
                    "product_id": source['product_id'],
                    "rows_moved": rows_moved,
                    "line_correction_ids": line_correction_ids,
                    "allocation_moves": allocation_moves,
                    "target_lot_new_balance": computed_balance,
                    "audit_note": req.reason
                }
    except HTTPException:
        raise
    except Exception as e:
        if _is_readonly_error(e): raise
        logger.error(f"Lot merge failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# PRODUCTION REQUIREMENTS
# ═══════════════════════════════════════════════════════════════

@app.get("/production/requirements")
def production_requirements(
    product_name: str = Query(..., description="Finished good or batch product name"),
    cases: Optional[int] = Query(None, description="Number of cases (for finished goods)"),
    batches: Optional[int] = Query(None, description="Number of batches (for batch products)"),
    _: bool = Depends(verify_api_key)
):
    """Given a finished good + cases OR batch product + batches, return the full ingredient breakdown."""
    try:
        with get_transaction() as cur:
            product = resolve_product_full(cur, product_name)
            pid = product['id']
            pname = product['name']

            # Determine if this is a finished good or batch product
            cur.execute("SELECT type FROM products WHERE id = %s", (pid,))
            ptype = cur.fetchone()['type']

            batch_product_id = None
            batch_product_name = None
            num_batches = batches
            total_output_lb = None

            if ptype == 'finished':
                if not cases:
                    raise HTTPException(400, "cases parameter required for finished goods")

                # Look up the batch product from product_bom
                cur.execute("""
                    SELECT pb.component_product_id AS batch_product_id, p.name AS batch_name, p.default_batch_lb
                    FROM product_bom pb
                    JOIN products p ON p.id = pb.component_product_id
                    WHERE pb.finished_product_id = %s AND p.type = 'batch'
                """, (pid,))
                link = cur.fetchone()

                if not link:
                    raise HTTPException(404, f"No batch product linked to '{pname}'. Add a product_bom mapping.")

                batch_product_id = link['batch_product_id']
                batch_product_name = link['batch_name']
                batch_size = float(link['default_batch_lb'] or 0)

                # Calculate how many lbs needed
                case_weight = float(product.get('case_size_lb') or 0)
                if case_weight <= 0:
                    raise HTTPException(400, f"No case_size_lb set for '{pname}'")

                total_output_lb = cases * case_weight
                if batch_size > 0:
                    import math
                    num_batches = math.ceil(total_output_lb / batch_size)
                else:
                    raise HTTPException(400, f"No default_batch_lb set for batch product '{batch_product_name}'")

            elif ptype == 'batch':
                batch_product_id = pid
                batch_product_name = pname
                if not num_batches:
                    num_batches = 1
                batch_size = float(product.get('default_batch_lb') or 0)
                total_output_lb = num_batches * batch_size
            else:
                raise HTTPException(400, f"Product '{pname}' is type '{ptype}', expected 'finished' or 'batch'")

            # Get the BOM for the batch product
            cur.execute("""
                SELECT bf.ingredient_product_id, p.name AS ingredient_name, bf.quantity_lb,
                       COALESCE(bf.exclude_from_inventory, false) AS exclude_from_inventory
                FROM batch_formulas bf
                JOIN products p ON p.id = bf.ingredient_product_id
                WHERE bf.product_id = %s
                ORDER BY bf.quantity_lb DESC
            """, (batch_product_id,))
            formula = cur.fetchall()

            if not formula:
                raise HTTPException(404, f"No BOM found for batch product '{batch_product_name}'")

            # Check if any ingredient is itself a batch product (nested BOM, e.g. PB Banana)
            ingredients = []
            for ing in formula:
                needed = float(ing['quantity_lb']) * num_batches
                excluded = ing['exclude_from_inventory']

                # Check if this ingredient has its own BOM (nested)
                cur.execute("SELECT COUNT(*) AS cnt FROM batch_formulas WHERE product_id = %s", (ing['ingredient_product_id'],))
                has_sub_bom = cur.fetchone()['cnt'] > 0

                # Get current inventory
                cur.execute(f"""
                    SELECT COALESCE(SUM(tl.quantity_lb), 0) AS available
                    FROM lots l
                    LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                    WHERE l.product_id = %s
                """, (ing['ingredient_product_id'],))
                available = float(cur.fetchone()['available'])

                ing_data = {
                    "ingredient_id": ing['ingredient_product_id'],
                    "ingredient_name": ing['ingredient_name'],
                    "per_batch_lb": float(ing['quantity_lb']),
                    "total_needed_lb": needed,
                    "available_lb": available,
                    "sufficient": available >= needed or excluded,
                    "excluded": excluded
                }

                if has_sub_bom:
                    # Expand the sub-BOM
                    sub_batches_needed = needed  # lbs needed of this sub-batch
                    cur.execute("SELECT default_batch_lb FROM products WHERE id = %s", (ing['ingredient_product_id'],))
                    sub_batch_size = float(cur.fetchone()['default_batch_lb'] or 0)
                    if sub_batch_size > 0:
                        import math
                        sub_num_batches = math.ceil(sub_batches_needed / sub_batch_size)
                    else:
                        sub_num_batches = 1

                    cur.execute("""
                        SELECT bf.ingredient_product_id, p.name AS ingredient_name, bf.quantity_lb,
                               COALESCE(bf.exclude_from_inventory, false) AS exclude_from_inventory
                        FROM batch_formulas bf
                        JOIN products p ON p.id = bf.ingredient_product_id
                        WHERE bf.product_id = %s
                        ORDER BY bf.quantity_lb DESC
                    """, (ing['ingredient_product_id'],))
                    sub_formula = cur.fetchall()

                    sub_ingredients = []
                    for sub in sub_formula:
                        sub_needed = float(sub['quantity_lb']) * sub_num_batches
                        cur.execute(f"""
                            SELECT COALESCE(SUM(tl.quantity_lb), 0) AS available
                            FROM lots l LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                            WHERE l.product_id = %s
                        """, (sub['ingredient_product_id'],))
                        sub_avail = float(cur.fetchone()['available'])
                        sub_excluded = sub['exclude_from_inventory']
                        sub_ingredients.append({
                            "ingredient_id": sub['ingredient_product_id'],
                            "ingredient_name": sub['ingredient_name'],
                            "per_batch_lb": float(sub['quantity_lb']),
                            "total_needed_lb": sub_needed,
                            "available_lb": sub_avail,
                            "sufficient": sub_avail >= sub_needed or sub_excluded,
                            "excluded": sub_excluded
                        })

                    ing_data["is_sub_batch"] = True
                    ing_data["sub_batches_needed"] = sub_num_batches
                    ing_data["sub_ingredients"] = sub_ingredients

                ingredients.append(ing_data)

            all_sufficient = all(
                i['sufficient'] and all(s['sufficient'] for s in i.get('sub_ingredients', []))
                for i in ingredients
            )

            result = {
                "product_name": pname,
                "product_type": ptype,
                "batch_product": batch_product_name,
                "batches_needed": num_batches,
                "total_output_lb": total_output_lb,
                "all_ingredients_sufficient": all_sufficient,
                "ingredients": ingredients
            }

            if ptype == 'finished':
                result["cases"] = cases
                result["case_weight_lb"] = case_weight

            return result

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Production requirements failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# TODAY SO FAR — DASHBOARD PRODUCTION TILE
# ═══════════════════════════════════════════════════════════════

@app.get("/production/today-tile")
def production_today_tile(
    date: Optional[str] = Query(None, description="Plant date in YYYY-MM-DD format; defaults to today"),
    _: bool = Depends(verify_api_key)
):
    """Canonical, posted-ledger production metrics for the dashboard Today So Far tile.

    Counts intentionally retain full precision.  Retail pack transactions do
    not currently carry a sellable bag/unit count, so the response exposes the
    honest ``granola_retail_cases`` fallback rather than inventing a bag count.
    """
    if date:
        try:
            target_date = datetime.strptime(date, "%Y-%m-%d").date()
        except ValueError:
            raise HTTPException(400, "date must be YYYY-MM-DD format")
    else:
        target_date = get_plant_now().date()

    day_start = datetime(target_date.year, target_date.month, target_date.day,
                         tzinfo=PLANT_TIMEZONE)
    day_end = day_start + timedelta(days=1)
    made = {
        "granola_batches": 0.0, "coconut_pans": 0.0, "graham_batches": 0.0,
        "other_batches": 0.0, "other_products": [],
    }
    packed = {
        "granola_bulk_10lb_cases": 0.0, "granola_bulk_25lb_cases": 0.0,
        "granola_retail_cases": 0.0, "granola_retail_bag_count_available": False,
        "coconut_cases": 0.0, "graham_cases": 0.0,
        "other_cases": 0.0, "other_products": [],
    }

    try:
        with get_transaction() as cur:
            cur.execute("""
                SELECT t.type AS transaction_type, p.name AS product_name,
                       p.default_batch_lb, p.yield_multiplier,
                       p.case_size_lb, p.pack_format,
                       SUM(tl.quantity_lb) AS total_lb
                FROM ledger_current_transactions t
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN products p ON p.id = tl.product_id
                WHERE t.type IN ('make', 'pack')
                  AND t.effective_status = 'posted'
                  AND tl.quantity_lb > 0
                  AND t.timestamp >= %s AND t.timestamp < %s
                GROUP BY t.type, p.id, p.name, p.default_batch_lb,
                         p.yield_multiplier, p.case_size_lb, p.pack_format
                ORDER BY t.type, p.name
            """, (day_start, day_end))
            rows = cur.fetchall()

        for row in rows:
            product_name = row["product_name"]
            family = _production_family_from_name(product_name)
            total_lb = float(row["total_lb"] or 0)
            if row["transaction_type"] == "make":
                unit_size = _made_unit_size_lbs(row["default_batch_lb"], row["yield_multiplier"])
                # A missing unit definition must remain visible under Other,
                # regardless of its classified family; its count is unavailable.
                if unit_size is None:
                    made["other_products"].append({
                        "product_name": product_name, "batches": 0.0,
                        "output_lb": total_lb, "count_available": False,
                    })
                    continue
                count = total_lb / unit_size
                if family == "granola":
                    made["granola_batches"] += count
                elif family == "coconut":
                    made["coconut_pans"] += count
                elif family == "graham":
                    made["graham_batches"] += count
                else:
                    made["other_batches"] += count
                    made["other_products"].append({
                        "product_name": product_name, "batches": count,
                        "output_lb": total_lb,
                        "count_available": unit_size is not None,
                    })
                continue

            case_size = float(row["case_size_lb"] or 0)
            # As with made activity, do not hide a classified product whose
            # ledger record lacks the unit definition needed for a count.
            if case_size <= 0:
                packed["other_products"].append({
                    "product_name": product_name, "cases": 0.0,
                    "output_lb": total_lb, "case_size_lb": None,
                    "count_available": False,
                })
                continue
            cases = total_lb / case_size
            pack_format = row["pack_format"]
            if family == "granola" and pack_format == "10lb":
                packed["granola_bulk_10lb_cases"] += cases
            elif family == "granola" and pack_format == "25lb":
                packed["granola_bulk_25lb_cases"] += cases
            elif family == "granola" and pack_format == "bagged":
                packed["granola_retail_cases"] += cases
            elif family == "coconut":
                packed["coconut_cases"] += cases
            elif family == "graham":
                packed["graham_cases"] += cases
            else:
                packed["other_cases"] += cases
                packed["other_products"].append({
                    "product_name": product_name, "cases": cases,
                    "output_lb": total_lb, "case_size_lb": case_size or None,
                    "count_available": case_size > 0,
                })

        return {"date": str(target_date), "made": made, "packed": packed}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Production today-tile failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# PRODUCTION DAY SUMMARY
# ═══════════════════════════════════════════════════════════════

@app.get("/production/day-summary")
def production_day_summary(
    date: Optional[str] = Query(None, description="Date in YYYY-MM-DD format; defaults to today"),
    _: bool = Depends(verify_api_key)
):
    """Return all make/pack/adjust activity for a given day, grouped by product with lot-level detail."""
    try:
        with get_transaction() as cur:
            if date:
                try:
                    target_date = datetime.strptime(date, "%Y-%m-%d").date()
                except ValueError:
                    raise HTTPException(400, "date must be YYYY-MM-DD format")
            else:
                target_date = get_plant_now().date()

            day_start = datetime(target_date.year, target_date.month, target_date.day,
                                 tzinfo=PLANT_TIMEZONE)
            day_end = day_start + timedelta(days=1)

            # ── Batch production (make transactions) ──
            cur.execute("""
                SELECT p.id as product_id, p.name as product_name,
                       l.id as lot_id, l.lot_code,
                       tl.quantity_lb, t.id as transaction_id
                FROM ledger_current_transactions t
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN products p ON p.id = tl.product_id
                JOIN lots l ON l.id = tl.lot_id
                WHERE t.type = 'make'
                  AND t.effective_status = 'posted'
                  AND t.timestamp >= %s AND t.timestamp < %s
                  AND tl.quantity_lb > 0
                ORDER BY t.timestamp
            """, (day_start, day_end))
            make_rows = cur.fetchall()

            # ── Pack consumption from batch lots (negative lines on batch products) ──
            cur.execute("""
                SELECT l.id as lot_id, l.lot_code,
                       ABS(tl.quantity_lb) as packed_lb,
                       t.id as transaction_id
                FROM ledger_current_transactions t
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN lots l ON l.id = tl.lot_id
                WHERE t.type = 'pack'
                  AND t.effective_status = 'posted'
                  AND t.timestamp >= %s AND t.timestamp < %s
                  AND tl.quantity_lb < 0
                ORDER BY t.timestamp
            """, (day_start, day_end))
            pack_consume_rows = cur.fetchall()

            # ── Pack output (finished goods produced) ──
            cur.execute("""
                SELECT p.id as product_id, p.name as product_name,
                       l.lot_code, tl.quantity_lb,
                       t.notes
                FROM ledger_current_transactions t
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN products p ON p.id = tl.product_id
                JOIN lots l ON l.id = tl.lot_id
                WHERE t.type = 'pack'
                  AND t.effective_status = 'posted'
                  AND t.timestamp >= %s AND t.timestamp < %s
                  AND tl.quantity_lb > 0
                ORDER BY t.timestamp
            """, (day_start, day_end))
            pack_output_rows = cur.fetchall()

            # ── Adjustments for the day ──
            cur.execute("""
                SELECT p.name as product_name, l.id as lot_id, l.lot_code,
                       tl.quantity_lb as adjustment_lb,
                       t.adjust_reason as reason
                FROM ledger_current_transactions t
                JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                JOIN products p ON p.id = tl.product_id
                JOIN lots l ON l.id = tl.lot_id
                WHERE t.type = 'adjust'
                  AND t.effective_status = 'posted'
                  AND t.timestamp >= %s AND t.timestamp < %s
                ORDER BY t.timestamp
            """, (day_start, day_end))
            adjust_rows = cur.fetchall()

            # ── Build batch lot summaries ──
            # Collect all lot_ids touched by make today
            batch_lots = {}  # lot_id -> {lot_code, product_name, produced_lb, packed_lb, adjusted_lb}
            for r in make_rows:
                lid = r['lot_id']
                if lid not in batch_lots:
                    batch_lots[lid] = {
                        "lot_code": r['lot_code'],
                        "product_id": r['product_id'],
                        "product_name": r['product_name'],
                        "produced_lb": 0.0,
                        "packed_lb": 0.0,
                        "adjusted_lb": 0.0,
                    }
                batch_lots[lid]["produced_lb"] += float(r['quantity_lb'])

            # Add pack consumption (include lots produced on previous days)
            for r in pack_consume_rows:
                lid = r['lot_id']
                if lid not in batch_lots:
                    cur.execute("""
                        SELECT l.lot_code, l.product_id, p.name as product_name
                        FROM lots l JOIN products p ON p.id = l.product_id
                        WHERE l.id = %s
                    """, (lid,))
                    lot_info = cur.fetchone()
                    if lot_info:
                        batch_lots[lid] = {
                            "lot_code": lot_info['lot_code'],
                            "product_id": lot_info['product_id'],
                            "product_name": lot_info['product_name'],
                            "produced_lb": 0.0,
                            "packed_lb": 0.0,
                            "adjusted_lb": 0.0,
                        }
                if lid in batch_lots:
                    batch_lots[lid]["packed_lb"] += float(r['packed_lb'])

            # Add adjustments (include lots produced on previous days)
            for r in adjust_rows:
                lid = r['lot_id']
                if lid not in batch_lots:
                    cur.execute("""
                        SELECT l.lot_code, l.product_id, p.name as product_name
                        FROM lots l JOIN products p ON p.id = l.product_id
                        WHERE l.id = %s
                    """, (lid,))
                    lot_info = cur.fetchone()
                    if lot_info:
                        batch_lots[lid] = {
                            "lot_code": lot_info['lot_code'],
                            "product_id": lot_info['product_id'],
                            "product_name": lot_info['product_name'],
                            "produced_lb": 0.0,
                            "packed_lb": 0.0,
                            "adjusted_lb": 0.0,
                        }
                if lid in batch_lots:
                    batch_lots[lid]["adjusted_lb"] += float(r['adjustment_lb'])

            # Get current on-hand for each batch lot
            if batch_lots:
                lot_ids = list(batch_lots.keys())
                cur.execute(f"""
                    SELECT l.id as lot_id, COALESCE(SUM(tl.quantity_lb), 0) as on_hand
                    FROM lots l
                    LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                    WHERE l.id = ANY(%s)
                    GROUP BY l.id
                """, (lot_ids,))
                for row in cur.fetchall():
                    if row['lot_id'] in batch_lots:
                        batch_lots[row['lot_id']]["current_on_hand_lb"] = float(row['on_hand'])

            # Group batch lots by product
            products_map = {}
            for lid, info in batch_lots.items():
                pid = info["product_id"]
                if pid not in products_map:
                    products_map[pid] = {
                        "product_id": pid,
                        "product_name": info["product_name"],
                        "total_produced_lb": 0.0,
                        "total_packed_lb": 0.0,
                        "lots": []
                    }
                products_map[pid]["total_produced_lb"] += info["produced_lb"]
                products_map[pid]["total_packed_lb"] += info["packed_lb"]
                products_map[pid]["lots"].append({
                    "lot_code": info["lot_code"],
                    "produced_lb": round(info["produced_lb"], 2),
                    "packed_lb": round(info["packed_lb"], 2),
                    "adjusted_lb": round(info["adjusted_lb"], 2),
                    "current_on_hand_lb": round(info.get("current_on_hand_lb", 0), 2)
                })

            # ── Build finished goods section ──
            finished_goods = []
            for r in pack_output_rows:
                finished_goods.append({
                    "product_name": r['product_name'],
                    "lot_code": r['lot_code'],
                    "total_lb": float(r['quantity_lb']),
                })

            # ── Adjustments list ──
            adjustments = [
                {
                    "product_name": r['product_name'],
                    "lot_code": r['lot_code'],
                    "adjustment_lb": float(r['adjustment_lb']),
                    "reason": r['reason']
                }
                for r in adjust_rows
            ]

            return {
                "date": str(target_date),
                "day_name": target_date.strftime("%A"),
                "batch_products": list(products_map.values()),
                "finished_goods": finished_goods,
                "adjustments": adjustments
            }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Production day-summary failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# PRODUCTION SCHEDULING — 7-Day Tactical Scheduler
# ═══════════════════════════════════════════════════════════════

def _build_schedule_calendar(start_date: date, horizon_days: int, friday_modifier: float):
    """Build list of working days with capacity modifiers."""
    days = []
    current = start_date
    working_days_added = 0
    max_scan = horizon_days * 3  # scan enough calendar days
    scanned = 0
    while working_days_added < horizon_days and scanned < max_scan:
        dow = current.strftime("%A")
        if dow in ("Saturday", "Sunday"):
            current += timedelta(days=1)
            scanned += 1
            continue
        modifier = friday_modifier if dow == "Friday" else 1.0
        days.append({
            "date": current,
            "day_of_week": dow,
            "capacity_modifier": modifier,
        })
        working_days_added += 1
        current += timedelta(days=1)
        scanned += 1
    return days


def _load_line_config(cur):
    """Load production lines and their capacity modes from DB."""
    cur.execute("""
        SELECT pl.id, pl.name, pl.line_code, pl.active,
               json_agg(json_build_object(
                   'mode_id', lcm.id, 'mode_name', lcm.mode_name,
                   'workers_required', lcm.workers_required,
                   'batches_per_day', lcm.batches_per_day,
                   'pallets_per_day', lcm.pallets_per_day,
                   'bags_per_day', lcm.bags_per_day,
                   'pack_size_lb', lcm.pack_size_lb,
                   'is_default', lcm.is_default
               ) ORDER BY lcm.is_default DESC, lcm.workers_required) AS modes
        FROM production_lines pl
        LEFT JOIN line_capacity_modes lcm ON lcm.line_id = pl.id
        WHERE pl.active = true
        GROUP BY pl.id, pl.name, pl.line_code, pl.active
        ORDER BY pl.name
    """)
    lines = {}
    for row in cur.fetchall():
        lines[row['line_code']] = {
            'id': row['id'],
            'name': row['name'],
            'line_code': row['line_code'],
            'modes': row['modes'] or [],
        }
    return lines


def _load_product_line_map(cur):
    """Load product→line assignments. Returns dict: product_id → line_code."""
    cur.execute("""
        SELECT pla.product_id, pl.line_code
        FROM product_line_assignments pla
        JOIN production_lines pl ON pl.id = pla.line_id
    """)
    return {row['product_id']: row['line_code'] for row in cur.fetchall()}


def _load_demand(cur, horizon_end: date):
    """Load open/confirmed sales orders within or overdue relative to horizon."""
    cur.execute("""
        SELECT so.id AS order_id, so.order_number, so.requested_ship_date, so.status,
               sol.id AS line_id, sol.product_id, sol.quantity_lb, sol.quantity_shipped_lb,
               p.name AS product_name, p.type AS product_type
        FROM sales_orders so
        JOIN sales_order_lines sol ON sol.sales_order_id = so.id
        JOIN products p ON p.id = sol.product_id
        WHERE so.status IN ('confirmed', 'in_production', 'ready')
          AND sol.line_status IN ('pending', 'partial')
          AND (so.requested_ship_date IS NULL OR so.requested_ship_date <= %s)
        ORDER BY so.requested_ship_date ASC NULLS LAST, so.id ASC
    """, (horizon_end,))
    return cur.fetchall()


def _load_finished_inventory(cur):
    """Load on-hand inventory for finished and batch products."""
    cur.execute(f"""
        SELECT p.id AS product_id, p.name AS product_name,
               COALESCE(SUM(tl.quantity_lb), 0) AS on_hand_lb
        FROM products p
        LEFT JOIN lots l ON l.product_id = p.id
        LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
        WHERE p.active = true AND p.type IN ('finished', 'batch')
        GROUP BY p.id, p.name
    """)
    return {row['product_id']: float(row['on_hand_lb']) for row in cur.fetchall()}


def _load_ingredient_inventory(cur):
    """Load on-hand inventory for ingredient products."""
    cur.execute(f"""
        SELECT p.id AS product_id, p.name AS product_name,
               COALESCE(SUM(tl.quantity_lb), 0) AS on_hand_lb
        FROM products p
        LEFT JOIN lots l ON l.product_id = p.id
        LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
        WHERE p.active = true AND p.type = 'ingredient'
        GROUP BY p.id, p.name
    """)
    return {row['product_id']: {'name': row['product_name'], 'on_hand': float(row['on_hand_lb'])} for row in cur.fetchall()}


def _load_bom_structure(cur):
    """Load full BOM: finished→batch (product_bom) and batch→ingredients (batch_formulas).
    Returns:
      fg_to_batch: {finished_product_id: {batch_product_id, batch_name, quantity, uom}}
      batch_to_ingredients: {batch_product_id: [{ingredient_product_id, ingredient_name, quantity_lb, exclude}]}
      batch_sizes: {batch_product_id: default_batch_lb}
    """
    # Finished good → batch product mapping
    cur.execute("""
        SELECT pb.finished_product_id, pb.component_product_id, p.name AS component_name,
               p.type AS component_type, pb.quantity, pb.uom, p.default_batch_lb
        FROM product_bom pb
        JOIN products p ON p.id = pb.component_product_id
    """)
    fg_to_batch = {}
    for row in cur.fetchall():
        fid = row['finished_product_id']
        if fid not in fg_to_batch:
            fg_to_batch[fid] = []
        fg_to_batch[fid].append({
            'component_product_id': row['component_product_id'],
            'component_name': row['component_name'],
            'component_type': row['component_type'],
            'quantity': float(row['quantity'] or 1),
            'uom': row['uom'] or 'unit',
            'default_batch_lb': float(row['default_batch_lb']) if row['default_batch_lb'] else None,
        })

    # Batch product → ingredients
    cur.execute("""
        SELECT bf.product_id, bf.ingredient_product_id, p.name AS ingredient_name,
               bf.quantity_lb, COALESCE(bf.exclude_from_inventory, false) AS exclude_from_inventory
        FROM batch_formulas bf
        JOIN products p ON p.id = bf.ingredient_product_id
        ORDER BY bf.product_id, bf.quantity_lb DESC
    """)
    batch_to_ingredients = {}
    for row in cur.fetchall():
        bid = row['product_id']
        if bid not in batch_to_ingredients:
            batch_to_ingredients[bid] = []
        batch_to_ingredients[bid].append({
            'ingredient_product_id': row['ingredient_product_id'],
            'ingredient_name': row['ingredient_name'],
            'quantity_lb': float(row['quantity_lb']),
            'exclude': row['exclude_from_inventory'],
        })

    # Batch sizes
    cur.execute("SELECT id, default_batch_lb FROM products WHERE type = 'batch' AND active = true AND default_batch_lb IS NOT NULL")
    batch_sizes = {row['id']: float(row['default_batch_lb']) for row in cur.fetchall()}

    return fg_to_batch, batch_to_ingredients, batch_sizes


def _simulated_allocation(demand_rows, inventory, fg_to_batch, batch_sizes, product_line_map):
    """
    Walk demand in ship-date order. For each order line, allocate from available
    finished goods inventory. Whatever remains becomes a production requirement.
    Returns list of production requirements (what to make).
    """
    available = dict(inventory)  # copy — will be mutated during simulation
    production_reqs = []  # [{product_id, product_name, batch_product_id, batch_name, needed_lb, batches, overproduction_lb, order_numbers, line_code}]

    # Group demand by product to consolidate
    product_demand = {}  # product_id → [{order_number, needed_lb, ship_date}]
    for row in demand_rows:
        pid = row['product_id']
        remaining = float(row['quantity_lb']) - float(row['quantity_shipped_lb'] or 0)
        if remaining <= 0:
            continue
        if pid not in product_demand:
            product_demand[pid] = []
        product_demand[pid].append({
            'order_number': row['order_number'],
            'needed_lb': remaining,
            'ship_date': row['requested_ship_date'],
            'product_name': row['product_name'],
            'product_type': row['product_type'],
        })

    # Process each product's demand
    for pid, demands in product_demand.items():
        total_needed = sum(d['needed_lb'] for d in demands)
        on_hand = available.get(pid, 0)

        # Allocate from inventory
        allocated = min(on_hand, total_needed)
        available[pid] = on_hand - allocated
        net_need = total_needed - allocated

        if net_need <= 0:
            continue

        order_numbers = [d['order_number'] for d in demands]
        product_name = demands[0]['product_name']
        product_type = demands[0]['product_type']
        earliest_ship = min((d['ship_date'] for d in demands if d['ship_date']), default=None)

        # Determine what to produce
        if product_type == 'finished':
            # Look up batch product from BOM
            bom_components = fg_to_batch.get(pid, [])
            batch_component = next((c for c in bom_components if c['component_type'] == 'batch'), None)

            if not batch_component:
                # No BOM — can't schedule production, will be flagged
                production_reqs.append({
                    'product_id': pid,
                    'product_name': product_name,
                    'product_type': 'finished',
                    'batch_product_id': None,
                    'batch_name': None,
                    'needed_lb': net_need,
                    'batches': None,
                    'batch_size_lb': None,
                    'overproduction_lb': 0,
                    'order_numbers': order_numbers,
                    'earliest_ship_date': earliest_ship,
                    'line_code': product_line_map.get(pid),
                    'warning': f"No batch product in BOM for '{product_name}'",
                })
                continue

            batch_pid = batch_component['component_product_id']
            batch_name = batch_component['component_name']
            batch_size = batch_sizes.get(batch_pid, 0)

            if batch_size <= 0:
                production_reqs.append({
                    'product_id': pid,
                    'product_name': product_name,
                    'product_type': 'finished',
                    'batch_product_id': batch_pid,
                    'batch_name': batch_name,
                    'needed_lb': net_need,
                    'batches': None,
                    'batch_size_lb': 0,
                    'overproduction_lb': 0,
                    'order_numbers': order_numbers,
                    'earliest_ship_date': earliest_ship,
                    'line_code': product_line_map.get(batch_pid),
                    'warning': f"No default_batch_lb for '{batch_name}'",
                })
                continue

            # Also check batch inventory — maybe we have batch product on hand
            batch_on_hand = available.get(batch_pid, 0)
            net_need_after_batch = max(0, net_need - batch_on_hand)
            if batch_on_hand > 0:
                used = min(batch_on_hand, net_need)
                available[batch_pid] = batch_on_hand - used

            if net_need_after_batch <= 0:
                continue

            num_batches = math.ceil(net_need_after_batch / batch_size)
            total_output = num_batches * batch_size
            overproduction = total_output - net_need_after_batch

            production_reqs.append({
                'product_id': pid,
                'product_name': product_name,
                'product_type': 'finished',
                'batch_product_id': batch_pid,
                'batch_name': batch_name,
                'needed_lb': net_need_after_batch,
                'batches': num_batches,
                'batch_size_lb': batch_size,
                'overproduction_lb': round(overproduction, 2),
                'order_numbers': order_numbers,
                'earliest_ship_date': earliest_ship,
                'line_code': product_line_map.get(batch_pid),
                'warning': None,
            })

        elif product_type == 'batch':
            # Direct batch product demand
            batch_size = batch_sizes.get(pid, 0)
            line_code = product_line_map.get(pid)

            if batch_size <= 0:
                production_reqs.append({
                    'product_id': pid,
                    'product_name': product_name,
                    'product_type': 'batch',
                    'batch_product_id': pid,
                    'batch_name': product_name,
                    'needed_lb': net_need,
                    'batches': None,
                    'batch_size_lb': 0,
                    'overproduction_lb': 0,
                    'order_numbers': order_numbers,
                    'earliest_ship_date': earliest_ship,
                    'line_code': line_code,
                    'warning': f"No default_batch_lb for '{product_name}'",
                })
                continue

            num_batches = math.ceil(net_need / batch_size)
            total_output = num_batches * batch_size
            overproduction = total_output - net_need

            production_reqs.append({
                'product_id': pid,
                'product_name': product_name,
                'product_type': 'batch',
                'batch_product_id': pid,
                'batch_name': product_name,
                'needed_lb': net_need,
                'batches': num_batches,
                'batch_size_lb': batch_size,
                'overproduction_lb': round(overproduction, 2),
                'order_numbers': order_numbers,
                'earliest_ship_date': earliest_ship,
                'line_code': line_code,
                'warning': None,
            })

    return production_reqs


def _explode_ingredients(production_reqs, batch_to_ingredients, ingredient_inventory):
    """Calculate total ingredient needs and check for shortages."""
    ingredient_needs = {}  # ingredient_product_id → total_lb_needed

    for req in production_reqs:
        if req['batches'] is None or req['batch_product_id'] is None:
            continue
        formula = batch_to_ingredients.get(req['batch_product_id'], [])
        for ing in formula:
            if ing['exclude']:
                continue
            iid = ing['ingredient_product_id']
            needed = ing['quantity_lb'] * req['batches']
            ingredient_needs[iid] = ingredient_needs.get(iid, 0) + needed

    # Check against inventory
    ingredient_summary = []
    for iid, needed in sorted(ingredient_needs.items(), key=lambda x: x[1], reverse=True):
        info = ingredient_inventory.get(iid, {'name': f'Unknown ({iid})', 'on_hand': 0})
        on_hand = info['on_hand']
        shortage = max(0, needed - on_hand)
        ingredient_summary.append({
            'ingredient_name': info['name'],
            'ingredient_id': iid,
            'required_lb': round(needed, 2),
            'on_hand_lb': round(on_hand, 2),
            'shortage_lb': round(shortage, 2),
            'status': '⚠️ Ingredient Risk' if shortage > 0 else '✅ OK',
        })

    return ingredient_summary


def _schedule_runs_to_days(production_reqs, calendar_days, line_config, total_workers, strategy='earliest'):
    """
    Assign production runs to days respecting capacity and labor constraints.
    strategy: 'earliest' = pull forward, 'latest' = push back (closer to ship date)
    Returns (scheduled_days, unscheduled_orders).
    """
    # Initialize day structures
    day_schedules = []
    for day_info in calendar_days:
        day_sched = {
            'date': day_info['date'].isoformat(),
            'day_of_week': day_info['day_of_week'],
            'capacity_modifier': day_info['capacity_modifier'],
            'total_labor_used': 0,
            'lines': {},
        }
        for lc, linfo in line_config.items():
            day_sched['lines'][lc] = {
                'line_name': linfo['name'],
                'workers_assigned': 0,
                'runs': [],
                'warnings': [],
                'remaining_batches': None,  # will be set when line activated
                'remaining_bags': None,
                'remaining_pallets': None,
            }
        day_schedules.append(day_sched)

    unscheduled = []

    # Sort reqs: by earliest_ship_date for 'earliest', reverse for 'latest'
    sorted_reqs = sorted(
        production_reqs,
        key=lambda r: (r['earliest_ship_date'] or date(2099, 1, 1), r.get('needed_lb', 0)),
        reverse=(strategy == 'latest')
    )

    for req in sorted_reqs:
        if req['batches'] is None or req['batches'] <= 0:
            if req.get('warning'):
                unscheduled.append({
                    'order_numbers': req['order_numbers'],
                    'product_name': req['product_name'],
                    'reason': req['warning'],
                })
            continue

        line_code = req['line_code']
        if not line_code or line_code not in line_config:
            unscheduled.append({
                'order_numbers': req['order_numbers'],
                'product_name': req['product_name'],
                'reason': f"No production line assigned for '{req.get('batch_name') or req['product_name']}'",
            })
            continue

        linfo = line_config[line_code]
        modes = linfo['modes']
        if not modes:
            unscheduled.append({
                'order_numbers': req['order_numbers'],
                'product_name': req['product_name'],
                'reason': f"No capacity modes configured for line '{linfo['name']}'",
            })
            continue

        remaining_batches = req['batches']
        day_indices = range(len(day_schedules)) if strategy == 'earliest' else reversed(range(len(day_schedules)))

        for di in day_indices:
            if remaining_batches <= 0:
                break

            day = day_schedules[di]
            modifier = day['capacity_modifier']
            line_day = day['lines'][line_code]

            # Pick the best capacity mode that fits labor
            available_labor = total_workers - day['total_labor_used']

            # If this line already has workers assigned today, use that mode
            if line_day['workers_assigned'] > 0:
                # Line already active — use remaining capacity
                can_do = line_day.get('remaining_batches') or 0
                if can_do <= 0:
                    continue

                run_batches = min(remaining_batches, can_do)
                batch_size = req['batch_size_lb']
                run_qty = run_batches * batch_size

                line_day['runs'].append({
                    'product_name': req.get('batch_name') or req['product_name'],
                    'product_id': req.get('batch_product_id') or req['product_id'],
                    'batches': run_batches,
                    'quantity_lb': round(run_qty, 2),
                    'for_orders': req['order_numbers'],
                    'overproduction_lb': round(req['overproduction_lb'], 2) if remaining_batches - run_batches <= 0 else 0,
                    'overproduction_reason': 'Batch Rounding' if (remaining_batches - run_batches <= 0 and req['overproduction_lb'] > 0) else None,
                })
                line_day['remaining_batches'] = can_do - run_batches
                remaining_batches -= run_batches
                continue

            # Line not active yet today — need to activate with a mode
            best_mode = None
            for mode in modes:
                w = mode['workers_required']
                if w <= available_labor:
                    best_mode = mode
                    break  # modes sorted by default first, then lowest workers

            if not best_mode:
                continue  # can't fit any mode on this day

            workers = best_mode['workers_required']
            raw_capacity = best_mode.get('batches_per_day') or 0
            day_capacity = max(1, int(raw_capacity * modifier))

            run_batches = min(remaining_batches, day_capacity)
            batch_size = req['batch_size_lb']
            run_qty = run_batches * batch_size

            # Activate the line
            day['total_labor_used'] += workers
            line_day['workers_assigned'] = workers
            line_day['remaining_batches'] = day_capacity - run_batches

            line_day['runs'].append({
                'product_name': req.get('batch_name') or req['product_name'],
                'product_id': req.get('batch_product_id') or req['product_id'],
                'batches': run_batches,
                'quantity_lb': round(run_qty, 2),
                'for_orders': req['order_numbers'],
                'overproduction_lb': round(req['overproduction_lb'], 2) if remaining_batches - run_batches <= 0 else 0,
                'overproduction_reason': 'Batch Rounding' if (remaining_batches - run_batches <= 0 and req['overproduction_lb'] > 0) else None,
            })
            remaining_batches -= run_batches

        if remaining_batches > 0:
            unscheduled.append({
                'order_numbers': req['order_numbers'],
                'product_name': req.get('batch_name') or req['product_name'],
                'reason': f"Insufficient capacity in window ({remaining_batches} batches remaining)",
            })

    # Format output
    formatted_days = []
    for day in day_schedules:
        lines_out = []
        for lc in ['granola', 'coconut', 'bulk_pack', 'pouch']:
            if lc in day['lines']:
                ld = day['lines'][lc]
                lines_out.append({
                    'line_name': ld['line_name'],
                    'workers_assigned': ld['workers_assigned'],
                    'runs': ld['runs'],
                    'warnings': ld['warnings'],
                })
        formatted_days.append({
            'date': day['date'],
            'day_of_week': day['day_of_week'],
            'capacity_modifier': day['capacity_modifier'],
            'total_labor_used': day['total_labor_used'],
            'lines': lines_out,
        })

    return formatted_days, unscheduled


def _handle_schedule_suggest(body: dict):
    """Generate a proposed 7-day production schedule based on open orders, inventory, and capacity."""
    start_date_str = body.get('start_date')
    horizon_days = body.get('horizon_days', 7)
    total_workers = body.get('total_workers', 10)
    friday_modifier = body.get('friday_modifier', 0.5)

    if start_date_str:
        try:
            start = date.fromisoformat(start_date_str)
        except ValueError:
            raise HTTPException(400, f"Invalid start_date format: '{start_date_str}'. Use YYYY-MM-DD.")
    else:
        start = get_plant_now().date() + timedelta(days=1)

    with get_transaction() as cur:
        # 1. Load config
        line_config = _load_line_config(cur)
        product_line_map = _load_product_line_map(cur)

        # 2. Build calendar
        calendar_days = _build_schedule_calendar(start, horizon_days, friday_modifier)
        if not calendar_days:
            return {"error": "No working days in the specified horizon"}
        horizon_end = calendar_days[-1]['date']

        # 3. Load demand
        demand = _load_demand(cur, horizon_end)

        # 4. Load supply
        inventory = _load_finished_inventory(cur)
        ingredient_inv = _load_ingredient_inventory(cur)

        # 5. Load BOM structure
        fg_to_batch, batch_to_ingredients, batch_sizes = _load_bom_structure(cur)

        # 6. Simulated allocation → net requirements
        production_reqs = _simulated_allocation(
            demand, inventory, fg_to_batch, batch_sizes, product_line_map
        )

        if not production_reqs:
            return {
                "schedule_id": str(uuid.uuid4()),
                "horizon": {"start": start.isoformat(), "end": horizon_end.isoformat()},
                "total_workers_available": total_workers,
                "message": "No production needed — all open orders covered by current inventory.",
                "scenarios": [],
            }

        # 7. Explode ingredients
        ingredient_summary = _explode_ingredients(
            production_reqs, batch_to_ingredients, ingredient_inv
        )

        # 8. Schedule runs — try earliest-first
        days_earliest, unscheduled_earliest = _schedule_runs_to_days(
            production_reqs, calendar_days, line_config, total_workers, strategy='earliest'
        )

        scenarios = []

        if not unscheduled_earliest:
            # Everything fits — single recommended scenario
            scenarios.append({
                'scenario_name': 'Recommended',
                'days': days_earliest,
                'ingredient_summary': [i for i in ingredient_summary if i['shortage_lb'] > 0],
                'unscheduled_orders': [],
            })
        else:
            # Conflicts — generate two scenarios
            scenarios.append({
                'scenario_name': 'Scenario A: Pull Production Earlier',
                'days': days_earliest,
                'ingredient_summary': [i for i in ingredient_summary if i['shortage_lb'] > 0],
                'unscheduled_orders': unscheduled_earliest,
            })

            # Scenario B: push production later
            days_latest, unscheduled_latest = _schedule_runs_to_days(
                production_reqs, calendar_days, line_config, total_workers, strategy='latest'
            )
            scenarios.append({
                'scenario_name': 'Scenario B: Push Production Later',
                'days': days_latest,
                'ingredient_summary': [i for i in ingredient_summary if i['shortage_lb'] > 0],
                'unscheduled_orders': unscheduled_latest,
            })

        # Build summary of production requirements for context
        production_summary = []
        for req_item in production_reqs:
            production_summary.append({
                'product': req_item.get('batch_name') or req_item['product_name'],
                'for_finished_good': req_item['product_name'] if req_item['product_type'] == 'finished' else None,
                'needed_lb': round(req_item['needed_lb'], 2),
                'batches': req_item['batches'],
                'total_output_lb': round(req_item['batches'] * req_item['batch_size_lb'], 2) if req_item['batches'] and req_item['batch_size_lb'] else None,
                'overproduction_lb': req_item['overproduction_lb'],
                'line': req_item['line_code'],
                'for_orders': req_item['order_numbers'],
                'warning': req_item.get('warning'),
            })

        return {
            "schedule_id": str(uuid.uuid4()),
            "horizon": {"start": start.isoformat(), "end": horizon_end.isoformat()},
            "total_workers_available": total_workers,
            "open_orders_in_window": len(set(r['order_number'] for r in demand)),
            "production_requirements": production_summary,
            "scenarios": scenarios,
            "all_ingredient_status": ingredient_summary,
        }


def _handle_schedule_confirm(body: dict):
    """Confirm a proposed (or edited) production schedule — saves to production_schedule table."""
    runs_data = body.get('runs')
    if not runs_data or not isinstance(runs_data, list):
        raise HTTPException(400, "confirm action requires a 'runs' array")

    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Load line lookup
            cur.execute("SELECT id, line_code FROM production_lines WHERE active = true")
            line_lookup = {r['line_code']: r['id'] for r in cur.fetchall()}

            confirmed_ids = []
            for run in runs_data:
                line_code = run.get('line_code')
                line_id = line_lookup.get(line_code)
                if not line_id:
                    raise HTTPException(400, f"Unknown line_code: '{line_code}'")

                run_date_str = run.get('date')
                try:
                    run_date = date.fromisoformat(run_date_str)
                except (ValueError, TypeError):
                    raise HTTPException(400, f"Invalid date: '{run_date_str}'")

                product_id = run.get('product_id')
                if not product_id:
                    raise HTTPException(400, "Each run requires a 'product_id'")

                cur.execute("""
                    INSERT INTO production_schedule
                        (schedule_date, line_id, product_id, planned_batches, planned_quantity_lb,
                         planned_bags, workers_assigned, status, linked_order_numbers,
                         overproduction_lb, overproduction_reason, notes, confirmed_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'confirmed', %s, %s, %s, %s, NOW())
                    ON CONFLICT (schedule_date, line_id, product_id)
                    DO UPDATE SET
                        planned_batches = EXCLUDED.planned_batches,
                        planned_quantity_lb = EXCLUDED.planned_quantity_lb,
                        planned_bags = EXCLUDED.planned_bags,
                        workers_assigned = EXCLUDED.workers_assigned,
                        linked_order_numbers = EXCLUDED.linked_order_numbers,
                        overproduction_lb = EXCLUDED.overproduction_lb,
                        overproduction_reason = EXCLUDED.overproduction_reason,
                        notes = EXCLUDED.notes,
                        status = 'confirmed',
                        confirmed_at = NOW()
                    RETURNING id
                """, (
                    run_date, line_id, product_id,
                    run.get('planned_batches'), run.get('planned_quantity_lb'),
                    run.get('planned_bags'), run.get('workers_assigned', 0),
                    run.get('linked_order_numbers'),
                    run.get('overproduction_lb', 0), run.get('overproduction_reason'),
                    run.get('notes'),
                ))
                row = cur.fetchone()
                if row:
                    confirmed_ids.append(row['id'])

            conn.commit()

    return {
        "confirmed": True,
        "runs_saved": len(confirmed_ids),
        "schedule_ids": confirmed_ids,
    }


def _handle_schedule_current(body: dict):
    """View the current confirmed production schedule."""
    start_date_str = body.get('start_date')
    days = body.get('days', 7)

    if start_date_str:
        try:
            start = date.fromisoformat(start_date_str)
        except ValueError:
            raise HTTPException(400, f"Invalid start_date: '{start_date_str}'")
    else:
        start = get_plant_now().date()

    end = start + timedelta(days=days)

    with get_transaction() as cur:
        cur.execute("""
            SELECT ps.id, ps.schedule_date, ps.planned_batches, ps.planned_quantity_lb,
                   ps.planned_bags, ps.workers_assigned, ps.status,
                   ps.linked_order_numbers, ps.overproduction_lb, ps.overproduction_reason,
                   ps.notes, ps.confirmed_at,
                   pl.name AS line_name, pl.line_code,
                   p.name AS product_name, p.id AS product_id
            FROM production_schedule ps
            JOIN production_lines pl ON pl.id = ps.line_id
            JOIN products p ON p.id = ps.product_id
            WHERE ps.schedule_date >= %s AND ps.schedule_date < %s
              AND ps.status != 'cancelled'
            ORDER BY ps.schedule_date, pl.name, p.name
        """, (start, end))
        rows = cur.fetchall()

    # Group by date
    by_date = {}
    for r in rows:
        d = r['schedule_date'].isoformat()
        if d not in by_date:
            by_date[d] = {
                'date': d,
                'day_of_week': r['schedule_date'].strftime("%A"),
                'runs': [],
                'total_workers': 0,
            }
        by_date[d]['runs'].append({
            'schedule_id': r['id'],
            'line_name': r['line_name'],
            'line_code': r['line_code'],
            'product_name': r['product_name'],
            'product_id': r['product_id'],
            'planned_batches': r['planned_batches'],
            'planned_quantity_lb': float(r['planned_quantity_lb']) if r['planned_quantity_lb'] else None,
            'planned_bags': r['planned_bags'],
            'workers_assigned': r['workers_assigned'],
            'status': r['status'],
            'linked_orders': r['linked_order_numbers'],
            'overproduction_lb': float(r['overproduction_lb']) if r['overproduction_lb'] else 0,
            'notes': r['notes'],
        })
        by_date[d]['total_workers'] += r['workers_assigned']

    return {
        "period": {"start": start.isoformat(), "end": end.isoformat()},
        "days": list(by_date.values()),
        "total_runs": len(rows),
    }


@app.post("/schedule")
def schedule_dispatch(request_body: dict, _: bool = Depends(verify_api_key)):
    """Unified scheduling endpoint. Dispatches based on the 'action' field.

    Actions:
      - suggest:  Generate a proposed schedule (optional: start_date, horizon_days, total_workers, friday_modifier)
      - confirm:  Confirm/save a schedule (requires: runs[])
      - current:  View the confirmed schedule (optional: start_date, days)
    """
    action = request_body.get("action")
    if not action:
        raise HTTPException(400, "Missing required field: 'action'. Use 'suggest', 'confirm', or 'current'.")

    action = action.strip().lower()

    try:
        if action == "suggest":
            return _handle_schedule_suggest(request_body)
        elif action == "confirm":
            return _handle_schedule_confirm(request_body)
        elif action == "current":
            return _handle_schedule_current(request_body)
        else:
            raise HTTPException(400, f"Unknown action: '{action}'. Use 'suggest', 'confirm', or 'current'.")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Schedule ({action}) failed: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# AUDIT / INTEGRITY CHECK ENDPOINT
# ═══════════════════════════════════════════════════════════════

@app.get("/audit/integrity")
def audit_integrity():
    """Run automated integrity checks and return structured results.
    No auth required — read-only diagnostic endpoint for dashboard."""
    try:
        with get_transaction() as cur:
            checks = []
            now = get_plant_now()

            # 1. Negative lot balances [CRITICAL]
            cur.execute(f"""
                SELECT l.id as lot_id, l.lot_code, p.name as product_name,
                       COALESCE(SUM(tl.quantity_lb), 0) as balance
                FROM lots l
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                JOIN products p ON p.id = l.product_id
                GROUP BY l.id, p.name
                HAVING COALESCE(SUM(tl.quantity_lb), 0) < -%s
            """, (BALANCE_EPSILON,))
            neg_lots = cur.fetchall()
            checks.append({
                "name": "negative_lot_balances",
                "severity": "critical",
                "status": "fail" if neg_lots else "pass",
                "details": [{"lot_id": r['lot_id'], "lot_code": r['lot_code'],
                             "product": r['product_name'], "balance": float(r['balance'])} for r in neg_lots]
            })

            # 2. Production missing ILC [CRITICAL]
            cur.execute("""
                SELECT t.id as transaction_id, t.notes,
                       COALESCE(SUM(tl.quantity_lb) FILTER (WHERE tl.quantity_lb > 0), 0) as output_lb
                FROM ledger_current_transactions t
                LEFT JOIN ledger_current_transaction_lines tl ON tl.transaction_id = t.id
                LEFT JOIN ingredient_lot_consumption ilc ON ilc.transaction_id = t.id
                WHERE t.type = 'make'
                  AND t.effective_status = 'posted'
                GROUP BY t.id, t.notes
                HAVING COUNT(ilc.id) = 0
                   AND COALESCE(SUM(tl.quantity_lb) FILTER (WHERE tl.quantity_lb > 0), 0) > 0
            """)
            missing_ilc = cur.fetchall()
            checks.append({
                "name": "production_missing_ilc",
                "severity": "critical",
                "status": "fail" if missing_ilc else "pass",
                "details": [{"transaction_id": r['transaction_id'], "output_lb": float(r['output_lb']),
                             "note": r['notes']} for r in missing_ilc]
            })

            # 3. Ship transactions missing shipment_lines after Feb 27 [MAJOR]
            cur.execute("""
                SELECT t.id as transaction_id, t.timestamp, t.customer_name
                FROM ledger_current_transactions t
                LEFT JOIN shipment_lines sl ON sl.transaction_id = t.id
                WHERE t.type = 'ship'
                  AND t.effective_status = 'posted'
                  AND t.timestamp >= '2026-02-27'
                  AND sl.id IS NULL
                GROUP BY t.id, t.timestamp, t.customer_name
            """)
            missing_sl = cur.fetchall()
            checks.append({
                "name": "ship_missing_shipment_lines",
                "severity": "major",
                "status": "fail" if missing_sl else "pass",
                "details": [{"transaction_id": r['transaction_id'],
                             "customer": r['customer_name']} for r in missing_sl]
            })

            # 4. Lots missing received_at [MAJOR]
            cur.execute("SELECT COUNT(*) as cnt FROM lots WHERE received_at IS NULL")
            null_received = cur.fetchone()['cnt']
            checks.append({
                "name": "lots_missing_received_at",
                "severity": "major",
                "status": "fail" if null_received > 0 else "pass",
                "details": [{"count": null_received}] if null_received > 0 else []
            })

            # 5. Lots missing supplier_lot_code on receive transactions [MAJOR]
            cur.execute("""
                SELECT l.id as lot_id, l.lot_code, p.name as product_name
                FROM lots l
                JOIN products p ON p.id = l.product_id
                WHERE l.entry_source = 'received'
                  AND l.supplier_lot_code IS NULL
            """)
            missing_slc = cur.fetchall()
            checks.append({
                "name": "lots_missing_supplier_lot_code",
                "severity": "major",
                "status": "fail" if missing_slc else "pass",
                "details": [{"lot_id": r['lot_id'], "lot_code": r['lot_code'],
                             "product": r['product_name']} for r in missing_slc]
            })

            # 6. Finished goods missing case_size_lb [MAJOR]
            cur.execute("""
                SELECT DISTINCT p.id as product_id, p.name
                FROM products p
                JOIN ledger_current_transaction_lines tl ON tl.product_id = p.id
                JOIN ledger_current_transactions t ON t.id = tl.transaction_id
                WHERE t.type = 'pack' AND tl.quantity_lb > 0
                  AND t.effective_status = 'posted'
                  AND p.case_size_lb IS NULL
            """)
            missing_cs = cur.fetchall()
            checks.append({
                "name": "finished_goods_missing_case_size",
                "severity": "major",
                "status": "fail" if missing_cs else "pass",
                "details": [{"product_id": r['product_id'], "name": r['name']} for r in missing_cs]
            })

            # 7. Floating point dust [MINOR]
            cur.execute(f"""
                SELECT l.id as lot_id, l.lot_code,
                       COALESCE(SUM(tl.quantity_lb), 0) as balance
                FROM lots l
                LEFT JOIN {POSTED_LINES} tl ON tl.lot_id = l.id
                GROUP BY l.id
                HAVING ABS(COALESCE(SUM(tl.quantity_lb), 0)) < 0.01
                   AND COALESCE(SUM(tl.quantity_lb), 0) != 0
            """)
            dust = cur.fetchall()
            checks.append({
                "name": "floating_point_dust",
                "severity": "minor",
                "status": "fail" if dust else "pass",
                "details": [{"lot_id": r['lot_id'], "lot_code": r['lot_code'],
                             "balance": float(r['balance'])} for r in dust]
            })

            # 8. Voided transaction count [INFO]
            cur.execute("SELECT COUNT(*) as cnt FROM ledger_current_transactions WHERE effective_status = 'voided'")
            voided_count = cur.fetchone()['cnt']
            checks.append({
                "name": "voided_transaction_count",
                "severity": "info",
                "status": "pass",
                "details": [{"count": voided_count}]
            })

            # Calculate score
            score = 100
            for check in checks:
                if check['status'] == 'fail':
                    if check['severity'] == 'critical':
                        score -= 10
                    elif check['severity'] == 'major':
                        score -= 5
                    elif check['severity'] == 'minor':
                        score -= 1
            score = max(0, score)

            return {
                "timestamp": now.isoformat(),
                "score": score,
                "checks": checks
            }
    except Exception as e:
        logger.error(f"Audit integrity check failed: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


# ═══════════════════════════════════════════════════════════════
# STATIC FILE SERVING — Dashboard UI (must be LAST)
# ═══════════════════════════════════════════════════════════════

_dashboard_dir = pathlib.Path(__file__).parent / "dashboard"
if _dashboard_dir.is_dir():
    app.mount("/dashboard", StaticFiles(directory=str(_dashboard_dir), html=True), name="dashboard-ui")
