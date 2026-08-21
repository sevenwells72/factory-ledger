"""Regression coverage for the dashboard-only styled orders matrix export."""

from contextlib import contextmanager
from datetime import date
from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import main


class _SeededCursor:
    def __init__(self, rows):
        self.rows = rows

    def execute(self, query, params=None):
        assert "/export/orders-matrix.xlsx" not in query

    def fetchall(self):
        return self.rows


def test_orders_matrix_export_workbook(monkeypatch):
    seeded_lines = [
        {
            "customer": "Sunday Customer", "order_id": "SO-TEST-001",
            "due_date": date(2026, 7, 12), "sku": "70050",
            "product_name": "Granola Classic 25 LB", "qty": Decimal("2"),
            "lb_per_case": Decimal("25"),
        },
        {
            "customer": "Monday Customer", "order_id": "SO-TEST-002",
            "due_date": date(2026, 7, 13), "sku": "10001",
            "product_name": "Coconut Sweetened Flake CNS 10 LB", "qty": Decimal("3"),
            "lb_per_case": Decimal("10"),
        },
        {
            "customer": "Sunday Customer", "order_id": "SO-TEST-001",
            "due_date": date(2026, 7, 12), "sku": "70073",
            "product_name": "BS Granola – Peanut Butter Banana – 6x7 OZ Case",
            "qty": Decimal("1.5"), "lb_per_case": Decimal("2.625"),
        },
        {
            "customer": "Monday Customer", "order_id": "SO-TEST-002",
            "due_date": date(2026, 7, 13), "sku": "31012",
            "product_name": "Graham Cracker Crumbs 10 LB Case",
            "qty": Decimal("4"), "lb_per_case": Decimal("10"),
        },
    ]

    @contextmanager
    def seeded_transaction():
        yield _SeededCursor(seeded_lines)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")

    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert workbook.sheetnames == ["Cases", "Pounds"]

    cases = workbook["Cases"]
    pounds = workbook["Pounds"]
    assert cases.column_dimensions["D"].hidden is True
    assert cases["A1"].font.sz == 7
    assert cases.freeze_panes == "E2"

    granola_col = next(cell.column for cell in cases[1] if cell.value == "Granola Classic 25#")
    fractional_col = next(cell.column for cell in cases[1] if cell.value == "BS Granola Peanut Butter Banana 6x7 OZ")
    coconut_col = next(cell.column for cell in cases[1] if cell.value == "Coco Swt Flake CNS 10#")
    graham_col = next(cell.column for cell in cases[1] if cell.value == "Graham Cracker Crumbs 10#")
    assert cases.cell(2, granola_col).value == 2
    assert pounds.cell(2, granola_col).value == 50
    assert cases.cell(3, coconut_col).value == 3
    assert pounds.cell(3, coconut_col).value == 30
    assert cases.cell(2, fractional_col).value == 1.5
    assert pounds.cell(2, fractional_col).value == 3.9375
    assert cases.cell(2, fractional_col).number_format == '#,##0.#;(#,##0.#);"—"'
    assert pounds.cell(2, fractional_col).number_format == '#,##0.#;(#,##0.#);"—"'
    assert cases.cell(2, granola_col).number_format == '#,##0;(#,##0);"—"'
    assert "0.2 pans" in cases.cell(2, granola_col).comment.text
    assert "0.2 pans" in pounds.cell(2, granola_col).comment.text
    assert "pans" in cases.cell(2, granola_col).comment.text
    assert "pans" in pounds.cell(2, granola_col).comment.text
    assert "repack" in cases.cell(3, graham_col).comment.text.lower()
    assert "repack" in pounds.cell(3, graham_col).comment.text.lower()
    assert "<0.1 pans" in cases.cell(2, fractional_col).comment.text
    assert "<0.1 pans" in pounds.cell(2, fractional_col).comment.text
    assert "0.0 pans" not in cases.cell(2, fractional_col).comment.text
    assert "0.0 pans" not in pounds.cell(2, fractional_col).comment.text
    assert cases.cell(2, granola_col).comment.author == "Factory Ledger"
    with ZipFile(BytesIO(response.content)) as archive:
        comment_shapes = "".join(
            archive.read(name).decode("utf-8")
            for name in archive.namelist()
            if name.endswith(".vml")
        )
    assert "width:260px;height:80px" in comment_shapes.replace(" ", "")

    total_col = cases.max_column
    total_row = 4
    assert pounds.cell(total_row, total_col).value == f"=SUM(E{total_row}:{pounds.cell(total_row, total_col - 1).column_letter}{total_row})"
    assert sum(
        pounds.cell(row, col).value or 0
        for row in (2, 3)
        for col in range(5, total_col)
    ) == 123.9375
    assert cases["A3"].border.top.style == "medium"
    assert cases.auto_filter.ref.endswith("3")

    input_row = 7
    batches_row = 8
    assert "CNS Production Source of Truth" in cases.cell(input_row, granola_col).comment.text
    assert cases.cell(batches_row, granola_col).comment.text == "50 lb ÷ 322.6 lb/pan"


def test_orders_matrix_uses_case_size_instead_of_uom_text(monkeypatch):
    rows = [{
        "customer": "Custom UOM Customer", "order_id": "SO-CUSTOM-UOM",
        "due_date": date(2026, 7, 13), "sku": "70999",
        "product_name": "Granola Unknown Pack", "qty": Decimal("1"),
        "lb_per_case": Decimal("12.5"),
    }]

    @contextmanager
    def seeded_transaction():
        yield _SeededCursor(rows)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")
    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    cases = workbook["Cases"]
    product_col = next(
        cell.column for cell in cases[1]
        if cell.value == "Granola Unknown Pack"
    )
    assert cases.cell(2, product_col).value == 1
    assert workbook["Pounds"].cell(2, product_col).value == 12.5


def test_orders_matrix_query_excludes_non_finished_products(monkeypatch):
    """Raw ingredients sold by weight must not enter the production matrix."""
    rows = [{
        "customer": "Finished Customer", "order_id": "SO-FINISHED",
        "due_date": date(2026, 8, 17), "sku": "70050",
        "product_name": "Granola Classic 25 LB", "qty": Decimal("2"),
        "lb_per_case": Decimal("25"),
    }]

    class _FinishedGoodsCursor(_SeededCursor):
        def execute(self, query, params=None):
            super().execute(query, params)
            assert "p.type = 'finished'" in query

    @contextmanager
    def seeded_transaction():
        yield _FinishedGoodsCursor(rows)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")
    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    assert "Granola Classic 25#" in [cell.value for cell in workbook["Cases"][1]]


def test_orders_matrix_exports_bulk_per_lb_line(monkeypatch):
    """Bulk per-lb SKUs carry NULL case_size_lb by design and must not 422."""
    rows = [
        {
            "customer": "Bulk Customer", "order_id": "SO-BULK-001",
            "due_date": date(2026, 8, 20), "sku": "70013",
            "product_name": "Granola SS Classic #9 Bulk per/lb",
            "qty": Decimal("450"), "lb_per_case": None,
        },
        {
            "customer": "Case Customer", "order_id": "SO-CASE-001",
            "due_date": date(2026, 8, 20), "sku": "70050",
            "product_name": "Granola Classic 25 LB",
            "qty": Decimal("2"), "lb_per_case": Decimal("25"),
        },
    ]

    @contextmanager
    def seeded_transaction():
        yield _SeededCursor(rows)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")
    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    cases = workbook["Cases"]
    pounds = workbook["Pounds"]

    headers = [cell.value for cell in cases[1]]
    # Name already ends in "per/lb", so it is not suffixed again.
    bulk_header = "Granola SS Classic #9 Bulk per/lb"
    assert bulk_header in headers, headers
    assert headers == [cell.value for cell in pounds[1]]
    # The case-priced SKU keeps its plain header.
    assert "Granola Classic 25#" in headers

    bulk_col = next(cell.column for cell in cases[1] if cell.value == bulk_header)
    case_col = next(cell.column for cell in cases[1] if cell.value == "Granola Classic 25#")

    # Rows sort by (due date, customer): Bulk Customer is row 2, Case Customer row 3.
    # lb_per_case falls back to 1.0, so both sheets show the same pounds figure.
    assert cases.cell(2, bulk_col).value == 450
    assert pounds.cell(2, bulk_col).value == 450
    assert cases.cell(3, case_col).value == 2
    assert pounds.cell(3, case_col).value == 50

    # 70013 has no pan yield on file, so no cases-to-pans note is attached.
    assert cases.cell(2, bulk_col).comment is None
    assert pounds.cell(2, bulk_col).comment is None
    # The case-priced SKU still gets its normal note.
    assert "cases" in cases.cell(3, case_col).comment.text


def test_orders_matrix_per_lb_note_uses_pounds_when_pan_yield_known(monkeypatch):
    """A per-lb SKU with a pan yield gets lb-based pan math, never case text."""
    rows = [{
        "customer": "Bulk Customer", "order_id": "SO-BULK-002",
        "due_date": date(2026, 8, 20), "sku": "70050",
        "product_name": "Granola Classic Bulk per/lb",
        "qty": Decimal("645.2"), "lb_per_case": None,
    }]

    @contextmanager
    def seeded_transaction():
        yield _SeededCursor(rows)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")
    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()

    assert response.status_code == 200, response.text
    cases = load_workbook(BytesIO(response.content), data_only=False)["Cases"]
    col = next(
        cell.column for cell in cases[1]
        if cell.value == "Granola Classic Bulk per/lb"
    )
    note = cases.cell(2, col).comment.text
    assert "2.0 pans" in note
    assert "sold per lb" in note
    assert "cases" not in note


def test_orders_matrix_still_rejects_zero_case_size(monkeypatch):
    """case_size_lb <= 0 is corrupt data and must still fail loudly."""
    rows = [
        {
            "customer": "Bad Customer", "order_id": "SO-BAD-001",
            "due_date": date(2026, 8, 20), "sku": "70999",
            "product_name": "Granola Zero Case", "qty": Decimal("120"),
            "lb_per_case": Decimal("0"),
        },
        {
            "customer": "Bulk Customer", "order_id": "SO-BULK-003",
            "due_date": date(2026, 8, 20), "sku": "70013",
            "product_name": "Granola SS Classic #9 Bulk per/lb",
            "qty": Decimal("450"), "lb_per_case": None,
        },
        {
            "customer": "Case Customer", "order_id": "SO-CASE-002",
            "due_date": date(2026, 8, 20), "sku": "70050",
            "product_name": "Granola Classic 25 LB", "qty": Decimal("2"),
            "lb_per_case": Decimal("25"),
        },
    ]

    @contextmanager
    def seeded_transaction():
        yield _SeededCursor(rows)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")
    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()

    assert response.status_code == 422, response.text
    detail = response.json()["detail"]
    assert detail["error_code"] == "INVALID_ORDER_EXPORT_CASE_SIZE"
    # Only the corrupt SKU is named — the bulk per-lb line is not offending.
    assert detail["offending_skus"] == ["70999"]


def test_orders_matrix_query_coalesces_qty_for_null_case_size(monkeypatch):
    """NULL case_size_lb must still yield a qty (pounds), not a NULL row."""
    rows = [{
        "customer": "Case Customer", "order_id": "SO-CASE-003",
        "due_date": date(2026, 8, 20), "sku": "70050",
        "product_name": "Granola Classic 25 LB", "qty": Decimal("2"),
        "lb_per_case": Decimal("25"),
    }]

    class _CoalesceCursor(_SeededCursor):
        def execute(self, query, params=None):
            super().execute(query, params)
            assert "COALESCE(sol.quantity_lb / NULLIF(p.case_size_lb, 0), sol.quantity_lb)" in query

    @contextmanager
    def seeded_transaction():
        yield _CoalesceCursor(rows)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")
    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()

    assert response.status_code == 200, response.text


def test_orders_matrix_per_lb_header_suffix_not_duplicated(monkeypatch):
    """"(lb)" marks per-lb columns, but never stutters on a name already saying so."""
    rows = [
        {
            "customer": "Bulk Customer", "order_id": "SO-BULK-004",
            "due_date": date(2026, 8, 20), "sku": "70013",
            "product_name": "Granola SS Classic #9 Bulk per/lb",
            "qty": Decimal("450"), "lb_per_case": None,
        },
        {
            "customer": "Bulk Customer", "order_id": "SO-BULK-004",
            "due_date": date(2026, 8, 20), "sku": "70777",
            "product_name": "Granola Hypothetical Bulk",
            "qty": Decimal("200"), "lb_per_case": None,
        },
        {
            "customer": "Case Customer", "order_id": "SO-CASE-004",
            "due_date": date(2026, 8, 20), "sku": "70050",
            "product_name": "Granola Classic 25 LB", "qty": Decimal("2"),
            "lb_per_case": Decimal("25"),
        },
    ]

    @contextmanager
    def seeded_transaction():
        yield _SeededCursor(rows)

    monkeypatch.setattr(main, "get_transaction", seeded_transaction)
    monkeypatch.setattr(main, "API_KEY", "matrix-test-key")
    client = TestClient(main.app)
    try:
        response = client.get(
            "/export/orders-matrix.xlsx",
            headers={"X-API-Key": "matrix-test-key"},
        )
    finally:
        client.close()

    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    for sheet in ("Cases", "Pounds"):
        headers = [cell.value for cell in workbook[sheet][1]]
        # Already ends in "per/lb" — left alone, and definitely not doubled up.
        assert "Granola SS Classic #9 Bulk per/lb" in headers, headers
        assert "Granola SS Classic #9 Bulk per/lb (lb)" not in headers
        # Per-lb SKU whose name does not say so — suffixed.
        assert "Granola Hypothetical Bulk (lb)" in headers, headers
        # Case-priced SKU — never suffixed.
        assert "Granola Classic 25#" in headers
        assert "Granola Classic 25# (lb)" not in headers


def test_matrix_column_header_suffix_rule():
    """Unit coverage for the per-lb header suffix rule itself."""
    def header(name, per_lb):
        return main._matrix_column_header({"product_name": name, "per_lb": per_lb})

    assert header("Granola SS Classic #9 Bulk per/lb", True) == "Granola SS Classic #9 Bulk per/lb"
    assert header("Granola Bulk per lb", True) == "Granola Bulk per lb"
    assert header("Granola Bulk perlb", True) == "Granola Bulk perlb"
    assert header("Granola Hypothetical Bulk", True) == "Granola Hypothetical Bulk (lb)"
    assert header("Granola Hypothetical Bulk", False) == "Granola Hypothetical Bulk"
